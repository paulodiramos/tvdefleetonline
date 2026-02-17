"""
Processador de Ficheiros Prio
Processa ficheiros XLS (combustível) e CSV (elétrico) extraídos da Prio
"""
import os
import csv
import logging
from datetime import datetime, timezone
from typing import Dict, List, Optional
from io import StringIO
import uuid

logger = logging.getLogger(__name__)


class PrioProcessor:
    """Processador de ficheiros Prio"""
    
    def __init__(self, db):
        self.db = db
    
    @staticmethod
    def _limpar_valor(valor: str) -> float:
        """Limpar e converter valor monetário para float"""
        if not valor or valor == "" or valor == "-":
            return 0.0
        # Remover símbolos de moeda, espaços e converter vírgula para ponto
        valor_limpo = str(valor).replace("€", "").replace(" ", "").replace(",", ".").strip()
        try:
            return float(valor_limpo) if valor_limpo else 0.0
        except:
            return 0.0
    
    @staticmethod
    def _limpar_litros(valor: str) -> float:
        """Limpar e converter litros para float"""
        if not valor or valor == "":
            return 0.0
        valor_limpo = str(valor).replace("L", "").replace("l", "").replace(",", ".").replace(" ", "").strip()
        try:
            return float(valor_limpo) if valor_limpo else 0.0
        except:
            return 0.0
    
    @staticmethod
    def _limpar_kwh(valor: str) -> float:
        """Limpar e converter kWh para float"""
        if not valor or valor == "":
            return 0.0
        valor_limpo = str(valor).replace("kWh", "").replace("KWH", "").replace(",", ".").replace(" ", "").strip()
        try:
            return float(valor_limpo) if valor_limpo else 0.0
        except:
            return 0.0
    
    @staticmethod
    def _parse_data(data_str: str) -> Optional[datetime]:
        """Parse de data em vários formatos"""
        if not data_str:
            return None
        
        formatos = [
            '%d/%m/%Y %H:%M:%S',
            '%d/%m/%Y %H:%M',
            '%d/%m/%Y',
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d',
            '%d-%m-%Y %H:%M:%S',
            '%d-%m-%Y'
        ]
        
        for fmt in formatos:
            try:
                return datetime.strptime(data_str.strip(), fmt)
            except:
                continue
        
        return None
    
    async def processar_combustivel_xls(
        self, 
        filepath: str, 
        parceiro_id: str,
        semana: int = None,
        ano: int = None
    ) -> Dict:
        """
        Processar ficheiro XLS de combustível da Prio
        
        Campos típicos do ficheiro:
        - Posto, Rede, Data, Cartão, Estado, Litros, Combustível, Recibo, KMS, V. Unit., Total
        """
        try:
            if not os.path.exists(filepath):
                return {"sucesso": False, "erro": f"Ficheiro não encontrado: {filepath}"}
            
            logger.info(f"📄 Processando ficheiro combustível: {filepath}")
            
            # Detectar tipo de ficheiro e usar biblioteca apropriada
            is_xlsx = filepath.endswith('.xlsx') or filepath.endswith('.XLSX')
            
            headers = []
            rows_data = []
            header_row_idx = 0
            
            if is_xlsx:
                # Usar openpyxl para ficheiros .xlsx (Excel 2007+)
                from openpyxl import load_workbook
                workbook = load_workbook(filepath, data_only=True)
                sheet = workbook.active
                
                # Encontrar a linha de cabeçalhos (procurar linha com 'DATA' ou 'POSTO')
                for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                    row_values = [str(cell).strip().upper() if cell else '' for cell in row]
                    if any('DATA' in val or 'POSTO' in val or 'CARTÃO' in val or 'CARTAO' in val for val in row_values):
                        header_row_idx = row_idx
                        headers = row_values
                        break
                
                if not header_row_idx:
                    # Fallback: usar primeira linha não vazia
                    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                        row_values = [str(cell).strip().upper() if cell else '' for cell in row]
                        if any(val for val in row_values):
                            header_row_idx = row_idx
                            headers = row_values
                            break
                
                logger.info(f"📊 Cabeçalhos encontrados na linha {header_row_idx} (xlsx): {headers}")
                
                # Obter dados das linhas seguintes após os cabeçalhos
                for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
                    # Ignorar linhas completamente vazias
                    if any(cell for cell in row):
                        rows_data.append(row)
            else:
                # Usar xlrd para ficheiros .xls (Excel 97-2003)
                import xlrd
                workbook = xlrd.open_workbook(filepath)
                sheet = workbook.sheet_by_index(0)
                
                # Encontrar a linha de cabeçalhos
                for row_idx in range(min(10, sheet.nrows)):
                    row_values = [str(sheet.cell_value(row_idx, col)).strip().upper() for col in range(sheet.ncols)]
                    if any('DATA' in val or 'POSTO' in val or 'CARTÃO' in val or 'CARTAO' in val for val in row_values):
                        header_row_idx = row_idx
                        headers = row_values
                        break
                
                if not header_row_idx and sheet.nrows > 0:
                    # Fallback: usar primeira linha não vazia
                    for row_idx in range(min(10, sheet.nrows)):
                        row_values = [str(sheet.cell_value(row_idx, col)).strip().upper() for col in range(sheet.ncols)]
                        if any(val for val in row_values):
                            header_row_idx = row_idx
                            headers = row_values
                            break
                
                logger.info(f"📊 Cabeçalhos encontrados na linha {header_row_idx} (xls): {headers}")
                
                # Obter dados das linhas seguintes após os cabeçalhos
                for row_idx in range(header_row_idx + 1, sheet.nrows):
                    row = [sheet.cell_value(row_idx, col) for col in range(sheet.ncols)]
                    # Ignorar linhas completamente vazias
                    if any(cell for cell in row):
                        rows_data.append(row)
            
            # Mapear colunas
            col_map = {}
            for idx, header in enumerate(headers):
                if 'DATA' in header and 'FATURA' not in header:
                    col_map['data'] = idx
                elif 'CARTÃO' in header or 'CARTAO' in header:
                    if 'DESC' in header:
                        # "DESC. CARTÃO" contém a matrícula do veículo
                        col_map['matricula'] = idx
                    elif 'GRUPO' not in header:
                        col_map['cartao'] = idx
                elif 'MATRÍCULA' in header or 'MATRICULA' in header:
                    col_map['matricula'] = idx
                elif 'LITROS' in header:
                    col_map['litros'] = idx
                elif 'COMBUSTÍVEL' in header or 'COMBUSTIVEL' in header:
                    col_map['tipo_combustivel'] = idx
                elif header == 'TOTAL':
                    col_map['total'] = idx
                elif 'V. UNIT' in header or 'VAL. UNIT' in header or 'PRECO' in header or 'PREÇO' in header:
                    col_map['preco_litro'] = idx
                elif 'POSTO' in header:
                    col_map['posto'] = idx
                elif 'KMS' in header or "KM'S" in header or 'KM' == header:
                    col_map['kms'] = idx
            
            logger.info(f"📊 Mapeamento de colunas: {col_map}")
            
            registos_processados = 0
            registos_inseridos = 0
            registos_atualizados = 0
            registos_fora_periodo = 0
            
            # Calcular período da semana se fornecido
            data_inicio_semana = None
            data_fim_semana = None
            if semana and ano:
                from datetime import date
                data_inicio_semana = date.fromisocalendar(ano, semana, 1)  # Segunda-feira
                data_fim_semana = date.fromisocalendar(ano, semana, 7)     # Domingo
                logger.info(f"📅 Filtro de período: {data_inicio_semana} a {data_fim_semana}")
            
            # Processar linhas de dados
            for row_idx, row in enumerate(rows_data):
                try:
                    # Função helper para obter valor da célula de forma segura
                    def get_cell_value(col_key):
                        if col_key in col_map:
                            idx = col_map[col_key]
                            if idx < len(row):
                                val = row[idx]
                                return str(val) if val is not None else ""
                        return ""
                    
                    # Extrair dados
                    data_str = get_cell_value('data')
                    data = self._parse_data(data_str)
                    
                    # Filtrar por período da semana se especificado
                    if data and data_inicio_semana and data_fim_semana:
                        data_registo = data.date() if hasattr(data, 'date') else data
                        if not (data_inicio_semana <= data_registo <= data_fim_semana):
                            registos_fora_periodo += 1
                            logger.debug(f"  Registo ignorado (fora do período): {data_registo}")
                            continue
                    
                    matricula = get_cell_value('matricula')
                    cartao = get_cell_value('cartao')
                    
                    litros = self._limpar_litros(get_cell_value('litros'))
                    total = self._limpar_valor(get_cell_value('total'))
                    preco_litro = self._limpar_valor(get_cell_value('preco_litro'))
                    
                    tipo_combustivel = get_cell_value('tipo_combustivel')
                    posto = get_cell_value('posto')
                    kms = self._limpar_valor(get_cell_value('kms'))
                    
                    if not matricula and not cartao:
                        continue  # Linha vazia ou inválida
                    
                    # Calcular semana/ano do registo baseado na sua data real
                    semana_registo = semana
                    ano_registo = ano
                    if data:
                        data_registo = data.date() if hasattr(data, 'date') else data
                        iso_cal = data_registo.isocalendar()
                        semana_registo = iso_cal[1]
                        ano_registo = iso_cal[0]
                    
                    # Normalizar matrícula
                    matricula_norm = matricula.upper().replace(" ", "").replace("-", "")
                    
                    # Encontrar veículo e motorista pela matrícula
                    vehicle_id = None
                    motorista_id = None
                    
                    # Procurar veículo que corresponde à matrícula
                    veiculo = await self.db.vehicles.find_one({
                        "parceiro_id": parceiro_id,
                        "$or": [
                            {"matricula": matricula},
                            {"matricula": matricula_norm},
                            {"matricula": {"$regex": matricula_norm, "$options": "i"}}
                        ]
                    })
                    
                    if veiculo:
                        vehicle_id = str(veiculo.get("_id"))
                        motorista_id = veiculo.get("motorista_id")
                        logger.debug(f"  Veículo encontrado: {vehicle_id}, Motorista: {motorista_id}")
                    else:
                        # Tentar procurar sem parceiro_id (matrícula pode ser global)
                        veiculo = await self.db.vehicles.find_one({
                            "$or": [
                                {"matricula": matricula},
                                {"matricula": matricula_norm},
                                {"matricula": {"$regex": matricula_norm, "$options": "i"}}
                            ]
                        })
                        if veiculo:
                            vehicle_id = str(veiculo.get("_id"))
                            motorista_id = veiculo.get("motorista_id")
                            logger.debug(f"  Veículo encontrado (global): {vehicle_id}, Motorista: {motorista_id}")
                    
                    # Criar registo
                    registo = {
                        "parceiro_id": parceiro_id,
                        "plataforma": "prio",
                        "tipo": "combustivel",
                        "data": data or datetime.now(timezone.utc),
                        "matricula": matricula,
                        "matricula_normalizada": matricula_norm,
                        "cartao": cartao,
                        "litros": litros,
                        "valor": total,
                        "preco_litro": preco_litro,
                        "tipo_combustivel": tipo_combustivel,
                        "posto": posto,
                        "kms": kms,
                        "semana": semana_registo,
                        "ano": ano_registo,
                        "fonte": "rpa_prio",
                        "importado_em": datetime.now(timezone.utc),
                        "ficheiro_origem": os.path.basename(filepath),
                        "vehicle_id": vehicle_id,
                        "motorista_id": motorista_id
                    }
                    
                    # Verificar duplicado (mesmo parceiro, matricula, data, litros)
                    query_duplicado = {
                        "parceiro_id": parceiro_id,
                        "matricula_normalizada": matricula_norm,
                        "litros": litros,
                        "valor": total
                    }
                    
                    if data:
                        query_duplicado["data"] = data
                    
                    existente = await self.db.abastecimentos_combustivel.find_one(query_duplicado)
                    
                    if existente:
                        # Actualizar
                        registo["id"] = existente.get("id", str(uuid.uuid4()))
                        await self.db.abastecimentos_combustivel.update_one(
                            {"_id": existente["_id"]},
                            {"$set": registo}
                        )
                        registos_atualizados += 1
                    else:
                        # Inserir novo
                        registo["id"] = str(uuid.uuid4())
                        await self.db.abastecimentos_combustivel.insert_one(registo)
                        registos_inseridos += 1
                    
                    registos_processados += 1
                    
                except Exception as e:
                    logger.warning(f"⚠️ Erro na linha {row_idx}: {e}")
                    continue
            
            logger.info(f"✅ Combustível processado: {registos_processados} registos ({registos_inseridos} novos, {registos_atualizados} atualizados, {registos_fora_periodo} fora do período)")
            
            return {
                "sucesso": True,
                "registos_processados": registos_processados,
                "registos_inseridos": registos_inseridos,
                "registos_atualizados": registos_atualizados,
                "registos_fora_periodo": registos_fora_periodo,
                "tipo": "combustivel"
            }
            
        except ImportError:
            logger.error("❌ Biblioteca xlrd não instalada")
            return {"sucesso": False, "erro": "Biblioteca xlrd não instalada. Execute: pip install xlrd"}
        except Exception as e:
            logger.error(f"❌ Erro ao processar combustível: {e}")
            return {"sucesso": False, "erro": str(e)}
    
    async def processar_eletrico_csv(
        self, 
        filepath: str, 
        parceiro_id: str,
        semana: int = None,
        ano: int = None
    ) -> Dict:
        """
        Processar ficheiro CSV de carregamentos elétricos da Prio
        
        Campos típicos do ficheiro (baseado no vídeo):
        - DATA, CARTÃO, NOME, MATRÍCULA, ID. CARREGAMENTO, ENERGIA, DURAÇÃO, CUSTO OPC, TOTAL, TOTAL C/ IVA
        """
        try:
            if not os.path.exists(filepath):
                return {"sucesso": False, "erro": f"Ficheiro não encontrado: {filepath}"}
            
            logger.info(f"📄 Processando ficheiro elétrico: {filepath}")
            
            # Detectar tipo de ficheiro
            is_xlsx = filepath.endswith('.xlsx') or filepath.endswith('.XLSX')
            is_xls = filepath.endswith('.xls') or filepath.endswith('.XLS')
            
            rows_data = []
            fieldnames = []
            
            if is_xlsx:
                # Usar openpyxl para ficheiros .xlsx
                from openpyxl import load_workbook
                workbook = load_workbook(filepath, data_only=True)
                sheet = workbook.active
                
                first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
                fieldnames = [str(cell).strip().upper() if cell else '' for cell in first_row]
                logger.info(f"📊 Cabeçalhos encontrados (xlsx): {fieldnames}")
                
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_dict = {fieldnames[i]: str(row[i]) if row[i] is not None else '' for i in range(min(len(fieldnames), len(row)))}
                    rows_data.append(row_dict)
                    
            elif is_xls:
                # Usar xlrd para ficheiros .xls
                import xlrd
                workbook = xlrd.open_workbook(filepath)
                sheet = workbook.sheet_by_index(0)
                
                fieldnames = [str(sheet.cell_value(0, col)).strip().upper() for col in range(sheet.ncols)]
                logger.info(f"📊 Cabeçalhos encontrados (xls): {fieldnames}")
                
                for row_idx in range(1, sheet.nrows):
                    row_dict = {fieldnames[col]: str(sheet.cell_value(row_idx, col)) for col in range(sheet.ncols)}
                    rows_data.append(row_dict)
            else:
                # Ler ficheiro CSV
                with open(filepath, 'r', encoding='utf-8-sig') as f:
                    content = f.read()
                
                # Detectar delimitador
                delimiter = ';' if ';' in content[:500] else ','
                
                reader = csv.DictReader(StringIO(content), delimiter=delimiter)
                fieldnames = reader.fieldnames if reader.fieldnames else []
                logger.info(f"📊 Cabeçalhos encontrados (csv): {fieldnames}")
                
                for row in reader:
                    rows_data.append({k.upper().strip(): v for k, v in row.items()})
            
            registos_processados = 0
            registos_inseridos = 0
            registos_atualizados = 0
            
            for row_norm in rows_data:
                try:
                    # Normalizar keys (já está uppercase para xlsx/xls, normalizar csv)
                    if not is_xlsx and not is_xls:
                        row_norm = {k.upper().strip(): v for k, v in row_norm.items()}
                    
                    # Extrair dados
                    data_str = row_norm.get('DATA', row_norm.get('DATE', ''))
                    data = self._parse_data(data_str)
                    
                    matricula = row_norm.get('MATRÍCULA', row_norm.get('MATRICULA', row_norm.get('LICENSE PLATE', '')))
                    cartao = row_norm.get('CARTÃO', row_norm.get('CARTAO', row_norm.get('CARD', '')))
                    nome = row_norm.get('NOME', row_norm.get('NAME', ''))
                    
                    energia_kwh = self._limpar_kwh(row_norm.get('ENERGIA', row_norm.get('ENERGY', '0')))
                    duracao = row_norm.get('DURAÇÃO', row_norm.get('DURACAO', row_norm.get('DURATION', '')))
                    
                    total = self._limpar_valor(row_norm.get('TOTAL', '0'))
                    total_iva = self._limpar_valor(row_norm.get('TOTAL C/ IVA', row_norm.get('TOTAL COM IVA', '0')))
                    custo_opc = self._limpar_valor(row_norm.get('CUSTO OPC', '0'))
                    
                    id_carregamento = row_norm.get('ID. CARREGAMENTO', row_norm.get('CHARGING ID', ''))
                    
                    if not matricula and not cartao:
                        continue  # Linha vazia ou inválida
                    
                    # Normalizar matrícula
                    matricula_norm = matricula.upper().replace(" ", "").replace("-", "")
                    
                    # Usar valor com IVA se disponível, senão usar total
                    valor_final = total_iva if total_iva > 0 else total
                    
                    # Encontrar veículo e motorista pela matrícula
                    vehicle_id = None
                    motorista_id = None
                    
                    if matricula_norm:
                        # Procurar veículo que corresponde à matrícula
                        veiculo = await self.db.vehicles.find_one({
                            "parceiro_id": parceiro_id,
                            "$or": [
                                {"matricula": matricula},
                                {"matricula": matricula_norm},
                                {"matricula": {"$regex": matricula_norm, "$options": "i"}}
                            ]
                        })
                        
                        if veiculo:
                            vehicle_id = str(veiculo.get("_id"))
                            motorista_id = veiculo.get("motorista_id")
                            logger.debug(f"  Elétrico - Veículo encontrado: {vehicle_id}, Motorista: {motorista_id}")
                        else:
                            # Tentar procurar sem parceiro_id
                            veiculo = await self.db.vehicles.find_one({
                                "$or": [
                                    {"matricula": matricula},
                                    {"matricula": matricula_norm},
                                    {"matricula": {"$regex": matricula_norm, "$options": "i"}}
                                ]
                            })
                            if veiculo:
                                vehicle_id = str(veiculo.get("_id"))
                                motorista_id = veiculo.get("motorista_id")
                    
                    # Criar registo
                    registo = {
                        "parceiro_id": parceiro_id,
                        "plataforma": "prio_electric",
                        "tipo": "eletrico",
                        "data": data or datetime.now(timezone.utc),
                        "matricula": matricula,
                        "matricula_normalizada": matricula_norm,
                        "cartao": cartao,
                        "nome": nome,
                        "energia_kwh": energia_kwh,
                        "duracao": duracao,
                        "valor": valor_final,
                        "kwh": energia_kwh,  # Adicionar campo kwh para filtro
                        "custo_opc": custo_opc,
                        "id_carregamento": id_carregamento,
                        "semana": semana,
                        "ano": ano,
                        "fonte": "rpa_prio",
                        "importado_em": datetime.now(timezone.utc),
                        "ficheiro_origem": os.path.basename(filepath),
                        "vehicle_id": vehicle_id,
                        "motorista_id": motorista_id
                    }
                    
                    # Verificar duplicado
                    query_duplicado = {
                        "parceiro_id": parceiro_id,
                        "matricula_normalizada": matricula_norm
                    }
                    
                    if id_carregamento:
                        query_duplicado["id_carregamento"] = id_carregamento
                    elif data:
                        query_duplicado["data"] = data
                        query_duplicado["energia_kwh"] = energia_kwh
                    
                    existente = await self.db.despesas_combustivel.find_one(query_duplicado)
                    
                    if existente:
                        # Actualizar
                        registo["id"] = existente.get("id", str(uuid.uuid4()))
                        await self.db.despesas_combustivel.update_one(
                            {"_id": existente["_id"]},
                            {"$set": registo}
                        )
                        registos_atualizados += 1
                    else:
                        # Inserir novo
                        registo["id"] = str(uuid.uuid4())
                        await self.db.despesas_combustivel.insert_one(registo)
                        registos_inseridos += 1
                    
                    registos_processados += 1
                    
                except Exception as e:
                    logger.warning(f"⚠️ Erro na linha: {e}")
                    continue
            
            logger.info(f"✅ Elétrico processado: {registos_processados} registos ({registos_inseridos} novos, {registos_atualizados} atualizados)")
            
            return {
                "sucesso": True,
                "registos_processados": registos_processados,
                "registos_inseridos": registos_inseridos,
                "registos_atualizados": registos_atualizados,
                "tipo": "eletrico"
            }
            
        except Exception as e:
            logger.error(f"❌ Erro ao processar elétrico: {e}")
            return {"sucesso": False, "erro": str(e)}
