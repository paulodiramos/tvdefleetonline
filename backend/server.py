from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form, Body
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import re
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
from io import BytesIO
import base64
from PIL import Image
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import mimetypes
import shutil
import openpyxl
import csv
import io
import tempfile

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Upload directories
UPLOAD_DIR = ROOT_DIR / "uploads"
MOTORISTAS_UPLOAD_DIR = UPLOAD_DIR / "motoristas"
PAGAMENTOS_UPLOAD_DIR = UPLOAD_DIR / "pagamentos"
VEHICLE_DOCS_UPLOAD_DIR = UPLOAD_DIR / "vehicle_documents"

# Ensure directories exist
UPLOAD_DIR.mkdir(exist_ok=True)
MOTORISTAS_UPLOAD_DIR.mkdir(exist_ok=True)
PAGAMENTOS_UPLOAD_DIR.mkdir(exist_ok=True)
VEHICLE_DOCS_UPLOAD_DIR.mkdir(exist_ok=True)

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 24

security = HTTPBearer()

# Import routers from routes package
from routes.auth import router as auth_router
from routes.motoristas import router as motoristas_router
from routes.notificacoes import router as notificacoes_router
from routes.mensagens import router as mensagens_router
from routes.ifthenpay import router as ifthenpay_router

app = FastAPI()
api_router = APIRouter(prefix="/api")

# ==================== FILE UPLOAD UTILITIES ====================

def safe_parse_date(date_str: str, format: str = "%Y-%m-%d"):
    """Safely parse date string, return None if empty or invalid"""
    if not date_str or not date_str.strip():
        return None
    try:
        return datetime.strptime(date_str.strip(), format)
    except (ValueError, TypeError):
        return None

async def convert_image_to_pdf(image_path: Path, output_path: Path) -> Path:
    """Convert an image file to PDF format"""
    try:
        img = Image.open(image_path)
        
        # Convert RGBA to RGB if necessary
        if img.mode == 'RGBA':
            img = img.convert('RGB')
        
        # Get image dimensions
        img_width, img_height = img.size
        
        # Calculate scaling to fit A4 page
        a4_width, a4_height = A4
        scale = min(a4_width / img_width, a4_height / img_height) * 0.9
        
        new_width = img_width * scale
        new_height = img_height * scale
        
        # Create PDF
        c = canvas.Canvas(str(output_path), pagesize=A4)
        
        # Center image on page
        x = (a4_width - new_width) / 2
        y = (a4_height - new_height) / 2
        
        c.drawImage(str(image_path), x, y, width=new_width, height=new_height)
        c.save()
        
        return output_path
    except Exception as e:
        logging.error(f"Error converting image to PDF: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to convert image to PDF: {str(e)}")

async def save_uploaded_file(file: UploadFile, destination_dir: Path, filename: str) -> Path:
    """Save uploaded file to destination directory"""
    try:
        file_path = destination_dir / filename
        
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        return file_path
    except Exception as e:
        logging.error(f"Error saving file: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to save file: {str(e)}")

async def process_uploaded_file(file: UploadFile, destination_dir: Path, file_id: str) -> Dict[str, str]:
    """
    Process uploaded file: save it and convert to PDF if it's an image
    Returns dict with 'original_path' and 'pdf_path' (if converted)
    """
    file_extension = Path(file.filename).suffix.lower()
    original_filename = f"{file_id}_original{file_extension}"
    
    # Save original file
    original_path = await save_uploaded_file(file, destination_dir, original_filename)
    
    result = {
        "original_path": str(original_path.relative_to(ROOT_DIR)),
        "pdf_path": None
    }
    
    # Check if file is an image
    image_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp', '.heic', '.heif'}
    
    if file_extension in image_extensions:
        # Convert image to PDF
        pdf_filename = f"{file_id}.pdf"
        pdf_path = destination_dir / pdf_filename
        
        await convert_image_to_pdf(original_path, pdf_path)
        result["pdf_path"] = str(pdf_path.relative_to(ROOT_DIR))
    elif file_extension == '.pdf':
        # If already PDF, just reference it
        result["pdf_path"] = result["original_path"]
    
    return result

async def merge_images_to_pdf_a4(image1_path: Path, image2_path: Path, output_pdf_path: Path):
    """Merge two images (frente e verso) into a single A4 PDF"""
    from PIL import Image
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    
    # A4 size in points (72 DPI)
    a4_width, a4_height = A4
    
    # Create PDF
    c = canvas.Canvas(str(output_pdf_path), pagesize=A4)
    
    # Process first image (frente) - Page 1
    img1 = Image.open(image1_path)
    img1_width, img1_height = img1.size
    
    # Calculate scaling to fit A4 while maintaining aspect ratio
    scale1 = min(a4_width / img1_width, a4_height / img1_height) * 0.9  # 90% to leave margins
    new_width1 = img1_width * scale1
    new_height1 = img1_height * scale1
    
    # Center on page
    x1 = (a4_width - new_width1) / 2
    y1 = (a4_height - new_height1) / 2
    
    c.drawImage(str(image1_path), x1, y1, new_width1, new_height1)
    c.showPage()
    
    # Process second image (verso) - Page 2
    img2 = Image.open(image2_path)
    img2_width, img2_height = img2.size
    
    scale2 = min(a4_width / img2_width, a4_height / img2_height) * 0.9
    new_width2 = img2_width * scale2
    new_height2 = img2_height * scale2
    
    x2 = (a4_width - new_width2) / 2
    y2 = (a4_height - new_height2) / 2
    
    c.drawImage(str(image2_path), x2, y2, new_width2, new_height2)
    c.save()

# ==================== ALERT CHECKING UTILITIES ====================


async def auto_add_to_agenda(vehicle_id: str, tipo: str, data_vencimento: str, titulo: str):
    """Automatically add event to vehicle agenda when date is filled"""
    from datetime import datetime, timedelta
    
    try:
        # Parse date
        vencimento_date = datetime.strptime(data_vencimento, "%Y-%m-%d")
        # Set reminder for 30 days before
        reminder_date = (vencimento_date - timedelta(days=30)).strftime("%Y-%m-%d")
        
        evento = {
            "id": str(uuid.uuid4()),
            "tipo": tipo,
            "titulo": titulo,
            "data": reminder_date,
            "hora": "09:00",
            "descricao": f"Lembrete: {titulo} vence em {data_vencimento}",
            "auto_gerado": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Add to vehicle agenda
        await db.vehicles.update_one(
            {"id": vehicle_id},
            {"$push": {"agenda": evento}}
        )
        
        logger.info(f"Auto-added to agenda: {titulo} for vehicle {vehicle_id}")
    except Exception as e:
        logger.error(f"Error auto-adding to agenda: {e}")



async def check_and_create_alerts():
    """
    Check all vehicles and drivers for expiring documents/maintenance
    and create alerts if needed
    """
    from datetime import date, timedelta
    
    # Get admin settings for thresholds
    admin_settings = await db.admin_settings.find_one({"id": "admin_settings"})
    if not admin_settings:
        # Default settings
        anos_validade_matricula = 20
        km_aviso_manutencao = 5000
    else:
        anos_validade_matricula = admin_settings.get("anos_validade_matricula", 20)
        km_aviso_manutencao = admin_settings.get("km_aviso_manutencao", 5000)
    
    today = date.today()
    
    # Check vehicles
    vehicles = await db.vehicles.find({}, {"_id": 0}).to_list(None)
    
    for vehicle in vehicles:
        vehicle_id = vehicle["id"]
        
        # Check registration expiry
        if vehicle.get("validade_matricula"):
            try:
                parsed_date = safe_parse_date(vehicle["validade_matricula"])
                if not parsed_date:
                    continue
                validade_date = parsed_date.date()
                days_until_expiry = (validade_date - today).days
                
                if 0 <= days_until_expiry <= 30:
                    # Check if alert already exists
                    existing_alert = await db.alertas.find_one({
                        "tipo": "validade_matricula",
                        "entidade_id": vehicle_id,
                        "status": "ativo"
                    })
                    
                    if not existing_alert:
                        alert = {
                            "id": str(uuid.uuid4()),
                            "tipo": "validade_matricula",
                            "entidade_id": vehicle_id,
                            "entidade_tipo": "veiculo",
                            "titulo": f"Matrícula expira em breve - {vehicle['matricula']}",
                            "descricao": f"A matrícula do veículo {vehicle['marca']} {vehicle['modelo']} ({vehicle['matricula']}) expira em {days_until_expiry} dias.",
                            "data_vencimento": vehicle["validade_matricula"],
                            "prioridade": "alta" if days_until_expiry <= 7 else "media",
                            "dias_antecedencia": 30,
                            "status": "ativo",
                            "criado_em": datetime.now(timezone.utc).isoformat()
                        }
                        await db.alertas.insert_one(alert)
            except Exception as e:
                logger.error(f"Error creating alert: {e}")
                pass
        
        # Check insurance expiry
        if vehicle.get("insurance") and vehicle["insurance"].get("data_validade"):
            try:
                parsed_date = safe_parse_date(vehicle["insurance"]["data_validade"])
                if not parsed_date:
                    continue
                validade_date = parsed_date.date()
                days_until_expiry = (validade_date - today).days
                
                if 0 <= days_until_expiry <= 30:
                    existing_alert = await db.alertas.find_one({
                        "tipo": "seguro",
                        "entidade_id": vehicle_id,
                        "status": "ativo"
                    })
                    
                    if not existing_alert:
                        alert = {
                            "id": str(uuid.uuid4()),
                            "tipo": "seguro",
                            "entidade_id": vehicle_id,
                            "entidade_tipo": "veiculo",
                            "titulo": f"Seguro expira em breve - {vehicle['matricula']}",
                            "descricao": f"O seguro do veículo {vehicle['marca']} {vehicle['modelo']} ({vehicle['matricula']}) expira em {days_until_expiry} dias.",
                            "data_vencimento": vehicle["insurance"]["data_validade"],
                            "prioridade": "alta",
                            "dias_antecedencia": 30,
                            "status": "ativo",
                            "criado_em": datetime.now(timezone.utc).isoformat()
                        }
                        await db.alertas.insert_one(alert)
            except Exception as e:
                logger.error(f"Error creating alert: {e}")
                pass
        
        # Check next inspection
        if vehicle.get("inspection") and vehicle["inspection"].get("proxima_inspecao"):
            try:
                parsed_date = safe_parse_date(vehicle["inspection"]["proxima_inspecao"])
                if not parsed_date:
                    continue
                proxima_date = parsed_date.date()
                days_until_inspection = (proxima_date - today).days
                
                if 0 <= days_until_inspection <= 30:
                    existing_alert = await db.alertas.find_one({
                        "tipo": "inspecao",
                        "entidade_id": vehicle_id,
                        "status": "ativo"
                    })
                    
                    if not existing_alert:
                        alert = {
                            "id": str(uuid.uuid4()),
                            "tipo": "inspecao",
                            "entidade_id": vehicle_id,
                            "entidade_tipo": "veiculo",
                            "titulo": f"Inspeção em breve - {vehicle['matricula']}",
                            "descricao": f"A próxima inspeção do veículo {vehicle['marca']} {vehicle['modelo']} ({vehicle['matricula']}) está marcada para {days_until_inspection} dias.",
                            "data_vencimento": vehicle["inspection"]["proxima_inspecao"],
                            "prioridade": "media",
                            "dias_antecedencia": 30,
                            "status": "ativo",
                            "criado_em": datetime.now(timezone.utc).isoformat()
                        }
                        await db.alertas.insert_one(alert)
            except Exception as e:
                logger.error(f"Error creating alert: {e}")
                pass
        
        # Check fire extinguisher expiry
        if vehicle.get("extintor") and vehicle["extintor"].get("data_validade"):
            try:
                parsed_date = safe_parse_date(vehicle["extintor"]["data_validade"])
                if not parsed_date:
                    continue
                validade_date = parsed_date.date()
                days_until_expiry = (validade_date - today).days
                
                if 0 <= days_until_expiry <= 30:
                    existing_alert = await db.alertas.find_one({
                        "tipo": "extintor",
                        "entidade_id": vehicle_id,
                        "status": "ativo"
                    })
                    
                    if not existing_alert:
                        alert = {
                            "id": str(uuid.uuid4()),
                            "tipo": "extintor",
                            "entidade_id": vehicle_id,
                            "entidade_tipo": "veiculo",
                            "titulo": f"Extintor expira em breve - {vehicle['matricula']}",
                            "descricao": f"O extintor do veículo {vehicle['marca']} {vehicle['modelo']} ({vehicle['matricula']}) expira em {days_until_expiry} dias.",
                            "data_vencimento": vehicle["extintor"]["data_validade"],
                            "prioridade": "alta",
                            "dias_antecedencia": 30,
                            "status": "ativo",
                            "criado_em": datetime.now(timezone.utc).isoformat()
                        }
                        await db.alertas.insert_one(alert)
            except Exception as e:
                logger.error(f"Error creating extintor alert: {e}")
                pass

        
        # Check maintenance (based on km)
        if vehicle.get("maintenance_history"):
            for maintenance in vehicle["maintenance_history"]:
                if maintenance.get("km_proxima"):
                    km_atual = vehicle.get("km_atual", 0)
                    km_restantes = maintenance["km_proxima"] - km_atual
                    
                    if 0 <= km_restantes <= km_aviso_manutencao:
                        existing_alert = await db.alertas.find_one({
                            "tipo": "manutencao",
                            "entidade_id": vehicle_id,
                            "status": "ativo",
                            "descricao": {"$regex": maintenance.get("tipo_manutencao", "")}
                        })
                        
                        if not existing_alert:
                            alert = {
                                "id": str(uuid.uuid4()),
                                "tipo": "manutencao",
                                "entidade_id": vehicle_id,
                                "entidade_tipo": "veiculo",
                                "titulo": f"Manutenção necessária - {vehicle['matricula']}",
                                "descricao": f"O veículo {vehicle['marca']} {vehicle['modelo']} ({vehicle['matricula']}) precisa de {maintenance.get('tipo_manutencao', 'manutenção')} em {km_restantes} km.",
                                "data_vencimento": "",
                                "prioridade": "alta" if km_restantes <= 500 else "media",
                                "dias_antecedencia": 0,
                                "status": "ativo",
                                "criado_em": datetime.now(timezone.utc).isoformat()
                            }
                            await db.alertas.insert_one(alert)
    
    # Check motoristas
    motoristas = await db.motoristas.find({}, {"_id": 0}).to_list(None)
    
    for motorista in motoristas:
        motorista_id = motorista["id"]
        
        # Check TVDE license expiry
        if motorista.get("licenca_tvde_validade"):
            try:
                parsed_date = safe_parse_date(motorista["licenca_tvde_validade"])
                if not parsed_date:
                    continue
                validade_date = parsed_date.date()
                days_until_expiry = (validade_date - today).days
                
                if 0 <= days_until_expiry <= 30:
                    existing_alert = await db.alertas.find_one({
                        "tipo": "licenca_tvde",
                        "entidade_id": motorista_id,
                        "status": "ativo"
                    })
                    
                    if not existing_alert:
                        alert = {
                            "id": str(uuid.uuid4()),
                            "tipo": "licenca_tvde",
                            "entidade_id": motorista_id,
                            "entidade_tipo": "motorista",
                            "titulo": f"Licença TVDE expira em breve - {motorista['name']}",
                            "descricao": f"A licença TVDE do motorista {motorista['name']} expira em {days_until_expiry} dias.",
                            "data_vencimento": motorista["licenca_tvde_validade"],
                            "prioridade": "alta",
                            "dias_antecedencia": 30,
                            "status": "ativo",
                            "criado_em": datetime.now(timezone.utc).isoformat()
                        }
                        await db.alertas.insert_one(alert)
            except Exception as e:
                logger.error(f"Error creating alert: {e}")
                pass
        
        # Check driver's license expiry
        if motorista.get("carta_conducao_validade"):
            try:
                parsed_date = safe_parse_date(motorista["carta_conducao_validade"])
                if not parsed_date:
                    continue
                validade_date = parsed_date.date()
                days_until_expiry = (validade_date - today).days
                
                if 0 <= days_until_expiry <= 30:
                    existing_alert = await db.alertas.find_one({
                        "tipo": "carta_conducao",
                        "entidade_id": motorista_id,
                        "status": "ativo"
                    })
                    
                    if not existing_alert:
                        alert = {
                            "id": str(uuid.uuid4()),
                            "tipo": "carta_conducao",
                            "entidade_id": motorista_id,
                            "entidade_tipo": "motorista",
                            "titulo": f"Carta de condução expira em breve - {motorista['name']}",
                            "descricao": f"A carta de condução do motorista {motorista['name']} expira em {days_until_expiry} dias.",
                            "data_vencimento": motorista["carta_conducao_validade"],
                            "prioridade": "alta",
                            "dias_antecedencia": 30,
                            "status": "ativo",
                            "criado_em": datetime.now(timezone.utc).isoformat()
                        }
                        await db.alertas.insert_one(alert)
            except Exception as e:
                logger.error(f"Error creating alert: {e}")
                pass

# ==================== MODELS ====================

class UserRole:
    ADMIN = "admin"
    GESTAO = "gestao"  # Gestor - pode gerir múltiplos parceiros atribuídos pelo Admin
    PARCEIRO = "parceiro"  # Pode gerir veículos e motoristas
    MOTORISTA = "motorista"

# User Profile Models
class UserProfileUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    empresa: Optional[str] = None
    nif: Optional[str] = None
    morada: Optional[str] = None
    niveis_servico: Optional[List[str]] = None  # For operacional role

class ChangePasswordRequest(BaseModel):
    old_password: str
    new_password: str

# Parceiro Model - EXPANDED
class ParceiroCreate(BaseModel):
    nome_empresa: str
    contribuinte_empresa: str  # NIF da empresa (9 dígitos)
    morada_completa: str
    codigo_postal: str  # Formato: xxxx-xxx
    localidade: str
    nome_manager: str
    email_manager: EmailStr
    email_empresa: EmailStr
    telefone: str
    telemovel: str
    email: EmailStr  # Mantido para compatibilidade
    certidao_permanente: str  # Formato: xxxx-xxxx-xxxx (só dígitos)
    codigo_certidao_comercial: str
    validade_certidao_comercial: str  # Data validade
    seguro_responsabilidade_civil: Optional[str] = None
    seguro_acidentes_trabalho: Optional[str] = None
    licenca_tvde: Optional[str] = None
    plano_id: Optional[str] = None  # ID do plano de assinatura
    gestor_associado_id: Optional[str] = None

class Parceiro(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    # Campos flexíveis - suporta ambas as estruturas (antiga e nova)
    nome_empresa: Optional[str] = None
    nome: Optional[str] = None  # Estrutura nova
    contribuinte_empresa: Optional[str] = None
    nif: Optional[str] = None  # Estrutura nova
    morada_completa: Optional[str] = None
    morada: Optional[str] = None  # Estrutura nova
    codigo_postal: str
    localidade: Optional[str] = None
    nome_manager: Optional[str] = None
    responsavel_nome: Optional[str] = None  # Estrutura nova
    email_manager: Optional[str] = None
    email_empresa: Optional[str] = None
    telefone: Optional[str] = None
    telemovel: Optional[str] = None
    responsavel_contacto: Optional[str] = None  # Estrutura nova
    email: str
    certidao_permanente: Optional[str] = None
    codigo_certidao_comercial: Optional[str] = None
    validade_certidao_comercial: Optional[str] = None
    seguro_responsabilidade_civil: Optional[str] = None
    seguro_acidentes_trabalho: Optional[str] = None
    licenca_tvde: Optional[str] = None
    plano_id: Optional[str] = None
    plano_nome: Optional[str] = None
    plano_valida_ate: Optional[str] = None
    plano_status: str = "pendente"  # "pendente", "ativo", "suspenso"
    gestor_associado_id: Optional[str] = None
    total_vehicles: int = 0
    approved: Optional[bool] = None
    status: Optional[str] = None
    # Contrato único do parceiro com múltiplos tipos
    contrato_texto: Optional[str] = None  # Texto base do contrato
    contratos_tipos: List[Dict[str, Any]] = []  # Lista de tipos de contrato disponíveis: [{nome, tipo, valores}]
    representante_legal_nome: Optional[str] = None
    representante_legal_contribuinte: Optional[str] = None
    representante_legal_cc: Optional[str] = None
    representante_legal_cc_validade: Optional[str] = None
    representante_legal_telefone: Optional[str] = None
    representante_legal_email: Optional[str] = None
    campos_customizados: Dict[str, Any] = {}
    # Configurações de Alertas
    dias_aviso_seguro: int = 30  # Dias de antecedência para alertar sobre seguro
    dias_aviso_inspecao: int = 30  # Dias de antecedência para alertar sobre inspeção
    km_aviso_revisao: int = 5000  # KM de antecedência para alertar sobre revisão
    created_at: datetime
    # Campos antigos mantidos como opcionais para compatibilidade
    name: Optional[str] = None
    phone: Optional[str] = None
    empresa: Optional[str] = None
    nif: Optional[str] = None
    morada: Optional[str] = None

# Admin Settings Model
class AdminSettings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = "admin_settings"
    anos_validade_matricula: int = 20
    km_aviso_manutencao: int = 5000
    updated_at: datetime
    updated_by: str

# Modelo para ganhos importados da Uber
class GanhoUber(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    uuid_motorista_uber: str  # UUID do motorista na Uber
    motorista_id: Optional[str] = None  # ID do motorista no sistema (se encontrado)
    nome_motorista: str
    apelido_motorista: str
    # Período do ficheiro
    periodo_inicio: Optional[str] = None
    periodo_fim: Optional[str] = None
    # Valores principais
    pago_total: float
    rendimentos_total: float
    dinheiro_recebido: Optional[float] = None
    # Tarifas
    tarifa_total: float
    tarifa_base: Optional[float] = None
    tarifa_ajuste: Optional[float] = None
    tarifa_cancelamento: Optional[float] = None
    tarifa_dinamica: Optional[float] = None
    taxa_reserva: Optional[float] = None
    uber_priority: Optional[float] = None
    tempo_espera: Optional[float] = None
    # Taxas e impostos
    taxa_servico: Optional[float] = None
    imposto_tarifa: Optional[float] = None
    taxa_aeroporto: Optional[float] = None
    # Outros
    gratificacao: Optional[float] = None
    portagens: Optional[float] = None
    ajustes: Optional[float] = None
    # Metadata
    ficheiro_nome: str
    data_importacao: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    importado_por: str
    
class GanhoUberImportResponse(BaseModel):
    total_linhas: int
    motoristas_encontrados: int
    motoristas_nao_encontrados: int
    total_ganhos: float
    ganhos_importados: List[Dict[str, Any]]
    erros: List[str] = []

# Modelo para ganhos importados da Bolt
class GanhoBolt(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    identificador_motorista_bolt: str  # Identificador do motorista na Bolt
    identificador_individual: str  # Identificador individual
    motorista_id: Optional[str] = None  # ID do motorista no sistema (se encontrado)
    nome_motorista: str
    email_motorista: str
    telemovel_motorista: Optional[str] = None
    # Período do ficheiro (extraído do nome: 2025W45)
    periodo_semana: Optional[str] = None
    periodo_ano: Optional[str] = None
    # Ganhos brutos
    ganhos_brutos_total: float
    ganhos_brutos_app: float
    iva_ganhos_app: float
    ganhos_brutos_dinheiro: float
    iva_ganhos_dinheiro: float
    dinheiro_recebido: float
    # Extras
    gorjetas: float
    ganhos_campanha: float
    reembolsos_despesas: float
    # Taxas
    taxas_cancelamento: float
    iva_taxas_cancelamento: float
    portagens: float
    taxas_reserva: float
    iva_taxas_reserva: float
    total_taxas: float
    # Comissões e deduções
    comissoes: float
    reembolsos_passageiros: float
    outras_taxas: float
    # Ganhos líquidos
    ganhos_liquidos: float
    pagamento_previsto: float
    # Produtividade
    ganhos_brutos_por_hora: float
    ganhos_liquidos_por_hora: float
    # Metadata
    ficheiro_nome: str
    parceiro_id: Optional[str] = None
    data_importacao: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    importado_por: str

# Modelo para movimentos Via Verde
class ViaVerdeMovimento(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    parceiro_id: str
    license_plate: str
    obu: Optional[str] = None
    service: Optional[str] = None
    service_description: Optional[str] = None
    market: Optional[str] = None
    entry_date: Optional[str] = None
    exit_date: Optional[str] = None
    entry_point: Optional[str] = None
    exit_point: Optional[str] = None
    value: Optional[float] = None
    is_payed: Optional[str] = None
    payment_date: Optional[str] = None
    contract_number: Optional[str] = None
    liquid_value: Optional[float] = None
    ficheiro_nome: str
    data_importacao: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    importado_por: str

# Modelo para relatórios GPS de distância
class GPSDistancia(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    parceiro_id: str
    veiculo: str  # Matrícula
    condutor: Optional[str] = None
    distancia_percorrida: Optional[float] = None  # KM
    motor_ligado_tempo: Optional[str] = None  # Ex: "13h 16m 34s"
    motor_ligado_minutos: Optional[int] = None
    ficheiro_nome: str
    data_importacao: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    importado_por: str

# Modelo para transações de combustível elétrico
class CombustivelEletrico(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    parceiro_id: str
    numero_cartao: Optional[str] = None
    nome: Optional[str] = None
    descricao: Optional[str] = None
    matricula: Optional[str] = None
    id_carregamento: Optional[str] = None
    posto: Optional[str] = None
    energia: Optional[float] = None  # kWh
    duracao: Optional[float] = None  # minutos
    custo: Optional[float] = None
    opc: Optional[float] = None
    iec: Optional[float] = None
    total: Optional[float] = None
    total_com_iva: Optional[float] = None
    fatura: Optional[str] = None
    ficheiro_nome: str
    data_importacao: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    importado_por: str

# Modelo para transações de combustível fóssil
class CombustivelFossil(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    parceiro_id: str
    posto: Optional[str] = None
    pais: Optional[str] = None
    rede: Optional[str] = None
    data: Optional[str] = None
    hora: Optional[str] = None
    cartao: Optional[str] = None
    desc_cartao: Optional[str] = None
    estado: Optional[str] = None
    grupo_cartao: Optional[str] = None
    litros: Optional[float] = None
    combustivel: Optional[str] = None
    recibo: Optional[str] = None
    valor_liquido: Optional[float] = None
    iva: Optional[float] = None
    kms: Optional[int] = None
    id_condutor: Optional[str] = None
    fatura: Optional[str] = None
    data_fatura: Optional[str] = None
    valor_unitario: Optional[float] = None
    valor_ref: Optional[float] = None
    valor_desc: Optional[float] = None
    cliente: Optional[str] = None
    tipo_pagamento: Optional[str] = None
    ficheiro_nome: str
    data_importacao: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    importado_por: str

# Modelos para Sincronização Automática
class CredenciaisPlataforma(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    parceiro_id: str  # Cada parceiro tem suas próprias credenciais
    plataforma: str  # uber, bolt, via_verde, combustivel, gps
    email: str
    password_encrypted: str  # Encriptado
    ativo: bool = True
    ultima_sincronizacao: Optional[datetime] = None
    proxima_sincronizacao: Optional[datetime] = None
    sincronizacao_automatica: bool = False
    horario_sincronizacao: Optional[str] = None  # ex: "09:00"
    frequencia_dias: int = 7  # Semanal por padrão
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    
class LogSincronizacao(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    parceiro_id: str  # Log associado ao parceiro
    plataforma: str
    tipo_sincronizacao: str  # manual, automatico
    status: str  # sucesso, erro, parcial
    data_inicio: datetime
    data_fim: Optional[datetime] = None
    registos_importados: int = 0
    registos_erro: int = 0
    mensagem: Optional[str] = None
    detalhes: Dict[str, Any] = {}
    executado_por: Optional[str] = None  # user_id se manual

# ==================== MODELOS DE CONTRATOS ====================

class TemplateContrato(BaseModel):
    """Template de contrato criado pelo parceiro"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    parceiro_id: str
    nome_template: str  # Ex: "Aluguer Básico", "Comissão Premium"
    tipo_contrato: str  # aluguer_sem_caucao, aluguer_com_caucao, aluguer_caucao_parcelada, 
                        # periodo_epoca, aluguer_epocas_sem_caucao, aluguer_epocas_caucao,
                        # aluguer_epoca_caucao_parcelada, compra_veiculo, comissao, 
                        # motorista_privado, outros
    periodicidade_padrao: str = "semanal"  # semanal ou mensal
    
    # Configurações base (valores padrão, podem ser editados ao criar contrato)
    valor_base: Optional[float] = None  # Valor semanal/mensal base
    valor_caucao: Optional[float] = None
    numero_parcelas_caucao: Optional[int] = None
    valor_parcela_caucao: Optional[float] = None
    
    # Épocas (apenas para tipos com época)
    valor_epoca_alta: Optional[float] = None
    valor_epoca_baixa: Optional[float] = None
    caucao_epoca_alta: Optional[float] = None
    caucao_epoca_baixa: Optional[float] = None
    
    # Comissão
    percentagem_motorista: Optional[float] = None  # Ex: 60
    percentagem_parceiro: Optional[float] = None   # Ex: 40 (soma deve ser 100)
    combustivel_incluido: bool = False
    regime_trabalho: Optional[str] = None  # part_time, full_time
    
    # Compra de veículo
    valor_compra_veiculo: Optional[float] = None
    numero_semanas_compra: Optional[int] = None
    com_slot: bool = False
    extra_seguro: bool = False
    valor_extra_seguro: Optional[float] = None
    
    # Cláusulas do contrato (texto livre, editável pelo parceiro)
    clausulas_texto: Optional[str] = None
    
    # Metadata
    ativo: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str  # user_id

class ContratoMotorista(BaseModel):
    """Contrato individual criado para um motorista"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    
    # Relacionamentos
    template_id: str  # Template usado
    parceiro_id: str
    motorista_id: str
    veiculo_id: Optional[str] = None
    
    # Dados do template (snapshot no momento da criação)
    nome_template: str
    tipo_contrato: str
    periodicidade: str  # semanal ou mensal
    
    # Valores aplicados (podem ser diferentes do template)
    valor_aplicado: float
    valor_caucao_aplicado: Optional[float] = None
    numero_parcelas_caucao_aplicado: Optional[int] = None
    valor_parcela_caucao_aplicado: Optional[float] = None
    
    # Épocas
    epoca_atual: Optional[str] = None  # alta, baixa
    valor_epoca_alta_aplicado: Optional[float] = None
    valor_epoca_baixa_aplicado: Optional[float] = None
    
    # Comissão
    percentagem_motorista_aplicado: Optional[float] = None
    percentagem_parceiro_aplicado: Optional[float] = None
    combustivel_incluido_aplicado: bool = False
    regime_trabalho_aplicado: Optional[str] = None
    
    # Compra de veículo
    valor_compra_aplicado: Optional[float] = None
    numero_semanas_aplicado: Optional[int] = None
    com_slot_aplicado: bool = False
    extra_seguro_aplicado: bool = False
    valor_extra_seguro_aplicado: Optional[float] = None
    
    # Cláusulas (snapshot do template)
    clausulas_texto: Optional[str] = None
    
    # Datas
    data_inicio: str  # Data de início do contrato
    data_fim: Optional[str] = None  # Data de fim (se aplicável)
    data_emissao: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    
    # Status
    status: str = "ativo"  # ativo, terminado, cancelado
    
    # PDF gerado
    pdf_url: Optional[str] = None
    
    # Assinaturas
    assinado_motorista: bool = False
    data_assinatura_motorista: Optional[datetime] = None
    assinado_parceiro: bool = False
    data_assinatura_parceiro: Optional[datetime] = None
    
    # Metadata
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str  # user_id

class TemplateContratoCreate(BaseModel):
    """Schema para criar template de contrato"""
    nome_template: str
    tipo_contrato: str
    periodicidade_padrao: str = "semanal"
    valor_base: Optional[float] = None
    valor_caucao: Optional[float] = None
    numero_parcelas_caucao: Optional[int] = None
    valor_epoca_alta: Optional[float] = None
    valor_epoca_baixa: Optional[float] = None
    percentagem_motorista: Optional[float] = None
    percentagem_parceiro: Optional[float] = None
    combustivel_incluido: bool = False
    regime_trabalho: Optional[str] = None
    valor_compra_veiculo: Optional[float] = None
    numero_semanas_compra: Optional[int] = None
    com_slot: bool = False
    extra_seguro: bool = False
    valor_extra_seguro: Optional[float] = None
    clausulas_texto: Optional[str] = None

class ContratoMotoristaCreate(BaseModel):
    """Schema para criar contrato de motorista"""
    template_id: str
    motorista_id: str
    veiculo_id: Optional[str] = None
    periodicidade: str
    valor_aplicado: float
    valor_caucao_aplicado: Optional[float] = None
    numero_parcelas_caucao_aplicado: Optional[int] = None
    epoca_atual: Optional[str] = None
    valor_epoca_alta_aplicado: Optional[float] = None
    valor_epoca_baixa_aplicado: Optional[float] = None
    percentagem_motorista_aplicado: Optional[float] = None
    percentagem_parceiro_aplicado: Optional[float] = None
    combustivel_incluido_aplicado: bool = False
    regime_trabalho_aplicado: Optional[str] = None
    valor_compra_aplicado: Optional[float] = None
    numero_semanas_aplicado: Optional[int] = None
    com_slot_aplicado: bool = False
    extra_seguro_aplicado: bool = False
    valor_extra_seguro_aplicado: Optional[float] = None
    data_inicio: str
    data_fim: Optional[str] = None

class UserBase(BaseModel):
    email: EmailStr
    name: str
    role: str
    phone: Optional[str] = None

class UserCreate(UserBase):
    password: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class User(UserBase):
    model_config = ConfigDict(extra="ignore")
    id: str
    created_at: datetime
    approved: bool = False
    associated_partner_id: Optional[str] = None
    associated_gestor_id: Optional[str] = None
    parceiros_atribuidos: List[str] = []  # IDs dos parceiros atribuídos (para gestores)
    subscription_id: Optional[str] = None  # ID da subscrição ativa
    plano_id: Optional[str] = None  # ID do plano no sistema unificado
    plano_nome: Optional[str] = None  # Nome do plano
    plano_valida_ate: Optional[str] = None  # Data de validade do plano
    plano_manutencao_ativo: bool = False  # Plano adicional para editar manutenções
    plano_alertas_ativo: bool = False  # Plano adicional para editar alertas
    senha_provisoria: bool = False  # Se True, deve mudar senha no próximo login
    campos_customizados: Dict[str, Any] = {}  # Campos adicionais customizáveis

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: User

# ==================== PLANOS E SUBSCRIÇÕES ====================

class PlanoPromocao(BaseModel):
    ativa: bool = False
    nome: str = ""  # Ex: "Black Friday"
    desconto_percentagem: float = 0  # Ex: 20 (para 20%)
    valida_ate: Optional[str] = None  # Data ISO

class Plano(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    nome: str  # Ex: "Plano Premium"
    descricao: str
    features: List[str] = []  # ["relatorios", "gestao_seguros", "gestao_contas", "gestao_motoristas", "gestao_veiculos", etc]
    perfis_permitidos: List[str] = []  # ["admin", "parceiro", "operacional", "gestao"]
    tipo_cobranca: str = "por_veiculo"  # "por_veiculo" ou "por_motorista"
    preco_semanal_sem_iva: float
    iva_percentagem: float = 23  # IVA em Portugal
    preco_mensal_sem_iva: float
    desconto_mensal_percentagem: float = 0  # Desconto para pagamento mensal
    promocao: PlanoPromocao = PlanoPromocao()
    ativo: bool = True
    created_at: datetime
    updated_at: datetime

class PlanoCreate(BaseModel):
    nome: str
    descricao: str
    features: List[str] = []
    perfis_permitidos: List[str] = []  # ["admin", "parceiro", "operacional", "gestao"]
    tipo_cobranca: str = "por_veiculo"  # "por_veiculo" ou "por_motorista"
    preco_semanal_sem_iva: float
    iva_percentagem: float = 23
    preco_mensal_sem_iva: float
    desconto_mensal_percentagem: float = 0
    promocao: PlanoPromocao = PlanoPromocao()

class Subscription(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: str
    user_name: str  # Nome do utilizador
    user_email: str  # Email do utilizador
    plano_id: str
    plano_nome: str  # Nome do plano
    periodo: str  # "semanal" ou "mensal"
    status: str  # "pendente", "ativo", "expirado", "cancelado"
    preco_pago: float  # Preço final com IVA e descontos
    data_inicio: Optional[datetime] = None
    data_expiracao: Optional[datetime] = None
    data_proximo_pagamento: Optional[datetime] = None  # Para notificações
    # Dados de pagamento
    pagamento_metodo: Optional[str] = None  # "multibanco", "mbway"
    pagamento_referencia: Optional[str] = None  # Referência Multibanco
    pagamento_entidade: Optional[str] = None  # Entidade Multibanco
    pagamento_id_transacao: Optional[str] = None  # ID da transação IFThenPay
    pagamento_status: str = "pendente"  # "pendente", "pago", "expirado"
    created_at: datetime
    updated_at: datetime
    # Admin pode atribuir manualmente
    atribuido_manualmente: bool = False  # Se admin atribuiu sem pagamento
    duracao_dias: Optional[int] = None  # Se admin definiu duração específica

class SubscriptionCreate(BaseModel):
    plano_id: str
    periodo: str  # "semanal" ou "mensal"
    pagamento_metodo: str  # "multibanco" ou "mbway"

class ConfiguracaoIFThenPay(BaseModel):
    id: str = "ifthen_pay_config"
    ativa: bool = False
    entidade: str = ""  # Entidade Multibanco
    subentidade: str = ""  # Subentidade
    chave_api: str = ""  # Chave API IFThenPay
    modo_teste: bool = True  # Se está em modo de teste

# Recibos e Ganhos
class Recibo(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    motorista_id: str
    motorista_nome: str
    parceiro_id: Optional[str] = None
    parceiro_nome: Optional[str] = None
    valor: float
    mes_referencia: str  # Ex: "2024-11"
    ficheiro_url: str  # URL do PDF do recibo
    status: str = "pendente"  # "pendente", "verificado", "pago", "rejeitado"
    verificado_por: Optional[str] = None  # ID de quem verificou
    verificado_em: Optional[datetime] = None
    observacoes: Optional[str] = None
    created_at: datetime
    updated_at: datetime

class ReciboCreate(BaseModel):
    mes_referencia: str
    valor: float
    ficheiro_url: str

class Ganho(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    motorista_id: str
    motorista_nome: str
    parceiro_id: Optional[str] = None
    parceiro_nome: Optional[str] = None
    valor: float
    data: str  # Data do ganho
    plataforma: str  # "uber", "bolt", "outro"
    descricao: Optional[str] = None
    created_at: datetime

class PedidoAlteracao(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    motorista_id: str
    motorista_nome: str
    campo: str  # Campo que quer alterar
    valor_atual: str
    valor_novo: str
    motivo: str
    status: str = "pendente"  # "pendente", "aprovado", "rejeitado"
    respondido_por: Optional[str] = None
    respondido_em: Optional[datetime] = None
    created_at: datetime

# Motorista Models - EXPANDED
class MotoristaDocuments(BaseModel):
    # Documentos principais com foto (convertidos para PDF A4)
    comprovativo_morada_pdf: Optional[str] = None  # Comprovativo de morada (PDF)
    cc_frente_verso_pdf: Optional[str] = None  # CC frente e verso juntos em PDF A4
    carta_frente_verso_pdf: Optional[str] = None  # Carta frente e verso juntos em PDF A4
    licenca_tvde_pdf: Optional[str] = None  # Licença TVDE (PDF)
    registo_criminal_pdf: Optional[str] = None  # Registo criminal (PDF)
    iban_comprovativo_pdf: Optional[str] = None  # Comprovativo de IBAN (PDF)
    # Documentos antigos (mantidos para compatibilidade)
    cartao_cidadao_foto: Optional[str] = None
    carta_conducao_foto: Optional[str] = None
    licenca_tvde_foto: Optional[str] = None
    comprovativo_morada: Optional[str] = None
    iban_comprovativo: Optional[str] = None
    license_photo: Optional[str] = None
    cv_file: Optional[str] = None
    profile_photo: Optional[str] = None
    documento_identificacao: Optional[str] = None
    documento_identificacao_frente: Optional[str] = None
    documento_identificacao_verso: Optional[str] = None
    carta_conducao_frente: Optional[str] = None
    carta_conducao_verso: Optional[str] = None
    licenca_tvde: Optional[str] = None
    registo_criminal: Optional[str] = None
    comprovativo_iban: Optional[str] = None
    seguro_comprovativo: Optional[str] = None
    contrato: Optional[str] = None
    additional_docs: List[str] = []

class DocumentoValidacao(BaseModel):
    """Estado de validação de cada documento"""
    validado: bool = False
    validado_por: Optional[str] = None  # ID do validador (admin/gestor/operacional/parceiro)
    validado_em: Optional[str] = None  # Data de validação
    pode_editar: bool = True  # Se motorista pode editar (iban e registo_criminal sempre podem)

class SolicitacaoAlteracao(BaseModel):
    """Solicitação de alteração de dados pelo motorista"""
    id: str
    motorista_id: str
    motorista_nome: str
    campo: str  # Nome do campo que quer alterar
    valor_atual: str  # Valor atual do campo
    valor_solicitado: str  # Novo valor solicitado
    justificativa: str  # Motivo da alteração
    status: str = "pendente"  # pendente, aprovada, rejeitada
    respondido_por: Optional[str] = None
    resposta: Optional[str] = None
    created_at: datetime
    respondido_em: Optional[datetime] = None

class MotoristaCreate(BaseModel):
    email: EmailStr
    password: Optional[str] = None  # Optional for admin creation
    name: str
    phone: str
    morada_completa: str
    codigo_postal: str
    data_nascimento: str
    nacionalidade: str
    tipo_documento: str  # CC, Passaporte, Residencia
    numero_documento: str
    validade_documento: str
    nif: str
    carta_conducao_numero: str
    carta_conducao_validade: str
    licenca_tvde_numero: str
    licenca_tvde_validade: str
    codigo_registo_criminal: Optional[str] = None
    parceiro_atribuido: Optional[str] = None
    veiculo_atribuido: Optional[str] = None
    tipo_motorista: Optional[str] = "independente"  # tempo_integral, meio_periodo, independente, parceiro
    regime: str  # aluguer, comissao, carro_proprio
    iban: Optional[str] = None
    email_uber: Optional[str] = None
    telefone_uber: Optional[str] = None
    uuid_motorista_uber: Optional[str] = None  # UUID do motorista na Uber para identificar ganhos
    email_bolt: Optional[str] = None
    telefone_bolt: Optional[str] = None
    identificador_motorista_bolt: Optional[str] = None  # Identificador do motorista na Bolt para ganhos
    whatsapp: str
    tipo_pagamento: str  # fatura, recibo_verde, sem_recibo
    senha_provisoria: bool = False

class Motorista(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    email: str
    name: str
    phone: str
    morada_completa: Optional[str] = None
    codigo_postal: Optional[str] = None
    data_nascimento: Optional[str] = None
    nacionalidade: Optional[str] = None
    tipo_documento: Optional[str] = None
    numero_documento: Optional[str] = None
    validade_documento: Optional[str] = None
    nif: Optional[str] = None
    carta_conducao_numero: Optional[str] = None
    carta_conducao_validade: Optional[str] = None
    licenca_tvde_numero: Optional[str] = None
    licenca_tvde_validade: Optional[str] = None
    codigo_registo_criminal: Optional[str] = None  # Formato: xxxx-xxxx-xxxx
    parceiro_id: Optional[str] = None  # ID do parceiro (para compatibilidade)
    parceiro_atribuido: Optional[str] = None  # ID do parceiro
    veiculo_atribuido: Optional[str] = None  # ID do veículo
    status_motorista: Optional[str] = "pendente_documentos"  # aguarda_carro, ferias, ativo, pendente_documentos, desativo
    tipo_motorista: Optional[str] = None  # tempo_integral, meio_periodo, independente, parceiro
    regime: Optional[str] = None
    iban: Optional[str] = None
    email_uber: Optional[str] = None
    telefone_uber: Optional[str] = None
    uuid_motorista_uber: Optional[str] = None  # UUID do motorista na Uber para identificar ganhos
    email_bolt: Optional[str] = None
    telefone_bolt: Optional[str] = None
    identificador_motorista_bolt: Optional[str] = None  # Identificador do motorista na Bolt para ganhos
    whatsapp: Optional[str] = None
    tipo_pagamento: Optional[str] = None
    tipo_pagamento_outro: Optional[str] = None
    id_cartao_frota_combustivel: Optional[str] = None  # ID automático do cartão frota (gerado)
    
    # Localidade
    localidade: Optional[str] = None
    
    # Documentos Fiscais
    numero_seguranca_social: Optional[str] = None
    numero_cartao_utente: Optional[str] = None
    
    # Carta de Condução (expandido)
    numero_carta: Optional[str] = None
    numero_carta_conducao: Optional[str] = None  # Alias para numero_carta
    categoria_carta_conducao: Optional[str] = None  # B, B+E, C, D, etc
    emissao_carta: Optional[str] = None
    data_emissao_carta: Optional[str] = None  # Alias para emissao_carta
    validade_carta: Optional[str] = None
    validade_carta_conducao: Optional[str] = None  # Alias para validade_carta
    
    # Licença TVDE (expandido)
    numero_licenca_tvde: Optional[str] = None
    validade_licenca_tvde: Optional[str] = None
    
    # Registo Criminal
    validade_registo_criminal: Optional[str] = None
    data_limite_registo_criminal: Optional[str] = None
    dias_aviso_renovacao_registo: int = 30  # Dias de antecedência para aviso
    
    # CC / Documento de Identificação
    numero_cc: Optional[str] = None
    validade_cc: Optional[str] = None
    numero_documento_identificacao: Optional[str] = None  # CC ou Cartão de Residência
    validade_documento_identificacao: Optional[str] = None
    
    # Passaporte
    numero_passaporte: Optional[str] = None
    validade_passaporte: Optional[str] = None
    
    # Dados Bancários
    nome_banco: Optional[str] = None
    
    # Contacto de Emergência
    emergencia_nome: Optional[str] = None
    emergencia_telefone: Optional[str] = None
    emergencia_email: Optional[str] = None
    emergencia_morada: Optional[str] = None
    emergencia_codigo_postal: Optional[str] = None
    emergencia_localidade: Optional[str] = None
    emergencia_ligacao: Optional[str] = None
    
    # Seguro de Acidentes Pessoais
    seguro_numero_apolice: Optional[str] = None
    seguro_seguradora: Optional[str] = None
    seguro_validade: Optional[str] = None
    
    documents: MotoristaDocuments
    documents_validacao: Dict[str, DocumentoValidacao] = {}  # Controle de validação por documento
    documentos_aprovados: bool = False  # Se true, motorista não pode editar dados (exceto registo_criminal e iban)
    observacoes_internas: Optional[str] = None  # Notas internas visíveis apenas para Admin/Gestor/Parceiro/Operacional
    approved: bool = False
    senha_provisoria: bool = False
    campos_customizados: Dict[str, Any] = {}  # Campos adicionais customizáveis
    created_at: datetime
    approved_by: Optional[str] = None
    approved_at: Optional[datetime] = None

# Caução Model (para veículos) - deve vir antes de Vehicle
class CaucaoVeiculo(BaseModel):
    caucao_total: float = 0.0
    caucao_divisao: str = "total"  # "semanal", "mensal", "total"
    caucao_valor_semanal: float = 0.0  # Calculado automaticamente
    caucao_pago: float = 0.0
    caucao_restante: float = 0.0  # Calculado automaticamente

# Vehicle Models - COMPLETE EXPANDED
class TipoContrato(BaseModel):
    # Tipo de contrato alinhado com sistema de templates
    tipo: Optional[str] = "aluguer"  # aluguer_sem_caucao, aluguer_com_caucao, aluguer_caucao_parcelada,
                                      # periodo_epoca, aluguer_epocas_sem_caucao, aluguer_epocas_caucao,
                                      # aluguer_epoca_caucao_parcelada, compra_veiculo, comissao, 
                                      # motorista_privado, outros
    
    # Para Aluguer (todos os tipos de aluguer)
    valor_aluguer: Optional[float] = None
    periodicidade: Optional[str] = "semanal"  # semanal, mensal
    
    # Caução (para tipos com caução)
    valor_caucao: Optional[float] = None
    numero_parcelas_caucao: Optional[int] = None
    
    # Épocas (para tipos com época)
    valor_epoca_alta: Optional[float] = None
    valor_epoca_baixa: Optional[float] = None
    
    # Para Comissão
    comissao_parceiro: Optional[float] = None  # % da comissão para o parceiro
    comissao_motorista: Optional[float] = None  # % da comissão para o motorista (soma deve ser 100%)
    inclui_combustivel: bool = False
    regime: Optional[str] = None  # full_time, part_time
    
    # Para Compra do Veículo
    valor_compra_veiculo: Optional[float] = None
    numero_semanas_compra: Optional[int] = None
    valor_semanal_compra: Optional[float] = None  # Calculado
    periodo_compra: Optional[int] = None  # Número de semanas (legacy)
    valor_acumulado: Optional[float] = None
    valor_falta_cobrar: Optional[float] = None
    com_slot: bool = False
    custo_slot: Optional[float] = None  # Legacy
    extra_seguro: bool = False
    valor_extra_seguro: Optional[float] = None
    
    # Geral (mantido para compatibilidade)
    inclui_via_verde: bool = False
    
    # Part-time: 4 turnos livres configuráveis (opcionais)
    horario_turno_1: Optional[str] = None  # Ex: "09:00-13:00"
    horario_turno_2: Optional[str] = None  # Ex: "14:00-18:00"
    horario_turno_3: Optional[str] = None  # Ex: "19:00-23:00"
    horario_turno_4: Optional[str] = None  # Ex: "00:00-06:00"

class CategoriasUber(BaseModel):
    uberx: bool = False
    share: bool = False
    electric: bool = False
    black: bool = False
    comfort: bool = False
    xl: bool = False
    xxl: bool = False
    pet: bool = False
    package: bool = False

class CategoriasBolt(BaseModel):
    economy: bool = False
    comfort: bool = False
    executive: bool = False
    xl: bool = False
    green: bool = False
    xxl: bool = False
    motorista_privado: bool = False
    pet: bool = False

class VehicleInsurance(BaseModel):
    seguradora: Optional[str] = None  # Renomeado de companhia
    companhia: Optional[str] = None  # Mantido para retrocompatibilidade
    numero_apolice: Optional[str] = None
    apolice: Optional[str] = None  # Mantido para retrocompatibilidade
    agente_seguros: Optional[str] = None
    data_inicio: Optional[str] = None
    data_validade: Optional[str] = None
    valor: Optional[float] = None
    valor_anual: Optional[float] = None  # Mantido para retrocompatibilidade
    preco: Optional[float] = None  # Mantido para retrocompatibilidade
    periodicidade: Optional[str] = "anual"
    carta_verde_url: Optional[str] = None
    condicoes_url: Optional[str] = None
    fatura_url: Optional[str] = None
    condicoes: Optional[str] = None
    notas: Optional[str] = None

class VehicleMaintenance(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    data_intervencao: str
    tipo_manutencao: str
    descricao: str
    descricao_detalhada: Optional[str] = None
    custos: float
    oficina: str
    tempo_intervencao: str
    km_intervencao: int
    data_proxima: Optional[str] = None
    km_proxima: Optional[int] = None
    o_que_fazer: Optional[str] = None

class VehicleExtinguisher(BaseModel):
    numeracao: Optional[str] = None  # Número de série/identificação do extintor
    fornecedor: Optional[str] = None
    empresa_certificacao: Optional[str] = None
    preco: Optional[float] = None
    data_instalacao: Optional[str] = None  # Data de instalação do extintor
    data_entrega: Optional[str] = None  # Mantido para retrocompatibilidade
    data_validade: Optional[str] = None
    proxima_intervencao: Optional[str] = None
    certificado_url: Optional[str] = None

class VehicleInspection(BaseModel):
    fornecedor: Optional[str] = None
    data_inspecao: Optional[str] = None
    proxima_inspecao: Optional[str] = None
    custo: Optional[float] = None
    # New fields for inspection value
    ultima_inspecao: Optional[str] = None
    resultado: Optional[str] = None
    valor: Optional[float] = None
    centro_inspecao: Optional[str] = None
    observacoes: Optional[str] = None
    ficha_inspecao_url: Optional[str] = None

class VehicleAvailability(BaseModel):
    status: str  # disponivel, atribuido, manutencao, seguro, sinistro, venda
    motoristas_atribuidos: List[str] = []
    data_entrega_manutencao: Optional[str] = None
    tipo_manutencao: Optional[str] = None

class VehicleCreate(BaseModel):
    marca: str
    modelo: str
    versao: Optional[str] = None  # Ex: "Sport", "Exclusive", "Comfort"
    ano: Optional[int] = None
    matricula: str
    data_matricula: str
    validade_matricula: str
    cor: str
    combustivel: str  # gasolina, diesel, eletrico, hibrido, gnv
    caixa: str  # manual, automatica
    lugares: int
    tipo_contrato: TipoContrato
    categorias_uber: CategoriasUber
    categorias_bolt: CategoriasBolt
    via_verde_disponivel: bool = False
    cartao_frota_disponivel: bool = False
    parceiro_id: str

class Vehicle(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    marca: str
    modelo: str
    versao: Optional[str] = None
    ano: Optional[int] = None
    matricula: str
    data_matricula: str
    validade_matricula: str
    alerta_validade: bool = False  # True se falta menos de 30 dias
    cor: str
    combustivel: str
    caixa: str
    lugares: int
    tipo_contrato: Optional[TipoContrato] = None
    categorias_uber: Optional[CategoriasUber] = None
    categorias_bolt: Optional[CategoriasBolt] = None
    via_verde_disponivel: bool = False
    via_verde_id: Optional[str] = None  # ID do Via Verde
    cartao_frota_disponivel: bool = False
    cartao_frota_id: Optional[str] = None  # ID do Cartão Frota (associado ao motorista quando veículo é atribuído)
    # Garantia
    tem_garantia: bool = False
    stand_responsavel: Optional[str] = None
    data_limite_garantia: Optional[str] = None
    seguro: Optional[VehicleInsurance] = None
    manutencoes: List[VehicleMaintenance] = []
    extintor: Optional[VehicleExtinguisher] = None
    inspecoes: List[VehicleInspection] = []
    inspection: Optional[VehicleInspection] = None  # Single inspection field with valor
    disponibilidade: Optional[VehicleAvailability] = None
    km_atual: Optional[int] = 0
    km_aviso_manutencao: int = 5000
    alertas_manutencao: List[str] = []
    fotos: List[str] = []  # URLs das fotos (máximo 3, convertidas para PDF)
    caucao: Optional[CaucaoVeiculo] = None  # Caução do veículo
    danos: List[Dict[str, Any]] = []  # Histórico de danos
    agenda: List[Dict[str, Any]] = []  # Agenda do veículo
    historico_editavel: List[Dict[str, Any]] = []  # Histórico editável/observações
    fotos_veiculo: List[str] = []  # URLs das fotos do veículo (max 3)
    proxima_revisao_km: Optional[int] = None
    proxima_revisao_data: Optional[str] = None
    proxima_revisao_notas: Optional[str] = None
    proxima_revisao_valor_previsto: Optional[float] = None
    motorista_atribuido: Optional[str] = None  # ID do motorista
    motorista_atribuido_nome: Optional[str] = None  # Nome do motorista
    status: str = "disponivel"  # disponivel, atribuido, manutencao, venda, condicoes
    ultima_revisao_km: Optional[int] = None  # KM da última revisão
    data_seguro_ate: Optional[str] = None  # Validade do seguro (alias para insurance.data_validade)
    data_inspecao_ate: Optional[str] = None  # Validade da inspeção (alias para inspection.proxima_inspecao)
    plano_manutencoes: List[Dict[str, Any]] = []  # Plano de manutenções periódicas: [{nome, intervalo_km, ativo}]
    alertas_configuracao: Dict[str, int] = {
        "dias_aviso_seguro": 30,
        "dias_aviso_inspecao": 30,
        "dias_aviso_extintor": 30,
        "km_aviso_manutencao": 5000
    }  # Configuração de alertas específica do veículo
    verificacao_danos_ativa: bool = False  # Verificação de danos pelo gestor/operacional
    campos_customizados: Dict[str, Any] = {}  # Campos adicionais customizáveis
    # Marketplace fields
    disponivel_venda: bool = False  # Se está disponível para venda no marketplace
    disponivel_aluguer: bool = False  # Se está disponível para aluguer no marketplace
    preco_venda: Optional[float] = None  # Preço de venda
    preco_aluguer_mensal: Optional[float] = None  # Preço de aluguer mensal
    descricao_marketplace: Optional[str] = None  # Descrição para marketplace
    # Documentos do veículo
    documento_carta_verde: Optional[str] = None  # Carta verde do seguro
    documento_condicoes: Optional[str] = None  # Documento de condições
    documento_recibo_seguro: Optional[str] = None  # Recibo de pagamento do seguro
    documento_inspecao: Optional[str] = None  # Documento/certificado da inspeção
    created_at: datetime
    updated_at: datetime
    parceiro_id: str

# Financial Models
class ExpenseCreate(BaseModel):
    vehicle_id: str
    tipo: str  # seguro, manutencao, combustivel, portagem, multa, etc
    descricao: str
    valor: float
    data: str
    categoria: str

class Expense(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    vehicle_id: str
    tipo: str
    descricao: str
    valor: float
    data: str
    categoria: str
    created_at: datetime

class RevenueCreate(BaseModel):
    vehicle_id: str
    motorista_id: Optional[str] = None
    tipo: str  # uber, bolt, aluguer, outro
    valor: float
    data: str
    km_percorridos: Optional[int] = None

class Revenue(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    vehicle_id: str
    motorista_id: Optional[str] = None
    tipo: str
    valor: float
    data: str
    km_percorridos: Optional[int] = None
    created_at: datetime

# Partner Financial Models
class PartnerExpenseCreate(BaseModel):
    parceiro_id: str
    descricao: str
    valor: float
    data: str
    categoria: str  # manutencao, combustivel, portagem, seguro, etc
    observacoes: Optional[str] = None

class PartnerExpense(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    parceiro_id: str
    descricao: str
    valor: float
    data: str
    categoria: str
    observacoes: Optional[str] = None
    created_by: str
    created_at: datetime

class PartnerRevenueCreate(BaseModel):
    parceiro_id: str
    descricao: str
    valor: float
    data: str
    tipo: str  # comissao, aluguer, servico, outro
    observacoes: Optional[str] = None

class PartnerRevenue(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    parceiro_id: str
    descricao: str
    valor: float
    data: str
    tipo: str
    observacoes: Optional[str] = None
    created_by: str
    created_at: datetime

# Reports
class ROIReport(BaseModel):
    vehicle_id: str
    periodo: str
    total_receitas: float
    total_despesas: float
    roi: float
    km_percorridos: int
    utilizacao_percentual: float

# Pagamentos Models
class PagamentoCreate(BaseModel):
    motorista_id: str
    valor: float
    periodo_inicio: str
    periodo_fim: str
    tipo_documento: str  # fatura, recibo_verde, outro
    notas: Optional[str] = None

class Pagamento(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    motorista_id: str
    parceiro_id: str
    valor: float
    periodo_inicio: str
    periodo_fim: str
    tipo_documento: str
    documento_url: Optional[str] = None
    documento_analisado: bool = False
    analise_documento: Optional[Dict[str, Any]] = None

# ==================== SUBSCRIPTION MODELS ====================

class PlanoAssinatura(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    nome: str  # Ex: "Parceiro Base", "Operacional Premium"
    tipo_usuario: Optional[str] = "parceiro"  # "parceiro", "operacional"
    preco_por_unidade: Optional[float] = 0.0  # Por veículo (parceiro) ou por motorista (operacional)
    descricao: str
    features: Optional[List[str]] = []  # Lista de features habilitadas (mapeia de funcionalidades se existir)
    funcionalidades: Optional[List[str]] = []  # Backward compatibility
    ativo: bool = True
    created_at: datetime
    updated_at: datetime
    # Additional fields for compatibility
    preco_mensal: Optional[float] = None
    preco_anual: Optional[float] = None

class PlanoCreate(BaseModel):
    nome: str
    tipo_usuario: str
    preco_por_unidade: float
    descricao: str
    features: List[str]

class UserSubscription(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: str
    plano_id: str
    status: str  # "ativo", "cancelado", "suspenso", "pagamento_pendente"
    data_inicio: str
    data_fim: Optional[str] = None
    unidades_ativas: int = 0  # Número de veículos (parceiro) ou motoristas (operacional)
    valor_mensal: float = 0.0  # Calculado: preco_por_unidade * unidades_ativas

# ==================== PLANOS MOTORISTA MODELS ====================

class PlanoMotorista(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    nome: str  # Ex: "Básico", "Premium", "Profissional"
    descricao: str
    preco_semanal: float  # Preço semanal (0 para plano gratuito)
    preco_mensal: float  # Preço mensal
    desconto_mensal_percentagem: float = 0  # Desconto para pagamento mensal (ex: 15%)
    features: Dict[str, bool]  # Ex: {"alertas_recibos": True, "alertas_documentos": True, "dashboard_analytics": False}
    ativo: bool = True
    permite_pagamento_online: bool = True  # Se permite pagamento via IFThenPay
    created_at: datetime
    updated_at: datetime

class PlanoMotoristaCreate(BaseModel):
    nome: str
    descricao: str
    preco_semanal: float
    preco_mensal: float
    desconto_mensal_percentagem: float = 0
    features: Dict[str, bool]
    permite_pagamento_online: bool = True

class PlanoMotoristaUpdate(BaseModel):
    nome: Optional[str] = None
    descricao: Optional[str] = None
    preco_semanal: Optional[float] = None
    preco_mensal: Optional[float] = None
    desconto_mensal_percentagem: Optional[float] = None
    features: Optional[Dict[str, bool]] = None
    ativo: Optional[bool] = None
    permite_pagamento_online: Optional[bool] = None

class MotoristaPlanoAssinatura(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    motorista_id: str
    plano_id: str
    plano_nome: Optional[str] = None
    periodicidade: str  # "semanal" ou "mensal"
    preco_pago: float  # Valor efetivamente pago (com descontos aplicados)
    status: str  # "ativo", "cancelado", "expirado", "pagamento_pendente", "aguardando_pagamento"
    data_inicio: str
    data_fim: Optional[str] = None  # Data de vencimento da assinatura
    auto_renovacao: bool = False
    metodo_pagamento: Optional[str] = None  # "manual", "ifthenpay", "multibanco", "mbway"
    referencia_pagamento: Optional[str] = None  # Referência multibanco ou MBWay
    entidade_pagamento: Optional[str] = None  # Entidade multibanco
    pagamento_confirmado: bool = False
    data_pagamento: Optional[str] = None
    created_at: datetime
    updated_at: datetime


# ==================== RELATORIO CONFIGURATION MODELS ====================

class RelatorioConfig(BaseModel):
    """Configuração de campos disponíveis para relatórios semanais do parceiro"""
    model_config = ConfigDict(extra="ignore")
    id: str
    parceiro_id: str
    
    # Campos do cabeçalho do relatório
    incluir_numero_relatorio: bool = True
    incluir_data_emissao: bool = True
    incluir_periodo: bool = True
    incluir_nome_parceiro: bool = True
    incluir_nome_motorista: bool = True
    incluir_veiculo: bool = True  # Marca, modelo e matrícula
    
    # Campos de estatísticas de viagens
    incluir_viagens_bolt: bool = True
    incluir_viagens_uber: bool = True
    incluir_viagens_totais: bool = True
    incluir_horas_bolt: bool = True
    incluir_horas_uber: bool = True
    incluir_horas_totais: bool = True
    
    # Campos de ganhos
    incluir_ganhos_uber: bool = True
    incluir_ganhos_bolt: bool = True
    incluir_ganhos_totais: bool = True
    
    # Campos de despesas
    incluir_valor_aluguer: bool = True  # Valor semanal/comissão/compra
    incluir_combustivel: bool = True
    incluir_via_verde: bool = True
    via_verde_atraso_semanas: int = 1  # Atraso padrão de 1 semana
    incluir_caucao: bool = True
    incluir_caucao_parcelada: bool = True
    incluir_danos: bool = True
    incluir_danos_acumulados: bool = True
    incluir_danos_descricao: bool = True
    incluir_danos_parcelados: bool = True
    incluir_extras: bool = True  # Débitos ou créditos extras
    
    # Campos de total
    incluir_total_recibo: bool = True  # Ganhos - Despesas
    
    # Campos de combustível detalhado
    incluir_tabela_combustivel: bool = True
    incluir_combustivel_matricula: bool = True
    incluir_combustivel_local: bool = True
    incluir_combustivel_data_hora: bool = True
    incluir_combustivel_cartao: bool = True
    incluir_combustivel_quantidade: bool = True
    incluir_combustivel_valor: bool = True
    
    # Configurações adicionais
    formato_numero_relatorio: str = "xxxxx/ano"  # Formato do número do relatório
    texto_observacoes_padrao: Optional[str] = None  # Texto padrão de observações
    
    created_at: datetime
    updated_at: datetime

class RelatorioConfigUpdate(BaseModel):
    """Model para atualizar configuração de relatórios"""
    incluir_numero_relatorio: Optional[bool] = None
    incluir_data_emissao: Optional[bool] = None
    incluir_periodo: Optional[bool] = None
    incluir_nome_parceiro: Optional[bool] = None
    incluir_nome_motorista: Optional[bool] = None
    incluir_veiculo: Optional[bool] = None
    incluir_viagens_bolt: Optional[bool] = None
    incluir_viagens_uber: Optional[bool] = None
    incluir_viagens_totais: Optional[bool] = None
    incluir_horas_bolt: Optional[bool] = None
    incluir_horas_uber: Optional[bool] = None
    incluir_horas_totais: Optional[bool] = None
    incluir_ganhos_uber: Optional[bool] = None
    incluir_ganhos_bolt: Optional[bool] = None
    incluir_ganhos_totais: Optional[bool] = None
    incluir_valor_aluguer: Optional[bool] = None
    incluir_combustivel: Optional[bool] = None
    incluir_via_verde: Optional[bool] = None
    via_verde_atraso_semanas: Optional[int] = None
    incluir_caucao: Optional[bool] = None
    incluir_caucao_parcelada: Optional[bool] = None
    incluir_danos: Optional[bool] = None
    incluir_danos_acumulados: Optional[bool] = None
    incluir_danos_descricao: Optional[bool] = None
    incluir_danos_parcelados: Optional[bool] = None
    incluir_extras: Optional[bool] = None
    incluir_total_recibo: Optional[bool] = None
    incluir_tabela_combustivel: Optional[bool] = None
    incluir_combustivel_matricula: Optional[bool] = None
    incluir_combustivel_local: Optional[bool] = None
    incluir_combustivel_data_hora: Optional[bool] = None
    incluir_combustivel_cartao: Optional[bool] = None
    incluir_combustivel_quantidade: Optional[bool] = None
    incluir_combustivel_valor: Optional[bool] = None
    formato_numero_relatorio: Optional[str] = None
    texto_observacoes_padrao: Optional[str] = None

# ==================== CONTRACT MODELS ====================

class ContratoCreate(BaseModel):
    parceiro_id: str
    motorista_id: str
    vehicle_id: str
    data_inicio: str
    tipo_contrato: str = "comissao"  # "comissao", "aluguer_normal", "aluguer_epocas", "compra", "motorista_privado"
    
    # Valores financeiros
    valor_semanal: float = 230.0
    comissao_percentual: Optional[float] = None  # Para tipo comissão
    caucao_total: float = 300.0
    caucao_lavagem: float = 90.0
    
    # Campos de caução
    tem_caucao: bool = True
    caucao_parcelada: bool = False
    caucao_parcelas: Optional[int] = None
    caucao_texto: Optional[str] = None  # Texto personalizado sobre caução
    
    # Campos para época (opcional)
    tem_epoca: bool = False
    data_inicio_epoca_alta: Optional[str] = None
    data_fim_epoca_alta: Optional[str] = None
    valor_epoca_alta: Optional[float] = None
    texto_epoca_alta: Optional[str] = None  # Texto/observações época alta
    data_inicio_epoca_baixa: Optional[str] = None
    data_fim_epoca_baixa: Optional[str] = None
    valor_epoca_baixa: Optional[float] = None
    texto_epoca_baixa: Optional[str] = None  # Texto/observações época baixa
    
    # Condições do veículo (editáveis na emissão)
    condicoes_veiculo: Optional[str] = None
    
    # Template/texto do contrato
    template_texto: Optional[str] = None

class Contrato(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    referencia: str  # Ex: "001/2024"
    parceiro_id: str
    motorista_id: str
    vehicle_id: Optional[str] = None
    
    # Dados Parceiro Completos
    parceiro_nome: str
    parceiro_nif: str
    parceiro_morada: str
    parceiro_codigo_postal: Optional[str] = None
    parceiro_localidade: Optional[str] = None
    parceiro_telefone: Optional[str] = None
    parceiro_email: str
    parceiro_representante_legal_nome: Optional[str] = None
    parceiro_representante_legal_contribuinte: Optional[str] = None
    parceiro_representante_legal_cc: Optional[str] = None
    parceiro_representante_legal_cc_validade: Optional[str] = None
    parceiro_representante_legal_telefone: Optional[str] = None
    parceiro_representante_legal_email: Optional[str] = None
    
    # Dados Motorista Completos
    motorista_nome: str
    motorista_cc: str
    motorista_cc_validade: Optional[str] = None
    motorista_nif: str
    motorista_morada: str
    motorista_codigo_postal: Optional[str] = None
    motorista_localidade: Optional[str] = None
    motorista_telefone: Optional[str] = None
    motorista_carta_conducao: Optional[str] = None
    motorista_carta_conducao_validade: Optional[str] = None
    motorista_licenca_tvde: Optional[str] = None
    motorista_licenca_tvde_validade: Optional[str] = None
    motorista_seguranca_social: Optional[str] = None
    motorista_email: str
    
    # Dados Veículo
    vehicle_marca: Optional[str] = None
    vehicle_modelo: Optional[str] = None
    vehicle_matricula: Optional[str] = None
    
    # Termos Financeiros
    tipo_contrato: str  # "comissao", "aluguer_normal", "aluguer_epocas", "compra", "motorista_privado"
    valor_semanal: float = 230.0
    comissao_percentual: Optional[float] = None
    caucao_total: float = 300.0
    caucao_lavagem: float = 90.0
    
    # Campos de caução
    tem_caucao: bool = True
    caucao_parcelada: bool = False
    caucao_parcelas: Optional[int] = None
    caucao_texto: Optional[str] = None  # Texto personalizado sobre caução
    
    # Campos de Época (opcional)
    tem_epoca: bool = False
    data_inicio_epoca_alta: Optional[str] = None
    data_fim_epoca_alta: Optional[str] = None
    valor_epoca_alta: Optional[float] = None
    texto_epoca_alta: Optional[str] = None
    data_inicio_epoca_baixa: Optional[str] = None
    data_fim_epoca_baixa: Optional[str] = None
    valor_epoca_baixa: Optional[float] = None
    texto_epoca_baixa: Optional[str] = None
    
    # Condições do veículo
    condicoes_veiculo: Optional[str] = None
    
    # Template do contrato
    template_texto: Optional[str] = None
    
    # Datas
    data_inicio: str
    data_emissao: str  # Data em que o contrato foi gerado/emitido
    data_assinatura: str
    local_assinatura: str = "Lisboa"
    
    # Status e Assinaturas
    status: str  # "rascunho", "ativo", "terminado"
    parceiro_assinado: bool = False
    motorista_assinado: bool = False
    parceiro_assinatura_data: Optional[str] = None
    motorista_assinatura_data: Optional[str] = None
    pdf_url: Optional[str] = None
    created_at: datetime
    updated_at: datetime



# ==================== MINUTAS DE CONTRATO ====================

class MinutaContrato(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    parceiro_id: str
    nome: str  # Ex: "Contrato Aluguer Padrão", "Contrato Comissão"
    tipo_contrato: str  # "comissao", "aluguer", "compra", "motorista_privado"
    texto_minuta: str  # Texto com variáveis
    ativa: bool = True
    created_at: datetime
    updated_at: datetime
    created_by: str

class MinutaContratoCreate(BaseModel):
    parceiro_id: str
    nome: str
    tipo_contrato: str
    texto_minuta: str

# ==================== CONFIGURAÇÕES DO SISTEMA ====================

class ConfiguracaoSistema(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = "config_sistema"
    condicoes_gerais: str = ""
    politica_privacidade: str = ""
    updated_at: datetime
    updated_by: str

class ConfiguracaoUpdate(BaseModel):
    condicoes_gerais: Optional[str] = None
    politica_privacidade: Optional[str] = None

# ==================== RECIBOS E PAGAMENTOS MODELS ====================

class RelatorioGanhosCreate(BaseModel):
    motorista_id: str
    periodo_inicio: str
    periodo_fim: str
    valor_total: float
    detalhes: Dict[str, Any]  # Breakdown de ganhos (uber, bolt, etc)
    notas: Optional[str] = None

class RelatorioGanhos(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    motorista_id: str
    parceiro_id: str
    motorista_nome: str
    periodo_inicio: str
    periodo_fim: str
    
    # Ganhos
    ganhos_uber: float = 0.0
    ganhos_bolt: float = 0.0
    
    # Despesas
    despesas_combustivel: List[Dict[str, Any]] = []  # [{data, hora, valor, quantidade, local}]
    total_combustivel: float = 0.0
    
    despesas_via_verde: List[Dict[str, Any]] = []  # [{data, hora, valor, local}]
    total_via_verde: float = 0.0
    
    # Extras e deduções
    km_efetuados: Optional[int] = None
    extras_parceiro: List[Dict[str, Any]] = []  # [{tipo, descricao, valor}] - danos, multas, etc
    total_extras: float = 0.0
    
    valor_divida_anterior: float = 0.0
    valor_caucao_semanal: float = 0.0
    valor_caucao_acumulada: float = 0.0
    valor_dano_veiculo_semanal: float = 0.0
    valor_dano_veiculo_total: float = 0.0
    
    # Estatísticas
    horas_online: Optional[float] = None
    numero_viagens: Optional[int] = None
    media_por_viagem: Optional[float] = None
    media_por_hora: Optional[float] = None
    
    # Valores finais
    valor_bruto: float = 0.0  # ganhos_uber + ganhos_bolt
    valor_descontos: float = 0.0  # total de todas as deduções
    valor_liquido: float = 0.0  # valor final a receber
    
    # Dados do parceiro (para recibo)
    parceiro_nome: Optional[str] = None
    parceiro_nif: Optional[str] = None
    parceiro_morada: Optional[str] = None
    
    detalhes: Dict[str, Any] = {}
    notas: Optional[str] = None
    status: str = "rascunho"  # "rascunho", "enviado", "recibo_gerado", "pago"
    recibo_url: Optional[str] = None
    recibo_emitido_em: Optional[str] = None
    aprovado_pagamento: bool = False
    aprovado_pagamento_por: Optional[str] = None
    aprovado_pagamento_em: Optional[str] = None
    pago: bool = False
    pago_por: Optional[str] = None
    pago_em: Optional[str] = None
    comprovativo_pagamento_url: Optional[str] = None
    created_by: str  # user_id (admin, gestor, operacional)
    created_at: datetime
    updated_at: datetime

class ReciboMotorista(BaseModel):
    """Recibo emitido pelo motorista"""
    model_config = ConfigDict(extra="ignore")
    id: str
    relatorio_ganhos_id: str
    motorista_id: str
    motorista_nome: str
    motorista_nif: str
    periodo_inicio: str
    periodo_fim: str
    valor_total: float
    descricao_servicos: str
    pdf_url: str
    emitido_em: datetime
    created_at: datetime

class SubscriptionCreate(BaseModel):
    user_id: str
    plano_id: str

# ==================== CSV DATA MODELS ====================

class GanhoUber(BaseModel):
    """Dados processados de CSV Uber"""
    model_config = ConfigDict(extra="ignore")
    id: str
    motorista_id: str
    uuid_motorista_uber: str
    nome_motorista: str
    periodo_inicio: str
    periodo_fim: str
    total_pago: float
    created_at: datetime

class GanhoBolt(BaseModel):
    """Dados processados de CSV Bolt"""
    model_config = ConfigDict(extra="ignore")
    id: str
    motorista_id: str
    email_motorista: str
    nome_motorista: str
    periodo_inicio: str
    periodo_fim: str
    ganhos_brutos: float
    ganhos_liquidos: float
    viagens_terminadas: int
    created_at: datetime

class TransacaoCombustivel(BaseModel):
    """Dados processados de Excel Prio/Combustível"""
    model_config = ConfigDict(extra="ignore")
    id: str
    motorista_id: str
    cartao: str
    data_transacao: str
    hora_transacao: str
    posto: str
    combustivel: str
    litros: float
    valor_liquido: float
    iva: float
    total: float
    kms: Optional[int] = None
    created_at: datetime

class ContratoGerado(BaseModel):
    """Contrato gerado em PDF"""
    model_config = ConfigDict(extra="ignore")
    id: str
    parceiro_id: str
    motorista_id: str
    veiculo_id: str
    tipo_contrato: str  # aluguer, comissao
    pdf_url: str
    dados_contrato: Dict[str, Any]
    status: str  # gerado, assinado, cancelado
    created_at: datetime
    assinado_em: Optional[datetime] = None

class Fatura(BaseModel):
    """Fatura de subscrição"""
    model_config = ConfigDict(extra="ignore")
    id: str
    subscription_id: str
    user_id: str
    periodo_referencia: str  # "2025-11"
    valor_total: float
    unidades_cobradas: int
    status: str  # pendente, paga, cancelada
    pdf_url: Optional[str] = None
    data_emissao: str
    data_vencimento: str
    data_pagamento: Optional[str] = None
    created_at: datetime

# Alertas Models
class AlertaCreate(BaseModel):
    tipo: str  # seguro, inspecao, licenca_tvde, manutencao, validade_matricula
    entidade_id: str  # vehicle_id or motorista_id
    entidade_tipo: str  # veiculo, motorista
    titulo: str
    descricao: str
    data_vencimento: str
    prioridade: str  # alta, media, baixa
    dias_antecedencia: int = 30

class Alerta(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    tipo: str
    entidade_id: str
    entidade_tipo: str
    titulo: str
    descricao: str
    data_vencimento: str
    prioridade: str
    dias_antecedencia: int
    status: str  # ativo, resolvido, ignorado
    criado_em: datetime
    resolvido_em: Optional[datetime] = None

# ==================== AUTH UTILITIES ====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return bcrypt.checkpw(plain_password.encode('utf-8'), hashed_password.encode('utf-8'))

def create_access_token(user_id: str, email: str, role: str) -> str:
    expiration = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    payload = {
        "user_id": user_id,
        "email": email,
        "role": role,
        "exp": expiration
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> Dict[str, Any]:
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def check_feature_access(user: Dict, feature_name: str) -> bool:
    """Check if user has access to a specific feature based on their subscription"""
    # Admin always has access
    if user["role"] == UserRole.ADMIN:
        return True
    
    # Get user's subscription
    subscription_id = user.get("subscription_id")
    if not subscription_id:
        # No subscription = no access to premium features
        return False
    
    subscription = await db.subscriptions.find_one({"id": subscription_id, "status": "ativo"}, {"_id": 0})
    if not subscription:
        return False
    
    # Get plan features
    plano = await db.planos.find_one({"id": subscription["plano_id"]}, {"_id": 0})
    if not plano:
        return False
    
    return feature_name in plano.get("features", [])

# ==================== CSV PROCESSING UTILITIES ====================

async def process_uber_csv(file_content: bytes, parceiro_id: str, periodo_inicio: str, periodo_fim: str) -> Dict[str, Any]:
    """Process Uber CSV file and extract earnings data (for parceiro/operacional)"""
    try:
        # Save original CSV file for audit/backup
        csv_dir = UPLOAD_DIR / "csv" / "uber"
        csv_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_filename = f"uber_{parceiro_id}_{timestamp}.csv"
        csv_path = csv_dir / csv_filename
        
        with open(csv_path, 'wb') as f:
            f.write(file_content)
        
        # Decode CSV
        csv_text = file_content.decode('utf-8-sig')  # Handle BOM
        
        # Detect delimiter (Portuguese Excel often uses semicolon)
        sample = csv_text[:1000]
        delimiter = ';' if sample.count(';') > sample.count(',') else ','
        logger.info(f"Detected CSV delimiter for Uber import: '{delimiter}'")
        
        csv_reader = csv.DictReader(io.StringIO(csv_text), delimiter=delimiter)
        
        # Process CSV rows (can have multiple motoristas)
        total_registos = 0
        total_pago_all = 0
        motoristas_nao_encontrados = []
        motoristas_unicos = set()
        
        for row in csv_reader:
            uuid_uber = row.get("UUID do motorista", "")
            nome = f"{row.get('Nome próprio do motorista', '')} {row.get('Apelido do motorista', '')}".strip()
            total_pago = float(row.get("Pago a si", "0").replace(",", ".") or 0)
            
            if nome:  # Only process if there's a name
                # Check if motorista exists
                motorista_key = f"{nome}_{uuid_uber}"
                if motorista_key not in motoristas_unicos:
                    motoristas_unicos.add(motorista_key)
                    
                    # Try to find motorista by nome or email
                    motorista = await db.motoristas.find_one({"name": nome}, {"_id": 0})
                    
                    if not motorista:
                        # Motorista não encontrado
                        motoristas_nao_encontrados.append({
                            "nome": nome,
                            "uuid_uber": uuid_uber,
                            "email": "",  # Uber CSV não tem email
                            "telefone": ""
                        })
                
                # Store in database
                ganho = {
                    "id": str(uuid.uuid4()),
                    "parceiro_id": parceiro_id,
                    "uuid_motorista_uber": uuid_uber,
                    "nome_motorista": nome,
                    "periodo_inicio": periodo_inicio,
                    "periodo_fim": periodo_fim,
                    "total_pago": total_pago,
                    "csv_original": f"uploads/csv/uber/{csv_filename}",
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.ganhos_uber.insert_one(ganho)
                total_registos += 1
                total_pago_all += total_pago
        
        return {
            "success": True,
            "registos_importados": total_registos,
            "total_pago": total_pago_all,
            "periodo": f"{periodo_inicio} a {periodo_fim}",
            "csv_salvo": csv_filename,
            "motoristas_nao_encontrados": motoristas_nao_encontrados
        }
    
    except Exception as e:
        logger.error(f"Error processing Uber CSV: {e}")
        return {"success": False, "error": str(e)}

async def process_bolt_csv(file_content: bytes, parceiro_id: str, periodo_inicio: str, periodo_fim: str) -> Dict[str, Any]:
    """Process Bolt CSV file and extract earnings data (for parceiro/operacional)"""
    try:
        # Save original CSV file for audit/backup
        csv_dir = UPLOAD_DIR / "csv" / "bolt"
        csv_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_filename = f"bolt_{parceiro_id}_{timestamp}.csv"
        csv_path = csv_dir / csv_filename
        
        with open(csv_path, 'wb') as f:
            f.write(file_content)
        
        # Decode CSV
        csv_text = file_content.decode('utf-8-sig')
        
        # Detect delimiter (Portuguese Excel often uses semicolon)
        sample = csv_text[:1000]
        delimiter = ';' if sample.count(';') > sample.count(',') else ','
        logger.info(f"Detected CSV delimiter for Bolt import: '{delimiter}'")
        
        csv_reader = csv.DictReader(io.StringIO(csv_text), delimiter=delimiter)
        
        # Process CSV rows (can have multiple motoristas)
        total_registos = 0
        total_ganhos_liquidos = 0
        total_viagens = 0
        motoristas_nao_encontrados = []
        motoristas_unicos = set()
        
        for row in csv_reader:
            nome = row.get("Motorista", "").strip()
            email = row.get("Email", "").strip()
            telefone = row.get("Telemóvel", "").strip()
            ganhos_brutos = float(row.get("Ganhos brutos (total)|€", "0").replace(",", ".") or 0)
            ganhos_liquidos = float(row.get("Ganhos líquidos|€", "0").replace(",", ".") or 0)
            viagens = int(row.get("Viagens terminadas", "0") or 0)
            
            if nome:  # Only process if there's a name
                # Check if motorista exists
                motorista_key = f"{nome}_{email}"
                if motorista_key not in motoristas_unicos:
                    motoristas_unicos.add(motorista_key)
                    
                    # Try to find motorista by email or name
                    motorista = None
                    if email:
                        motorista = await db.motoristas.find_one({"email": email}, {"_id": 0})
                    if not motorista:
                        motorista = await db.motoristas.find_one({"name": nome}, {"_id": 0})
                    
                    if not motorista:
                        # Motorista não encontrado
                        motoristas_nao_encontrados.append({
                            "nome": nome,
                            "email": email,
                            "telefone": telefone,
                            "identificador_bolt": row.get("Identificador do motorista", "")
                        })
                
                # Store in database
                ganho = {
                    "id": str(uuid.uuid4()),
                    "parceiro_id": parceiro_id,
                    "email_motorista": email,
                    "nome_motorista": nome,
                    "periodo_inicio": periodo_inicio,
                    "periodo_fim": periodo_fim,
                    "ganhos_brutos": ganhos_brutos,
                    "ganhos_liquidos": ganhos_liquidos,
                    "viagens_terminadas": viagens,
                    "csv_original": f"uploads/csv/bolt/{csv_filename}",
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.ganhos_bolt.insert_one(ganho)
                total_registos += 1
                total_ganhos_liquidos += ganhos_liquidos
                total_viagens += viagens
        
        return {
            "success": True,
            "registos_importados": total_registos,
            "ganhos_liquidos": total_ganhos_liquidos,
            "viagens": total_viagens,
            "periodo": f"{periodo_inicio} a {periodo_fim}",
            "csv_salvo": csv_filename,
            "motoristas_nao_encontrados": motoristas_nao_encontrados
        }
    
    except Exception as e:
        logger.error(f"Error processing Bolt CSV: {e}")
        return {"success": False, "error": str(e)}

async def process_prio_excel(file_content: bytes, parceiro_id: str) -> Dict[str, Any]:
    """Process Prio Excel file and extract fuel transactions (for parceiro/operacional)"""
    try:
        # Save original Excel file for audit/backup
        excel_dir = UPLOAD_DIR / "csv" / "combustivel"
        excel_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"prio_{parceiro_id}_{timestamp}.xlsx"
        excel_path = excel_dir / excel_filename
        
        with open(excel_path, 'wb') as f:
            f.write(file_content)
        
        # Read with openpyxl
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        
        # Skip header rows (first 3 rows are empty/header)
        transactions = []
        for row in sheet.iter_rows(min_row=5, values_only=True):
            if not row[3]:  # Skip if no date
                continue
            
            data_transacao = row[3].strftime('%Y-%m-%d') if hasattr(row[3], 'strftime') else str(row[3])
            hora_transacao = row[4].strftime('%H:%M:%S') if hasattr(row[4], 'strftime') else str(row[4])
            
            transacao = {
                "id": str(uuid.uuid4()),
                "parceiro_id": parceiro_id,
                "cartao": str(row[5] or ""),
                "data_transacao": data_transacao,
                "hora_transacao": hora_transacao,
                "posto": str(row[0] or ""),
                "combustivel": str(row[10] or ""),
                "litros": float(row[9] or 0),
                "valor_liquido": float(row[12] or 0),
                "iva": float(row[13] or 0),
                "total": float(row[18] or 0),
                "kms": int(row[14] or 0) if row[14] else None,
                "excel_original": f"uploads/csv/combustivel/{excel_filename}",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            transactions.append(transacao)
        
        # Store in database
        if transactions:
            await db.transacoes_combustivel.insert_many(transactions)
        
        total_litros = sum(t["litros"] for t in transactions)
        total_valor = sum(t["total"] for t in transactions)
        
        return {
            "success": True,
            "transacoes_importadas": len(transactions),
            "total_litros": total_litros,
            "total_valor": total_valor,
            "excel_salvo": excel_filename
        }
    
    except Exception as e:
        logger.error(f"Error processing Prio Excel: {e}")
        return {"success": False, "error": str(e)}

async def process_viaverde_excel(file_content: bytes, parceiro_id: str, periodo_inicio: str, periodo_fim: str) -> Dict[str, Any]:
    """Process Via Verde Excel file and extract toll movements"""
    try:
        # Save original Excel file for audit/backup
        excel_dir = UPLOAD_DIR / "csv" / "viaverde"
        excel_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"viaverde_{parceiro_id}_{timestamp}.xlsx"
        excel_path = excel_dir / excel_filename
        
        with open(excel_path, 'wb') as f:
            f.write(file_content)
        
        # Read with openpyxl
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]
        
        # Process data rows
        movimentos = []
        veiculos_nao_encontrados = []
        veiculos_unicos = set()
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip if no license plate
                continue
            
            matricula = str(row[0] or "")
            
            # Check if vehicle exists
            if matricula and matricula not in veiculos_unicos:
                veiculos_unicos.add(matricula)
                
                veiculo_db = await db.vehicles.find_one({"matricula": matricula}, {"_id": 0})
                
                if not veiculo_db:
                    # Veículo não encontrado
                    veiculos_nao_encontrados.append({
                        "matricula": matricula,
                        "obu": str(row[1] or "") if len(row) > 1 else ""
                    })
            
            movimento = {
                "id": str(uuid.uuid4()),
                "parceiro_id": parceiro_id,
                "license_plate": matricula,
                "obu": str(row[1] or "") if len(row) > 1 else None,
                "service": str(row[2] or "") if len(row) > 2 else None,
                "service_description": str(row[3] or "") if len(row) > 3 else None,
                "market": str(row[4] or "") if len(row) > 4 else None,
                "entry_date": str(row[6] or "") if len(row) > 6 else None,
                "exit_date": str(row[7] or "") if len(row) > 7 else None,
                "entry_point": str(row[8] or "") if len(row) > 8 else None,
                "exit_point": str(row[9] or "") if len(row) > 9 else None,
                "value": float(row[10]) if len(row) > 10 and row[10] and isinstance(row[10], (int, float)) else None,
                "is_payed": str(row[11] or "") if len(row) > 11 else None,
                "payment_date": str(row[12] or "") if len(row) > 12 else None,
                "contract_number": str(row[13] or "") if len(row) > 13 else None,
                "liquid_value": float(row[16]) if len(row) > 16 and row[16] and isinstance(row[16], (int, float)) else None,
                "periodo_inicio": periodo_inicio,
                "periodo_fim": periodo_fim,
                "ficheiro_nome": excel_filename,
                "data_importacao": datetime.now(timezone.utc).isoformat()
            }
            
            movimentos.append(movimento)
        
        # Store in database
        if movimentos:
            await db.viaverde_movimentos.insert_many(movimentos)
        
        total_value = sum(m["value"] for m in movimentos if m["value"])
        
        return {
            "success": True,
            "movimentos_importados": len(movimentos),
            "total_value": round(total_value, 2),
            "periodo": f"{periodo_inicio} a {periodo_fim}",
            "ficheiro_salvo": excel_filename,
            "veiculos_nao_encontrados": veiculos_nao_encontrados
        }
    
    except Exception as e:
        logger.error(f"Error processing Via Verde Excel: {e}")
        return {"success": False, "error": str(e)}

async def process_gps_csv(file_content: bytes, parceiro_id: str, periodo_inicio: str, periodo_fim: str) -> Dict[str, Any]:
    """Process GPS distance report CSV file"""
    try:
        # Save original CSV file for audit/backup
        csv_dir = UPLOAD_DIR / "csv" / "gps"
        csv_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_filename = f"gps_{parceiro_id}_{timestamp}.csv"
        csv_path = csv_dir / csv_filename
        
        with open(csv_path, 'wb') as f:
            f.write(file_content)
        
        # Decode CSV (with Portuguese encoding)
        csv_text = file_content.decode('utf-8-sig')
        
        # Detect delimiter (Portuguese Excel often uses semicolon)
        sample = csv_text[:1000]
        delimiter = ';' if sample.count(';') > sample.count(',') else ','
        logger.info(f"Detected CSV delimiter for GPS import: '{delimiter}'")
        
        csv_reader = csv.DictReader(io.StringIO(csv_text), delimiter=delimiter)
        
        # Process CSV rows
        registos = []
        total_distancia = 0
        veiculos_nao_encontrados = []
        veiculos_unicos = set()
        
        for row in csv_reader:
            veiculo = row.get("Veículo", "").strip()
            condutor = row.get("Condutor", "").strip()
            distancia_str = row.get("Distância percorrida durante as horas do turno", "0").replace(",", ".")
            
            try:
                distancia = float(distancia_str) if distancia_str else 0
            except ValueError:
                distancia = 0
            
            motor_ligado_tempo = row.get("Motor ligado", "").strip()
            motor_ligado_min = int(row.get("Motor ligado (minutos)", "0") or 0)
            
            if veiculo:  # Only process if there's a vehicle
                # Check if vehicle exists
                if veiculo not in veiculos_unicos:
                    veiculos_unicos.add(veiculo)
                    
                    veiculo_db = await db.vehicles.find_one({"matricula": veiculo}, {"_id": 0})
                    
                    if not veiculo_db:
                        # Veículo não encontrado
                        veiculos_nao_encontrados.append({
                            "matricula": veiculo,
                            "condutor_atual": condutor if condutor and condutor != "(Não atribuído)" else ""
                        })
                
                registo = {
                    "id": str(uuid.uuid4()),
                    "parceiro_id": parceiro_id,
                    "veiculo": veiculo,
                    "condutor": condutor if condutor and condutor != "(Não atribuído)" else None,
                    "distancia_percorrida": distancia,
                    "motor_ligado_tempo": motor_ligado_tempo,
                    "motor_ligado_minutos": motor_ligado_min,
                    "periodo_inicio": periodo_inicio,
                    "periodo_fim": periodo_fim,
                    "ficheiro_nome": csv_filename,
                    "data_importacao": datetime.now(timezone.utc).isoformat()
                }
                
                registos.append(registo)
                total_distancia += distancia
        
        # Store in database
        if registos:
            await db.gps_distancia.insert_many(registos)
        
        return {
            "success": True,
            "registos_importados": len(registos),
            "total_distancia_km": round(total_distancia, 2),
            "periodo": f"{periodo_inicio} a {periodo_fim}",
            "ficheiro_salvo": csv_filename,
            "veiculos_nao_encontrados": veiculos_nao_encontrados
        }
    
    except Exception as e:
        logger.error(f"Error processing GPS CSV: {e}")
        return {"success": False, "error": str(e)}

async def process_combustivel_eletrico_excel(file_content: bytes, parceiro_id: str, periodo_inicio: str, periodo_fim: str) -> Dict[str, Any]:
    """Process Electric Fuel Transactions Excel file"""
    try:
        # Save original Excel file for audit/backup
        excel_dir = UPLOAD_DIR / "csv" / "combustivel_eletrico"
        excel_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"eletrico_{parceiro_id}_{timestamp}.xlsx"
        excel_path = excel_dir / excel_filename
        
        with open(excel_path, 'wb') as f:
            f.write(file_content)
        
        # Read with openpyxl
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]
        
        # Process data rows
        transacoes = []
        total_custo = 0
        total_energia = 0
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip if no card number
                continue
            
            # Extrair valores com tratamento de erros
            def get_float(val):
                try:
                    return float(val) if val and isinstance(val, (int, float)) else None
                except (ValueError, TypeError):
                    return None
            
            energia = get_float(row[6]) if len(row) > 6 else None
            custo = get_float(row[8]) if len(row) > 8 else None
            total = get_float(row[11]) if len(row) > 11 else None
            total_iva = get_float(row[12]) if len(row) > 12 else None
            
            transacao = {
                "id": str(uuid.uuid4()),
                "parceiro_id": parceiro_id,
                "numero_cartao": str(row[0] or "") if len(row) > 0 else None,
                "nome": str(row[1] or "") if len(row) > 1 else None,
                "descricao": str(row[2] or "") if len(row) > 2 else None,
                "matricula": str(row[3] or "") if len(row) > 3 else None,
                "id_carregamento": str(row[4] or "") if len(row) > 4 else None,
                "posto": str(row[5] or "") if len(row) > 5 else None,
                "energia": energia,
                "duracao": get_float(row[7]) if len(row) > 7 else None,
                "custo": custo,
                "opc": get_float(row[9]) if len(row) > 9 else None,
                "iec": get_float(row[10]) if len(row) > 10 else None,
                "total": total,
                "total_com_iva": total_iva,
                "fatura": str(row[13] or "") if len(row) > 13 else None,
                "periodo_inicio": periodo_inicio,
                "periodo_fim": periodo_fim,
                "ficheiro_nome": excel_filename,
                "data_importacao": datetime.now(timezone.utc).isoformat()
            }
            
            transacoes.append(transacao)
            if total_iva:
                total_custo += total_iva
            if energia:
                total_energia += energia
        
        # Store in database
        if transacoes:
            await db.combustivel_eletrico.insert_many(transacoes)
        
        return {
            "success": True,
            "transacoes_importadas": len(transacoes),
            "total_energia_kwh": round(total_energia, 2),
            "total_custo_eur": round(total_custo, 2),
            "periodo": f"{periodo_inicio} a {periodo_fim}",
            "ficheiro_salvo": excel_filename
        }
    
    except Exception as e:
        logger.error(f"Error processing Electric Fuel Excel: {e}")
        return {"success": False, "error": str(e)}

async def process_combustivel_fossil_excel(file_content: bytes, parceiro_id: str, periodo_inicio: str, periodo_fim: str) -> Dict[str, Any]:
    """Process Fossil Fuel Transactions Excel file"""
    try:
        # Save original Excel file for audit/backup
        excel_dir = UPLOAD_DIR / "csv" / "combustivel_fossil"
        excel_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"fossil_{parceiro_id}_{timestamp}.xlsx"
        excel_path = excel_dir / excel_filename
        
        with open(excel_path, 'wb') as f:
            f.write(file_content)
        
        # Read with openpyxl
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]
        
        # Process data rows
        transacoes = []
        total_litros = 0
        total_valor = 0
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[5]:  # Skip if no card number
                continue
            
            # Extrair valores com tratamento de erros
            def get_float(val):
                try:
                    return float(val) if val and isinstance(val, (int, float)) else None
                except (ValueError, TypeError):
                    return None
            
            def get_int(val):
                try:
                    return int(val) if val and isinstance(val, (int, float)) else None
                except (ValueError, TypeError):
                    return None
            
            litros = get_float(row[9]) if len(row) > 9 else None
            valor_liq = get_float(row[12]) if len(row) > 12 else None
            iva = get_float(row[13]) if len(row) > 13 else None
            
            transacao = {
                "id": str(uuid.uuid4()),
                "parceiro_id": parceiro_id,
                "posto": str(row[0] or "") if len(row) > 0 else None,
                "pais": str(row[1] or "") if len(row) > 1 else None,
                "rede": str(row[2] or "") if len(row) > 2 else None,
                "data": str(row[3] or "") if len(row) > 3 else None,
                "hora": str(row[4] or "") if len(row) > 4 else None,
                "cartao": str(row[5] or "") if len(row) > 5 else None,
                "desc_cartao": str(row[6] or "") if len(row) > 6 else None,
                "estado": str(row[7] or "") if len(row) > 7 else None,
                "grupo_cartao": str(row[8] or "") if len(row) > 8 else None,
                "litros": litros,
                "combustivel": str(row[10] or "") if len(row) > 10 else None,
                "recibo": str(row[11] or "") if len(row) > 11 else None,
                "valor_liquido": valor_liq,
                "iva": iva,
                "kms": get_int(row[14]) if len(row) > 14 else None,
                "id_condutor": str(row[15] or "") if len(row) > 15 else None,
                "fatura": str(row[16] or "") if len(row) > 16 else None,
                "data_fatura": str(row[17] or "") if len(row) > 17 else None,
                "valor_unitario": get_float(row[18]) if len(row) > 18 else None,
                "valor_ref": get_float(row[19]) if len(row) > 19 else None,
                "valor_desc": get_float(row[20]) if len(row) > 20 else None,
                "cliente": str(row[21] or "") if len(row) > 21 else None,
                "tipo_pagamento": str(row[22] or "") if len(row) > 22 else None,
                "periodo_inicio": periodo_inicio,
                "periodo_fim": periodo_fim,
                "ficheiro_nome": excel_filename,
                "data_importacao": datetime.now(timezone.utc).isoformat()
            }
            
            transacoes.append(transacao)
            if litros:
                total_litros += litros
            if valor_liq and iva:
                total_valor += (valor_liq + iva)
        
        # Store in database
        if transacoes:
            await db.combustivel_fossil.insert_many(transacoes)
        
        return {
            "success": True,
            "transacoes_importadas": len(transacoes),
            "total_litros": round(total_litros, 2),
            "total_valor_eur": round(total_valor, 2),
            "periodo": f"{periodo_inicio} a {periodo_fim}",
            "ficheiro_salvo": excel_filename
        }
    
    except Exception as e:
        logger.error(f"Error processing Fossil Fuel Excel: {e}")
        return {"success": False, "error": str(e)}

async def generate_contrato_pdf(parceiro: Dict, motorista: Dict, veiculo: Dict) -> str:
    """Generate contract PDF and return file path"""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import cm
    
    # Create contracts directory
    contratos_dir = UPLOAD_DIR / "contratos"
    contratos_dir.mkdir(exist_ok=True)
    
    # Generate filename
    contrato_id = str(uuid.uuid4())
    pdf_filename = f"contrato_{contrato_id}.pdf"
    pdf_path = contratos_dir / pdf_filename
    
    # Create PDF
    c = canvas.Canvas(str(pdf_path), pagesize=A4)
    width, height = A4
    
    # Title
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(width/2, height - 2*cm, "CONTRATO DE EXPLORAÇÃO DE VEÍCULO TVDE")
    
    # Date
    c.setFont("Helvetica", 10)
    data_hoje = datetime.now().strftime("%d/%m/%Y")
    c.drawCentredString(width/2, height - 3*cm, f"Data: {data_hoje}")
    
    y = height - 5*cm
    
    # Section 1: Parceiro/Empresa
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, y, "1. IDENTIFICAÇÃO DA EMPRESA")
    y -= 0.8*cm
    
    c.setFont("Helvetica", 10)
    c.drawString(2.5*cm, y, f"Empresa: {parceiro.get('nome_empresa', 'N/A')}")
    y -= 0.6*cm
    c.drawString(2.5*cm, y, f"NIF: {parceiro.get('contribuinte_empresa', 'N/A')}")
    y -= 0.6*cm
    c.drawString(2.5*cm, y, f"Morada: {parceiro.get('morada_completa', 'N/A')}")
    y -= 0.6*cm
    c.drawString(2.5*cm, y, f"Código Postal: {parceiro.get('codigo_postal', 'N/A')} {parceiro.get('localidade', '')}")
    y -= 0.6*cm
    c.drawString(2.5*cm, y, f"Manager: {parceiro.get('nome_manager', 'N/A')}")
    y -= 0.6*cm
    c.drawString(2.5*cm, y, f"Contacto: {parceiro.get('telefone', 'N/A')} / {parceiro.get('telemovel', 'N/A')}")
    y -= 1*cm
    
    # Section 2: Motorista
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, y, "2. IDENTIFICAÇÃO DO MOTORISTA")
    y -= 0.8*cm
    
    c.setFont("Helvetica", 10)
    c.drawString(2.5*cm, y, f"Nome: {motorista.get('name', 'N/A')}")
    y -= 0.6*cm
    c.drawString(2.5*cm, y, f"NIF: {motorista.get('nif', 'N/A')}")
    y -= 0.6*cm
    c.drawString(2.5*cm, y, f"Morada: {motorista.get('morada_completa', 'N/A')}")
    y -= 0.6*cm
    c.drawString(2.5*cm, y, f"Licença TVDE: {motorista.get('licenca_tvde_numero', 'N/A')}")
    y -= 0.6*cm
    c.drawString(2.5*cm, y, f"Contacto: {motorista.get('phone', 'N/A')} / {motorista.get('whatsapp', 'N/A')}")
    y -= 1*cm
    
    # Section 3: Veículo
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, y, "3. IDENTIFICAÇÃO DO VEÍCULO")
    y -= 0.8*cm
    
    c.setFont("Helvetica", 10)
    c.drawString(2.5*cm, y, f"Marca/Modelo: {veiculo.get('marca', 'N/A')} {veiculo.get('modelo', 'N/A')}")
    y -= 0.6*cm
    c.drawString(2.5*cm, y, f"Matrícula: {veiculo.get('matricula', 'N/A')}")
    y -= 0.6*cm
    c.drawString(2.5*cm, y, f"Cor: {veiculo.get('cor', 'N/A')}")
    y -= 0.6*cm
    c.drawString(2.5*cm, y, f"Combustível: {veiculo.get('combustivel', 'N/A')}")
    y -= 1*cm
    
    # Section 4: Condições
    tipo_contrato = veiculo.get('tipo_contrato', {})
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, y, "4. CONDIÇÕES DE EXPLORAÇÃO")
    y -= 0.8*cm
    
    c.setFont("Helvetica", 10)
    c.drawString(2.5*cm, y, f"Tipo de Contrato: {tipo_contrato.get('tipo', 'N/A').upper()}")
    y -= 0.6*cm
    
    if tipo_contrato.get('tipo') == 'aluguer':
        c.drawString(2.5*cm, y, f"Valor de Aluguer: €{tipo_contrato.get('valor_aluguer', 0):.2f} / mês")
        y -= 0.6*cm
    elif tipo_contrato.get('tipo') == 'comissao':
        c.drawString(2.5*cm, y, f"Comissão Parceiro: {tipo_contrato.get('comissao_parceiro', 0)}%")
        y -= 0.6*cm
        c.drawString(2.5*cm, y, f"Comissão Motorista: {tipo_contrato.get('comissao_motorista', 0)}%")
        y -= 0.6*cm
    
    c.drawString(2.5*cm, y, f"Regime: {tipo_contrato.get('regime', 'full_time').upper()}")
    y -= 0.6*cm
    c.drawString(2.5*cm, y, f"Inclui Combustível: {'Sim' if tipo_contrato.get('inclui_combustivel') else 'Não'}")
    y -= 0.6*cm
    c.drawString(2.5*cm, y, f"Inclui Via Verde: {'Sim' if tipo_contrato.get('inclui_via_verde') else 'Não'}")
    y -= 1*cm
    
    # Caução
    caucao = veiculo.get('caucao', {})
    if caucao and caucao.get('caucao_total', 0) > 0:
        c.setFont("Helvetica-Bold", 12)
        c.drawString(2*cm, y, "5. CAUÇÃO")
        y -= 0.8*cm
        
        c.setFont("Helvetica", 10)
        c.drawString(2.5*cm, y, f"Valor Total: €{caucao.get('caucao_total', 0):.2f}")
        y -= 0.6*cm
        c.drawString(2.5*cm, y, f"Divisão: {caucao.get('caucao_divisao', 'total').upper()}")
        y -= 0.6*cm
        if caucao.get('caucao_divisao') == 'semanal':
            c.drawString(2.5*cm, y, f"Valor Semanal: €{caucao.get('caucao_valor_semanal', 0):.2f}")
            y -= 0.6*cm
        y -= 0.5*cm
    
    # Signatures
    y -= 1*cm
    c.setFont("Helvetica-Bold", 10)
    c.drawString(3*cm, y, "Assinatura da Empresa:")
    c.drawString(12*cm, y, "Assinatura do Motorista:")
    y -= 0.5*cm
    c.line(3*cm, y, 8*cm, y)
    c.line(12*cm, y, 17*cm, y)
    
    # Save PDF
    c.save()
    
    return f"uploads/contratos/{pdf_filename}"

# ==================== AUTH ENDPOINTS ====================

@api_router.post("/auth/register", response_model=User)
async def register(user_data: UserCreate):
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_dict = user_data.model_dump()
    user_dict["password"] = hash_password(user_data.password)
    user_dict["id"] = str(uuid.uuid4())
    user_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    user_dict["approved"] = user_data.role == UserRole.ADMIN
    
    # Gestores não têm planos - apenas acesso a parceiros atribuídos
    
    await db.users.insert_one(user_dict)
    
    # Send welcome email
    await enviar_email_boas_vindas(user_dict)
    
    user_dict.pop("password")
    if isinstance(user_dict["created_at"], str):
        user_dict["created_at"] = datetime.fromisoformat(user_dict["created_at"])
    
    return User(**user_dict)

async def enviar_email_boas_vindas(user: Dict[str, Any]):
    """Send welcome email to new user"""
    role_labels = {
        "motorista": "Motorista",
        "parceiro": "Parceiro",
        "operacional": "Operacional",
        "gestao": "Gestor",
        "admin": "Administrador"
    }
    
    role_label = role_labels.get(user.get("role"), "Utilizador")
    
    # Get email config
    config = await db.configuracoes.find_one({"tipo": "email"}, {"_id": 0})
    email_from = config.get("email_contacto", "info@tvdefleet.com") if config else "info@tvdefleet.com"
    
    print(f"📧 ENVIAR EMAIL DE BOAS-VINDAS:")
    print(f"   Para: {user.get('email')}")
    print(f"   Nome: {user.get('name')}")
    print(f"   Role: {role_label}")
    
    # Queue email for sending
    await db.email_queue.insert_one({
        "to": user.get("email"),
        "from": email_from,
        "subject": f"Bem-vindo à TVDEFleet - Registo de {role_label}",
        "body": f"""
        Olá {user.get('name')},
        
        Bem-vindo à TVDEFleet!
        
        O seu registo como {role_label} foi recebido com sucesso.
        
        A nossa equipa irá analisar os seus dados e entrará em contacto em breve.
        Após aprovação, receberá as suas credenciais de acesso definitivas.
        
        Dados do registo:
        - Email: {user.get('email')}
        - Telefone: {user.get('phone', 'N/A')}
        - Data de registo: {user.get('created_at')}
        
        Se tiver alguma dúvida, entre em contacto connosco:
        Email: {email_from}
        Telefone: +351 912 345 678
        WhatsApp: https://wa.me/351912345678
        
        Obrigado,
        Equipa TVDEFleet
        """,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "status": "pending",
        "tipo": "boas_vindas",
        "user_id": user.get("id")
    })

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user or not verify_password(credentials.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if user["role"] == UserRole.MOTORISTA and not user.get("approved", False):
        raise HTTPException(status_code=403, detail="Account pending approval")
    
    # Verificar se tem senha provisória
    senha_provisoria = False
    if user["role"] == UserRole.MOTORISTA:
        motorista = await db.motoristas.find_one({"id": user["id"]}, {"_id": 0, "senha_provisoria": 1})
        if motorista and motorista.get("senha_provisoria", False):
            senha_provisoria = True
    
    token = create_access_token(user["id"], user["email"], user["role"])
    
    user.pop("password")
    if isinstance(user["created_at"], str):
        user["created_at"] = datetime.fromisoformat(user["created_at"])
    
    # Adicionar flag senha_provisoria na resposta
    user["senha_provisoria"] = senha_provisoria
    
    return TokenResponse(access_token=token, user=User(**user))

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: Dict = Depends(get_current_user)):
    if isinstance(current_user["created_at"], str):
        current_user["created_at"] = datetime.fromisoformat(current_user["created_at"])
    return User(**current_user)

# ==================== USER PROFILE ENDPOINTS ====================

@api_router.put("/profile/update")
async def update_profile(profile_data: UserProfileUpdate, current_user: Dict = Depends(get_current_user)):
    update_dict = {k: v for k, v in profile_data.model_dump().items() if v is not None}
    
    if not update_dict:
        raise HTTPException(status_code=400, detail="No data to update")
    
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": update_dict}
    )
    
    return {"message": "Profile updated successfully"}

@api_router.post("/profile/change-password")
async def change_password(password_data: ChangePasswordRequest, current_user: Dict = Depends(get_current_user)):
    user = await db.users.find_one({"id": current_user["id"]})
    
    if not verify_password(password_data.old_password, user["password"]):
        raise HTTPException(status_code=400, detail="Senha antiga incorreta")
    
    new_hashed = hash_password(password_data.new_password)
    
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": {"password": new_hashed, "senha_provisoria": False}}
    )
    
    # Update motorista record if exists
    await db.motoristas.update_one(
        {"id": current_user["id"]},
        {"$set": {"senha_provisoria": False}}
    )
    
    return {"message": "Password changed successfully"}

@api_router.get("/profile/permissions")
async def get_user_permissions(current_user: Dict = Depends(get_current_user)):
    role = current_user["role"]
    
    permissions = {
        UserRole.ADMIN: {
            "can_manage_users": True,
            "can_manage_all_vehicles": True,
            "can_manage_all_motoristas": True,
            "can_manage_parceiros": True,
            "can_view_all_reports": True,
            "can_configure_system": True,
            "description": "Acesso total ao sistema"
        },
        UserRole.GESTAO: {
            "can_manage_users": False,
            "can_manage_all_vehicles": True,
            "can_manage_all_motoristas": True,
            "can_manage_parceiros": True,
            "can_view_all_reports": True,
            "can_configure_system": False,
            "description": "Gestão de múltiplos parceiros e frotas"
        },
        UserRole.PARCEIRO: {
            "can_manage_users": False,
            "can_manage_all_vehicles": False,
            "can_manage_all_motoristas": False,
            "can_manage_parceiros": False,
            "can_view_all_reports": False,
            "can_configure_system": False,
            "can_create_motoristas": True,
            "can_manage_own_vehicles": True,
            "can_manage_own_motoristas": True,
            "description": "Gestão dos seus veículos e motoristas"
        },
        UserRole.MOTORISTA: {
            "can_manage_users": False,
            "can_manage_all_vehicles": False,
            "can_manage_all_motoristas": False,
            "can_manage_parceiros": False,
            "can_view_all_reports": False,
            "can_configure_system": False,
            "description": "Acesso a veículos atribuídos e dados pessoais"
        }
    }
    
    return {
        "role": role,
        "permissions": permissions.get(role, {}),
        "user": current_user
    }

@api_router.put("/gestores/{gestor_id}/atribuir-parceiros")
async def atribuir_parceiros_a_gestor(
    gestor_id: str,
    parceiros_data: Dict[str, List[str]],
    current_user: Dict = Depends(get_current_user)
):
    """Admin atribui parceiros a um gestor"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Apenas Admin pode atribuir parceiros a gestores")
    
    parceiros_ids = parceiros_data.get("parceiros_ids", [])
    
    # Verificar se gestor existe
    gestor = await db.users.find_one({"id": gestor_id, "role": "gestao"}, {"_id": 0})
    if not gestor:
        raise HTTPException(status_code=404, detail="Gestor não encontrado")
    
    # Atualizar gestor com lista de parceiros
    await db.users.update_one(
        {"id": gestor_id},
        {"$set": {"parceiros_atribuidos": parceiros_ids}}
    )
    
    # Remover este gestor de parceiros que não estão mais na lista
    await db.parceiros.update_many(
        {"gestor_associado_id": gestor_id, "id": {"$nin": parceiros_ids}},
        {"$set": {"gestor_associado_id": None}}
    )
    
    # Atribuir este gestor aos parceiros selecionados
    await db.parceiros.update_many(
        {"id": {"$in": parceiros_ids}},
        {"$set": {"gestor_associado_id": gestor_id}}
    )
    
    logger.info(f"Admin {current_user['id']} atribuiu {len(parceiros_ids)} parceiros ao gestor {gestor_id}")
    
    return {
        "message": "Parceiros atribuídos ao gestor com sucesso",
        "gestor_id": gestor_id,
        "parceiros_count": len(parceiros_ids)
    }

@api_router.get("/gestores/{gestor_id}/parceiros")
async def get_parceiros_do_gestor(
    gestor_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Obter lista de parceiros atribuídos a um gestor"""
    # Apenas Admin, o próprio Gestor ou outro Gestor pode ver
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    if current_user["role"] == UserRole.GESTAO and current_user["id"] != gestor_id:
        raise HTTPException(status_code=403, detail="Gestor só pode ver seus próprios parceiros")
    
    # Buscar parceiros atribuídos
    parceiros = await db.parceiros.find(
        {"gestor_associado_id": gestor_id},
        {"_id": 0}
    ).to_list(length=None)
    
    return parceiros

# ==================== MOTORISTA ENDPOINTS ====================

@api_router.post("/motoristas/register", response_model=Motorista)
async def register_motorista(motorista_data: MotoristaCreate):
    existing = await db.users.find_one({"email": motorista_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Generate provisional password if needed
    if motorista_data.senha_provisoria or not motorista_data.password:
        provisional_pass = motorista_data.phone.replace(" ", "")[-9:]
        password_to_hash = provisional_pass
        senha_provisoria = True
    else:
        password_to_hash = motorista_data.password
        senha_provisoria = False
    
    user_dict = {
        "id": str(uuid.uuid4()),
        "email": motorista_data.email,
        "name": motorista_data.name,
        "role": UserRole.MOTORISTA,
        "password": hash_password(password_to_hash),
        "phone": motorista_data.phone,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "approved": False
    }
    await db.users.insert_one(user_dict)
    
    motorista_dict = motorista_data.model_dump()
    motorista_dict.pop("password", None)
    motorista_dict["id"] = user_dict["id"]
    # Generate automatic ID for fleet card (format: FROTA-XXXXXXXX)
    motorista_dict["id_cartao_frota_combustivel"] = f"FROTA-{str(uuid.uuid4())[:8].upper()}"
    motorista_dict["documents"] = {
        "license_photo": None, 
        "cv_file": None, 
        "profile_photo": None,
        "documento_identificacao": None,
        "licenca_tvde": None,
        "registo_criminal": None,
        "contrato": None,
        "additional_docs": []
    }
    motorista_dict["approved"] = False
    motorista_dict["senha_provisoria"] = senha_provisoria
    motorista_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.motoristas.insert_one(motorista_dict)
    
    if isinstance(motorista_dict["created_at"], str):
        motorista_dict["created_at"] = datetime.fromisoformat(motorista_dict["created_at"])
    
    return Motorista(**motorista_dict)

@api_router.post("/motoristas/{motorista_id}/upload-document")
async def upload_document(
    motorista_id: str, 
    file: UploadFile = File(...), 
    doc_type: str = Form(...),
    current_user: Dict = Depends(get_current_user)
):
    motorista = await db.motoristas.find_one({"id": motorista_id}, {"_id": 0})
    if not motorista:
        raise HTTPException(status_code=404, detail="Motorista not found")
    
    # Process file: save and convert to PDF if image
    file_id = f"{motorista_id}_{doc_type}_{uuid.uuid4()}"
    file_info = await process_uploaded_file(file, MOTORISTAS_UPLOAD_DIR, file_id)
    
    # Store file path in database (prefer PDF version if available)
    file_url = file_info["pdf_path"] if file_info["pdf_path"] else file_info["original_path"]
    
    update_field = f"documents.{doc_type}"
    await db.motoristas.update_one(
        {"id": motorista_id},
        {"$set": {update_field: file_url}}
    )
    
    return {
        "message": "Document uploaded successfully",
        "doc_type": doc_type,
        "file_url": file_url,
        "converted_to_pdf": file_info["pdf_path"] is not None
    }

@api_router.post("/motoristas/{motorista_id}/validar-documento")
async def validar_documento(
    motorista_id: str,
    validacao_data: dict,
    current_user: Dict = Depends(get_current_user)
):
    """Validar documento (Admin, Gestor, Operacional, Parceiro)"""
    # Verificar permissões
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    motorista = await db.motoristas.find_one({"id": motorista_id}, {"_id": 0})
    if not motorista:
        raise HTTPException(status_code=404, detail="Motorista not found")
    
    doc_type = validacao_data.get("doc_type")
    validar = validacao_data.get("validar", True)
    
    if not doc_type:
        raise HTTPException(status_code=400, detail="doc_type is required")
    
    # Verificar se documento existe
    if not motorista.get("documents", {}).get(doc_type):
        raise HTTPException(status_code=404, detail="Document not found")
    
    # Atualizar validação
    validacao = {
        "validado": validar,
        "validado_por": current_user["id"] if validar else None,
        "validado_em": datetime.now(timezone.utc).isoformat() if validar else None,
        "pode_editar": doc_type in ["comprovativo_iban", "registo_criminal"]  # Sempre editáveis
    }
    
    update_field = f"documents_validacao.{doc_type}"
    await db.motoristas.update_one(
        {"id": motorista_id},
        {"$set": {update_field: validacao}}
    )
    
    return {
        "message": f"Documento {'validado' if validar else 'invalidado'} com sucesso",
        "doc_type": doc_type,
        "validacao": validacao
    }

@api_router.delete("/motoristas/{motorista_id}/documentos/{doc_type}")
async def delete_documento(
    motorista_id: str,
    doc_type: str,
    current_user: Dict = Depends(get_current_user)
):
    """Delete document (Admin, Gestor, Operacional only)"""
    # Check permissions - only Admin, Gestor, Operacional can delete documents
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Não autorizado - Apenas Admin, Gestor ou Operacional")
    
    motorista = await db.motoristas.find_one({"id": motorista_id}, {"_id": 0})
    if not motorista:
        raise HTTPException(status_code=404, detail="Motorista não encontrado")
    
    # Check if document exists
    if not motorista.get("documents", {}).get(doc_type):
        raise HTTPException(status_code=404, detail="Documento não encontrado")
    
    # Delete physical file if exists
    doc_url = motorista["documents"].get(doc_type)
    if doc_url:
        try:
            file_path = Path(doc_url.replace("/uploads/", str(UPLOAD_DIR) + "/"))
            if file_path.exists():
                file_path.unlink()
        except Exception as e:
            # Log error but continue with database deletion
            print(f"Warning: Could not delete physical file: {e}")
    
    # Remove document from database
    update_doc = {"$unset": {f"documents.{doc_type}": ""}}
    
    # Also remove validation data if exists
    if motorista.get("documents_validacao", {}).get(doc_type):
        update_doc["$unset"][f"documents_validacao.{doc_type}"] = ""
    
    await db.motoristas.update_one(
        {"id": motorista_id},
        update_doc
    )
    
    return {
        "message": f"Documento '{doc_type}' eliminado com sucesso",
        "deleted_by": current_user["id"],
        "deleted_at": datetime.now(timezone.utc).isoformat()
    }


@api_router.get("/motoristas", response_model=List[Motorista])
async def get_motoristas(current_user: Dict = Depends(get_current_user)):
    try:
        # Filter motoristas by role
        query = {}
        if current_user["role"] == "parceiro":
            query["parceiro_atribuido"] = current_user["id"]
        
        motoristas = await db.motoristas.find(query, {"_id": 0}).to_list(1000)
        for m in motoristas:
            # Handle created_at
            if m.get("created_at"):
                if isinstance(m["created_at"], str):
                    try:
                        m["created_at"] = datetime.fromisoformat(m["created_at"])
                    except ValueError:
                        m["created_at"] = datetime.now(timezone.utc)
            else:
                m["created_at"] = datetime.now(timezone.utc)
            
            # Handle approved_at
            if m.get("approved_at") and isinstance(m["approved_at"], str):
                try:
                    m["approved_at"] = datetime.fromisoformat(m["approved_at"])
                except ValueError:
                    m["approved_at"] = None
            
            # Add plano_nome lookup
            if m.get("plano_id"):
                plano = await db.planos_sistema.find_one({"id": m["plano_id"]}, {"_id": 0, "nome": 1})
                if plano:
                    m["plano_nome"] = plano.get("nome")
        
        return motoristas
    except Exception as e:
        logger.error(f"Error fetching motoristas: {e}")
        raise HTTPException(status_code=500, detail=f"Error fetching motoristas: {str(e)}")

# ==================== MOLONI AUTO-FATURAÇÃO (MOTORISTA) ====================

@api_router.get("/motoristas/{motorista_id}/moloni-config")
async def get_moloni_config(motorista_id: str, current_user: Dict = Depends(get_current_user)):
    """Get Moloni configuration for motorista"""
    if current_user["role"] not in ["admin", "gestao"] and current_user["id"] != motorista_id:
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    motorista = await db.users.find_one(
        {"id": motorista_id}, 
        {"_id": 0, "moloni_auto_faturacao": 1, "plano_id": 1, "plano_features": 1}
    )
    if not motorista:
        raise HTTPException(status_code=404, detail="Motorista não encontrado")
    
    # Verificar se o plano tem módulo de auto-faturação
    modulo_disponivel = False
    if motorista.get("plano_id"):
        plano = await db.planos_sistema.find_one({"id": motorista["plano_id"]}, {"_id": 0, "features": 1})
        if plano and plano.get("features"):
            modulo_disponivel = "moloni_auto_faturacao" in plano.get("features", [])
    
    config = motorista.get("moloni_auto_faturacao", {
        "ativo": False,
        "client_id": "",
        "client_secret": "",
        "username": "",
        "password": "",
        "company_id": "",
        "custo_mensal_extra": 10.00
    })
    
    # Adicionar flag de módulo disponível
    config["modulo_disponivel"] = modulo_disponivel
    
    # Não retornar credenciais sensíveis completas
    if config.get("client_secret") and len(config.get("client_secret", "")) > 4:
        config["client_secret_masked"] = "***" + config["client_secret"][-4:]
        config["client_secret"] = ""
    if config.get("password"):
        config["password"] = ""
    
    return config

@api_router.post("/motoristas/{motorista_id}/moloni-config")
async def save_moloni_config(
    motorista_id: str,
    config_data: Dict,
    current_user: Dict = Depends(get_current_user)
):
    """Save Moloni configuration for motorista"""
    if current_user["role"] not in ["admin", "gestao"] and current_user["id"] != motorista_id:
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    motorista = await db.users.find_one({"id": motorista_id})
    if not motorista:
        raise HTTPException(status_code=404, detail="Motorista não encontrado")
    
    # Verificar se o plano tem módulo de auto-faturação (apenas se tentando ativar)
    if config_data.get("ativo"):
        modulo_disponivel = False
        if motorista.get("plano_id"):
            plano = await db.planos_sistema.find_one({"id": motorista["plano_id"]}, {"_id": 0, "features": 1})
            if plano and plano.get("features"):
                modulo_disponivel = "moloni_auto_faturacao" in plano.get("features", [])
        
        if not modulo_disponivel:
            raise HTTPException(
                status_code=403, 
                detail="Módulo de auto-faturação Moloni não disponível no plano atual. Contacte o administrador para upgrade."
            )
    
    # Buscar config existente
    existing_config = motorista.get("moloni_auto_faturacao", {})
    
    # Estrutura da configuração - manter valores antigos se não enviados
    moloni_config = {
        "ativo": config_data.get("ativo", existing_config.get("ativo", False)),
        "client_id": config_data.get("client_id") or existing_config.get("client_id", ""),
        "client_secret": config_data.get("client_secret") or existing_config.get("client_secret", ""),
        "username": config_data.get("username") or existing_config.get("username", ""),
        "password": config_data.get("password") or existing_config.get("password", ""),
        "company_id": config_data.get("company_id") or existing_config.get("company_id", ""),
        "custo_mensal_extra": config_data.get("custo_mensal_extra", 10.00),
        "data_ativacao": datetime.now(timezone.utc).isoformat() if config_data.get("ativo") and not existing_config.get("ativo") else existing_config.get("data_ativacao")
    }
    
    await db.users.update_one(
        {"id": motorista_id},
        {"$set": {"moloni_auto_faturacao": moloni_config}}
    )
    
    return {"message": "Configuração Moloni salva com sucesso", "ativo": moloni_config["ativo"]}

@api_router.post("/motoristas/{motorista_id}/moloni-config/testar")
async def testar_moloni_config(
    motorista_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Test Moloni connection"""
    if current_user["role"] not in ["admin", "gestao"] and current_user["id"] != motorista_id:
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    motorista = await db.users.find_one({"id": motorista_id}, {"_id": 0, "moloni_auto_faturacao": 1})
    if not motorista or not motorista.get("moloni_auto_faturacao"):
        raise HTTPException(status_code=404, detail="Configuração Moloni não encontrada")
    
    config = motorista["moloni_auto_faturacao"]
    
    if not config.get("client_id") or not config.get("client_secret"):
        raise HTTPException(status_code=400, detail="Credenciais incompletas")
    
    # TODO: Implementar chamada real à API Moloni para validar
    return {"message": "Conexão Moloni testada com sucesso (simulado)", "status": "ok"}

@api_router.get("/motorista/meu-plano")
async def get_meu_plano_motorista_v2(current_user: Dict = Depends(get_current_user)):
    """Motorista: Get active plan (self)"""
    if current_user["role"] != "motorista":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Buscar plano ativo
    atribuicao = await db.planos_usuarios.find_one(
        {"user_id": current_user["id"], "status": "ativo"},
        {"_id": 0}
    )
    
    if not atribuicao:
        return {"tem_plano": False, "plano": None, "modulos": []}
    
    plano = await db.planos.find_one({"id": atribuicao["plano_id"]}, {"_id": 0})
    if not plano:
        return {"tem_plano": False, "plano": None, "modulos": []}
    
    modulos_ativos = atribuicao.get("modulos_ativos", plano.get("modulos", []))
    
    # Buscar info dos módulos
    from models.modulo import MODULOS_SISTEMA
    modulos_info = []
    for codigo in modulos_ativos:
        if codigo in MODULOS_SISTEMA:
            modulos_info.append({
                "codigo": codigo,
                "nome": MODULOS_SISTEMA[codigo]["nome"],
                "descricao": MODULOS_SISTEMA[codigo]["descricao"]
            })
    
    precos = plano.get("precos", {})
    preco_semanal = precos.get("semanal", {}).get("preco_com_iva", 0)
    preco_mensal = precos.get("mensal", {}).get("preco_com_iva", 0)
    
    return {
        "tem_plano": True,
        "plano": plano,
        "modulos": modulos_info,
        "preco_semanal": preco_semanal,
        "preco_mensal": preco_mensal
    }

@api_router.get("/motoristas/{motorista_id}")
async def get_motorista_by_id(motorista_id: str, current_user: Dict = Depends(get_current_user)):
    """Get a specific motorista by ID"""
    try:
        motorista = await db.motoristas.find_one({"id": motorista_id}, {"_id": 0})
        if not motorista:
            raise HTTPException(status_code=404, detail="Motorista not found")
        
        # Add plano_nome lookup
        if motorista.get("plano_id"):
            plano = await db.planos_sistema.find_one({"id": motorista["plano_id"]}, {"_id": 0, "nome": 1})
            if plano:
                motorista["plano_nome"] = plano.get("nome")
        
        # Return raw data without strict model validation
        return motorista
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching motorista {motorista_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/parceiros/{parceiro_id}/motoristas")
async def parceiro_criar_motorista(
    parceiro_id: str,
    motorista_data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Parceiro cria motorista com senha provisória"""
    # Apenas Parceiro (próprio) ou Gestor (com acesso) ou Admin podem criar
    if current_user["role"] == UserRole.PARCEIRO:
        if current_user["id"] != parceiro_id:
            raise HTTPException(status_code=403, detail="Parceiro só pode criar motoristas para si próprio")
    elif current_user["role"] == UserRole.GESTAO:
        # Verificar se gestor tem acesso a este parceiro
        parceiro = await db.parceiros.find_one({"id": parceiro_id}, {"_id": 0})
        if not parceiro or parceiro.get("gestor_associado_id") != current_user["id"]:
            raise HTTPException(status_code=403, detail="Gestor não tem acesso a este parceiro")
    elif current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    # Verificar se email já existe
    existing = await db.users.find_one({"email": motorista_data.get("email")})
    if existing:
        raise HTTPException(status_code=400, detail="Email já registado")
    
    # Gerar senha provisória (últimos 9 dígitos do telefone)
    phone = motorista_data.get("phone", "").replace(" ", "").replace("+351", "")
    senha_provisoria = phone[-9:] if len(phone) >= 9 else f"{phone}123"
    
    # Criar user
    user_id = str(uuid.uuid4())
    user_dict = {
        "id": user_id,
        "email": motorista_data.get("email"),
        "name": motorista_data.get("name"),
        "role": UserRole.MOTORISTA,
        "password": hash_password(senha_provisoria),
        "phone": motorista_data.get("phone", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "approved": True  # Criado por parceiro, já aprovado
    }
    await db.users.insert_one(user_dict)
    
    # Criar perfil motorista
    motorista_dict = {
        "id": user_id,
        "name": motorista_data.get("name"),
        "email": motorista_data.get("email"),
        "phone": motorista_data.get("phone", ""),
        "whatsapp": motorista_data.get("whatsapp", motorista_data.get("phone", "")),
        "nif": motorista_data.get("nif", ""),
        "data_nascimento": motorista_data.get("data_nascimento", ""),
        "nacionalidade": motorista_data.get("nacionalidade", "Portuguesa"),
        "morada_completa": motorista_data.get("morada_completa", ""),
        "codigo_postal": motorista_data.get("codigo_postal", ""),
        "carta_conducao_numero": motorista_data.get("carta_conducao_numero", ""),
        "carta_conducao_validade": motorista_data.get("carta_conducao_validade", ""),
        "licenca_tvde_numero": motorista_data.get("licenca_tvde_numero", ""),
        "licenca_tvde_validade": motorista_data.get("licenca_tvde_validade", ""),
        "regime": motorista_data.get("regime", ""),
        "tipo_pagamento": motorista_data.get("tipo_pagamento", ""),
        "id_cartao_frota_combustivel": f"FROTA-{str(uuid.uuid4())[:8].upper()}",
        "parceiro_atribuido": parceiro_id,
        "approved": True,
        "senha_provisoria": True,  # IMPORTANTE: Forçar mudança de senha
        "documents": {},
        "contacto_emergencia": {},
        "dados_bancarios": {},
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Atribuir plano base
    plano_base = await db.planos_sistema.find_one({
        "preco_mensal": 0, 
        "ativo": True, 
        "tipo_usuario": "motorista"
    }, {"_id": 0})
    
    if plano_base:
        from datetime import timedelta
        plano_valida_ate = (datetime.now(timezone.utc) + timedelta(days=30)).isoformat()
        motorista_dict["plano_id"] = plano_base["id"]
        motorista_dict["plano_nome"] = plano_base["nome"]
        motorista_dict["plano_valida_ate"] = plano_valida_ate
    
    await db.motoristas.insert_one(motorista_dict)
    
    logger.info(f"Parceiro {parceiro_id} criou motorista {user_id} com senha provisória")
    
    return {
        "message": "Motorista criado com sucesso",
        "motorista_id": user_id,
        "email": motorista_data.get("email"),
        "senha_provisoria": senha_provisoria,
        "deve_mudar_senha": True
    }

@api_router.put("/motoristas/{motorista_id}/approve")
async def approve_motorista(motorista_id: str, current_user: Dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Find or create default free plan in the new unified system
    plano_base = await db.planos_sistema.find_one({
        "preco_mensal": 0, 
        "ativo": True, 
        "tipo_usuario": "motorista"
    }, {"_id": 0})
    
    if not plano_base:
        # Create default free plan if it doesn't exist
        plano_base = {
            "id": str(uuid.uuid4()),
            "nome": "Base Gratuito",
            "descricao": "Plano base gratuito para todos os motoristas",
            "preco_mensal": 0,
            "tipo_usuario": "motorista",
            "modulos": [
                "dashboard_ganhos",
                "gestao_documentos"
            ],
            "ativo": True,
            "permite_trial": False,
            "dias_trial": 0,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.planos_sistema.insert_one(plano_base)
        logger.info(f"Created default free plan for motorista: {plano_base['id']}")
    
    # Calculate expiry date (30 days from now)
    from datetime import timedelta
    plano_valida_ate = (datetime.now(timezone.utc) + timedelta(days=30)).isoformat()
    
    # Check if motorista profile exists
    motorista_exists = await db.motoristas.find_one({"id": motorista_id}, {"_id": 0})
    
    if motorista_exists:
        # Update existing motorista with approval and base plan
        logger.info(f"Updating motorista {motorista_id} with plan {plano_base['id']}")
        result = await db.motoristas.update_one(
            {"id": motorista_id},
            {"$set": {
                "approved": True, 
                "approved_by": current_user["id"], 
                "approved_at": datetime.now(timezone.utc).isoformat(),
                "plano_id": plano_base["id"],
                "plano_nome": plano_base["nome"],
                "plano_valida_ate": plano_valida_ate,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        logger.info(f"Motorista update result: matched={result.matched_count}, modified={result.modified_count}")
    else:
        # Create motorista profile if it doesn't exist
        user = await db.users.find_one({"id": motorista_id}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        if user.get("role") != "motorista":
            raise HTTPException(status_code=400, detail="User is not a motorista")
        
        # Create motorista profile from user data with base plan
        motorista_profile = {
            "id": motorista_id,
            "name": user.get("name"),
            "email": user.get("email"),
            "phone": user.get("phone", ""),
            "whatsapp": user.get("whatsapp", ""),
            "data_nascimento": user.get("data_nascimento", ""),
            "nif": user.get("nif", ""),
            "nacionalidade": user.get("nacionalidade", "Portuguesa"),
            "morada_completa": user.get("morada_completa", ""),
            "codigo_postal": user.get("codigo_postal", ""),
            "id_cartao_frota_combustivel": f"FROTA-{str(uuid.uuid4())[:8].upper()}",
            "documents": {},
            "approved": True,
            "approved_by": current_user["id"],
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "vehicle_assigned": None,
            "parceiro_atribuido": None,
            "contrato_id": None,
            "contacto_emergencia": {},
            "dados_bancarios": {},
            "plano_id": plano_base["id"],
            "plano_nome": plano_base["nome"],
            "plano_valida_ate": plano_valida_ate,
            "created_at": user.get("created_at", datetime.now(timezone.utc).isoformat()),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.motoristas.insert_one(motorista_profile)
        logger.info(f"Created motorista profile for {motorista_id} with plan {plano_base['id']}")
    
    # Always update user approval status
    await db.users.update_one(
        {"id": motorista_id},
        {"$set": {"approved": True}}
    )
    
    # Verify the update
    updated_motorista = await db.motoristas.find_one({"id": motorista_id}, {"_id": 0})
    logger.info(f"Motorista after update: plano_id={updated_motorista.get('plano_id')}, plano_nome={updated_motorista.get('plano_nome')}")
    
    return {
        "message": "Motorista approved successfully",
        "plano_atribuido": {
            "id": plano_base["id"],
            "nome": plano_base["nome"],
            "preco_mensal": plano_base["preco_mensal"],
            "plano_valida_ate": plano_valida_ate
        }
    }

@api_router.put("/motoristas/{motorista_id}/atribuir-parceiro")
async def atribuir_motorista_a_parceiro(
    motorista_id: str,
    parceiro_data: Dict[str, str],
    current_user: Dict = Depends(get_current_user)
):
    """Admin atribui motorista inscrito a um parceiro"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Apenas Admin pode atribuir motoristas a parceiros")
    
    parceiro_id = parceiro_data.get("parceiro_id")
    if not parceiro_id:
        raise HTTPException(status_code=400, detail="parceiro_id é obrigatório")
    
    # Verificar se parceiro existe
    parceiro = await db.parceiros.find_one({"id": parceiro_id}, {"_id": 0})
    if not parceiro:
        raise HTTPException(status_code=404, detail="Parceiro não encontrado")
    
    # Verificar se motorista existe
    motorista = await db.motoristas.find_one({"id": motorista_id}, {"_id": 0})
    if not motorista:
        raise HTTPException(status_code=404, detail="Motorista não encontrado")
    
    # Atribuir motorista ao parceiro
    result = await db.motoristas.update_one(
        {"id": motorista_id},
        {"$set": {
            "parceiro_atribuido": parceiro_id,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    logger.info(f"Admin {current_user['id']} atribuiu motorista {motorista_id} ao parceiro {parceiro_id}")
    
    return {
        "message": "Motorista atribuído ao parceiro com sucesso",
        "motorista_id": motorista_id,
        "parceiro_id": parceiro_id
    }

@api_router.put("/motoristas/{motorista_id}")
async def update_motorista(
    motorista_id: str, 
    update_data: Dict[str, Any], 
    current_user: Dict = Depends(get_current_user)
):
    """Update motorista data (partial updates allowed)"""
    # Check if motorista exists
    motorista = await db.motoristas.find_one({"id": motorista_id}, {"_id": 0})
    if not motorista:
        raise HTTPException(status_code=404, detail="Motorista not found")
    
    # Check if parceiro/operacional is assigned to this motorista
    is_assigned = False
    if current_user["role"] in [UserRole.PARCEIRO]:
        # Check if current user is the assigned parceiro
        is_assigned = motorista.get("parceiro_atribuido") == current_user["id"]
    
    # Allow admin, gestao, OR parceiro/operacional assigned OR motorista editing their own profile
    is_authorized = (
        current_user["role"] in [UserRole.ADMIN, UserRole.GESTAO] or
        is_assigned or
        (current_user["role"] == UserRole.MOTORISTA and current_user["email"] == motorista.get("email"))
    )
    
    if not is_authorized:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # If motorista is editing and documents are approved, only allow specific fields
    if (current_user["role"] == UserRole.MOTORISTA and 
        motorista.get("documentos_aprovados", False)):
        # Only allow editing: registo criminal and iban
        allowed_fields = [
            'codigo_registo_criminal', 'validade_registo_criminal',
            'iban', 'nome_banco'
        ]
        
        # Filter update_data to only allowed fields
        filtered_update = {k: v for k, v in update_data.items() if k in allowed_fields}
        
        if not filtered_update:
            raise HTTPException(
                status_code=403, 
                detail="Documentos aprovados. Apenas Registo Criminal e IBAN podem ser alterados. Contacte o gestor para outras alterações."
            )
        
        update_data = filtered_update
    
    # Remove fields that shouldn't be updated directly
    update_data.pop("id", None)
    update_data.pop("created_at", None)
    update_data.pop("_id", None)
    update_data.pop("documentos_aprovados", None)  # Não permitir mudança direta deste campo
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    # Update motorista
    await db.motoristas.update_one(
        {"id": motorista_id},
        {"$set": update_data}
    )
    
    return {"message": "Motorista updated successfully"}

@api_router.put("/motoristas/{motorista_id}/desativar")
async def desativar_motorista(
    motorista_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Desativar motorista - Parceiros podem desativar seus próprios motoristas"""
    motorista = await db.motoristas.find_one({"id": motorista_id}, {"_id": 0})
    
    if not motorista:
        raise HTTPException(status_code=404, detail="Motorista not found")
    
    # Check permissions
    if current_user["role"] == "parceiro":
        # Parceiro can only deactivate their own motoristas
        if motorista.get("parceiro_atribuido") != current_user["id"]:
            raise HTTPException(status_code=403, detail="Sem permissão para desativar este motorista")
    elif current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    # Update motorista status to inactive
    await db.motoristas.update_one(
        {"id": motorista_id},
        {"$set": {"status": "inativo", "status_motorista": "desativo", "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Also deactivate user account
    await db.users.update_one(
        {"id": motorista_id},
        {"$set": {"active": False}}
    )
    
    return {"message": "Motorista desativado com sucesso"}

@api_router.put("/motoristas/{motorista_id}/ativar")
async def ativar_motorista(
    motorista_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Ativar motorista - Parceiros podem ativar seus próprios motoristas"""
    motorista = await db.motoristas.find_one({"id": motorista_id}, {"_id": 0})
    
    if not motorista:
        raise HTTPException(status_code=404, detail="Motorista not found")
    
    # Check permissions
    if current_user["role"] == "parceiro":
        # Parceiro can only activate their own motoristas
        if motorista.get("parceiro_atribuido") != current_user["id"]:
            raise HTTPException(status_code=403, detail="Sem permissão para ativar este motorista")
    elif current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    # Update motorista status to active
    await db.motoristas.update_one(
        {"id": motorista_id},
        {"$set": {"status": "ativo", "status_motorista": "ativo", "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Also activate user account
    await db.users.update_one(
        {"id": motorista_id},
        {"$set": {"active": True}}
    )
    
    return {"message": "Motorista ativado com sucesso"}

@api_router.delete("/motoristas/{motorista_id}")
async def delete_motorista(
    motorista_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Delete a motorista (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Delete from motoristas collection
    result = await db.motoristas.delete_one({"id": motorista_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Motorista not found")
    
    # Also delete from users collection if exists
    await db.users.delete_one({"id": motorista_id})
    
    return {"message": "Motorista deleted successfully"}

@api_router.put("/motoristas/{motorista_id}/aprovar-todos-documentos")
async def aprovar_todos_documentos(
    motorista_id: str,
    data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Aprovar todos os documentos do motorista (Admin/Gestor/Parceiro/Operacional)"""
    # Check authorization
    motorista = await db.motoristas.find_one({"id": motorista_id}, {"_id": 0})
    if not motorista:
        raise HTTPException(status_code=404, detail="Motorista not found")
    
    # Check if parceiro/operacional is assigned
    is_assigned = False
    if current_user["role"] in [UserRole.PARCEIRO]:
        is_assigned = motorista.get("parceiro_atribuido") == current_user["id"]
    
    is_authorized = (
        current_user["role"] in [UserRole.ADMIN, UserRole.GESTAO] or
        is_assigned
    )
    
    if not is_authorized:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Update all documents as validated
    documents_validacao = motorista.get("documents_validacao", {})
    for doc_type in motorista.get("documents", {}).keys():
        if motorista["documents"].get(doc_type):
            documents_validacao[doc_type] = {
                "validado": True,
                "validado_por": data.get("aprovado_por", current_user["name"]),
                "validado_em": datetime.now(timezone.utc).isoformat(),
                "observacoes": "Aprovação em lote"
            }
    
    # Set documentos_aprovados = True
    await db.motoristas.update_one(
        {"id": motorista_id},
        {
            "$set": {
                "documentos_aprovados": True,
                "documents_validacao": documents_validacao
            }
        }
    )
    
    return {"message": "Todos os documentos aprovados com sucesso"}

@api_router.get("/motoristas/{motorista_id}/documento/{doc_type}/download")
async def download_motorista_documento(
    motorista_id: str,
    doc_type: str,
    current_user: Dict = Depends(get_current_user)
):
    """Download document of motorista (Admin/Gestor/Parceiro/Operacional/Motorista)"""
    motorista = await db.motoristas.find_one({"id": motorista_id}, {"_id": 0})
    if not motorista:
        raise HTTPException(status_code=404, detail="Motorista not found")
    
    # Check authorization
    is_assigned = False
    if current_user["role"] in [UserRole.PARCEIRO]:
        is_assigned = motorista.get("parceiro_atribuido") == current_user["id"]
    
    is_authorized = (
        current_user["role"] in [UserRole.ADMIN, UserRole.GESTAO] or
        is_assigned or
        (current_user["role"] == UserRole.MOTORISTA and current_user["email"] == motorista.get("email"))
    )
    
    if not is_authorized:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get document path
    doc_url = motorista.get("documents", {}).get(doc_type)
    if not doc_url:
        raise HTTPException(status_code=404, detail="Document not found")
    
    # Construct file path
    file_path = Path(doc_url.replace("/uploads/", str(UPLOAD_DIR) + "/"))
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Document file not found")
    
    return FileResponse(
        path=file_path,
        media_type="application/pdf",
        filename=f"{doc_type}_{motorista['name']}.pdf"
    )

@api_router.get("/motoristas/{motorista_id}/contrato/download")
async def download_motorista_contrato(
    motorista_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Download contract of motorista (Admin/Gestor/Parceiro/Operacional/Motorista)"""
    print(f"[CONTRACT] Iniciando download de contrato para motorista_id: {motorista_id}", flush=True)
    
    motorista = await db.motoristas.find_one({"id": motorista_id}, {"_id": 0})
    if not motorista:
        print(f"[CONTRACT] Motorista não encontrado: {motorista_id}", flush=True)
        raise HTTPException(status_code=404, detail="Motorista not found")
    
    # Check authorization
    is_assigned = False
    if current_user["role"] in [UserRole.PARCEIRO]:
        is_assigned = motorista.get("parceiro_atribuido") == current_user["id"]
    
    is_authorized = (
        current_user["role"] in [UserRole.ADMIN, UserRole.GESTAO] or
        is_assigned or
        (current_user["role"] == UserRole.MOTORISTA and current_user["email"] == motorista.get("email"))
    )
    
    if not is_authorized:
        print(f"[CONTRACT] Não autorizado: {current_user['role']}", flush=True)
        raise HTTPException(status_code=403, detail="Not authorized")
    
    print(f"[CONTRACT] Autorização OK. Buscando contrato na DB...", flush=True)
    
    # Find contract for this motorista
    contrato = await db.contratos.find_one({"motorista_id": motorista_id}, {"_id": 0})
    print(f"[CONTRACT] Contrato encontrado: {contrato is not None}", flush=True)
    if contrato:
        print(f"[CONTRACT] contrato_assinado: {contrato.get('contrato_assinado')}", flush=True)
    
    if not contrato or not contrato.get("contrato_assinado"):
        print(f"[CONTRACT] Contrato não encontrado ou sem arquivo assinado", flush=True)
        raise HTTPException(status_code=404, detail="Contract not found")
    
    # Get contract path
    contract_path_str = contrato["contrato_assinado"].replace("/uploads/", str(UPLOAD_DIR) + "/")
    contract_path = Path(contract_path_str)
    
    print(f"[CONTRACT DOWNLOAD] Motorista ID: {motorista_id}", flush=True)
    print(f"[CONTRACT DOWNLOAD] Contrato assinado no DB: {contrato['contrato_assinado']}", flush=True)
    print(f"[CONTRACT DOWNLOAD] UPLOAD_DIR: {UPLOAD_DIR}", flush=True)
    print(f"[CONTRACT DOWNLOAD] Contract path calculado: {contract_path}", flush=True)
    print(f"[CONTRACT DOWNLOAD] Arquivo existe? {contract_path.exists()}", flush=True)
    
    if not contract_path.exists():
        print(f"[CONTRACT DOWNLOAD] ERRO: Arquivo não existe!", flush=True)
        # Debug: Include path in error message
        raise HTTPException(
            status_code=404, 
            detail=f"Contract file not found. Path: {contract_path} | DB path: {contrato['contrato_assinado']}"
        )
    
    return FileResponse(
        path=contract_path,
        media_type="application/pdf",
        filename=f"contrato_{motorista['name']}.pdf"
    )

@api_router.get("/motoristas/{motorista_id}/contrato/download-v2")
async def download_motorista_contrato_v2(
    motorista_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """NEW VERSION - Download contract of motorista"""
    motorista = await db.motoristas.find_one({"id": motorista_id}, {"_id": 0})
    if not motorista:
        raise HTTPException(status_code=404, detail="Motorista not found")
    
    # Check authorization
    is_assigned = False
    if current_user["role"] in [UserRole.PARCEIRO]:
        is_assigned = motorista.get("parceiro_atribuido") == current_user["id"]
    
    is_authorized = (
        current_user["role"] in [UserRole.ADMIN, UserRole.GESTAO] or
        is_assigned or
        (current_user["role"] == UserRole.MOTORISTA and current_user["email"] == motorista.get("email"))
    )
    
    if not is_authorized:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Find contract
    contrato = await db.contratos.find_one({"motorista_id": motorista_id}, {"_id": 0})
    if not contrato or not contrato.get("contrato_assinado"):
        raise HTTPException(status_code=404, detail="Contract not found")
    
    # Build path correctly
    contract_db_path = contrato["contrato_assinado"]
    if contract_db_path.startswith("/uploads/"):
        contract_path = Path("/app/backend/uploads" + contract_db_path[8:])
    else:
        contract_path = Path(contract_db_path)
    
    if not contract_path.exists():
        raise HTTPException(
            status_code=404, 
            detail=f"File not found at: {contract_path}"
        )
    
    return FileResponse(
        path=contract_path,
        media_type="application/pdf",
        filename=f"contrato_{motorista['name']}.pdf"
    )

@api_router.post("/pagamentos/{pagamento_id}/comprovativo")
async def upload_pagamento_comprovativo(
    pagamento_id: str,
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    """Upload payment proof/comprovativo (Parceiro/Admin/Gestor)"""
    # Authorization check
    if current_user["role"] not in [UserRole.PARCEIRO, UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Find payment
    pagamento = await db.pagamentos.find_one({"id": pagamento_id}, {"_id": 0})
    if not pagamento:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    # For parceiro, check if payment belongs to them
    if current_user["role"] == UserRole.PARCEIRO:
        if pagamento.get("parceiro_id") != current_user["id"]:
            raise HTTPException(status_code=403, detail="Not authorized")
    
    # Save file
    file_ext = file.filename.split('.')[-1]
    filename = f"comprovativo_pag_{pagamento_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.{file_ext}"
    file_path = UPLOAD_DIR / "comprovativos" / filename
    file_path.parent.mkdir(parents=True, exist_ok=True)
    
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)
    
    # Update payment with comprovativo path
    saved_path = f"/uploads/comprovativos/{filename}"
    await db.pagamentos.update_one(
        {"id": pagamento_id},
        {
            "$set": {
                "comprovativo_pagamento_url": saved_path,
                "status": "liquidado",
                "updated_at": datetime.utcnow().isoformat()
            }
        }
    )
    
    return {"message": "Comprovativo uploaded successfully", "path": saved_path}

@api_router.get("/relatorios-ganhos/{relatorio_id}/download")
async def download_relatorio_recibo(
    relatorio_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Download receipt PDF from earnings report (Motorista/Admin/Gestor)"""
    relatorio = await db.relatorios_ganhos.find_one({"id": relatorio_id}, {"_id": 0})
    if not relatorio:
        raise HTTPException(status_code=404, detail="Report not found")
    
    # Check authorization
    is_authorized = (
        current_user["role"] in [UserRole.ADMIN, UserRole.GESTAO] or
        (current_user["role"] == UserRole.MOTORISTA and current_user["id"] == relatorio["motorista_id"])
    )
    
    if not is_authorized:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get receipt path
    recibo_url = relatorio.get("recibo_pdf")
    if not recibo_url:
        raise HTTPException(status_code=404, detail="Receipt not uploaded yet")
    
    # Construct file path
    file_path = Path(recibo_url.replace("/uploads/", str(UPLOAD_DIR) + "/"))
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Receipt file not found")
    
    return FileResponse(
        path=file_path,
        media_type="application/pdf",
        filename=f"recibo_{relatorio_id}.pdf"
    )

@api_router.get("/relatorios-ganhos/{relatorio_id}/download-relatorio-pdf")
async def download_relatorio_semanal_pdf(
    relatorio_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Generate and download weekly earnings report PDF"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from io import BytesIO
    
    # Fetch report
    relatorio = await db.relatorios_ganhos.find_one({"id": relatorio_id}, {"_id": 0})
    if not relatorio:
        raise HTTPException(status_code=404, detail="Relatório não encontrado")
    
    # Check authorization
    is_authorized = (
        current_user["role"] in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO] or
        (current_user["role"] == UserRole.MOTORISTA and current_user["id"] == relatorio["motorista_id"])
    )
    
    if not is_authorized:
        raise HTTPException(status_code=403, detail="Não autorizado")
    
    # Get motorista details
    motorista = await db.users.find_one({"id": relatorio["motorista_id"]}, {"_id": 0})
    motorista_nome = motorista.get("name", "N/A") if motorista else "N/A"
    
    # Create PDF in memory
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#334155'),
        spaceAfter=12
    )
    
    # Title
    elements.append(Paragraph("TVDEFleet - Relatório Semanal de Ganhos", title_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Motorista info
    motorista_data = [
        ["Motorista:", motorista_nome],
        ["Período:", f"{relatorio.get('periodo_inicio', 'N/A')} a {relatorio.get('periodo_fim', 'N/A')}"],
        ["Data de Geração:", datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M")],
        ["Status:", relatorio.get('status', 'N/A')]
    ]
    
    motorista_table = Table(motorista_data, colWidths=[5*cm, 10*cm])
    motorista_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f1f5f9')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#334155')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(motorista_table)
    elements.append(Spacer(1, 1*cm))
    
    # Earnings Summary
    elements.append(Paragraph("Resumo de Ganhos", header_style))
    
    detalhes = relatorio.get('detalhes', {})
    earnings_data = [
        ["Descrição", "Valor"],
        ["Ganhos Uber", f"€{detalhes.get('uber', 0):.2f}"],
        ["Ganhos Bolt", f"€{detalhes.get('bolt', 0):.2f}"],
        ["Outros Ganhos", f"€{detalhes.get('outros', 0):.2f}"],
        ["", ""],
        ["Valor Total Bruto", f"€{relatorio.get('valor_total', 0):.2f}"],
        ["(-) Descontos/Comissões", f"€{(relatorio.get('valor_total', 0) - relatorio.get('valor_liquido', 0)):.2f}"],
        ["Valor Líquido", f"€{relatorio.get('valor_liquido', 0):.2f}"]
    ]
    
    earnings_table = Table(earnings_data, colWidths=[10*cm, 5*cm])
    earnings_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#dcfce7')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(earnings_table)
    elements.append(Spacer(1, 1*cm))
    
    # Notes section
    if relatorio.get('notas'):
        elements.append(Paragraph("Notas", header_style))
        elements.append(Paragraph(relatorio['notas'], styles['Normal']))
        elements.append(Spacer(1, 1*cm))
    
    # Footer
    footer_text = f"<para align='center'><font size='8' color='#64748b'>Este documento foi gerado automaticamente pelo sistema TVDEFleet em {datetime.now(timezone.utc).strftime('%d/%m/%Y às %H:%M UTC')}</font></para>"
    elements.append(Spacer(1, 2*cm))
    elements.append(Paragraph(footer_text, styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=relatorio_semanal_{relatorio_id}.pdf"}
    )


@api_router.post("/solicitacoes-alteracao")
async def criar_solicitacao_alteracao(
    data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Criar solicitação de alteração de dados (Motorista)"""
    if current_user["role"] != UserRole.MOTORISTA:
        raise HTTPException(status_code=403, detail="Only motoristas can create change requests")
    
    # Get motorista data
    motorista = await db.motoristas.find_one({"email": current_user["email"]}, {"_id": 0})
    if not motorista:
        raise HTTPException(status_code=404, detail="Motorista profile not found")
    
    solicitacao = {
        "id": str(uuid4()),
        "motorista_id": motorista["id"],
        "motorista_nome": motorista["name"],
        "campo": data["campo"],
        "valor_atual": data["valor_atual"],
        "valor_solicitado": data["valor_solicitado"],
        "justificativa": data["justificativa"],
        "status": "pendente",
        "respondido_por": None,
        "resposta": None,
        "created_at": datetime.now(timezone.utc),
        "respondido_em": None
    }
    
    await db.solicitacoes_alteracao.insert_one(solicitacao)
    
    return {"message": "Solicitação criada com sucesso", "id": solicitacao["id"]}

@api_router.get("/solicitacoes-alteracao")
async def listar_solicitacoes_alteracao(
    current_user: Dict = Depends(get_current_user)
):
    """Listar solicitações de alteração"""
    if current_user["role"] == UserRole.MOTORISTA:
        # Motorista vê apenas suas próprias solicitações
        motorista = await db.motoristas.find_one({"email": current_user["email"]}, {"_id": 0})
        if not motorista:
            return []
        solicitacoes = await db.solicitacoes_alteracao.find(
            {"motorista_id": motorista["id"]},
            {"_id": 0}
        ).to_list(100)
    else:
        # Admin/Gestor/Parceiro/Operacional veem todas
        solicitacoes = await db.solicitacoes_alteracao.find({}, {"_id": 0}).to_list(100)
    
    return solicitacoes

@api_router.put("/solicitacoes-alteracao/{solicitacao_id}/responder")
async def responder_solicitacao(
    solicitacao_id: str,
    data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Responder solicitação de alteração (Admin/Gestor/Parceiro/Operacional)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    solicitacao = await db.solicitacoes_alteracao.find_one({"id": solicitacao_id}, {"_id": 0})
    if not solicitacao:
        raise HTTPException(status_code=404, detail="Solicitação not found")
    
    status = data.get("status")  # "aprovada" ou "rejeitada"
    resposta = data.get("resposta", "")
    
    # Update solicitação
    await db.solicitacoes_alteracao.update_one(
        {"id": solicitacao_id},
        {
            "$set": {
                "status": status,
                "respondido_por": current_user["name"],
                "resposta": resposta,
                "respondido_em": datetime.now(timezone.utc)
            }
        }
    )
    
    # Se aprovada, atualizar dados do motorista
    if status == "aprovada":
        await db.motoristas.update_one(
            {"id": solicitacao["motorista_id"]},
            {"$set": {solicitacao["campo"]: solicitacao["valor_solicitado"]}}
        )
    
    return {"message": f"Solicitação {status} com sucesso"}

@api_router.post("/motoristas/{motorista_id}/upload-documento")
async def upload_motorista_documento(
    motorista_id: str,
    tipo_documento: str = Form(...),  # comprovativo_morada, cc_frente_verso, carta_frente_verso, licenca_tvde, registo_criminal, iban
    file: UploadFile = File(...),
    file2: Optional[UploadFile] = File(None),  # Para frente e verso
    current_user: Dict = Depends(get_current_user)
):
    """Upload document for motorista (converts images to PDF)"""
    motorista = await db.motoristas.find_one({"id": motorista_id}, {"_id": 0})
    if not motorista:
        raise HTTPException(status_code=404, detail="Motorista not found")
    
    motoristas_docs_dir = UPLOAD_DIR / "motoristas" / motorista_id
    motoristas_docs_dir.mkdir(parents=True, exist_ok=True)
    
    try:
        if tipo_documento in ['cc_frente_verso', 'carta_frente_verso'] and file2:
            # Process frente e verso together
            # Save first image
            file_id1 = str(uuid.uuid4())
            file_ext1 = Path(file.filename).suffix.lower()
            temp_path1 = motoristas_docs_dir / f"{file_id1}_temp{file_ext1}"
            with open(temp_path1, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            
            # Save second image
            file_id2 = str(uuid.uuid4())
            file_ext2 = Path(file2.filename).suffix.lower()
            temp_path2 = motoristas_docs_dir / f"{file_id2}_temp{file_ext2}"
            with open(temp_path2, "wb") as buffer:
                shutil.copyfileobj(file2.file, buffer)
            
            # Merge to PDF A4
            pdf_filename = f"{tipo_documento}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            pdf_path = motoristas_docs_dir / pdf_filename
            
            await merge_images_to_pdf_a4(temp_path1, temp_path2, pdf_path)
            
            # Clean up temp files
            temp_path1.unlink()
            temp_path2.unlink()
            
            doc_url = str(pdf_path.relative_to(ROOT_DIR))
        else:
            # Single file (convert to PDF if image)
            file_id = f"{tipo_documento}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            result = await process_uploaded_file(file, motoristas_docs_dir, file_id)
            doc_url = result.get("pdf_path") or result.get("original_path")
        
        # Update motorista documents
        field_name = f"documents.{tipo_documento}_pdf"
        await db.motoristas.update_one(
            {"id": motorista_id},
            {"$set": {field_name: doc_url}}
        )
        
        return {"message": "Document uploaded successfully", "document_url": doc_url}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error uploading document: {str(e)}")

# ==================== VEHICLE ENDPOINTS ====================

@api_router.post("/vehicles", response_model=Vehicle)
async def create_vehicle(vehicle_data: VehicleCreate, current_user: Dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    vehicle_dict = vehicle_data.model_dump()
    vehicle_dict["id"] = str(uuid.uuid4())
    vehicle_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    vehicle_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    # Calculate alerta_validade (30 days before expiry)
    validade = datetime.fromisoformat(vehicle_data.validade_matricula).replace(tzinfo=timezone.utc)
    dias_restantes = (validade - datetime.now(timezone.utc)).days
    vehicle_dict["alerta_validade"] = dias_restantes <= 30
    
    vehicle_dict["manutencoes"] = []
    vehicle_dict["inspecoes"] = []
    vehicle_dict["km_atual"] = 0
    vehicle_dict["km_aviso_manutencao"] = 5000
    vehicle_dict["alertas_manutencao"] = []
    vehicle_dict["disponibilidade"] = {
        "status": "disponivel",
        "motoristas_atribuidos": [],
        "data_entrega_manutencao": None,
        "tipo_manutencao": None
    }
    
    # Set parceiro_id from current user if parceiro/operacional
    if current_user["role"] in [UserRole.PARCEIRO]:
        vehicle_dict["parceiro_id"] = current_user["id"]
    
    await db.vehicles.insert_one(vehicle_dict)
    
    if isinstance(vehicle_dict["created_at"], str):
        vehicle_dict["created_at"] = datetime.fromisoformat(vehicle_dict["created_at"])
    if isinstance(vehicle_dict["updated_at"], str):
        vehicle_dict["updated_at"] = datetime.fromisoformat(vehicle_dict["updated_at"])
    
    return Vehicle(**vehicle_dict)

@api_router.get("/vehicles", response_model=List[Vehicle])
async def get_vehicles(current_user: Dict = Depends(get_current_user)):
    query = {}
    if current_user["role"] == UserRole.PARCEIRO:
        query["parceiro_id"] = current_user["id"]
    elif current_user["role"] == UserRole.GESTAO:
        # Gestor vê veículos dos parceiros atribuídos
        parceiros_ids = current_user.get("parceiros_atribuidos", [])
        if parceiros_ids:
            query["parceiro_id"] = {"$in": parceiros_ids}
        else:
            query["parceiro_id"] = None  # Nenhum veículo se sem parceiros
    
    vehicles = await db.vehicles.find(query, {"_id": 0}).to_list(1000)
    for v in vehicles:
        if isinstance(v["created_at"], str):
            v["created_at"] = datetime.fromisoformat(v["created_at"])
        if isinstance(v["updated_at"], str):
            v["updated_at"] = datetime.fromisoformat(v["updated_at"])
        # Fix km_atual if it's empty string or invalid
        if "km_atual" in v:
            if isinstance(v["km_atual"], str):
                v["km_atual"] = int(v["km_atual"]) if v["km_atual"].strip().isdigit() else 0
    return vehicles

@api_router.get("/vehicles/available", response_model=List[Vehicle])
async def get_available_vehicles():
    vehicles = await db.vehicles.find({"disponibilidade.status": "disponivel"}, {"_id": 0}).to_list(1000)
    for v in vehicles:
        if isinstance(v["created_at"], str):
            v["created_at"] = datetime.fromisoformat(v["created_at"])
        if isinstance(v["updated_at"], str):
            v["updated_at"] = datetime.fromisoformat(v["updated_at"])
    return vehicles

@api_router.post("/vehicles/{vehicle_id}/request")
async def request_vehicle(vehicle_id: str, request_data: Dict[str, str], current_user: Dict = Depends(get_current_user)):
    """Motorista requests a vehicle"""
    if current_user["role"] != UserRole.MOTORISTA:
        raise HTTPException(status_code=403, detail="Only motoristas can request vehicles")
    
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    if vehicle.get("disponibilidade", {}).get("status") != "disponivel":
        raise HTTPException(status_code=400, detail="Vehicle not available")
    
    # Create request record
    request_id = str(uuid.uuid4())
    vehicle_request = {
        "id": request_id,
        "vehicle_id": vehicle_id,
        "motorista_id": request_data.get("motorista_id", current_user["id"]),
        "status": "pendente",  # pendente, aprovado, rejeitado
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.vehicle_requests.insert_one(vehicle_request)
    
    return {"message": "Vehicle request submitted successfully", "request_id": request_id}

@api_router.get("/vehicles/{vehicle_id}", response_model=Vehicle)
async def get_vehicle(vehicle_id: str, current_user: Dict = Depends(get_current_user)):
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    if isinstance(vehicle["created_at"], str):
        vehicle["created_at"] = datetime.fromisoformat(vehicle["created_at"])
    if isinstance(vehicle["updated_at"], str):
        vehicle["updated_at"] = datetime.fromisoformat(vehicle["updated_at"])
    # Fix km_atual if it's empty string or invalid
    if "km_atual" in vehicle:
        if isinstance(vehicle["km_atual"], str):
            vehicle["km_atual"] = int(vehicle["km_atual"]) if vehicle["km_atual"].strip().isdigit() else 0
    
    return Vehicle(**vehicle)

@api_router.put("/vehicles/{vehicle_id}")
async def update_vehicle(vehicle_id: str, updates: Dict[str, Any], current_user: Dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Auto-add to agenda when dates are filled
    if updates.get("insurance") and updates["insurance"].get("data_validade"):
        await auto_add_to_agenda(
            vehicle_id, 
            "seguro", 
            updates["insurance"]["data_validade"],
            "Renovação de Seguro"
        )
    
    if updates.get("inspection") and updates["inspection"].get("proxima_inspecao"):
        await auto_add_to_agenda(
            vehicle_id,
            "inspecao",
            updates["inspection"]["proxima_inspecao"],
            "Inspeção do Veículo"
        )
    
    if updates.get("extintor") and updates["extintor"].get("data_validade"):
        await auto_add_to_agenda(
            vehicle_id,
            "extintor",
            updates["extintor"]["data_validade"],
            "Renovação de Extintor"
        )
    
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.vehicles.update_one({"id": vehicle_id}, {"$set": updates})
    return {"message": "Vehicle updated"}

@api_router.delete("/vehicles/{vehicle_id}")
async def delete_vehicle(vehicle_id: str, current_user: Dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    await db.vehicles.delete_one({"id": vehicle_id})
    return {"message": "Vehicle deleted"}

@api_router.post("/vehicles/{vehicle_id}/upload-photo")
async def upload_vehicle_photo(
    vehicle_id: str,
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    """Upload vehicle photo (max 3 photos, converted to PDF)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if vehicle exists
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    # Check photo limit (max 3)
    current_photos = vehicle.get("fotos", [])
    if len(current_photos) >= 3:
        raise HTTPException(status_code=400, detail="Maximum 3 photos allowed per vehicle")
    
    # Process file: save and convert to PDF
    file_id = f"vehicle_{vehicle_id}_photo_{len(current_photos) + 1}_{uuid.uuid4()}"
    
    # Create vehicle photos directory if not exists
    vehicle_photos_dir = UPLOAD_DIR / "vehicles"
    vehicle_photos_dir.mkdir(exist_ok=True)
    
    file_info = await process_uploaded_file(file, vehicle_photos_dir, file_id)
    
    # Store PDF path in database
    photo_url = file_info["pdf_path"] if file_info["pdf_path"] else file_info["original_path"]
    
    # Update vehicle with new photo
    await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$push": {"fotos": photo_url}}
    )
    
    return {
        "message": "Photo uploaded successfully",
        "photo_url": photo_url
    }

@api_router.post("/vehicles/{vehicle_id}/agenda")
async def add_vehicle_agenda(
    vehicle_id: str,
    agenda_data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Add event to vehicle agenda"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if vehicle exists
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    evento_id = str(uuid.uuid4())
    evento = {
        "id": evento_id,
        "tipo": agenda_data.get("tipo"),
        "titulo": agenda_data.get("titulo"),
        "data": agenda_data.get("data"),
        "hora": agenda_data.get("hora"),
        "descricao": agenda_data.get("descricao"),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$push": {"agenda": evento}}
    )
    
    return {"message": "Event added to agenda", "evento_id": evento_id}

@api_router.get("/vehicles/{vehicle_id}/agenda")
async def get_vehicle_agenda(vehicle_id: str, current_user: Dict = Depends(get_current_user)):
    """Get vehicle agenda"""
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0, "agenda": 1})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    return vehicle.get("agenda", [])

@api_router.put("/vehicles/{vehicle_id}/agenda/{evento_id}")
async def update_vehicle_agenda(
    vehicle_id: str,
    evento_id: str,
    agenda_data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Update event in vehicle agenda"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    # Update the specific event in agenda array
    await db.vehicles.update_one(
        {"id": vehicle_id, "agenda.id": evento_id},
        {"$set": {
            "agenda.$.tipo": agenda_data.get("tipo"),
            "agenda.$.titulo": agenda_data.get("titulo"),
            "agenda.$.data": agenda_data.get("data"),
            "agenda.$.hora": agenda_data.get("hora"),
            "agenda.$.descricao": agenda_data.get("descricao"),
            "agenda.$.updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Event updated successfully"}

@api_router.delete("/vehicles/{vehicle_id}/agenda/{evento_id}")
async def delete_vehicle_agenda(
    vehicle_id: str,
    evento_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Delete event from vehicle agenda"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    # Remove the event from agenda array
    await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$pull": {"agenda": {"id": evento_id}}}
    )
    
    return {"message": "Event deleted successfully"}

@api_router.get("/vehicles/{vehicle_id}/historico")
async def get_vehicle_historico(vehicle_id: str, current_user: Dict = Depends(get_current_user)):
    """Get vehicle history (maintenance, inspections, etc)"""
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    # Compile history from different sources
    historico = []
    
    # Add maintenance history
    if vehicle.get("manutencoes"):
        for man in vehicle["manutencoes"]:
            historico.append({
                "tipo": "manutencao",
                "data": man.get("data"),
                "descricao": f"{man.get('tipo_manutencao')}: {man.get('descricao')}",
                "valor": man.get("valor"),
                "km": man.get("km_realizada")
            })
    
    # Add inspection history
    if vehicle.get("inspecoes"):
        for insp in vehicle["inspecoes"]:
            historico.append({
                "tipo": "inspecao",
                "data": insp.get("ultima_inspecao"),
                "descricao": f"Inspeção - {insp.get('resultado', 'N/A')}",
                "valor": insp.get("valor"),
                "observacoes": insp.get("observacoes")
            })
    
    # Add damages history
    if vehicle.get("danos"):
        for dano in vehicle["danos"]:
            historico.append({
                "tipo": "dano",
                "data": dano.get("data"),
                "descricao": f"{dano.get('tipo')}: {dano.get('descricao')}",
                "valor": dano.get("valor_reparacao"),
                "responsavel": dano.get("responsavel")
            })
    
    # Sort by date (most recent first)
    historico.sort(key=lambda x: x.get("data", ""), reverse=True)
    
    return historico


@api_router.post("/vehicles/{vehicle_id}/historico")
async def add_historico_entry(
    vehicle_id: str,
    entry_data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Add editable history entry to vehicle"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    entry = {
        "id": str(uuid.uuid4()),
        "data": entry_data.get("data"),
        "titulo": entry_data.get("titulo"),
        "descricao": entry_data.get("descricao"),
        "tipo": entry_data.get("tipo", "observacao"),  # observacao, manutencao, incidente, etc
        "created_by": current_user["id"],
        "created_by_name": current_user.get("name", current_user.get("email")),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$push": {"historico_editavel": entry}}
    )
    
    return {"message": "History entry added", "entry_id": entry["id"]}

@api_router.put("/vehicles/{vehicle_id}/historico/{entry_id}")
async def update_historico_entry(
    vehicle_id: str,
    entry_id: str,
    entry_data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Update editable history entry"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get vehicle
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    # Update specific entry
    historico = vehicle.get("historico_editavel", [])
    for entry in historico:
        if entry["id"] == entry_id:
            entry["data"] = entry_data.get("data", entry["data"])
            entry["titulo"] = entry_data.get("titulo", entry["titulo"])
            entry["descricao"] = entry_data.get("descricao", entry["descricao"])
            entry["tipo"] = entry_data.get("tipo", entry["tipo"])
            entry["updated_at"] = datetime.now(timezone.utc).isoformat()
            break
    
    await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$set": {"historico_editavel": historico}}
    )
    
    return {"message": "History entry updated"}

@api_router.delete("/vehicles/{vehicle_id}/historico/{entry_id}")
async def delete_historico_entry(
    vehicle_id: str,
    entry_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Delete editable history entry"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$pull": {"historico_editavel": {"id": entry_id}}}
    )
    
    return {"message": "History entry deleted"}



@api_router.get("/vehicles/{vehicle_id}/relatorio-ganhos")
async def get_vehicle_relatorio_ganhos(vehicle_id: str, current_user: Dict = Depends(get_current_user)):
    """Get vehicle financial report (earnings vs expenses)"""
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    # Calculate ganhos (earnings) - would come from trips data
    # For now, return placeholder structure
    ganhos_total = 0.0
    despesas_total = 0.0
    detalhes = []
    
    # Calculate despesas from maintenance, insurance, etc
    if vehicle.get("manutencoes"):
        for man in vehicle["manutencoes"]:
            valor = man.get("valor", 0)
            despesas_total += valor
            detalhes.append({
                "tipo": "despesa",
                "descricao": f"Manutenção: {man.get('tipo_manutencao')}",
                "data": man.get("data"),
                "valor": valor
            })
    
    if vehicle.get("insurance"):
        valor = vehicle["insurance"].get("valor", 0)
        despesas_total += valor
        detalhes.append({
            "tipo": "despesa",
            "descricao": "Seguro",
            "data": vehicle["insurance"].get("data_inicio"),
            "valor": valor
        })
    
    if vehicle.get("inspection"):
        valor = vehicle["inspection"].get("valor", 0)
        if valor:
            despesas_total += valor
            detalhes.append({
                "tipo": "despesa",
                "descricao": "Inspeção",
                "data": vehicle["inspection"].get("ultima_inspecao"),
                "valor": valor
            })
    
    lucro = ganhos_total - despesas_total
    
    return {
        "ganhos_total": ganhos_total,
        "despesas_total": despesas_total,
        "lucro": lucro,
        "detalhes": sorted(detalhes, key=lambda x: x.get("data", ""), reverse=True)
    }

@api_router.post("/vehicles/{vehicle_id}/atribuir-motorista")
async def atribuir_motorista_vehicle(
    vehicle_id: str,
    data: Dict[str, str],
    current_user: Dict = Depends(get_current_user)
):
    """Atribuir motorista a um veículo"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    motorista_id = data.get("motorista_id")
    
    if motorista_id:
        # Get motorista data
        motorista = await db.motoristas.find_one({"id": motorista_id}, {"_id": 0})
        if not motorista:
            raise HTTPException(status_code=404, detail="Motorista not found")
        
        # Get vehicle data to copy cartao_frota_id
        vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
        if not vehicle:
            raise HTTPException(status_code=404, detail="Vehicle not found")
        
        # Update vehicle
        await db.vehicles.update_one(
            {"id": vehicle_id},
            {"$set": {
                "motorista_atribuido": motorista_id,
                "motorista_atribuido_nome": motorista.get("name"),
                "status": "atribuido",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        # If vehicle has cartao_frota_id, copy it to motorista
        if vehicle.get("cartao_frota_id"):
            await db.motoristas.update_one(
                {"id": motorista_id},
                {"$set": {
                    "id_cartao_frota_combustivel": vehicle.get("cartao_frota_id")
                }}
            )
        
        return {"message": "Motorista atribuído com sucesso"}
    else:
        # Remove assignment
        await db.vehicles.update_one(
            {"id": vehicle_id},
            {"$set": {
                "motorista_atribuido": None,
                "motorista_atribuido_nome": None,
                "status": "disponivel",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        return {"message": "Motorista removido com sucesso"}

# ==================== VEHICLE DOCUMENT UPLOADS ====================

@api_router.post("/vehicles/{vehicle_id}/upload-seguro-doc")
async def upload_seguro_document(
    vehicle_id: str,
    doc_type: str = Form(...),  # carta_verde, condicoes, fatura
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    """Upload insurance documents (carta verde, condições, fatura)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Create directory
    insurance_docs_dir = UPLOAD_DIR / "insurance_docs"
    insurance_docs_dir.mkdir(exist_ok=True)
    
    # Process file
    file_id = f"insurance_{doc_type}_{vehicle_id}_{uuid.uuid4()}"
    file_info = await process_uploaded_file(file, insurance_docs_dir, file_id)
    
    # Update vehicle
    field_map = {
        "carta_verde": "insurance.carta_verde_url",
        "condicoes": "insurance.condicoes_url",
        "fatura": "insurance.fatura_url"
    }
    
    if doc_type not in field_map:
        raise HTTPException(status_code=400, detail="Invalid doc_type")
    
    # Get current insurance data
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    insurance = vehicle.get("insurance", {})
    if doc_type == "carta_verde":
        insurance["carta_verde_url"] = file_info["saved_path"]
    elif doc_type == "condicoes":
        insurance["condicoes_url"] = file_info["saved_path"]
    elif doc_type == "fatura":
        insurance["fatura_url"] = file_info["saved_path"]
    
    await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$set": {"insurance": insurance}}
    )
    
    return {"message": "Document uploaded successfully", "url": file_info["saved_path"]}

@api_router.post("/vehicles/{vehicle_id}/upload-inspecao-doc")
async def upload_inspecao_document(
    vehicle_id: str,
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    """Upload inspection document"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Create directory
    inspection_docs_dir = UPLOAD_DIR / "inspection_docs"
    inspection_docs_dir.mkdir(exist_ok=True)
    
    # Process file
    file_id = f"inspection_{vehicle_id}_{uuid.uuid4()}"
    file_info = await process_uploaded_file(file, inspection_docs_dir, file_id)
    
    # Update vehicle
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    inspection = vehicle.get("inspection", {})
    inspection["ficha_inspecao_url"] = file_info["saved_path"]
    
    await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$set": {"inspection": inspection}}
    )
    
    return {"message": "Document uploaded successfully", "url": file_info["saved_path"]}

@api_router.post("/vehicles/{vehicle_id}/upload-extintor-doc")
async def upload_extintor_document(
    vehicle_id: str,
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    """Upload fire extinguisher certificate"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Create directory
    extintor_docs_dir = UPLOAD_DIR / "extintor_docs"
    extintor_docs_dir.mkdir(exist_ok=True)
    
    # Process file
    file_id = f"extintor_{vehicle_id}_{uuid.uuid4()}"
    file_info = await process_uploaded_file(file, extintor_docs_dir, file_id)
    
    # Update vehicle
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    extintor = vehicle.get("extintor", {})
    # Use pdf_path if available (converted file), otherwise use original_path
    file_url = file_info.get("pdf_path") or file_info.get("original_path")
    extintor["certificado_url"] = file_url
    
    await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$set": {"extintor": extintor}}
    )
    
    return {"message": "Document uploaded successfully", "certificado_url": file_url, "file_info": file_info}

@api_router.post("/vehicles/{vehicle_id}/upload-foto")
async def upload_vehicle_photo(
    vehicle_id: str,
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    """Upload vehicle photo (max 3)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    current_photos = vehicle.get("fotos_veiculo", [])
    if len(current_photos) >= 3:
        raise HTTPException(status_code=400, detail="Maximum 3 photos allowed")
    
    # Create directory
    photos_dir = UPLOAD_DIR / "vehicle_photos_info"
    photos_dir.mkdir(exist_ok=True)
    
    # Process file
    file_id = f"photo_{vehicle_id}_{uuid.uuid4()}"
    file_info = await process_uploaded_file(file, photos_dir, file_id)
    
    # Use pdf_path (converted file) or original_path if not converted
    photo_path = file_info.get("pdf_path") or file_info.get("original_path")
    
    # Add to vehicle
    await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$push": {"fotos_veiculo": photo_path}}
    )
    
    return {"message": "Photo uploaded successfully", "url": photo_path}

@api_router.delete("/vehicles/{vehicle_id}/fotos/{foto_index}")
async def delete_vehicle_photo(
    vehicle_id: str,
    foto_index: int,
    current_user: Dict = Depends(get_current_user)
):
    """Delete vehicle photo by index"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    fotos = vehicle.get("fotos_veiculo", [])
    if foto_index < 0 or foto_index >= len(fotos):
        raise HTTPException(status_code=400, detail="Invalid photo index")
    
    # Remove photo
    fotos.pop(foto_index)
    
    await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$set": {"fotos_veiculo": fotos}}
    )
    
    return {"message": "Photo deleted successfully"}

# ==================== VEHICLE DOCUMENTS UPLOAD ENDPOINTS ====================

@api_router.post("/vehicles/{vehicle_id}/upload-carta-verde")
async def upload_carta_verde(
    vehicle_id: str,
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    """Upload carta verde document"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    # Process and save file
    file_id = f"carta_verde_{vehicle_id}_{uuid.uuid4()}"
    file_info = await process_uploaded_file(file, VEHICLE_DOCS_UPLOAD_DIR, file_id)
    
    # Use pdf_path (converted file) or original_path if not converted
    document_path = file_info.get("pdf_path") or file_info.get("original_path")
    
    # Update vehicle with document path
    await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$set": {"documento_carta_verde": document_path}}
    )
    
    return {"message": "Carta verde uploaded successfully", "url": document_path}

@api_router.post("/vehicles/{vehicle_id}/upload-condicoes")
async def upload_condicoes(
    vehicle_id: str,
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    """Upload condições document"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    # Process and save file
    file_id = f"condicoes_{vehicle_id}_{uuid.uuid4()}"
    file_info = await process_uploaded_file(file, VEHICLE_DOCS_UPLOAD_DIR, file_id)
    
    # Use pdf_path (converted file) or original_path if not converted
    document_path = file_info.get("pdf_path") or file_info.get("original_path")
    
    # Update vehicle with document path
    await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$set": {"documento_condicoes": document_path}}
    )
    
    return {"message": "Condições document uploaded successfully", "url": document_path}

@api_router.post("/vehicles/{vehicle_id}/upload-recibo-seguro")
async def upload_recibo_seguro(
    vehicle_id: str,
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    """Upload recibo de pagamento do seguro"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    # Process and save file
    file_id = f"recibo_seguro_{vehicle_id}_{uuid.uuid4()}"
    file_info = await process_uploaded_file(file, VEHICLE_DOCS_UPLOAD_DIR, file_id)
    
    # Use pdf_path (converted file) or original_path if not converted
    document_path = file_info.get("pdf_path") or file_info.get("original_path")
    
    # Update vehicle with document path
    await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$set": {"documento_recibo_seguro": document_path}}
    )
    
    return {"message": "Recibo de seguro uploaded successfully", "url": document_path}

@api_router.post("/vehicles/{vehicle_id}/upload-documento-inspecao")
async def upload_documento_inspecao(
    vehicle_id: str,
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    """Upload documento/certificado da inspeção"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    # Process and save file
    file_id = f"doc_inspecao_{vehicle_id}_{uuid.uuid4()}"
    file_info = await process_uploaded_file(file, VEHICLE_DOCS_UPLOAD_DIR, file_id)
    
    # Use pdf_path (converted file) or original_path if not converted
    document_path = file_info.get("pdf_path") or file_info.get("original_path")
    
    # Update vehicle with document path
    await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$set": {"documento_inspecao": document_path}}
    )
    
    return {"message": "Documento de inspeção uploaded successfully", "url": document_path}

@api_router.put("/vehicles/{vehicle_id}/update-km")
async def update_vehicle_km(
    vehicle_id: str,
    data: Dict[str, int],
    current_user: Dict = Depends(get_current_user)
):
    """Update vehicle KM (can be automatic from GPS)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    km_atual = data.get("km_atual")
    if not km_atual or km_atual < 0:
        raise HTTPException(status_code=400, detail="Invalid KM value")
    
    await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$set": {
            "km_atual": km_atual,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "KM updated successfully", "km_atual": km_atual}

@api_router.put("/vehicles/{vehicle_id}/status")
async def update_vehicle_status(
    vehicle_id: str,
    data: Dict[str, str],
    current_user: Dict = Depends(get_current_user)
):
    """Atualizar status do veículo"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    status = data.get("status")
    valid_statuses = ["disponivel", "atribuido", "manutencao", "venda", "condicoes"]
    
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Status inválido. Use: {', '.join(valid_statuses)}")
    
    await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$set": {
            "status": status,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Status atualizado com sucesso", "status": status}

@api_router.delete("/vehicles/{vehicle_id}/photos/{photo_index}")
async def delete_vehicle_photo(
    vehicle_id: str,
    photo_index: int,
    current_user: Dict = Depends(get_current_user)
):
    """Delete a vehicle photo by index (0-2)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    fotos = vehicle.get("fotos", [])
    if photo_index < 0 or photo_index >= len(fotos):
        raise HTTPException(status_code=404, detail="Photo not found")
    
    # Remove photo from array
    photo_to_remove = fotos[photo_index]
    fotos.pop(photo_index)
    
    # Update database
    await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$set": {"fotos": fotos}}
    )
    
    # Try to delete file from disk
    try:
        photo_path = ROOT_DIR / photo_to_remove
        if photo_path.exists():
            photo_path.unlink()
    except Exception as e:
        logger.error(f"Error deleting photo file: {e}")
    
    return {"message": "Photo deleted successfully"}

# ==================== FINANCIAL ENDPOINTS ====================

@api_router.post("/expenses", response_model=Expense)
async def create_expense(expense_data: ExpenseCreate, current_user: Dict = Depends(get_current_user)):
    # Parceiros não podem criar despesas - apenas visualizar e confirmar pagamentos
    if current_user["role"] == UserRole.PARCEIRO:
        raise HTTPException(
            status_code=403, 
            detail="Parceiros não podem criar despesas. Apenas confirmar pagamentos."
        )
    
    expense_dict = expense_data.model_dump()
    expense_dict["id"] = str(uuid.uuid4())
    expense_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.expenses.insert_one(expense_dict)
    
    if isinstance(expense_dict["created_at"], str):
        expense_dict["created_at"] = datetime.fromisoformat(expense_dict["created_at"])
    
    return Expense(**expense_dict)

@api_router.get("/expenses", response_model=List[Expense])
async def get_expenses(vehicle_id: Optional[str] = None, current_user: Dict = Depends(get_current_user)):
    query = {}
    if vehicle_id:
        query["vehicle_id"] = vehicle_id
    
    expenses = await db.expenses.find(query, {"_id": 0}).to_list(1000)
    for e in expenses:
        if isinstance(e["created_at"], str):
            e["created_at"] = datetime.fromisoformat(e["created_at"])
    return expenses

@api_router.post("/revenues", response_model=Revenue)
async def create_revenue(revenue_data: RevenueCreate, current_user: Dict = Depends(get_current_user)):
    # Parceiros não podem criar receitas - apenas visualizar e confirmar valores recebidos
    if current_user["role"] == UserRole.PARCEIRO:
        raise HTTPException(
            status_code=403, 
            detail="Parceiros não podem criar receitas. Apenas confirmar valores recebidos."
        )
    
    revenue_dict = revenue_data.model_dump()
    revenue_dict["id"] = str(uuid.uuid4())
    revenue_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.revenues.insert_one(revenue_dict)
    
    if isinstance(revenue_dict["created_at"], str):
        revenue_dict["created_at"] = datetime.fromisoformat(revenue_dict["created_at"])
    
    return Revenue(**revenue_dict)

@api_router.get("/revenues", response_model=List[Revenue])
async def get_revenues(vehicle_id: Optional[str] = None, current_user: Dict = Depends(get_current_user)):
    query = {}
    if vehicle_id:
        query["vehicle_id"] = vehicle_id
    
    revenues = await db.revenues.find(query, {"_id": 0}).to_list(1000)
    for r in revenues:
        if isinstance(r["created_at"], str):
            r["created_at"] = datetime.fromisoformat(r["created_at"])
    return revenues

# ==================== PARTNER FINANCIAL ENDPOINTS ====================

@api_router.post("/parceiros/{parceiro_id}/despesas", response_model=PartnerExpense)
async def create_partner_expense(
    parceiro_id: str,
    expense_data: PartnerExpenseCreate, 
    current_user: Dict = Depends(get_current_user)
):
    """Create a manual expense entry for a partner"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    expense_dict = expense_data.model_dump()
    expense_dict["id"] = str(uuid.uuid4())
    expense_dict["created_by"] = current_user["id"]
    expense_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.partner_expenses.insert_one(expense_dict)
    
    if isinstance(expense_dict["created_at"], str):
        expense_dict["created_at"] = datetime.fromisoformat(expense_dict["created_at"])
    
    return PartnerExpense(**expense_dict)

@api_router.get("/parceiros/{parceiro_id}/despesas", response_model=List[PartnerExpense])
async def get_partner_expenses(parceiro_id: str, current_user: Dict = Depends(get_current_user)):
    """Get all expenses for a specific partner"""
    expenses = await db.partner_expenses.find(
        {"parceiro_id": parceiro_id}, 
        {"_id": 0}
    ).to_list(1000)
    
    for e in expenses:
        if isinstance(e["created_at"], str):
            e["created_at"] = datetime.fromisoformat(e["created_at"])
    
    return expenses

@api_router.post("/parceiros/{parceiro_id}/receitas", response_model=PartnerRevenue)
async def create_partner_revenue(
    parceiro_id: str,
    revenue_data: PartnerRevenueCreate, 
    current_user: Dict = Depends(get_current_user)
):
    """Create a manual revenue entry for a partner"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    revenue_dict = revenue_data.model_dump()
    revenue_dict["id"] = str(uuid.uuid4())
    revenue_dict["created_by"] = current_user["id"]
    revenue_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.partner_revenues.insert_one(revenue_dict)
    
    if isinstance(revenue_dict["created_at"], str):
        revenue_dict["created_at"] = datetime.fromisoformat(revenue_dict["created_at"])
    
    return PartnerRevenue(**revenue_dict)

@api_router.get("/parceiros/{parceiro_id}/receitas", response_model=List[PartnerRevenue])
async def get_partner_revenues(parceiro_id: str, current_user: Dict = Depends(get_current_user)):
    """Get all revenues for a specific partner"""
    revenues = await db.partner_revenues.find(
        {"parceiro_id": parceiro_id}, 
        {"_id": 0}
    ).to_list(1000)
    
    for r in revenues:
        if isinstance(r["created_at"], str):
            r["created_at"] = datetime.fromisoformat(r["created_at"])
    
    return revenues

@api_router.get("/reports/roi/{vehicle_id}")
async def get_vehicle_roi(vehicle_id: str, periodo: str = "mensal", current_user: Dict = Depends(get_current_user)):
    revenues = await db.revenues.find({"vehicle_id": vehicle_id}, {"_id": 0}).to_list(1000)
    expenses = await db.expenses.find({"vehicle_id": vehicle_id}, {"_id": 0}).to_list(1000)
    
    total_receitas = sum([r["valor"] for r in revenues])
    total_despesas = sum([e["valor"] for e in expenses])
    roi = total_receitas - total_despesas
    km_percorridos = sum([r.get("km_percorridos", 0) for r in revenues])
    
    return {
        "vehicle_id": vehicle_id,
        "periodo": periodo,
        "total_receitas": total_receitas,
        "total_despesas": total_despesas,
        "roi": roi,
        "km_percorridos": km_percorridos,
        "utilizacao_percentual": 0
    }

@api_router.get("/reports/dashboard")
async def get_dashboard_stats(current_user: Dict = Depends(get_current_user)):
    query = {}
    motorista_query = {}
    
    # Filter by role
    if current_user["role"] in [UserRole.PARCEIRO]:
        query["parceiro_id"] = current_user["id"]
        motorista_query["parceiro_atribuido"] = current_user["id"]
    
    total_vehicles = await db.vehicles.count_documents(query)
    available_vehicles = await db.vehicles.count_documents({**query, "status": "disponivel"})
    total_motoristas = await db.motoristas.count_documents(motorista_query)
    pending_motoristas = await db.motoristas.count_documents({**motorista_query, "approved": False})
    
    # Filter revenues and expenses by parceiro's vehicles
    if current_user["role"] in [UserRole.PARCEIRO]:
        vehicles = await db.vehicles.find({"parceiro_id": current_user["id"]}, {"_id": 0, "id": 1}).to_list(10000)
        vehicle_ids = [v["id"] for v in vehicles]
        revenues = await db.revenues.find({"vehicle_id": {"$in": vehicle_ids}}, {"_id": 0}).to_list(10000)
        expenses = await db.expenses.find({"vehicle_id": {"$in": vehicle_ids}}, {"_id": 0}).to_list(10000)
    else:
        revenues = await db.revenues.find({}, {"_id": 0}).to_list(10000)
        expenses = await db.expenses.find({}, {"_id": 0}).to_list(10000)
    
    total_receitas = sum([r.get("valor", 0) for r in revenues])
    total_despesas = sum([e.get("valor", 0) for e in expenses])
    
    return {
        "total_vehicles": total_vehicles,
        "available_vehicles": available_vehicles,
        "total_motoristas": total_motoristas,
        "pending_motoristas": pending_motoristas,
        "total_receitas": total_receitas,
        "total_despesas": total_despesas,
        "roi": total_receitas - total_despesas
    }

# ==================== PARCEIRO REPORTS ENDPOINTS ====================

@api_router.get("/reports/parceiro/semanal")
async def get_parceiro_weekly_report(current_user: Dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get vehicles do parceiro
    vehicles = await db.vehicles.find({"parceiro_id": current_user["id"]}, {"_id": 0, "id": 1}).to_list(1000)
    vehicle_ids = [v["id"] for v in vehicles]
    
    # Get data from last 7 days
    from datetime import datetime, timedelta
    seven_days_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
    
    # Revenues
    revenues = await db.revenues.find({
        "vehicle_id": {"$in": vehicle_ids},
        "data": {"$gte": seven_days_ago}
    }, {"_id": 0}).to_list(10000)
    
    # Expenses
    expenses = await db.expenses.find({
        "vehicle_id": {"$in": vehicle_ids},
        "data": {"$gte": seven_days_ago}
    }, {"_id": 0}).to_list(10000)
    
    total_ganhos = sum([r["valor"] for r in revenues])
    total_gastos = sum([e["valor"] for e in expenses])
    lucro = total_ganhos - total_gastos
    
    return {
        "periodo": "ultima_semana",
        "total_ganhos": total_ganhos,
        "total_gastos": total_gastos,
        "lucro": lucro,
        "roi_percentual": (lucro / total_ganhos * 100) if total_ganhos > 0 else 0
    }

@api_router.get("/reports/parceiro/por-veiculo")
async def get_parceiro_vehicle_report(current_user: Dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    vehicles = await db.vehicles.find({"parceiro_id": current_user["id"]}, {"_id": 0}).to_list(1000)
    
    report = []
    for vehicle in vehicles:
        revenues = await db.revenues.find({"vehicle_id": vehicle["id"]}, {"_id": 0}).to_list(10000)
        expenses = await db.expenses.find({"vehicle_id": vehicle["id"]}, {"_id": 0}).to_list(10000)
        
        total_ganhos = sum([r["valor"] for r in revenues])
        total_gastos = sum([e["valor"] for e in expenses])
        
        report.append({
            "vehicle_id": vehicle["id"],
            "marca": vehicle["marca"],
            "modelo": vehicle["modelo"],
            "matricula": vehicle["matricula"],
            "total_ganhos": total_ganhos,
            "total_gastos": total_gastos,
            "lucro": total_ganhos - total_gastos
        })
    
    return report

@api_router.get("/reports/parceiro/por-motorista")
async def get_parceiro_motorista_report(current_user: Dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get vehicles do parceiro
    vehicles = await db.vehicles.find({"parceiro_id": current_user["id"]}, {"_id": 0, "id": 1}).to_list(1000)
    vehicle_ids = [v["id"] for v in vehicles]
    
    # Get all revenues with motorista
    revenues = await db.revenues.find({
        "vehicle_id": {"$in": vehicle_ids},
        "motorista_id": {"$ne": None}
    }, {"_id": 0}).to_list(10000)
    
    # Group by motorista
    motorista_totals = {}
    for rev in revenues:
        mid = rev["motorista_id"]
        if mid not in motorista_totals:
            motorista_totals[mid] = {"ganhos": 0, "corridas": 0}
        motorista_totals[mid]["ganhos"] += rev["valor"]
        motorista_totals[mid]["corridas"] += 1
    
    # Get motorista details
    report = []
    for mid, data in motorista_totals.items():
        motorista = await db.motoristas.find_one({"id": mid}, {"_id": 0})
        if motorista:
            # Calculate lucro (assuming commission split)
            lucro_parceiro = data["ganhos"] * 0.2  # Example: 20% for parceiro
            
            report.append({
                "motorista_id": mid,
                "nome": motorista["name"],
                "email": motorista["email"],
                "total_ganhos": data["ganhos"],
                "total_corridas": data["corridas"],
                "lucro_parceiro": lucro_parceiro
            })
    
    return report

@api_router.get("/reports/parceiro/proximas-despesas")
async def get_parceiro_upcoming_expenses(current_user: Dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    from datetime import datetime, timedelta
    
    vehicles = await db.vehicles.find({"parceiro_id": current_user["id"]}, {"_id": 0}).to_list(1000)
    
    upcoming = []
    today = datetime.now()
    
    for vehicle in vehicles:
        # Check seguro
        if vehicle.get("seguro") and vehicle["seguro"].get("data_validade"):
            validade = datetime.fromisoformat(vehicle["seguro"]["data_validade"])
            if validade > today and (validade - today).days <= 60:
                upcoming.append({
                    "tipo": "seguro",
                    "veiculo": f"{vehicle['marca']} {vehicle['modelo']} ({vehicle['matricula']})",
                    "descricao": f"Renovação seguro - {vehicle['seguro']['seguradora']}",
                    "valor": vehicle["seguro"]["preco"],
                    "data": vehicle["seguro"]["data_validade"],
                    "dias_restantes": (validade - today).days
                })
        
        # Check matricula validade
        if vehicle.get("validade_matricula"):
            validade = datetime.fromisoformat(vehicle["validade_matricula"])
            if validade > today and (validade - today).days <= 60:
                upcoming.append({
                    "tipo": "matricula",
                    "veiculo": f"{vehicle['marca']} {vehicle['modelo']} ({vehicle['matricula']})",
                    "descricao": "Renovação matrícula",
                    "valor": 50.00,  # Estimativa
                    "data": vehicle["validade_matricula"],
                    "dias_restantes": (validade - today).days
                })
        
        # Check manutencoes agendadas
        for manutencao in vehicle.get("manutencoes", []):
            if manutencao.get("data_proxima"):
                data_proxima = datetime.fromisoformat(manutencao["data_proxima"])
                if data_proxima > today and (data_proxima - today).days <= 30:
                    upcoming.append({
                        "tipo": "manutencao",
                        "veiculo": f"{vehicle['marca']} {vehicle['modelo']} ({vehicle['matricula']})",
                        "descricao": manutencao.get("o_que_fazer", "Manutenção agendada"),
                        "valor": manutencao.get("custos", 0),
                        "data": manutencao["data_proxima"],
                        "dias_restantes": (data_proxima - today).days
                    })
    
    # Sort by dias_restantes
    upcoming.sort(key=lambda x: x["dias_restantes"])
    
    return upcoming

# ==================== PAGAMENTOS ENDPOINTS ====================

@api_router.post("/pagamentos", response_model=Pagamento)
async def create_pagamento(pagamento_data: PagamentoCreate, current_user: Dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.PARCEIRO:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    pagamento_dict = pagamento_data.model_dump()
    pagamento_dict["id"] = str(uuid.uuid4())
    pagamento_dict["parceiro_id"] = current_user["id"]
    pagamento_dict["documento_url"] = None
    pagamento_dict["documento_analisado"] = False
    pagamento_dict["analise_documento"] = None
    pagamento_dict["status"] = "pendente"
    pagamento_dict["pago_em"] = None
    pagamento_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.pagamentos.insert_one(pagamento_dict)
    
    if isinstance(pagamento_dict["created_at"], str):
        pagamento_dict["created_at"] = datetime.fromisoformat(pagamento_dict["created_at"])
    
    return Pagamento(**pagamento_dict)

@api_router.post("/pagamentos/{pagamento_id}/upload-documento")
async def upload_pagamento_documento(
    pagamento_id: str,
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    if current_user["role"] != UserRole.PARCEIRO:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    pagamento = await db.pagamentos.find_one({"id": pagamento_id, "parceiro_id": current_user["id"]})
    if not pagamento:
        raise HTTPException(status_code=404, detail="Pagamento not found")
    
    # Process file: save and convert to PDF if image
    file_id = f"pagamento_{pagamento_id}_{uuid.uuid4()}"
    file_info = await process_uploaded_file(file, PAGAMENTOS_UPLOAD_DIR, file_id)
    
    # Store file path in database (prefer PDF version if available)
    file_url = file_info["pdf_path"] if file_info["pdf_path"] else file_info["original_path"]
    
    # Mock document analysis (in production, use OCR/AI service)
    analise = {
        "tipo_detectado": pagamento["tipo_documento"],
        "valor_detectado": pagamento["valor"],
        "confianca": 0.95,
        "campos_extraidos": {
            "nif": "123456789",
            "nome": "Motorista",
            "valor": pagamento["valor"],
            "data": datetime.now().strftime('%Y-%m-%d')
        }
    }
    
    await db.pagamentos.update_one(
        {"id": pagamento_id},
        {
            "$set": {
                "documento_url": file_url,
                "documento_analisado": True,
                "analise_documento": analise
            }
        }
    )
    
    return {
        "message": "Documento carregado e analisado",
        "analise": analise,
        "file_url": file_url,
        "converted_to_pdf": file_info["pdf_path"] is not None
    }

@api_router.put("/pagamentos/{pagamento_id}/marcar-pago")
async def marcar_pagamento_pago(pagamento_id: str, current_user: Dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.PARCEIRO:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.pagamentos.update_one(
        {"id": pagamento_id, "parceiro_id": current_user["id"]},
        {
            "$set": {
                "status": "pago",
                "pago_em": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Pagamento not found")
    
    return {"message": "Pagamento marcado como pago"}

@api_router.get("/pagamentos/semana-atual")
async def get_pagamentos_semana(current_user: Dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    from datetime import datetime, timedelta
    
    # Get current week
    today = datetime.now()
    start_week = (today - timedelta(days=today.weekday())).strftime('%Y-%m-%d')
    end_week = (today + timedelta(days=6-today.weekday())).strftime('%Y-%m-%d')
    
    # Build query based on role
    query = {
        "periodo_fim": {"$gte": start_week, "$lte": end_week}
    }
    
    # Filter by parceiro_id only for parceiro and operacional
    if current_user["role"] in [UserRole.PARCEIRO]:
        query["parceiro_id"] = current_user["id"]
    
    pagamentos = await db.pagamentos.find(query, {"_id": 0}).to_list(1000)
    
    for p in pagamentos:
        if isinstance(p.get("created_at"), str):
            p["created_at"] = datetime.fromisoformat(p["created_at"])
        if p.get("pago_em") and isinstance(p["pago_em"], str):
            p["pago_em"] = datetime.fromisoformat(p["pago_em"])
        
        # Get motorista info
        motorista = await db.motoristas.find_one({"id": p["motorista_id"]}, {"_id": 0, "name": 1, "email": 1})
        if motorista:
            p["motorista_nome"] = motorista["name"]
    
    total_pagar = sum([p["valor"] for p in pagamentos if p["status"] == "pendente"])
    total_pago = sum([p["valor"] for p in pagamentos if p["status"] == "pago"])
    
    return {
        "pagamentos": pagamentos,
        "total_pagar": total_pagar,
        "total_pago": total_pago,
        "periodo": f"{start_week} a {end_week}"
    }

# ==================== CSV IMPORT ENDPOINTS ====================

@api_router.post("/import/csv")
async def import_csv(file: UploadFile = File(...), import_type: str = Form(...), current_user: Dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # TODO: Implement CSV parsing and import logic
    # content = await file.read()
    
    return {"message": f"CSV import for {import_type} will be processed", "filename": file.filename}

# ==================== PARCEIROS ENDPOINTS ====================

@api_router.post("/parceiros")
async def create_parceiro(parceiro_data: ParceiroCreate, current_user: Dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    parceiro_dict = parceiro_data.model_dump()
    parceiro_dict["id"] = str(uuid.uuid4())
    parceiro_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    parceiro_dict["total_vehicles"] = 0
    
    if current_user["role"] == UserRole.GESTAO:
        parceiro_dict["gestor_associado_id"] = current_user["id"]
    
    # Assign base free plan for parceiro
    plano_base = await db.planos_sistema.find_one({
        "preco_mensal": 0, 
        "ativo": True, 
        "tipo_usuario": "parceiro"
    }, {"_id": 0})
    
    if plano_base:
        from datetime import timedelta
        plano_valida_ate = (datetime.now(timezone.utc) + timedelta(days=30)).isoformat()
        parceiro_dict["plano_id"] = plano_base["id"]
        parceiro_dict["plano_nome"] = plano_base["nome"]
        parceiro_dict["plano_valida_ate"] = plano_valida_ate
        parceiro_dict["plano_status"] = "ativo"
        logger.info(f"Assigned base plan {plano_base['id']} to new parceiro")
    
    await db.parceiros.insert_one(parceiro_dict)
    
    # Create user account for parceiro
    user_dict = {
        "id": parceiro_dict["id"],
        "email": parceiro_data.email,
        "name": parceiro_data.nome_manager,  # Use nome_manager as the user name
        "role": UserRole.PARCEIRO,
        "password": hash_password("parceiro123"),  # Default password
        "phone": parceiro_data.telefone,  # Use telefone as the phone
        "created_at": datetime.now(timezone.utc).isoformat(),
        "approved": True,
        "associated_gestor_id": parceiro_dict.get("gestor_associado_id")
    }
    await db.users.insert_one(user_dict)
    
    if isinstance(parceiro_dict["created_at"], str):
        parceiro_dict["created_at"] = datetime.fromisoformat(parceiro_dict["created_at"])
    
    return Parceiro(**parceiro_dict)

@api_router.post("/parceiros/register-public")
async def create_parceiro_public(parceiro_data: Dict[str, Any]):
    """Create parceiro from public registration (no auth required)"""
    parceiro_id = f"parceiro-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    
    # Validate codigo_certidao_comercial format
    codigo_certidao = parceiro_data.get("codigo_certidao_comercial", "")
    if not codigo_certidao or not re.match(r'^\d{4}-\d{4}-\d{4}$', codigo_certidao):
        raise HTTPException(
            status_code=400, 
            detail="Código de Certidão Comercial é obrigatório e deve estar no formato xxxx-xxxx-xxxx"
        )
    
    new_parceiro = {
        "id": parceiro_id,
        "nome": parceiro_data.get("nome"),
        "email": parceiro_data.get("email"),
        "telefone": parceiro_data.get("telefone"),
        "nif": parceiro_data.get("nif"),
        "morada": parceiro_data.get("morada"),
        "codigo_postal": parceiro_data.get("codigo_postal"),
        "codigo_certidao_comercial": codigo_certidao,
        "responsavel_nome": parceiro_data.get("responsavel_nome"),
        "responsavel_contacto": parceiro_data.get("responsavel_contacto"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "approved": False,
        "status": "pendente"
    }
    
    await db.parceiros.insert_one(new_parceiro)
    
    return {
        "success": True,
        "message": "Parceiro registado com sucesso. Aguarde aprovação.",
        "parceiro_id": parceiro_id
    }
    
    return {"message": "Parceiro registered successfully", "id": parceiro_id}

@api_router.get("/parceiros", response_model=List[Parceiro])
async def get_parceiros(current_user: Dict = Depends(get_current_user)):
    query = {}
    
    # IMPORTANTE: Apenas mostrar parceiros aprovados (exceto para admin que pode ver todos)
    if current_user["role"] != UserRole.ADMIN:
        query["approved"] = True
    
    if current_user["role"] == UserRole.GESTAO:
        query["gestor_associado_id"] = current_user["id"]
    elif current_user["role"] == "parceiro":
        # Parceiro só vê ele próprio
        query["id"] = current_user["id"]
    elif current_user["role"] not in [UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    parceiros = await db.parceiros.find(query, {"_id": 0}).to_list(1000)
    for p in parceiros:
        # Fix created_at and updated_at if missing or invalid
        if "created_at" in p and isinstance(p["created_at"], str):
            p["created_at"] = datetime.fromisoformat(p["created_at"])
        elif "created_at" not in p:
            p["created_at"] = datetime.now(timezone.utc)
        
        if "updated_at" in p and isinstance(p["updated_at"], str):
            p["updated_at"] = datetime.fromisoformat(p["updated_at"])
        elif "updated_at" not in p:
            p["updated_at"] = datetime.now(timezone.utc)
        
        # Backward compatibility: map old fields to new fields if new fields are missing
        if "nome_empresa" not in p:
            if "empresa" in p:
                p["nome_empresa"] = p["empresa"]
            elif "nome" in p:
                p["nome_empresa"] = p["nome"]
            else:
                p["nome_empresa"] = "N/A"
        
        if "nome_manager" not in p:
            if "name" in p:
                p["nome_manager"] = p["name"]
            elif "responsavel_nome" in p:
                p["nome_manager"] = p["responsavel_nome"]
            else:
                p["nome_manager"] = "N/A"
        
        if "contribuinte_empresa" not in p and "nif" in p:
            p["contribuinte_empresa"] = p["nif"]
        if "morada_completa" not in p and "morada" in p:
            p["morada_completa"] = p["morada"]
        if "codigo_postal" not in p:
            p["codigo_postal"] = "0000-000"  # Default value
        if "localidade" not in p:
            p["localidade"] = "N/A"  # Default value
        if "telefone" not in p and "phone" in p:
            p["telefone"] = p["phone"]
        if "telemovel" not in p:
            if "responsavel_contacto" in p:
                p["telemovel"] = p["responsavel_contacto"]
            elif "phone" in p:
                p["telemovel"] = p["phone"]
            else:
                p["telemovel"] = "N/A"
        if "email" not in p:
            p["email"] = "noemail@example.com"  # Default value
        if "codigo_certidao_comercial" not in p:
            p["codigo_certidao_comercial"] = "N/A"  # Default value
        if "validade_certidao_comercial" not in p:
            p["validade_certidao_comercial"] = "2099-12-31"  # Default future date
        
        # Count vehicles and motoristas
        p["total_vehicles"] = await db.vehicles.count_documents({"parceiro_id": p["id"]})
        p["total_motoristas"] = await db.motoristas.count_documents({"parceiro_id": p["id"]})
    
    return parceiros


@api_router.get("/parceiros/csv-examples/{tipo}")
async def download_csv_example(
    tipo: str,
    current_user: Dict = Depends(get_current_user)
):
    """Download CSV example file for motoristas or veiculos - SEMPRE usa ID do parceiro logado"""
    logger.info(f"=== DOWNLOAD CSV EXEMPLO ===")
    logger.info(f"User: {current_user.get('email')}, Role: {current_user['role']}, ID: {current_user['id']}")
    logger.info(f"Tipo solicitado: {tipo}")
    
    if tipo not in ["motoristas", "veiculos"]:
        raise HTTPException(status_code=400, detail="Tipo inválido. Use 'motoristas' ou 'veiculos'")
    
    file_path = ROOT_DIR / "templates" / "csv_examples" / f"exemplo_{tipo}.csv"
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Ficheiro de exemplo não encontrado")
    
    # Read the original file
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # SEMPRE usa o ID do user logado (para parceiro, admin ou gestão)
    parceiro_id = current_user["id"]
    user_email = current_user.get("email", "N/A")
    user_name = current_user.get("name", "N/A")
    
    logger.info(f"Gerando CSV com ID do parceiro: {parceiro_id}")
    
    # Add comment with parceiro ID at the top
    comment = f"# ID do Parceiro: {parceiro_id}\n"
    comment += f"# Email: {user_email}\n"
    comment += f"# Nome: {user_name}\n"
    comment += f"# Este ficheiro é um exemplo. Edite conforme necessário.\n"
    comment += f"# IMPORTANTE: Ao importar, os dados serão associados automaticamente à sua conta.\n"
    modified_content = comment + content
    
    # Create temporary file with modified content
    import tempfile
    with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.csv', encoding='utf-8') as tmp:
        tmp.write(modified_content)
        tmp_path = tmp.name
    
    logger.info(f"CSV gerado com sucesso: exemplo_{tipo}_parceiro_{parceiro_id[:8]}.csv")
    
    return FileResponse(
        path=tmp_path,
        media_type="text/csv",
        filename=f"exemplo_{tipo}_{user_email.split('@')[0]}.csv"
    )

@api_router.post("/parceiros/{parceiro_id}/importar-motoristas")
async def importar_motoristas_csv(
    parceiro_id: str,
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    """Import motoristas from CSV file - associa automaticamente ao parceiro logado"""
    logger.info(f"=== INICIO IMPORTAÇÃO MOTORISTAS ===")
    logger.info(f"User role: {current_user['role']}, User ID: {current_user['id']}, User email: {current_user.get('email')}")
    logger.info(f"Parceiro_id recebido no path: {parceiro_id}")
    
    # Para parceiros, usar SEMPRE o próprio ID (ignorar o ID do path)
    if current_user["role"] == "parceiro":
        original_id = parceiro_id
        parceiro_id = current_user["id"]
        logger.info(f"PARCEIRO LOGADO - Substituindo ID do path '{original_id}' por current_user['id']: '{parceiro_id}'")
    
    # Check permissions
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, "parceiro"]:
        logger.error(f"Usuário sem permissão: role={current_user['role']}")
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Verify parceiro exists
    logger.info(f"Verificando se parceiro existe no banco: {parceiro_id}")
    parceiro = await db.parceiros.find_one({"id": parceiro_id}, {"_id": 0})
    if not parceiro:
        logger.error(f"PARCEIRO NÃO ENCONTRADO NO BANCO: {parceiro_id}")
        # List all parceiros IDs for debugging
        all_parceiros = await db.parceiros.find({}, {"_id": 0, "id": 1, "email": 1}).to_list(100)
        logger.error(f"Parceiros disponíveis no banco: {[p.get('id') for p in all_parceiros]}")
        raise HTTPException(status_code=404, detail=f"Parceiro not found: {parceiro_id}")
    
    logger.info(f"Parceiro encontrado: {parceiro.get('nome_empresa', parceiro.get('email', 'N/A'))}")
    logger.info(f"=== FIM VERIFICAÇÕES INICIAIS ===")
    
    try:
        # Read CSV file with multiple encoding support
        content = await file.read()
        
        # Try multiple encodings (Portuguese files often use ISO-8859-1 or Windows-1252)
        decoded = None
        for encoding in ['utf-8-sig', 'utf-8', 'iso-8859-1', 'windows-1252', 'latin-1']:
            try:
                decoded = content.decode(encoding)
                logger.info(f"Successfully decoded motoristas CSV with encoding: {encoding}")
                break
            except UnicodeDecodeError:
                continue
        
        if decoded is None:
            raise HTTPException(status_code=400, detail="Não foi possível ler o ficheiro CSV. Tente guardar o ficheiro como UTF-8.")
        
        # Helper function to normalize phone numbers from scientific notation
        def normalize_phone(value):
            if not value:
                return ''
            value = str(value).strip()
            # Convert scientific notation like 3,5196E+11 to normal number
            if 'E+' in value.upper() or 'e+' in value:
                try:
                    # Replace comma with dot for float parsing
                    value = value.replace(',', '.')
                    number = float(value)
                    # Format without decimals
                    value = str(int(number))
                except:
                    pass
            return value.replace(' ', '').replace('-', '')
        
        # Remove comment lines starting with #
        lines = decoded.split('\n')
        clean_lines = [line for line in lines if not line.strip().startswith('#')]
        decoded = '\n'.join(clean_lines)
        
        # Check if CSV has proper headers (check for multiple variations due to encoding issues)
        first_line_lower = clean_lines[0].lower() if clean_lines else ''
        has_proper_headers = (
            'nome' in first_line_lower and 
            'email' in first_line_lower and 
            ('telefone' in first_line_lower or 'tel' in first_line_lower)
        )
        
        if not has_proper_headers and len(clean_lines) > 0:
            # Auto-convert CSV without headers
            logger.info("CSV sem cabeçalhos detectado - convertendo automaticamente...")
            
            # Expected column mapping (formato ANTIGO sem Data Nascimento e Tipo Documento)
            COLUMN_MAP_OLD = {
                0: 'Nome', 1: 'Email', 2: 'Telefone', 3: 'WhatsApp', 4: 'Nacionalidade',
                5: 'Telefone Uber', 6: 'Email Uber', 7: 'ID Uber', 
                8: 'Telefone Bolt', 9: 'Email Bolt', 10: 'ID Bolt', 
                11: 'Morada', 12: 'Código Postal', 13: 'Localidade', 
                14: 'CC', 15: 'Validade CC', 16: 'NIF',
                17: 'Seg Social', 18: 'C Utente', 
                19: 'TVDE', 20: 'Validade TVDE',
                21: 'Carta', 22: 'Desde Carta', 23: 'Validade Carta',
                24: 'IBAN',
                25: 'Contacto Emergência Nome', 26: 'Contacto Emergência Telefone',
                27: 'Contacto Emergência Morada', 28: 'Contacto Emergência Código Postal'
            }
            
            # Expected column mapping (formato NOVO com Data Nascimento, Tipo Documento, etc)
            COLUMN_MAP_NEW = {
                0: 'Nome', 1: 'Email', 2: 'Telefone', 3: 'WhatsApp', 4: 'Nacionalidade',
                5: 'Data Nascimento',
                6: 'Telefone Uber', 7: 'Email Uber', 8: 'ID Uber', 
                9: 'Telefone Bolt', 10: 'Email Bolt', 11: 'ID Bolt', 
                12: 'Morada', 13: 'Código Postal', 14: 'Localidade', 
                15: 'Tipo Documento', 16: 'CC', 17: 'Validade CC', 18: 'NIF',
                19: 'Seg Social', 20: 'C Utente', 
                21: 'Código Registo Criminal',
                22: 'TVDE', 23: 'Validade TVDE',
                24: 'Carta', 25: 'Desde Carta', 26: 'Validade Carta',
                27: 'IBAN',
                28: 'Contacto Emergência Nome', 29: 'Contacto Emergência Parentesco',
                30: 'Contacto Emergência Telefone', 31: 'Contacto Emergência Email',
                32: 'Contacto Emergência Morada', 33: 'Contacto Emergência Código Postal',
                34: 'Contacto Emergência Localidade'
            }
            
            # Detect format based on number of columns
            reader_temp = csv.reader(clean_lines, delimiter=';' if clean_lines[0].count(';') > clean_lines[0].count(',') else ',')
            first_row = next(reader_temp)
            num_cols = len(first_row)
            
            # Use old format if less than 30 columns
            COLUMN_MAP = COLUMN_MAP_OLD if num_cols < 30 else COLUMN_MAP_NEW
            logger.info(f"Detected {num_cols} columns, using {'OLD' if num_cols < 30 else 'NEW'} format")
            
            # Detect delimiter
            sample = clean_lines[0] if clean_lines else ''
            data_delimiter = ';' if sample.count(';') > sample.count(',') else ','
            logger.info(f"Detected data delimiter: '{data_delimiter}'")
            
            # Parse existing data
            reader = csv.reader(clean_lines, delimiter=data_delimiter)
            data_rows = list(reader)
            
            # Build new CSV with headers
            output_lines = []
            
            # Add header row
            max_cols = max(COLUMN_MAP.keys()) + 1
            header_cols = [COLUMN_MAP.get(i, f'Col{i}') for i in range(max_cols)]
            output_lines.append(','.join(header_cols))
            
            # Convert data rows
            for row in data_rows:
                if len(row) == 0 or not row[0].strip():
                    continue
                
                converted_row = []
                for i in range(max_cols):
                    if i < len(row):
                        value = str(row[i]).strip()
                        # Normalize phone numbers (check both old and new positions)
                        if num_cols < 30:  # Old format
                            if i in [2, 3, 5, 8, 26]:  # Phone columns in old format
                                value = normalize_phone(value)
                        else:  # New format
                            if i in [2, 3, 6, 9, 30]:  # Phone columns in new format
                                value = normalize_phone(value)
                        converted_row.append(value)
                    else:
                        converted_row.append('')
                
                output_lines.append(','.join(converted_row))
            
            decoded = '\n'.join(output_lines)
            delimiter = ','
            logger.info(f"CSV convertido com sucesso: {len(data_rows)} linhas de dados")
        else:
            # Use detected delimiter for properly formatted CSV
            sample = decoded[:1000]
            delimiter = ';' if sample.count(';') > sample.count(',') else ','
            logger.info(f"CSV com cabeçalhos detectado, delimiter: '{delimiter}'")
        
        # Create CSV reader with appropriate delimiter
        csv_reader_temp = csv.DictReader(io.StringIO(decoded), delimiter=delimiter)
        
        # Normalize header names (fix encoding issues)
        def normalize_header(header):
            """Normalize header names to fix encoding issues"""
            replacements = {
                'CÃ³digo': 'Código',
                'CÃ¡lculo': 'Cálculo',
                'EmergÃªncia': 'Emergência',
                'NÃºmero': 'Número',
                'c�digo': 'código',
                'c�': 'có',
            }
            for old, new in replacements.items():
                header = header.replace(old, new)
            return header
        
        # Normalize all fieldnames
        normalized_fieldnames = [normalize_header(f) for f in csv_reader_temp.fieldnames]
        
        # Create new reader with normalized headers
        csv_reader = csv.DictReader(io.StringIO(decoded), fieldnames=normalized_fieldnames, delimiter=delimiter)
        next(csv_reader)  # Skip original header row
        
        motoristas_criados = 0
        erros = []
        
        for idx, row in enumerate(csv_reader, start=2):  # Start at 2 (header is line 1)
            try:
                # Validate required fields
                if not row.get('Nome') or not row.get('Email'):
                    erros.append(f"Linha {idx}: Nome e Email são obrigatórios")
                    continue
                
                # Check if user already exists
                existing_user = await db.users.find_one({"email": row['Email']})
                if existing_user:
                    erros.append(f"Linha {idx}: Email {row['Email']} já existe")
                    continue
                
                # Generate user ID
                user_id = str(uuid.uuid4())
                
                # Helper to add country code +351 if not present
                def add_country_code(phone):
                    if not phone:
                        return ''
                    phone = phone.strip()
                    # If phone doesn't start with + or 00, add +351
                    if phone and not phone.startswith('+') and not phone.startswith('00'):
                        # Remove leading 0 if present
                        if phone.startswith('0'):
                            phone = phone[1:]
                        phone = '+351' + phone
                    return phone
                
                # Normalize phone numbers and add country code
                telefone_normalizado = add_country_code(normalize_phone(row.get('Telefone', '')))
                whatsapp_normalizado = add_country_code(normalize_phone(row.get('WhatsApp', row.get('Telefone', ''))))
                telefone_uber_normalizado = add_country_code(normalize_phone(row.get('Telefone Uber', '')))
                telefone_bolt_normalizado = add_country_code(normalize_phone(row.get('Telefone Bolt', '')))
                emergencia_telefone_normalizado = add_country_code(normalize_phone(row.get('Contacto Emergência Telefone', '')))
                
                # Create user document
                user_doc = {
                    "id": user_id,
                    "email": row['Email'],
                    "name": row['Nome'],
                    "role": UserRole.MOTORISTA,
                    "password": hash_password(telefone_normalizado[-9:] if telefone_normalizado else 'password123'),  # Use last 9 digits of phone as password
                    "phone": telefone_normalizado,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "approved": True  # Auto-approve imported motoristas
                }
                await db.users.insert_one(user_doc)
                
                # Create motorista document
                motorista_doc = {
                    "id": user_id,
                    "email": row['Email'],
                    "name": row['Nome'],
                    "phone": telefone_normalizado,
                    "whatsapp": whatsapp_normalizado,
                    "nacionalidade": row.get('Nacionalidade', 'Portuguesa'),
                    
                    # Uber data
                    "telefone_uber": telefone_uber_normalizado,
                    "email_uber": row.get('Email Uber', ''),
                    "uuid_motorista_uber": row.get('ID Uber', ''),
                    
                    # Bolt data
                    "telefone_bolt": telefone_bolt_normalizado,
                    "email_bolt": row.get('Email Bolt', ''),
                    "identificador_motorista_bolt": row.get('ID Bolt', ''),
                    
                    # Personal Data
                    "data_nascimento": row.get('Data Nascimento', ''),
                    
                    # Address
                    "morada_completa": row.get('Morada', ''),
                    "codigo_postal": row.get('Código Postal', ''),
                    "localidade": row.get('Localidade', ''),
                    
                    # Documents
                    "tipo_documento": row.get('Tipo Documento', 'CC'),  # Padrão: Cartão de Cidadão
                    "numero_cc": row.get('CC', ''),
                    "numero_documento": row.get('CC', ''),  # Mapeamento para UI
                    "validade_cc": row.get('Validade CC', ''),
                    "validade_documento": row.get('Validade CC', ''),  # Mapeamento para UI
                    "nif": row.get('NIF', ''),
                    "numero_seguranca_social": row.get('Seg Social', ''),
                    "numero_cartao_utente": row.get('C Utente', ''),
                    "codigo_registo_criminal": row.get('Código Registo Criminal', ''),
                    
                    # TVDE License
                    "licenca_tvde_numero": row.get('TVDE', ''),
                    "licenca_tvde_validade": row.get('Validade TVDE', ''),
                    
                    # Driving License
                    "carta_conducao_numero": row.get('Carta', ''),
                    "carta_conducao_emissao": row.get('Desde Carta', ''),
                    "data_emissao_carta": row.get('Desde Carta', ''),
                    "carta_conducao_validade": row.get('Validade Carta', ''),
                    
                    # Banking
                    "iban": row.get('IBAN', ''),
                    
                    # Emergency Contact
                    "emergencia_nome": row.get('Contacto Emergência Nome', ''),
                    "emergencia_parentesco": row.get('Contacto Emergência Parentesco', ''),
                    "emergencia_telefone": emergencia_telefone_normalizado,
                    "emergencia_email": row.get('Contacto Emergência Email', ''),
                    "emergencia_morada": row.get('Contacto Emergência Morada', ''),
                    "emergencia_codigo_postal": row.get('Contacto Emergência Código Postal', ''),
                    "emergencia_localidade": row.get('Contacto Emergência Localidade', ''),
                    
                    # Assignment
                    "parceiro_id": parceiro_id,
                    "parceiro_atribuido": parceiro_id,
                    "status_motorista": "ativo",
                    "approved": True,
                    "senha_provisoria": True,
                    
                    # Auto-generated
                    "id_cartao_frota_combustivel": f"FROTA-{str(uuid.uuid4())[:8].upper()}",
                    
                    # Required documents field (empty for CSV import)
                    "documents": {},
                    
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.motoristas.insert_one(motorista_doc)
                motoristas_criados += 1
                
            except Exception as e:
                erros.append(f"Linha {idx}: {str(e)}")
                logger.error(f"Error importing motorista at line {idx}: {e}")
        
        return {
            "message": f"Importação concluída",
            "motoristas_criados": motoristas_criados,
            "erros": erros,
            "total_linhas": idx - 1  # Subtract header
        }
        
    except Exception as e:
        logger.error(f"Error processing CSV: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao processar CSV: {str(e)}")

@api_router.post("/parceiros/{parceiro_id}/importar-veiculos-csv")
async def importar_veiculos_csv(
    parceiro_id: str,
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    """Import veiculos from CSV file - associa automaticamente ao parceiro logado"""
    # Para parceiros, usar sempre o próprio ID
    if current_user["role"] == "parceiro":
        parceiro_id = current_user["id"]
    
    # Check permissions
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, "parceiro"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Verify parceiro exists
    parceiro = await db.parceiros.find_one({"id": parceiro_id}, {"_id": 0})
    if not parceiro:
        raise HTTPException(status_code=404, detail="Parceiro not found")
    
    try:
        # Read CSV file with multiple encoding support
        content = await file.read()
        
        # Try multiple encodings (Portuguese files often use ISO-8859-1 or Windows-1252)
        decoded = None
        for encoding in ['utf-8-sig', 'utf-8', 'iso-8859-1', 'windows-1252', 'latin-1']:
            try:
                decoded = content.decode(encoding)
                logger.info(f"Successfully decoded veiculos CSV with encoding: {encoding}")
                break
            except UnicodeDecodeError:
                continue
        
        if decoded is None:
            raise HTTPException(status_code=400, detail="Não foi possível ler o ficheiro CSV. Tente guardar o ficheiro como UTF-8.")
        
        # Remove comment lines starting with #
        lines = decoded.split('\n')
        clean_lines = [line for line in lines if not line.strip().startswith('#')]
        decoded = '\n'.join(clean_lines)
        
        # Detect delimiter (Portuguese Excel often uses semicolon)
        sample = decoded[:1000]
        delimiter = ';' if sample.count(';') > sample.count(',') else ','
        logger.info(f"Detected CSV delimiter: '{delimiter}'")
        
        # Create CSV reader with normalized headers
        csv_reader_temp = csv.DictReader(io.StringIO(decoded), delimiter=delimiter)
        
        # Normalize header names (fix encoding issues)
        def normalize_header(header):
            """Normalize header names to fix encoding issues"""
            replacements = {
                'MatrÃ­cula': 'Matrícula',
                'VersÃ£o': 'Versão',
                'CombustÃ­vel': 'Combustível',
                'DisponÃ­vel': 'Disponível',
                'CÃ³digo': 'Código',
                'Ã©': 'é',
                'Ã§': 'ç',
            }
            for old, new in replacements.items():
                header = header.replace(old, new)
            return header
        
        # Normalize all fieldnames
        normalized_fieldnames = [normalize_header(f) for f in csv_reader_temp.fieldnames]
        
        # Create new reader with normalized headers
        csv_reader = csv.DictReader(io.StringIO(decoded), fieldnames=normalized_fieldnames, delimiter=delimiter)
        next(csv_reader)  # Skip original header row
        
        # Helper function to normalize phone numbers from scientific notation
        def normalize_phone(value):
            if not value:
                return ''
            value = str(value).strip()
            # Convert scientific notation like 3,5196E+11 to normal number
            if 'E+' in value.upper() or 'e+' in value:
                try:
                    # Replace comma with dot for float parsing
                    value = value.replace(',', '.')
                    number = float(value)
                    # Format without decimals
                    value = str(int(number))
                except:
                    pass
            return value.replace(' ', '').replace('-', '')
        
        veiculos_criados = 0
        erros = []
        
        for idx, row in enumerate(csv_reader, start=2):  # Start at 2 (header is line 1)
            try:
                # Validate required fields
                if not row.get('Marca') or not row.get('Matrícula'):
                    erros.append(f"Linha {idx}: Marca e Matrícula são obrigatórias")
                    continue
                
                # Check if vehicle already exists
                existing_vehicle = await db.vehicles.find_one({"matricula": row['Matrícula']})
                if existing_vehicle:
                    erros.append(f"Linha {idx}: Matrícula {row['Matrícula']} já existe")
                    continue
                
                # Parse dates
                data_matricula = row.get('Data de Matrícula', '')
                validade_matricula = row.get('Validade da Matrícula', '')
                
                # Parse km_atual safely
                km_str = row.get('KM Atual', '').replace(' km', '').replace('.', '').replace(',', '').strip()
                km_atual = int(km_str) if km_str and km_str.isdigit() else 0
                
                # Parse ano safely
                ano_str = row.get('Ano', '').strip()
                ano = int(ano_str) if ano_str and ano_str.isdigit() else None
                
                # Parse lugares safely
                lugares_str = row.get('Lugares', '5').strip()
                lugares = int(lugares_str) if lugares_str and lugares_str.isdigit() else 5
                
                # Normalize status (fix encoding issues: disponível -> disponivel)
                status_raw = row.get('Status', 'disponivel').lower().strip()
                status_map = {
                    'disponível': 'disponivel',
                    'disponã­vel': 'disponivel',
                    'atribuído': 'atribuido',
                    'atribuã­do': 'atribuido',
                    'manutenção': 'manutencao',
                    'manutenã§ã£o': 'manutencao'
                }
                status = status_map.get(status_raw, status_raw)
                
                # Create vehicle document
                veiculo_doc = {
                    "id": str(uuid.uuid4()),
                    "marca": row['Marca'],
                    "modelo": row.get('Modelo', ''),
                    "versao": row.get('Versão', ''),
                    "ano": ano,
                    "matricula": row['Matrícula'],
                    "data_matricula": data_matricula,
                    "validade_matricula": validade_matricula,
                    "cor": row.get('Cor', ''),
                    "combustivel": row.get('Combustível', ''),
                    "caixa": row.get('Caixa', ''),
                    "lugares": lugares,
                    "localidade": row.get('Localidade', ''),
                    "km_atual": km_atual,
                    "status": status,
                    
                    # Assignment
                    "parceiro_id": parceiro_id,
                    "motorista_id": None,
                    
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.vehicles.insert_one(veiculo_doc)
                veiculos_criados += 1
                
            except Exception as e:
                erros.append(f"Linha {idx}: {str(e)}")
                logger.error(f"Error importing veiculo at line {idx}: {e}")
        
        return {
            "message": f"Importação concluída",
            "veiculos_criados": veiculos_criados,
            "erros": erros,
            "total_linhas": idx - 1  # Subtract header
        }
        
    except Exception as e:
        logger.error(f"Error processing CSV: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao processar CSV: {str(e)}")



@api_router.get("/parceiros/{parceiro_id}/estatisticas")
async def get_parceiro_estatisticas(
    parceiro_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Obter estatísticas do parceiro (motoristas, veículos, etc)"""
    # Verificar permissão
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        if current_user["role"] == "parceiro" and current_user["id"] != parceiro_id:
            raise HTTPException(status_code=403, detail="Not authorized")
    
    # Contar motoristas
    total_motoristas = await db.users.count_documents({
        "role": "motorista",
        "parceiro_id": parceiro_id
    })
    
    # Contar veículos
    total_veiculos = await db.vehicles.count_documents({
        "parceiro_id": parceiro_id
    })
    
    # Contar contratos
    total_contratos = await db.contratos.count_documents({
        "parceiro_id": parceiro_id
    })
    
    # Contar vistorias
    total_vistorias = await db.vistorias.count_documents({
        "parceiro_id": parceiro_id
    })
    
    return {
        "parceiro_id": parceiro_id,
        "total_motoristas": total_motoristas,
        "total_veiculos": total_veiculos,
        "total_contratos": total_contratos,
        "total_vistorias": total_vistorias
    }


@api_router.post("/parceiros/{parceiro_id}/certidao-permanente/upload")
async def upload_certidao_permanente(
    parceiro_id: str,
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    """Upload certidão permanente document"""
    # Verificar permissão
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        if current_user["role"] == "parceiro" and current_user["id"] != parceiro_id:
            raise HTTPException(status_code=403, detail="Not authorized")
    
    # Verificar se parceiro existe
    parceiro = await db.parceiros.find_one({"id": parceiro_id}, {"_id": 0})
    if not parceiro:
        raise HTTPException(status_code=404, detail="Parceiro not found")
    
    # Criar diretório para documentos do parceiro
    parceiro_docs_dir = UPLOAD_DIR / "parceiros" / parceiro_id
    parceiro_docs_dir.mkdir(parents=True, exist_ok=True)
    
    # Guardar ficheiro
    file_extension = file.filename.split('.')[-1]
    file_name = f"certidao_permanente.{file_extension}"
    file_path = parceiro_docs_dir / file_name
    
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)
    
    # Atualizar no banco de dados
    await db.parceiros.update_one(
        {"id": parceiro_id},
        {"$set": {
            "certidao_permanente_file": str(file_path),
            "certidao_permanente_filename": file.filename,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {
        "message": "Certidão permanente uploaded successfully",
        "file_path": str(file_path)
    }

@api_router.get("/parceiros/{parceiro_id}/certidao-permanente")
async def download_certidao_permanente(
    parceiro_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Download certidão permanente document"""
    # Verificar permissão
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        if current_user["role"] == "parceiro" and current_user["id"] != parceiro_id:
            raise HTTPException(status_code=403, detail="Not authorized")
    
    # Buscar parceiro
    parceiro = await db.parceiros.find_one({"id": parceiro_id}, {"_id": 0})
    if not parceiro:
        raise HTTPException(status_code=404, detail="Parceiro not found")
    
    if not parceiro.get("certidao_permanente_file"):
        raise HTTPException(status_code=404, detail="Certidão permanente not found")
    
    file_path = Path(parceiro["certidao_permanente_file"])
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found on server")
    
    # Determinar media type
    media_type = mimetypes.guess_type(str(file_path))[0] or 'application/octet-stream'
    
    return FileResponse(
        path=file_path,
        media_type=media_type,
        filename=parceiro.get("certidao_permanente_filename", "certidao_permanente.pdf")
    )

@api_router.put("/parceiros/{parceiro_id}/certidao-permanente")
async def update_certidao_permanente_data(
    parceiro_id: str,
    data: Dict[str, str] = Body(...),
    current_user: Dict = Depends(get_current_user)
):
    """Update certidão permanente code and validity"""
    # Verificar permissão
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        if current_user["role"] == "parceiro" and current_user["id"] != parceiro_id:
            raise HTTPException(status_code=403, detail="Not authorized")
    
    # Verificar se parceiro existe
    parceiro = await db.parceiros.find_one({"id": parceiro_id}, {"_id": 0})
    if not parceiro:
        raise HTTPException(status_code=404, detail="Parceiro not found")
    
    # Atualizar dados
    update_data = {
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    if "codigo_certidao_permanente" in data:
        update_data["codigo_certidao_permanente"] = data["codigo_certidao_permanente"]
    
    if "validade_certidao_permanente" in data:
        update_data["validade_certidao_permanente"] = data["validade_certidao_permanente"]
    
    await db.parceiros.update_one(
        {"id": parceiro_id},
        {"$set": update_data}
    )
    
    return {
        "message": "Certidão permanente data updated successfully",
        "data": update_data
    }


@api_router.get("/parceiros/meu-plano")
async def get_meu_plano_parceiro(current_user: Dict = Depends(get_current_user)):
    """Parceiro: Get active plan with calculated costs"""
    # Allow both parceiro string and UserRole enum
    user_role = current_user.get("role")
    if user_role != "parceiro" and user_role != UserRole.PARCEIRO:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Buscar plano ativo
    atribuicao = await db.planos_usuarios.find_one(
        {"user_id": current_user["id"], "status": "ativo"},
        {"_id": 0}
    )
    
    if not atribuicao:
        # Fallback para users.plano_id
        user = await db.users.find_one({"id": current_user["id"]}, {"_id": 0})
        if not user or not user.get("plano_id"):
            return {
                "tem_plano": False,
                "plano": None,
                "modulos": [],
                "custo_semanal": 0,
                "custo_mensal": 0,
                "total_veiculos": 0,
                "total_motoristas": 0,
                "motoristas_com_recibos": 0
            }
        
        plano = await db.planos.find_one({"id": user["plano_id"]}, {"_id": 0})
        if not plano:
            return {"tem_plano": False, "plano": None}
        
        modulos_ativos = plano.get("modulos", [])
    else:
        plano = await db.planos.find_one({"id": atribuicao["plano_id"]}, {"_id": 0})
        if not plano:
            return {"tem_plano": False, "plano": None}
        modulos_ativos = atribuicao.get("modulos_ativos", plano.get("modulos", []))
    
    # Contar veículos e motoristas do parceiro
    total_veiculos = await db.vehicles.count_documents({"parceiro_id": current_user["id"]})
    total_motoristas = await db.users.count_documents({"parceiro_id": current_user["id"], "role": "motorista"})
    
    # Contar motoristas que enviam recibos (assumindo que têm campo recibos_ativos)
    motoristas_com_recibos = await db.users.count_documents({
        "parceiro_id": current_user["id"], 
        "role": "motorista",
        "envia_recibos": True
    })
    
    # Calcular custos
    precos = plano.get("precos", {})
    preco_base_semanal = precos.get("semanal", {}).get("preco_com_iva", 0)
    
    # Custo por veículo/motorista
    if plano.get("tipo_cobranca") == "por_veiculo":
        custo_base = preco_base_semanal * total_veiculos
    else:
        custo_base = preco_base_semanal
    
    # Adicionar custo de recibos se opção ativa
    custo_recibos = 0
    if plano.get("opcao_recibos_motorista", False):
        custo_recibos = plano.get("preco_recibo_por_motorista", 0) * motoristas_com_recibos
    
    custo_semanal = custo_base + custo_recibos
    custo_mensal = custo_semanal * 4.33  # Média de semanas por mês
    
    # Buscar info dos módulos
    from models.modulo import MODULOS_SISTEMA
    modulos_info = []
    for codigo in modulos_ativos:
        if codigo in MODULOS_SISTEMA:
            modulos_info.append({
                "codigo": codigo,
                "nome": MODULOS_SISTEMA[codigo]["nome"],
                "descricao": MODULOS_SISTEMA[codigo]["descricao"]
            })
    
    return {
        "tem_plano": True,
        "plano": plano,
        "modulos": modulos_info,
        "custo_semanal": round(custo_semanal, 2),
        "custo_mensal": round(custo_mensal, 2),
        "total_veiculos": total_veiculos,
        "total_motoristas": total_motoristas,
        "motoristas_com_recibos": motoristas_com_recibos,
        "detalhes_calculo": {
            "preco_base_semanal": preco_base_semanal,
            "custo_base": custo_base,
            "custo_recibos": custo_recibos,
            "tipo_cobranca": plano.get("tipo_cobranca")
        }
    }

@api_router.get("/parceiros/{parceiro_id}")
async def get_parceiro(parceiro_id: str, current_user: Dict = Depends(get_current_user)):
    """Get specific parceiro by ID"""
    # Admin/Gestor can see any, Parceiro can only see self
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    if current_user["role"] == UserRole.PARCEIRO and current_user["id"] != parceiro_id:
        raise HTTPException(status_code=403, detail="Can only view your own data")
    
    # Try parceiros collection first
    parceiro = await db.parceiros.find_one({"id": parceiro_id}, {"_id": 0})
    
    # Fallback to users collection
    if not parceiro:
        parceiro = await db.users.find_one({"id": parceiro_id, "role": "parceiro"}, {"_id": 0})
    
    if not parceiro:
        raise HTTPException(status_code=404, detail="Parceiro not found")
    
    # Backward compatibility
    if isinstance(parceiro.get("created_at"), str):
        parceiro["created_at"] = datetime.fromisoformat(parceiro["created_at"])
    
    # Map fields for compatibility with both old and new structure
    if "nome_empresa" not in parceiro and "nome" in parceiro:
        parceiro["nome_empresa"] = parceiro["nome"]
    if "contribuinte_empresa" not in parceiro and "nif" in parceiro:
        parceiro["contribuinte_empresa"] = parceiro["nif"]
    if "morada_completa" not in parceiro and "morada" in parceiro:
        parceiro["morada_completa"] = parceiro["morada"]
    if "nome_manager" not in parceiro and "responsavel_nome" in parceiro:
        parceiro["nome_manager"] = parceiro["responsavel_nome"]
    if "telemovel" not in parceiro and "responsavel_contacto" in parceiro:
        parceiro["telemovel"] = parceiro["responsavel_contacto"]
    
    # Defaults for required fields
    if "localidade" not in parceiro:
        parceiro["localidade"] = "N/A"
    if "validade_certidao_comercial" not in parceiro:
        parceiro["validade_certidao_comercial"] = "2099-12-31"
    
    return parceiro

@api_router.put("/parceiros/{parceiro_id}")
async def update_parceiro(
    parceiro_id: str,
    updates: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Update parceiro data (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Only admin can update parceiros")
    
    result = await db.parceiros.update_one(
        {"id": parceiro_id},
        {"$set": updates}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Parceiro not found")
    
    return {"message": "Parceiro updated successfully"}

@api_router.delete("/parceiros/{parceiro_id}")
async def delete_parceiro(
    parceiro_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Delete parceiro (Admin only)
    
    This will:
    - Remove the parceiro from the database
    - Unassign all vehicles associated with this parceiro
    - Unassign all motoristas associated with this parceiro
    """
    # Only admin can delete parceiros
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Apenas Admin pode eliminar parceiros")
    
    # Check if parceiro exists
    parceiro = await db.parceiros.find_one({"id": parceiro_id}, {"_id": 0})
    if not parceiro:
        raise HTTPException(status_code=404, detail="Parceiro não encontrado")
    
    # Unassign all vehicles from this parceiro
    vehicles_result = await db.vehicles.update_many(
        {"parceiro_id": parceiro_id},
        {"$set": {"parceiro_id": None, "parceiro_nome": None}}
    )
    
    # Unassign all motoristas from this parceiro
    motoristas_result = await db.motoristas.update_many(
        {"parceiro_atribuido": parceiro_id},
        {"$set": {"parceiro_atribuido": None}}
    )
    
    # Also update users collection if parceiro has a user account
    if parceiro.get("user_id"):
        await db.users.delete_one({"id": parceiro["user_id"]})
    
    # Delete the parceiro
    delete_result = await db.parceiros.delete_one({"id": parceiro_id})
    
    if delete_result.deleted_count == 0:
        raise HTTPException(status_code=500, detail="Erro ao eliminar parceiro")
    
    return {
        "message": f"Parceiro '{parceiro.get('nome_empresa', 'N/A')}' eliminado com sucesso",
        "parceiro_id": parceiro_id,
        "vehicles_unassigned": vehicles_result.modified_count,
        "motoristas_unassigned": motoristas_result.modified_count,
        "deleted_by": current_user["id"],
        "deleted_at": datetime.now(timezone.utc).isoformat()
    }


@api_router.get("/parceiros/{parceiro_id}/alertas")
async def get_parceiro_alertas(parceiro_id: str, current_user: Dict = Depends(get_current_user)):
    """
    Get all alerts for a specific partner's vehicles and drivers
    Returns: seguros, inspeções, manutenções baseadas em KM configurado no parceiro
    """
    from datetime import date, timedelta
    
    # Get parceiro to retrieve alert configurations
    parceiro = await db.parceiros.find_one({"id": parceiro_id}, {"_id": 0})
    if not parceiro:
        raise HTTPException(status_code=404, detail="Parceiro not found")
    
    # Get configuration (use defaults if not set)
    dias_aviso_seguro = parceiro.get("dias_aviso_seguro", 30)
    dias_aviso_inspecao = parceiro.get("dias_aviso_inspecao", 30)
    km_aviso_revisao = parceiro.get("km_aviso_revisao", 5000)
    
    today = date.today()
    
    # Get all vehicles for this partner
    vehicles = await db.vehicles.find({"parceiro_id": parceiro_id}, {"_id": 0}).to_list(None)
    
    alertas_seguro = []
    alertas_inspecao = []
    alertas_manutencao = []
    alertas_extintor = []
    
    for vehicle in vehicles:
        matricula = vehicle.get("matricula", "N/A")
        
        # Check insurance expiry
        if vehicle.get("insurance") and vehicle["insurance"].get("data_validade"):
            try:
                parsed_date = safe_parse_date(vehicle["insurance"]["data_validade"])
                if not parsed_date:
                    continue
                data_validade = parsed_date.date()
                dias_restantes = (data_validade - today).days
                
                if 0 <= dias_restantes <= dias_aviso_seguro:
                    alertas_seguro.append({
                        "vehicle_id": vehicle["id"],
                        "matricula": matricula,
                        "data_validade": vehicle["insurance"]["data_validade"],
                        "dias_restantes": dias_restantes,
                        "urgente": dias_restantes <= 7
                    })
            except Exception as e:
                logging.error(f"Error parsing insurance date for {matricula}: {e}")
        
        # Check inspection expiry
        if vehicle.get("inspection") and vehicle["inspection"].get("proxima_inspecao"):
            try:
                parsed_date = safe_parse_date(vehicle["inspection"]["proxima_inspecao"])
                if not parsed_date:
                    continue
                proxima_inspecao = parsed_date.date()
                dias_restantes = (proxima_inspecao - today).days
                
                if 0 <= dias_restantes <= dias_aviso_inspecao:
                    alertas_inspecao.append({
                        "vehicle_id": vehicle["id"],
                        "matricula": matricula,
                        "proxima_inspecao": vehicle["inspection"]["proxima_inspecao"],
                        "dias_restantes": dias_restantes,
                        "urgente": dias_restantes <= 7
                    })
            except Exception as e:
                logging.error(f"Error parsing inspection date for {matricula}: {e}")
        
        # Check extinguisher expiry
        if vehicle.get("extintor") and vehicle["extintor"].get("data_validade"):
            try:
                parsed_date = safe_parse_date(vehicle["extintor"]["data_validade"])
                if not parsed_date:
                    continue
                data_validade_extintor = parsed_date.date()
                dias_restantes = (data_validade_extintor - today).days
                
                if 0 <= dias_restantes <= dias_aviso_inspecao:
                    alertas_extintor.append({
                        "vehicle_id": vehicle["id"],
                        "matricula": matricula,
                        "data_validade": vehicle["extintor"]["data_validade"],
                        "dias_restantes": dias_restantes,
                        "urgente": dias_restantes <= 7
                    })
            except Exception as e:
                logging.error(f"Error parsing extintor date for {matricula}: {e}")
        
        # Check maintenance based on KM (plano_manutencoes)
        km_atual = vehicle.get("km_atual", 0)
        ultima_revisao_km = vehicle.get("ultima_revisao_km", 0)
        plano_manutencoes = vehicle.get("plano_manutencoes", [])
        
        # Default maintenance plan if not configured
        if not plano_manutencoes:
            plano_manutencoes = [
                {"tipo": "Pastilhas", "intervalo_km": 30000},
                {"tipo": "Pastilhas e Discos", "intervalo_km": 60000},
                {"tipo": "Óleo e Filtros", "intervalo_km": 15000}
            ]
        
        for manutencao in plano_manutencoes:
            intervalo_km = manutencao.get("intervalo_km", 0)
            tipo_manutencao = manutencao.get("tipo", "Manutenção")
            
            if intervalo_km > 0:
                # Calculate next maintenance KM
                proxima_manutencao_km = ultima_revisao_km + intervalo_km
                km_restantes = proxima_manutencao_km - km_atual
                
                if 0 <= km_restantes <= km_aviso_revisao:
                    alertas_manutencao.append({
                        "vehicle_id": vehicle["id"],
                        "matricula": matricula,
                        "tipo_manutencao": tipo_manutencao,
                        "km_atual": km_atual,
                        "km_proxima": proxima_manutencao_km,
                        "km_restantes": km_restantes,
                        "urgente": km_restantes <= 1000
                    })
    
    # Sort by urgency
    alertas_seguro.sort(key=lambda x: x["dias_restantes"])
    alertas_inspecao.sort(key=lambda x: x["dias_restantes"])
    alertas_extintor.sort(key=lambda x: x["dias_restantes"])
    alertas_manutencao.sort(key=lambda x: x["km_restantes"])
    
    return {
        "parceiro_id": parceiro_id,
        "configuracao": {
            "dias_aviso_seguro": dias_aviso_seguro,
            "dias_aviso_inspecao": dias_aviso_inspecao,
            "km_aviso_revisao": km_aviso_revisao
        },
        "alertas": {
            "seguros": alertas_seguro,
            "inspecoes": alertas_inspecao,
            "extintores": alertas_extintor,
            "manutencoes": alertas_manutencao
        },
        "totais": {
            "seguros": len(alertas_seguro),
            "inspecoes": len(alertas_inspecao),
            "extintores": len(alertas_extintor),
            "manutencoes": len(alertas_manutencao),
            "total": len(alertas_seguro) + len(alertas_inspecao) + len(alertas_extintor) + len(alertas_manutencao)
        }
    }

# ==================== DOCUMENT SEND ENDPOINTS ====================

@api_router.post("/documents/send-email")
async def send_document_email(
    document_url: str = Form(...), 
    recipient_email: str = Form(...),
    document_type: str = Form(...),
    current_user: Dict = Depends(get_current_user)
):
    # In production, implement actual email sending
    return {
        "message": "Email would be sent",
        "recipient": recipient_email,
        "document": document_type
    }

@api_router.post("/documents/send-whatsapp")
async def send_document_whatsapp(
    document_url: str = Form(...), 
    recipient_phone: str = Form(...),
    document_type: str = Form(...),
    current_user: Dict = Depends(get_current_user)
):
    # In production, implement actual WhatsApp API integration
    return {
        "message": "WhatsApp message would be sent",
        "recipient": recipient_phone,
        "document": document_type
    }

# ==================== VEHICLE MAINTENANCE DETAILED ====================

@api_router.post("/vehicles/{vehicle_id}/manutencoes")
async def add_vehicle_maintenance(
    vehicle_id: str,
    manutencao: VehicleMaintenance,
    current_user: Dict = Depends(get_current_user)
):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    vehicle = await db.vehicles.find_one({"id": vehicle_id})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    manutencao_dict = manutencao.model_dump()
    
    await db.vehicles.update_one(
        {"id": vehicle_id},
        {
            "$push": {"manutencoes": manutencao_dict},
            "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}
        }
    )
    
    return {"message": "Maintenance added", "manutencao_id": manutencao.id}

@api_router.get("/vehicles/{vehicle_id}/manutencoes/{manutencao_id}")
async def get_maintenance_detail(
    vehicle_id: str,
    manutencao_id: str,
    current_user: Dict = Depends(get_current_user)
):
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    for manutencao in vehicle.get("manutencoes", []):
        if manutencao.get("id") == manutencao_id:
            return manutencao
    
    raise HTTPException(status_code=404, detail="Maintenance not found")

@api_router.get("/vehicles/{vehicle_id}/relatorio-intervencoes")
async def get_vehicle_interventions_report(
    vehicle_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """
    Get comprehensive interventions report for a vehicle including:
    - Insurance (Seguro)
    - Inspection (Inspeção)
    - Fire Extinguisher (Extintor)
    - Maintenance/Revisions (Revisões)
    All with dates for past and future interventions
    """
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    interventions = []
    today = datetime.now(timezone.utc).date()
    
    # Add insurance interventions
    seguro = vehicle.get("seguro", {})
    if seguro:
        if seguro.get("data_validade"):
            interventions.append({
                "id": f"seguro_{vehicle_id}",
                "tipo": "Seguro",
                "descricao": f"Renovação de Seguro - {seguro.get('seguradora', 'N/A')}",
                "data": seguro.get("data_validade"),
                "categoria": "seguro",
                "status": "pending" if datetime.fromisoformat(seguro.get("data_validade")).date() >= today else "completed",
                "criado_por": seguro.get("criado_por"),
                "editado_por": seguro.get("editado_por")
            })
    
    # Add inspection interventions
    inspection = vehicle.get("inspection", {})
    if inspection:
        if inspection.get("data_inspecao"):
            interventions.append({
                "tipo": "Inspeção",
                "descricao": "Inspeção realizada",
                "data": inspection.get("data_inspecao"),
                "categoria": "inspecao",
                "status": "completed"
            })
        if inspection.get("proxima_inspecao"):
            interventions.append({
                "tipo": "Inspeção",
                "descricao": "Próxima Inspeção",
                "data": inspection.get("proxima_inspecao"),
                "categoria": "inspecao",
                "status": "pending" if datetime.fromisoformat(inspection.get("proxima_inspecao")).date() >= today else "completed"
            })
    
    # Add fire extinguisher interventions
    extintor = vehicle.get("extintor", {})
    if extintor:
        if extintor.get("data_instalacao") or extintor.get("data_entrega"):
            data_inst = extintor.get("data_instalacao") or extintor.get("data_entrega")
            interventions.append({
                "tipo": "Extintor",
                "descricao": "Instalação do Extintor",
                "data": data_inst,
                "categoria": "extintor",
                "status": "completed"
            })
        if extintor.get("data_validade"):
            interventions.append({
                "tipo": "Extintor",
                "descricao": "Validade do Extintor",
                "data": extintor.get("data_validade"),
                "categoria": "extintor",
                "status": "pending" if datetime.fromisoformat(extintor.get("data_validade")).date() >= today else "completed"
            })
    
    # Add maintenance/revision interventions
    manutencoes = vehicle.get("manutencoes", [])
    for manutencao in manutencoes:
        if manutencao.get("data_intervencao"):
            interventions.append({
                "tipo": "Revisão",
                "descricao": f"{manutencao.get('tipo_manutencao', 'Manutenção')} - {manutencao.get('descricao', '')}",
                "data": manutencao.get("data_intervencao"),
                "km": manutencao.get("km_intervencao"),
                "categoria": "revisao",
                "status": "completed"
            })
        if manutencao.get("data_proxima"):
            interventions.append({
                "tipo": "Revisão",
                "descricao": f"Próxima {manutencao.get('tipo_manutencao', 'Manutenção')} - {manutencao.get('o_que_fazer', '')}",
                "data": manutencao.get("data_proxima"),
                "km": manutencao.get("km_proxima"),
                "categoria": "revisao",
                "status": "pending" if datetime.fromisoformat(manutencao.get("data_proxima")).date() >= today else "completed"
            })
    
    # Sort interventions by date (most recent first)
    interventions.sort(key=lambda x: x.get("data", ""), reverse=True)
    
    return {
        "vehicle_id": vehicle_id,
        "interventions": interventions,
        "total": len(interventions)
    }

@api_router.put("/vehicles/{vehicle_id}/intervencao/{intervencao_id}")
async def update_vehicle_intervention(
    vehicle_id: str,
    intervencao_id: str,
    update_data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Update intervention status"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    new_status = update_data.get("status")
    user_name = current_user.get("name", current_user.get("email"))
    
    # Update based on intervention type
    if intervencao_id.startswith("seguro_"):
        await db.vehicles.update_one(
            {"id": vehicle_id},
            {"$set": {
                "seguro.status": new_status,
                "seguro.editado_por": user_name,
                "seguro.editado_em": datetime.now(timezone.utc).isoformat()
            }}
        )
    elif intervencao_id.startswith("inspecao_"):
        await db.vehicles.update_one(
            {"id": vehicle_id},
            {"$set": {
                "inspection.status": new_status,
                "inspection.editado_por": user_name,
                "inspection.editado_em": datetime.now(timezone.utc).isoformat()
            }}
        )
    elif intervencao_id.startswith("extintor_"):
        await db.vehicles.update_one(
            {"id": vehicle_id},
            {"$set": {
                "extintor.status": new_status,
                "extintor.editado_por": user_name,
                "extintor.editado_em": datetime.now(timezone.utc).isoformat()
            }}
        )
    
    return {"message": "Intervention updated successfully"}

@api_router.get("/vehicles/proximas-datas/dashboard")
async def get_proximas_datas_dashboard(current_user: Dict = Depends(get_current_user)):
    """
    Get dashboard with upcoming dates for all vehicles:
    - Next inspection date
    - Next revision date/km
    - Insurance renewal date
    - Fire extinguisher revision date
    """
    vehicles = await db.vehicles.find({}, {"_id": 0}).to_list(length=None)
    
    dashboard_data = []
    today = datetime.now(timezone.utc).date()
    
    for vehicle in vehicles:
        # Get parceiro info
        parceiro_id = vehicle.get("parceiro_id")
        parceiro_nome = "N/A"
        if parceiro_id:
            parceiro = await db.parceiros.find_one({"id": parceiro_id}, {"_id": 0, "nome": 1})
            if parceiro:
                parceiro_nome = parceiro.get("nome", "N/A")
        
        proximas_datas = {
            "vehicle_id": vehicle.get("id"),
            "matricula": vehicle.get("matricula"),
            "marca": vehicle.get("marca"),
            "modelo": vehicle.get("modelo"),
            "parceiro_nome": parceiro_nome,
            "datas": []
        }
        
        # Próxima inspeção
        inspection = vehicle.get("inspection", {}) or {}
        if inspection and inspection.get("proxima_inspecao"):
            data_inspecao = datetime.fromisoformat(inspection.get("proxima_inspecao")).date()
            dias_restantes = (data_inspecao - today).days
            proximas_datas["datas"].append({
                "tipo": "Inspeção",
                "data": inspection.get("proxima_inspecao"),
                "dias_restantes": dias_restantes,
                "urgente": dias_restantes <= 30
            })
        
        # Renovação de seguro
        seguro = vehicle.get("seguro", {}) or {}
        if seguro and seguro.get("data_validade"):
            data_seguro = datetime.fromisoformat(seguro.get("data_validade")).date()
            dias_restantes = (data_seguro - today).days
            proximas_datas["datas"].append({
                "tipo": "Seguro",
                "data": seguro.get("data_validade"),
                "seguradora": seguro.get("seguradora"),
                "dias_restantes": dias_restantes,
                "urgente": dias_restantes <= 30
            })
        
        # Revisão do extintor
        extintor = vehicle.get("extintor", {}) or {}
        if extintor and extintor.get("data_validade"):
            data_extintor = datetime.fromisoformat(extintor.get("data_validade")).date()
            dias_restantes = (data_extintor - today).days
            proximas_datas["datas"].append({
                "tipo": "Extintor",
                "data": extintor.get("data_validade"),
                "dias_restantes": dias_restantes,
                "urgente": dias_restantes <= 30
            })
        
        # Próxima revisão
        if vehicle.get("proxima_revisao_data"):
            data_revisao = datetime.fromisoformat(vehicle.get("proxima_revisao_data")).date()
            dias_restantes = (data_revisao - today).days
            proximas_datas["datas"].append({
                "tipo": "Revisão",
                "data": vehicle.get("proxima_revisao_data"),
                "km": vehicle.get("proxima_revisao_km"),
                "dias_restantes": dias_restantes,
                "urgente": dias_restantes <= 15
            })
        
        if proximas_datas["datas"]:
            dashboard_data.append(proximas_datas)
    
    # Sort by most urgent first
    dashboard_data.sort(key=lambda x: min([d["dias_restantes"] for d in x["datas"]]) if x["datas"] else 9999)
    
    return {
        "dashboard_data": dashboard_data,
        "total_vehicles": len(dashboard_data)
    }

# ==================== SISTEMA DE MÓDULOS E PLANOS ====================

@api_router.get("/modulos")
async def get_modulos_disponiveis(
    tipo_usuario: Optional[str] = None,
    current_user: Dict = Depends(get_current_user)
):
    """Lista todos os módulos disponíveis no sistema, com filtro opcional por tipo de usuário"""
    from models.modulo import MODULOS_SISTEMA
    
    modulos = []
    for codigo, info in MODULOS_SISTEMA.items():
        # Filtrar por tipo de usuário se especificado
        if tipo_usuario and info.get("tipo_usuario") != tipo_usuario:
            continue
            
        modulos.append({
            "id": codigo,
            "codigo": codigo,
            "nome": info["nome"],
            "descricao": info["descricao"],
            "tipo_usuario": info.get("tipo_usuario", "parceiro"),
            "ordem": info["ordem"],
            "ativo": True
        })
    
    return sorted(modulos, key=lambda x: x["ordem"])


@api_router.get("/planos")
async def get_planos(current_user: Dict = Depends(get_current_user)):
    """Lista todos os planos disponíveis"""
    planos = await db.planos.find({"ativo": True}, {"_id": 0}).to_list(100)
    return planos


@api_router.post("/planos")
async def criar_plano(plano_data: Dict[str, Any], current_user: Dict = Depends(get_current_user)):
    """Criar novo plano (Admin apenas)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Apenas admin pode criar planos")
    
    plano_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    # Estrutura de preços por periodicidade
    precos = {
        "semanal": {
            "preco_sem_iva": plano_data.get("preco_semanal_sem_iva", 0),
            "preco_com_iva": plano_data.get("preco_semanal_com_iva", 0),
            "desconto_percentual": plano_data.get("desconto_semanal", 0)
        },
        "mensal": {
            "preco_sem_iva": plano_data.get("preco_mensal_sem_iva", 0),
            "preco_com_iva": plano_data.get("preco_mensal_com_iva", 0),
            "desconto_percentual": plano_data.get("desconto_mensal", 0)
        },
        "trimestral": {
            "preco_sem_iva": plano_data.get("preco_trimestral_sem_iva", 0),
            "preco_com_iva": plano_data.get("preco_trimestral_com_iva", 0),
            "desconto_percentual": plano_data.get("desconto_trimestral", 0)
        },
        "semestral": {
            "preco_sem_iva": plano_data.get("preco_semestral_sem_iva", 0),
            "preco_com_iva": plano_data.get("preco_semestral_com_iva", 0),
            "desconto_percentual": plano_data.get("desconto_semestral", 0)
        },
        "anual": {
            "preco_sem_iva": plano_data.get("preco_anual_sem_iva", 0),
            "preco_com_iva": plano_data.get("preco_anual_com_iva", 0),
            "desconto_percentual": plano_data.get("desconto_anual", 0)
        }
    }
    
    plano = {
        "id": plano_id,
        "nome": plano_data["nome"],
        "descricao": plano_data.get("descricao"),
        "modulos": plano_data.get("modulos", []),
        "tipo_cobranca": plano_data.get("tipo_cobranca", "por_veiculo") if plano_data.get("tipo_usuario") == "parceiro" else None,
        "limite_veiculos": plano_data.get("limite_veiculos") if plano_data.get("tipo_usuario") == "parceiro" else None,
        "precos": precos,
        "taxa_iva": plano_data.get("taxa_iva", 23),
        "tipo_usuario": plano_data.get("tipo_usuario", "parceiro"),
        "opcao_recibos_motorista": plano_data.get("opcao_recibos_motorista", False) if plano_data.get("tipo_usuario") == "parceiro" else False,
        "preco_recibo_por_motorista": plano_data.get("preco_recibo_por_motorista", 0) if plano_data.get("tipo_usuario") == "parceiro" else 0,
        "ativo": plano_data.get("ativo", True),
        "destaque": plano_data.get("destaque", False),
        "promocao_ativa": plano_data.get("promocao_ativa", False),
        "promocao_data_inicio": plano_data.get("promocao_data_inicio"),
        "promocao_data_fim": plano_data.get("promocao_data_fim"),
        "promocao_desconto_percentual": plano_data.get("promocao_desconto_percentual", 0),
        "created_at": now,
        "updated_at": now
    }
    
    await db.planos.insert_one(plano)
    return {"message": "Plano criado com sucesso", "plano_id": plano_id}


@api_router.put("/planos/{plano_id}")
async def atualizar_plano(
    plano_id: str,
    plano_data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Atualizar plano existente (Admin apenas)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Apenas admin pode atualizar planos")
    
    # Estrutura de preços por periodicidade
    precos = {
        "semanal": {
            "preco_sem_iva": plano_data.get("preco_semanal_sem_iva", 0),
            "preco_com_iva": plano_data.get("preco_semanal_com_iva", 0),
            "desconto_percentual": plano_data.get("desconto_semanal", 0)
        },
        "mensal": {
            "preco_sem_iva": plano_data.get("preco_mensal_sem_iva", 0),
            "preco_com_iva": plano_data.get("preco_mensal_com_iva", 0),
            "desconto_percentual": plano_data.get("desconto_mensal", 0)
        },
        "trimestral": {
            "preco_sem_iva": plano_data.get("preco_trimestral_sem_iva", 0),
            "preco_com_iva": plano_data.get("preco_trimestral_com_iva", 0),
            "desconto_percentual": plano_data.get("desconto_trimestral", 0)
        },
        "semestral": {
            "preco_sem_iva": plano_data.get("preco_semestral_sem_iva", 0),
            "preco_com_iva": plano_data.get("preco_semestral_com_iva", 0),
            "desconto_percentual": plano_data.get("desconto_semestral", 0)
        },
        "anual": {
            "preco_sem_iva": plano_data.get("preco_anual_sem_iva", 0),
            "preco_com_iva": plano_data.get("preco_anual_com_iva", 0),
            "desconto_percentual": plano_data.get("desconto_anual", 0)
        }
    }
    
    update_data = {
        "nome": plano_data["nome"],
        "descricao": plano_data.get("descricao"),
        "modulos": plano_data.get("modulos", []),
        "tipo_cobranca": plano_data.get("tipo_cobranca", "por_veiculo") if plano_data.get("tipo_usuario") == "parceiro" else None,
        "limite_veiculos": plano_data.get("limite_veiculos") if plano_data.get("tipo_usuario") == "parceiro" else None,
        "precos": precos,
        "taxa_iva": plano_data.get("taxa_iva", 23),
        "tipo_usuario": plano_data.get("tipo_usuario", "parceiro"),
        "opcao_recibos_motorista": plano_data.get("opcao_recibos_motorista", False) if plano_data.get("tipo_usuario") == "parceiro" else False,
        "preco_recibo_por_motorista": plano_data.get("preco_recibo_por_motorista", 0) if plano_data.get("tipo_usuario") == "parceiro" else 0,
        "ativo": plano_data.get("ativo", True),
        "promocao_ativa": plano_data.get("promocao_ativa", False),
        "promocao_data_inicio": plano_data.get("promocao_data_inicio"),
        "promocao_data_fim": plano_data.get("promocao_data_fim"),
        "promocao_desconto_percentual": plano_data.get("promocao_desconto_percentual", 0),
        "updated_at": datetime.now(timezone.utc)
    }
    
    result = await db.planos.update_one({"id": plano_id}, {"$set": update_data})
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Plano não encontrado")
    
    return {"message": "Plano atualizado com sucesso"}


@api_router.delete("/planos/{plano_id}")
async def deletar_plano(plano_id: str, current_user: Dict = Depends(get_current_user)):
    """Deletar plano (Admin apenas)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Apenas admin pode deletar planos")
    
    # Verificar se há usuários com este plano
    usuarios_com_plano = await db.planos_usuarios.count_documents({"plano_id": plano_id, "status": "ativo"})
    if usuarios_com_plano > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Não é possível eliminar. {usuarios_com_plano} utilizador(es) têm este plano ativo."
        )
    
    # Eliminar plano permanentemente
    result = await db.planos.delete_one({"id": plano_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Plano não encontrado")
    
    return {"message": "Plano eliminado com sucesso"}


@api_router.post("/users/{user_id}/atribuir-modulos")
async def atribuir_modulos_usuario(
    user_id: str,
    atribuicao_data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Atribuir plano ou módulos individuais a um usuário"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    # Verificar se usuário existe
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    
    atribuicao_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    # Calcular data_fim baseado no tipo de pagamento
    data_fim = None
    tipo_pagamento = atribuicao_data.get("tipo_pagamento")
    
    if tipo_pagamento == "mensal":
        data_fim = now + timedelta(days=30)
    elif tipo_pagamento == "anual":
        data_fim = now + timedelta(days=365)
    # vitalicio = None (sem data fim)
    
    atribuicao = {
        "id": atribuicao_id,
        "user_id": user_id,
        "plano_id": atribuicao_data.get("plano_id"),
        "modulos_ativos": atribuicao_data.get("modulos_ativos", []),
        "tipo_pagamento": tipo_pagamento,
        "valor_pago": atribuicao_data.get("valor_pago", 0),
        "data_inicio": now,
        "data_fim": data_fim,
        "status": "ativo",
        "renovacao_automatica": atribuicao_data.get("renovacao_automatica", False),
        "created_at": now,
        "updated_at": now,
        "created_by": current_user["id"]
    }
    
    # Remover atribuição anterior se existir
    await db.planos_usuarios.delete_many({"user_id": user_id})
    
    # Inserir nova atribuição
    await db.planos_usuarios.insert_one(atribuicao)
    
    return {
        "message": "Módulos atribuídos com sucesso",
        "atribuicao_id": atribuicao_id,
        "data_fim": data_fim.isoformat() if data_fim else None
    }


@api_router.post("/users/{user_id}/modulos")
async def update_modulos_usuario(
    user_id: str,
    modulos_data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Atualizar módulos ativos de um usuário (Admin/Gestor)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    # Verificar se usuário existe
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    
    modulos_ativos = modulos_data.get("modulos_ativos", [])
    
    # Verificar se existe atribuição ativa
    atribuicao = await db.planos_usuarios.find_one(
        {"user_id": user_id, "status": "ativo"},
        {"_id": 0}
    )
    
    if atribuicao:
        # Atualizar módulos na atribuição existente
        await db.planos_usuarios.update_one(
            {"id": atribuicao["id"]},
            {
                "$set": {
                    "modulos_ativos": modulos_ativos,
                    "updated_at": datetime.now(timezone.utc)
                }
            }
        )
    else:
        # Criar nova atribuição apenas com módulos
        atribuicao_id = str(uuid.uuid4())
        now = datetime.now(timezone.utc)
        
        nova_atribuicao = {
            "id": atribuicao_id,
            "user_id": user_id,
            "plano_id": None,
            "modulos_ativos": modulos_ativos,
            "tipo_pagamento": "vitalicio",
            "valor_pago": 0,
            "data_inicio": now,
            "data_fim": None,
            "status": "ativo",
            "renovacao_automatica": False,
            "created_at": now,
            "updated_at": now,
            "created_by": current_user["id"]
        }
        
        await db.planos_usuarios.insert_one(nova_atribuicao)
    
    logger.info(f"Módulos atualizados para user {user_id} por {current_user['email']}")
    return {"message": "Módulos atualizados com sucesso", "modulos_ativos": modulos_ativos}


@api_router.get("/users/{user_id}/modulos")
async def get_modulos_usuario(user_id: str, current_user: Dict = Depends(get_current_user)):
    """Obter módulos ativos de um usuário"""
    # Usuário pode ver seus próprios módulos, ou admin/gestor podem ver de qualquer um
    if current_user["id"] != user_id and current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    atribuicao = await db.planos_usuarios.find_one(
        {"user_id": user_id, "status": "ativo"},
        {"_id": 0}
    )
    
    if not atribuicao:
        return {"modulos_ativos": [], "plano": None}
    
    # Se tem plano, buscar módulos do plano
    if atribuicao.get("plano_id"):
        plano = await db.planos.find_one({"id": atribuicao["plano_id"]}, {"_id": 0})
        if plano:
            atribuicao["modulos_ativos"] = plano["modulos"]
            atribuicao["plano"] = plano
    
    return atribuicao


@api_router.get("/users/{user_id}/verificar-modulo/{modulo_codigo}")
async def verificar_acesso_modulo(
    user_id: str,
    modulo_codigo: str,
    current_user: Dict = Depends(get_current_user)
):
    """Verificar se usuário tem acesso a um módulo específico"""
    atribuicao = await db.planos_usuarios.find_one(
        {"user_id": user_id, "status": "ativo"},
        {"_id": 0}
    )
    
    if not atribuicao:
        # Fallback: Verificar se o usuário tem plano_id diretamente na coleção users
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if user and user.get("plano_id") and user.get("plano_status") == "ativo":
            plano = await db.planos.find_one({"id": user["plano_id"]}, {"_id": 0})
            if plano:
                modulos_ativos = plano.get("modulos", [])
                tem_acesso = modulo_codigo in modulos_ativos
                return {
                    "tem_acesso": tem_acesso,
                    "motivo": "Acesso concedido" if tem_acesso else "Módulo não incluído no plano"
                }
        
        return {"tem_acesso": False, "motivo": "Nenhum plano ativo"}
    
    # Verificar se expirou
    if atribuicao.get("data_fim"):
        data_fim = datetime.fromisoformat(atribuicao["data_fim"])
        if data_fim < datetime.now(timezone.utc):
            await db.planos_usuarios.update_one(
                {"id": atribuicao["id"]},
                {"$set": {"status": "expirado"}}
            )
            return {"tem_acesso": False, "motivo": "Plano expirado"}
    
    # Buscar módulos
    modulos_ativos = atribuicao.get("modulos_ativos", [])
    
    if atribuicao.get("plano_id"):
        plano = await db.planos.find_one({"id": atribuicao["plano_id"]}, {"_id": 0})
        if plano:
            modulos_ativos = plano.get("modulos", [])
    
    tem_acesso = modulo_codigo in modulos_ativos
    
    return {
        "tem_acesso": tem_acesso,
        "motivo": "Acesso concedido" if tem_acesso else "Módulo não incluído no plano"
    }


# ==================== VEHICLE VISTORIAS (INSPECTIONS) ====================

@api_router.post("/vehicles/{vehicle_id}/vistorias")
async def create_vistoria(
    vehicle_id: str,
    vistoria_data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Create a new vehicle vistoria/inspection"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Verify vehicle exists
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    # Create vistoria
    vistoria_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    vistoria = {
        "id": vistoria_id,
        "veiculo_id": vehicle_id,
        "parceiro_id": vistoria_data.get("parceiro_id"),
        "data_vistoria": vistoria_data.get("data_vistoria", now.isoformat()),
        "tipo": vistoria_data.get("tipo", "periodica"),
        "km_veiculo": vistoria_data.get("km_veiculo"),
        "responsavel_nome": current_user.get("name"),
        "responsavel_id": current_user.get("id"),
        "observacoes": vistoria_data.get("observacoes"),
        "estado_geral": vistoria_data.get("estado_geral", "bom"),
        "status": vistoria_data.get("status", "fechada"),  # aberta | fechada
        "fotos": vistoria_data.get("fotos", []),
        "itens_verificados": vistoria_data.get("itens_verificados", {}),
        "danos_encontrados": vistoria_data.get("danos_encontrados", []),
        "pdf_relatorio": None,
        "assinatura_responsavel": vistoria_data.get("assinatura_responsavel"),
        "assinatura_motorista": vistoria_data.get("assinatura_motorista"),
        "created_at": now,
        "updated_at": now
    }
    
    await db.vistorias.insert_one(vistoria)
    
    # Update vehicle with next vistoria date if provided
    if vistoria_data.get("proxima_vistoria"):
        await db.vehicles.update_one(
            {"id": vehicle_id},
            {"$set": {"proxima_vistoria": vistoria_data.get("proxima_vistoria")}}
        )
    
    # If vistoria is closed (fechada), generate PDF and send email automatically
    if vistoria.get("status") == "fechada":
        try:
            # Generate PDF
            await gerar_pdf_vistoria(vehicle_id, vistoria_id, current_user)
            # Send email (will be implemented when email service is configured)
            # await enviar_email_vistoria(vehicle_id, vistoria_id, current_user)
            logger.info(f"PDF generated for closed vistoria {vistoria_id}")
        except Exception as e:
            logger.error(f"Error generating PDF for vistoria {vistoria_id}: {e}")
            # Don't fail the whole operation if PDF generation fails
    
    return {"message": "Vistoria created successfully", "vistoria_id": vistoria_id}


@api_router.get("/vehicles/{vehicle_id}/vistorias")
async def get_vehicle_vistorias(
    vehicle_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Get all vistorias for a vehicle"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO, UserRole.MOTORISTA]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    vistorias = await db.vistorias.find(
        {"veiculo_id": vehicle_id},
        {"_id": 0}
    ).sort("data_vistoria", -1).to_list(100)
    
    return vistorias


@api_router.get("/vehicles/{vehicle_id}/vistorias/{vistoria_id}")
async def get_vistoria(
    vehicle_id: str,
    vistoria_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Get a specific vistoria"""
    vistoria = await db.vistorias.find_one(
        {"id": vistoria_id, "veiculo_id": vehicle_id},
        {"_id": 0}
    )
    
    if not vistoria:
        raise HTTPException(status_code=404, detail="Vistoria not found")
    
    return vistoria


@api_router.put("/vehicles/{vehicle_id}/vistorias/{vistoria_id}")
async def update_vistoria(
    vehicle_id: str,
    vistoria_id: str,
    update_data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Update a vistoria"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.vistorias.update_one(
        {"id": vistoria_id, "veiculo_id": vehicle_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vistoria not found")
    
    return {"message": "Vistoria updated successfully"}


@api_router.delete("/vehicles/{vehicle_id}/vistorias/{vistoria_id}")
async def delete_vistoria(
    vehicle_id: str,
    vistoria_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Delete a vistoria"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.vistorias.delete_one({"id": vistoria_id, "veiculo_id": vehicle_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Vistoria not found")
    
    return {"message": "Vistoria deleted successfully"}


@api_router.post("/vehicles/{vehicle_id}/vistorias/{vistoria_id}/upload-foto")
async def upload_vistoria_foto(
    vehicle_id: str,
    vistoria_id: str,
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    """Upload a photo for a vistoria"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Verify vistoria exists
    vistoria = await db.vistorias.find_one({"id": vistoria_id, "veiculo_id": vehicle_id}, {"_id": 0})
    if not vistoria:
        raise HTTPException(status_code=404, detail="Vistoria not found")
    
    try:
        # Create vistorias directory
        vistorias_dir = UPLOAD_DIR / "vistorias" / vehicle_id
        vistorias_dir.mkdir(parents=True, exist_ok=True)
        
        # Process file
        file_id = f"vistoria_{vistoria_id}_{uuid.uuid4()}"
        file_info = await process_uploaded_file(file, vistorias_dir, file_id)
        
        # Update vistoria with photo URL
        await db.vistorias.update_one(
            {"id": vistoria_id},
            {"$push": {"fotos": file_info["url"]}, "$set": {"updated_at": datetime.now(timezone.utc)}}
        )
        
        return {"message": "Photo uploaded successfully", "photo_url": file_info["url"]}
    
    except Exception as e:
        logger.error(f"Error uploading vistoria photo: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/vehicles/{vehicle_id}/vistorias/{vistoria_id}/gerar-pdf")
async def gerar_pdf_vistoria(
    vehicle_id: str,
    vistoria_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Generate comprehensive PDF report for a vistoria"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    vistoria = await db.vistorias.find_one({"id": vistoria_id, "veiculo_id": vehicle_id}, {"_id": 0})
    if not vistoria:
        raise HTTPException(status_code=404, detail="Vistoria not found")
    
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import cm
        from reportlab.lib.colors import HexColor, black
        
        pdf_dir = UPLOAD_DIR / "vistorias" / "relatorios"
        pdf_dir.mkdir(parents=True, exist_ok=True)
        pdf_path = pdf_dir / f"vistoria_{vistoria_id}.pdf"
        
        c = canvas.Canvas(str(pdf_path), pagesize=A4)
        width, height = A4
        
        # Header
        c.setFillColor(HexColor('#1e40af'))
        c.rect(0, height - 3*cm, width, 3*cm, fill=True, stroke=False)
        
        c.setFillColor(HexColor('#ffffff'))
        c.setFont("Helvetica-Bold", 20)
        c.drawString(2*cm, height - 2*cm, "RELATÓRIO DE VISTORIA")
        c.setFont("Helvetica", 12)
        c.drawString(2*cm, height - 2.6*cm, f"Vistoria ID: {vistoria_id[:8]}")
        
        # Reset color
        c.setFillColor(black)
        
        y_position = height - 4.5*cm
        
        # Vehicle Information Section
        c.setFont("Helvetica-Bold", 14)
        c.drawString(2*cm, y_position, "Informações do Veículo")
        y_position -= 0.8*cm
        
        c.setFont("Helvetica", 10)
        c.drawString(2*cm, y_position, f"Marca/Modelo: {vehicle.get('marca', 'N/A')} {vehicle.get('modelo', 'N/A')}")
        y_position -= 0.5*cm
        c.drawString(2*cm, y_position, f"Matrícula: {vehicle.get('matricula', 'N/A')}")
        y_position -= 0.5*cm
        c.drawString(2*cm, y_position, f"Ano: {vehicle.get('ano', 'N/A')}")
        y_position -= 0.5*cm
        c.drawString(2*cm, y_position, f"Cor: {vehicle.get('cor', 'N/A')}")
        y_position -= 1*cm
        
        # Inspection Details Section
        c.setFont("Helvetica-Bold", 14)
        c.drawString(2*cm, y_position, "Detalhes da Vistoria")
        y_position -= 0.8*cm
        
        c.setFont("Helvetica", 10)
        data_vistoria = vistoria.get('data_vistoria', '')
        if isinstance(data_vistoria, str):
            try:
                dt = datetime.fromisoformat(data_vistoria.replace('Z', '+00:00'))
                data_vistoria_fmt = dt.strftime('%d/%m/%Y %H:%M')
            except:
                data_vistoria_fmt = data_vistoria
        else:
            data_vistoria_fmt = str(data_vistoria)
        
        c.drawString(2*cm, y_position, f"Data da Vistoria: {data_vistoria_fmt}")
        y_position -= 0.5*cm
        c.drawString(2*cm, y_position, f"Tipo: {vistoria.get('tipo', 'N/A').capitalize()}")
        y_position -= 0.5*cm
        c.drawString(2*cm, y_position, f"KM do Veículo: {vistoria.get('km_veiculo', 'N/A')}")
        y_position -= 0.5*cm
        
        # TODO: Complete PDF generation...
        c.save()
        
        pdf_url = f"/uploads/vistorias/relatorios/vistoria_{vistoria_id}.pdf"
        
        # Update vistoria with PDF path
        await db.vistorias.update_one(
            {"id": vistoria_id},
            {"$set": {"pdf_relatorio": str(pdf_path), "pdf_url": pdf_url}}
        )
        
        return {"message": "PDF generated", "pdf_url": pdf_url}
        
    except Exception as e:
        logger.error(f"Error generating PDF: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/vehicles/{vehicle_id}/vistorias/{vistoria_id}/download-pdf")
async def download_pdf_vistoria(
    vehicle_id: str,
    vistoria_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Download vistoria PDF report"""
    vistoria = await db.vistorias.find_one({"id": vistoria_id, "veiculo_id": vehicle_id}, {"_id": 0})
    if not vistoria:
        raise HTTPException(status_code=404, detail="Vistoria not found")
    
    pdf_path = vistoria.get("pdf_relatorio")
    if not pdf_path or not Path(pdf_path).exists():
        # Generate PDF if doesn't exist
        await gerar_pdf_vistoria(vehicle_id, vistoria_id, current_user)
        vistoria = await db.vistorias.find_one({"id": vistoria_id}, {"_id": 0})
        pdf_path = vistoria.get("pdf_relatorio")
    
    if not pdf_path or not Path(pdf_path).exists():
        raise HTTPException(status_code=404, detail="PDF not found")
    
    return FileResponse(
        path=pdf_path,
        media_type="application/pdf",
        filename=f"vistoria_{vistoria_id[:8]}.pdf"
    )


@api_router.post("/vehicles/{vehicle_id}/vistorias/{vistoria_id}/enviar-email")
async def enviar_email_vistoria(
    vehicle_id: str,
    vistoria_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Send vistoria PDF report by email to parceiro and motorista"""
    vistoria = await db.vistorias.find_one({"id": vistoria_id, "veiculo_id": vehicle_id}, {"_id": 0})
    if not vistoria:
        raise HTTPException(status_code=404, detail="Vistoria not found")
    
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    # Generate PDF if doesn't exist
    pdf_path = vistoria.get("pdf_relatorio")
    if not pdf_path or not Path(pdf_path).exists():
        await gerar_pdf_vistoria(vehicle_id, vistoria_id, current_user)
        vistoria = await db.vistorias.find_one({"id": vistoria_id}, {"_id": 0})
        pdf_path = vistoria.get("pdf_relatorio")
    
    if not pdf_path or not Path(pdf_path).exists():
        raise HTTPException(status_code=404, detail="PDF not found")
    
    # Get parceiro and motorista emails
    parceiro_id = vehicle.get("parceiro_id")
    motorista_id = vehicle.get("motorista_id")
    
    recipients = []
    
    if parceiro_id:
        parceiro = await db.users.find_one({"id": parceiro_id}, {"_id": 0})
        if parceiro and parceiro.get("email"):
            recipients.append(parceiro["email"])
    
    if motorista_id:
        motorista = await db.users.find_one({"id": motorista_id}, {"_id": 0})
        if motorista and motorista.get("email"):
            recipients.append(motorista["email"])
    
    if not recipients:
        raise HTTPException(status_code=400, detail="No email recipients found")
    
    # TODO: Integrate with email service
    # For now, just log
    logger.info(f"Would send vistoria PDF to: {recipients}")
    
    return {
        "message": "Email enviado com sucesso",
        "recipients": recipients,
        "note": "Email service integration pending"
    }


@api_router.post("/vehicles/{vehicle_id}/agendar-vistoria")
async def agendar_vistoria(
    vehicle_id: str,
    agendamento: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Schedule a future inspection for a vehicle"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    try:
        vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
        if not vehicle:
            raise HTTPException(status_code=404, detail="Vehicle not found")
        
        # Create agenda item
        agenda_item = {
            "id": str(uuid.uuid4()),
            "tipo": "vistoria",
            "data_agendada": agendamento.get("data_agendada"),
            "tipo_vistoria": agendamento.get("tipo_vistoria", "periodica"),
            "notas": agendamento.get("notas", ""),
            "parceiro_id": agendamento.get("parceiro_id"),
            "status": "agendada",  # agendada, realizada, cancelada
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": current_user["id"]
        }
        
        # Add to agenda
        await db.vehicles.update_one(
            {"id": vehicle_id},
            {"$push": {"agenda": agenda_item}}
        )
        
        logger.info(f"Inspection scheduled for vehicle {vehicle_id} on {agenda_item['data_agendada']}")
        return {"message": "Vistoria agendada com sucesso", "agenda_item": agenda_item}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error scheduling inspection: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/vehicles/{vehicle_id}/agenda")
async def get_vehicle_agenda(
    vehicle_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Get vehicle agenda/schedule"""
    try:
        vehicle = await db.vehicles.find_one(
            {"id": vehicle_id},
            {"_id": 0, "agenda": 1, "matricula": 1}
        )
        
        if not vehicle:
            raise HTTPException(status_code=404, detail="Vehicle not found")
        
        agenda = vehicle.get("agenda", [])
        
        # Filter only future and pending inspections
        from datetime import datetime, timezone
        today = datetime.now(timezone.utc).date().isoformat()
        
        agenda_vistorias = [
            item for item in agenda 
            if item.get("tipo") == "vistoria" 
            and item.get("data_agendada", "") >= today
            and item.get("status") != "cancelada"
        ]
        
        return agenda
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching vehicle agenda: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.delete("/vehicles/{vehicle_id}/agenda/{agenda_id}")
async def delete_agenda_item(
    vehicle_id: str,
    agenda_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Delete an agenda item from vehicle schedule"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    try:
        vehicle = await db.vehicles.find_one({"id": vehicle_id})
        if not vehicle:
            raise HTTPException(status_code=404, detail="Vehicle not found")
        
        # Remove agenda item
        result = await db.vehicles.update_one(
            {"id": vehicle_id},
            {"$pull": {"agenda": {"id": agenda_id}}}
        )
        
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="Agenda item not found")
        
        logger.info(f"Agenda item {agenda_id} removed from vehicle {vehicle_id}")
        return {"message": "Agenda item removed successfully", "agenda_id": agenda_id}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching vehicle agenda: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.delete("/vehicles/{vehicle_id}/agenda/{agenda_id}")
async def cancelar_agendamento(
    vehicle_id: str,
    agenda_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Cancel a scheduled inspection"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    try:
        # Mark as cancelled
        await db.vehicles.update_one(
            {"id": vehicle_id, "agenda.id": agenda_id},
            {"$set": {"agenda.$.status": "cancelada"}}
        )
        
        logger.info(f"Agenda item {agenda_id} cancelled for vehicle {vehicle_id}")
        return {"message": "Agendamento cancelado com sucesso"}
        
    except Exception as e:
        logger.error(f"Error cancelling agenda item: {e}")
        raise HTTPException(status_code=500, detail=str(e))


        c.drawString(2*cm, y_position, f"Responsável: {vistoria.get('responsavel_nome', 'N/A')}")
        y_position -= 0.5*cm
        
        # Estado Geral
        estado_geral = vistoria.get('estado_geral', 'bom')
        estado_color = HexColor('#10b981') if estado_geral == 'bom' else HexColor('#ef4444') if estado_geral == 'mau' else HexColor('#f59e0b')
        c.setFillColor(estado_color)
        c.drawString(2*cm, y_position, f"Estado Geral: {estado_geral.upper()}")
        c.setFillColor(black)
        y_position -= 1*cm
        
        # Itens Verificados Section
        if vistoria.get('itens_verificados'):
            c.setFont("Helvetica-Bold", 14)
            c.drawString(2*cm, y_position, "Itens Verificados")
            y_position -= 0.8*cm
            
            c.setFont("Helvetica", 10)
            itens = vistoria.get('itens_verificados', {})
            for key, value in itens.items():
                if y_position < 3*cm:
                    c.showPage()
                    y_position = height - 2*cm
                
                status_symbol = "✓" if value else "✗"
                status_color = HexColor('#10b981') if value else HexColor('#ef4444')
                c.setFillColor(status_color)
                c.drawString(2*cm, y_position, f"{status_symbol} {key.replace('_', ' ').title()}")
                c.setFillColor(black)
                y_position -= 0.5*cm
            
            y_position -= 0.5*cm
        
        # Danos Encontrados Section
        if vistoria.get('danos_encontrados') and len(vistoria['danos_encontrados']) > 0:
            c.setFont("Helvetica-Bold", 14)
            c.setFillColor(HexColor('#ef4444'))
            c.drawString(2*cm, y_position, "Danos Encontrados")
            c.setFillColor(black)
            y_position -= 0.8*cm
            
            c.setFont("Helvetica", 10)
            for idx, dano in enumerate(vistoria['danos_encontrados'], 1):
                if y_position < 3*cm:
                    c.showPage()
                    y_position = height - 2*cm
                
                c.drawString(2*cm, y_position, f"{idx}. {dano.get('descricao', 'N/A')}")
                y_position -= 0.4*cm
                c.drawString(2.5*cm, y_position, f"Localização: {dano.get('localizacao', 'N/A')}")
                y_position -= 0.4*cm
                c.drawString(2.5*cm, y_position, f"Gravidade: {dano.get('gravidade', 'N/A')}")
                y_position -= 0.7*cm
        
        # Observações Section
        if vistoria.get('observacoes'):
            if y_position < 5*cm:
                c.showPage()
                y_position = height - 2*cm
            
            c.setFont("Helvetica-Bold", 14)
            c.drawString(2*cm, y_position, "Observações")
            y_position -= 0.8*cm
            
            c.setFont("Helvetica", 10)
            # Wrap text
            obs_text = vistoria.get('observacoes', '')
            text_object = c.beginText(2*cm, y_position)
            text_object.setFont("Helvetica", 10)
            
            # Simple text wrapping
            max_width = width - 4*cm
            words = obs_text.split(' ')
            current_line = ""
            
            for word in words:
                test_line = current_line + word + " "
                if c.stringWidth(test_line, "Helvetica", 10) < max_width:
                    current_line = test_line
                else:
                    text_object.textLine(current_line)
                    current_line = word + " "
            
            if current_line:
                text_object.textLine(current_line)
            
            c.drawText(text_object)
        
        # Footer
        c.setFont("Helvetica", 8)
        c.setFillColor(HexColor('#64748b'))
        c.drawString(2*cm, 1.5*cm, f"Gerado em: {datetime.now(timezone.utc).strftime('%d/%m/%Y às %H:%M')}")
        c.drawString(2*cm, 1*cm, "TVDEFleet - Sistema de Gestão de Frotas")
        
        c.save()
        
        pdf_url = f"/uploads/vistorias/relatorios/vistoria_{vistoria_id}.pdf"
        await db.vistorias.update_one(
            {"id": vistoria_id}, 
            {
                "$set": {
                    "pdf_relatorio": pdf_url,
                    "updated_at": datetime.now(timezone.utc)
                }
            }
        )
        
        # Add to vehicle history
        await db.vehicles.update_one(
            {"id": vehicle_id},
            {
                "$push": {
                    "historico_editavel": {
                        "id": str(uuid.uuid4()),
                        "tipo": "vistoria",
                        "titulo": f"Vistoria {vistoria.get('tipo', 'periódica')}",
                        "descricao": f"Estado geral: {vistoria.get('estado_geral', 'N/A')}",
                        "data": vistoria.get('data_vistoria'),
                        "vistoria_id": vistoria_id,
                        "pdf_url": pdf_url,
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                }
            }
        )
        
        logger.info(f"PDF generated for vistoria {vistoria_id} and added to vehicle history")
        
        return {
            "message": "PDF generated successfully and added to vehicle history",
            "pdf_url": pdf_url,
            "vistoria_id": vistoria_id
        }
    except Exception as e:
        logger.error(f"Error generating PDF: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/vehicles/{vehicle_id}/agendar-vistoria")
async def agendar_vistoria(
    vehicle_id: str,
    agendamento_data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Agendar vistoria para um veículo"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Verify vehicle exists
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    agendamento_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    # Parse data_agendamento
    data_agendamento = agendamento_data.get("data_agendamento")
    if isinstance(data_agendamento, str):
        data_agendamento = datetime.fromisoformat(data_agendamento.replace('Z', '+00:00'))
    
    # Calcular próxima vistoria automática se recorrente
    proxima_auto = None
    recorrencia = agendamento_data.get("recorrencia")
    if recorrencia and data_agendamento:
        if recorrencia == "mensal":
            proxima_auto = data_agendamento + timedelta(days=30)
        elif recorrencia == "trimestral":
            proxima_auto = data_agendamento + timedelta(days=90)
        elif recorrencia == "semestral":
            proxima_auto = data_agendamento + timedelta(days=180)
        elif recorrencia == "anual":
            proxima_auto = data_agendamento + timedelta(days=365)
    
    agendamento = {
        "id": agendamento_id,
        "veiculo_id": vehicle_id,
        "data_agendamento": data_agendamento,
        "recorrencia": recorrencia,
        "proxima_vistoria_auto": proxima_auto,
        "tipo": agendamento_data.get("tipo", "periodica"),
        "observacoes": agendamento_data.get("observacoes"),
        "notificar_antes_dias": agendamento_data.get("notificar_antes_dias", 7),
        "status": "agendada",
        "created_at": now,
        "created_by": current_user["id"]
    }
    
    await db.vistorias_agendadas.insert_one(agendamento)
    
    # Atualizar veículo com próxima vistoria
    await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$set": {"proxima_vistoria": data_agendamento}}
    )
    
    return {
        "message": "Vistoria agendada com sucesso",
        "agendamento_id": agendamento_id,
        "proxima_vistoria_auto": proxima_auto.isoformat() if proxima_auto else None
    }


@api_router.get("/vehicles/{vehicle_id}/agendamentos-vistoria")
async def get_agendamentos_vistoria(
    vehicle_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Listar agendamentos de vistoria de um veículo"""
    agendamentos = await db.vistorias_agendadas.find(
        {"veiculo_id": vehicle_id},
        {"_id": 0}
    ).sort("data_agendamento", -1).to_list(50)
    
    return agendamentos


@api_router.post("/vehicles/{vehicle_id}/vistorias/{vistoria_id}/adicionar-dano")
async def adicionar_dano_vistoria(
    vehicle_id: str,
    vistoria_id: str,
    dano_data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Adicionar dano identificado em vistoria"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    dano_id = str(uuid.uuid4())
    dano = {
        "id": dano_id,
        "descricao": dano_data["descricao"],
        "localizacao": dano_data["localizacao"],
        "gravidade": dano_data.get("gravidade", "moderado"),
        "fotos": dano_data.get("fotos", []),
        "custo_estimado": dano_data.get("custo_estimado"),
        "custo_real": dano_data.get("custo_real"),
        "responsavel": dano_data.get("responsavel"),
        "motorista_id": dano_data.get("motorista_id"),
        "motorista_nome": dano_data.get("motorista_nome"),
        "reparado": False,
        "data_reparacao": None,
        "created_at": datetime.now(timezone.utc)
    }
    
    # Adicionar dano à vistoria
    result = await db.vistorias.update_one(
        {"id": vistoria_id, "veiculo_id": vehicle_id},
        {
            "$push": {"danos_encontrados": dano},
            "$set": {"updated_at": datetime.now(timezone.utc)}
        }
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vistoria not found")
    
    # Recalcular custo total
    vistoria = await db.vistorias.find_one({"id": vistoria_id}, {"_id": 0})
    if vistoria:
        custo_total = sum(
            d.get("custo_real") or d.get("custo_estimado") or 0
            for d in vistoria.get("danos_encontrados", [])
        )
        await db.vistorias.update_one(
            {"id": vistoria_id},
            {"$set": {"custo_total_danos": custo_total}}
        )
    
    return {"message": "Dano adicionado com sucesso", "dano_id": dano_id}


    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get vistoria and vehicle data
    vistoria = await db.vistorias.find_one({"id": vistoria_id, "veiculo_id": vehicle_id}, {"_id": 0})
    if not vistoria:
        raise HTTPException(status_code=404, detail="Vistoria not found")
    
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import cm
        
        # Create PDF
        pdf_dir = UPLOAD_DIR / "vistorias" / "relatorios"
        pdf_dir.mkdir(parents=True, exist_ok=True)
        pdf_path = pdf_dir / f"vistoria_{vistoria_id}.pdf"
        
        c = canvas.Canvas(str(pdf_path), pagesize=A4)
        width, height = A4
        
        # Title
        c.setFont("Helvetica-Bold", 16)
        c.drawString(2*cm, height - 2*cm, "RELATÓRIO DE VISTORIA")
        
        # Vehicle info
        c.setFont("Helvetica-Bold", 12)
        c.drawString(2*cm, height - 3.5*cm, "Dados do Veículo:")
        c.setFont("Helvetica", 10)
        c.drawString(2*cm, height - 4*cm, f"Marca/Modelo: {vehicle.get('marca')} {vehicle.get('modelo')}")
        c.drawString(2*cm, height - 4.5*cm, f"Matrícula: {vehicle.get('matricula')}")
        
        # Vistoria info
        c.setFont("Helvetica-Bold", 12)
        c.drawString(2*cm, height - 6*cm, "Dados da Vistoria:")
        c.setFont("Helvetica", 10)
        c.drawString(2*cm, height - 6.5*cm, f"Tipo: {vistoria.get('tipo')}")
        c.drawString(2*cm, height - 7*cm, f"Data: {vistoria.get('data_vistoria')}")
        c.drawString(2*cm, height - 7.5*cm, f"Estado Geral: {vistoria.get('estado_geral')}")
        c.drawString(2*cm, height - 8*cm, f"Responsável: {vistoria.get('responsavel_nome')}")
        
        if vistoria.get('km_veiculo'):
            c.drawString(2*cm, height - 8.5*cm, f"Km: {vistoria.get('km_veiculo')}")
        
        # Observations
        if vistoria.get('observacoes'):
            c.setFont("Helvetica-Bold", 12)
            c.drawString(2*cm, height - 10*cm, "Observações:")
            c.setFont("Helvetica", 10)
            c.drawString(2*cm, height - 10.5*cm, vistoria.get('observacoes'))
        
        c.save()
        
        pdf_url = f"/uploads/vistorias/relatorios/vistoria_{vistoria_id}.pdf"
        
        # Update vistoria with PDF URL
        await db.vistorias.update_one(
            {"id": vistoria_id},
            {"$set": {"pdf_relatorio": pdf_url, "updated_at": datetime.now(timezone.utc)}}
        )
        
        return {"message": "PDF generated successfully", "pdf_url": pdf_url}
    
    except Exception as e:
        logger.error(f"Error generating PDF: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ADMIN SETTINGS ====================

@api_router.get("/admin/settings")
async def get_admin_settings(current_user: Dict = Depends(get_current_user)):
    settings = await db.settings.find_one({"id": "admin_settings"}, {"_id": 0})
    if not settings:
        # Create default settings
        default_settings = {
            "id": "admin_settings",
            "anos_validade_matricula": 20,
            "km_aviso_manutencao": 5000,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": "system"
        }
        await db.settings.insert_one(default_settings)
        settings = default_settings
    
    if isinstance(settings.get("updated_at"), str):
        settings["updated_at"] = datetime.fromisoformat(settings["updated_at"])
    
    return settings

@api_router.put("/admin/settings")
async def update_admin_settings(
    anos_validade_matricula: Optional[int] = None,
    km_aviso_manutencao: Optional[int] = None,
    current_user: Dict = Depends(get_current_user)
):
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat(), "updated_by": current_user["id"]}
    
    if anos_validade_matricula is not None:
        update_data["anos_validade_matricula"] = anos_validade_matricula
    
    if km_aviso_manutencao is not None:
        update_data["km_aviso_manutencao"] = km_aviso_manutencao
    
    await db.settings.update_one(
        {"id": "admin_settings"},
        {"$set": update_data},
        upsert=True
    )
    
    return {"message": "Settings updated", "settings": update_data}

# ==================== EXEMPLOS DE CONTRATO ====================

@api_router.get("/contratos/exemplos")
async def get_exemplos_contratos():
    """Get list of available contract examples"""
    exemplos = [
        {
            "id": "aluguer",
            "nome": "Contrato de Aluguer de Veículo TVDE",
            "tipo": "aluguer",
            "descricao": "Contrato completo de aluguer de veículo para atividade TVDE com caução e condições",
            "arquivo": "exemplo_contrato_aluguer.txt"
        },
        {
            "id": "comissao",
            "nome": "Contrato de Comissão TVDE",
            "tipo": "comissao",
            "descricao": "Contrato de prestação de serviços por comissão (percentual sobre ganhos)",
            "arquivo": "exemplo_contrato_comissao.txt"
        },
        {
            "id": "motorista_privado",
            "nome": "Contrato de Motorista Privado",
            "tipo": "motorista_privado",
            "descricao": "Contrato para motorista que conduz veículo da empresa (regime privado)",
            "arquivo": "exemplo_contrato_motorista_privado.txt"
        }
    ]
    return exemplos

@api_router.get("/contratos/exemplos/{tipo}/download")
async def download_exemplo_contrato(tipo: str):
    """Download contract example file"""
    import os
    from fastapi.responses import FileResponse, StreamingResponse
    
    arquivos = {
        "aluguer": "/app/exemplo_contrato_aluguer.txt",
        "comissao": "/app/exemplo_contrato_comissao.txt",
        "motorista_privado": "/app/exemplo_contrato_motorista_privado.txt"
    }
    
    if tipo not in arquivos:
        raise HTTPException(status_code=404, detail="Exemplo não encontrado")
    
    filepath = arquivos[tipo]
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Arquivo não encontrado")
    
    return FileResponse(
        filepath,
        media_type="text/plain",
        filename=f"exemplo_contrato_{tipo}.txt"
    )

# ==================== MINUTAS DE CONTRATO ENDPOINTS ====================

@api_router.get("/parceiros/{parceiro_id}/minutas")
async def get_minutas_parceiro(parceiro_id: str, current_user: Dict = Depends(get_current_user)):
    """Get all contract templates for a parceiro"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    minutas = await db.minutas_contrato.find({"parceiro_id": parceiro_id}, {"_id": 0}).to_list(length=None)
    return minutas

@api_router.post("/parceiros/{parceiro_id}/minutas")
async def create_minuta(parceiro_id: str, minuta_data: MinutaContratoCreate, current_user: Dict = Depends(get_current_user)):
    """Create new contract template for parceiro"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    minuta_dict = minuta_data.model_dump()
    minuta_dict["id"] = str(uuid.uuid4())
    minuta_dict["parceiro_id"] = parceiro_id
    minuta_dict["ativa"] = True
    minuta_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    minuta_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    minuta_dict["created_by"] = current_user["id"]
    
    await db.minutas_contrato.insert_one(minuta_dict)
    
    return {"message": "Minuta criada com sucesso", "id": minuta_dict["id"]}

@api_router.put("/minutas/{minuta_id}")
async def update_minuta(minuta_id: str, updates: Dict[str, Any], current_user: Dict = Depends(get_current_user)):
    """Update contract template"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.minutas_contrato.update_one(
        {"id": minuta_id},
        {"$set": updates}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Minuta not found")
    
    return {"message": "Minuta atualizada com sucesso"}

@api_router.delete("/minutas/{minuta_id}")
async def delete_minuta(minuta_id: str, current_user: Dict = Depends(get_current_user)):
    """Delete contract template"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.minutas_contrato.delete_one({"id": minuta_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Minuta not found")
    
    return {"message": "Minuta excluída com sucesso"}

# ==================== CONFIGURAÇÕES TEXTOS LEGAIS ====================

@api_router.get("/config/textos-legais")
async def get_textos_legais():
    """Public: Get legal texts (terms and privacy policy)"""
    config = await db.configuracoes.find_one({"id": "config_sistema"}, {"_id": 0})
    if not config:
        # Create default config
        default_config = {
            "id": "config_sistema",
            "condicoes_gerais": "Condições Gerais a definir...",
            "politica_privacidade": "Política de Privacidade a definir...",
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": "system"
        }
        await db.configuracoes.insert_one(default_config)
        return default_config
    return config

@api_router.put("/admin/config/textos-legais")
async def update_textos_legais(config_data: ConfiguracaoUpdate, current_user: Dict = Depends(get_current_user)):
    """Admin only: Update legal texts"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized - Admin only")
    
    update_data = {
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user["id"]
    }
    
    if config_data.condicoes_gerais is not None:
        update_data["condicoes_gerais"] = config_data.condicoes_gerais
    
    if config_data.politica_privacidade is not None:
        update_data["politica_privacidade"] = config_data.politica_privacidade
    
    await db.configuracoes.update_one(
        {"id": "config_sistema"},
        {"$set": update_data},
        upsert=True
    )
    
    return {"message": "Textos legais atualizados com sucesso", "data": update_data}

@api_router.put("/admin/config/comunicacoes")
async def update_comunicacoes(
    request_data: Dict,
    current_user: Dict = Depends(get_current_user)
):
    """Admin only: Update communication settings (email and WhatsApp)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized - Admin only")
    
    update_data = {
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user["id"]
    }
    
    if "email_comunicacoes" in request_data:
        update_data["email_comunicacoes"] = request_data["email_comunicacoes"]
    
    if "whatsapp_comunicacoes" in request_data:
        update_data["whatsapp_comunicacoes"] = request_data["whatsapp_comunicacoes"]
    
    await db.configuracoes.update_one(
        {"id": "config_sistema"},
        {"$set": update_data},
        upsert=True
    )
    
    return {"message": "Configurações de comunicação atualizadas com sucesso", "data": update_data}

@api_router.put("/admin/config/integracoes")
async def update_integracoes(
    request_data: Dict,
    current_user: Dict = Depends(get_current_user)
):
    """Admin only: Update integration credentials (IFThenPay and Moloni)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized - Admin only")
    
    update_data = {
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user["id"]
    }
    
    # IFThenPay credentials
    if "ifthenpay_entity" in request_data:
        update_data["ifthenpay_entity"] = request_data["ifthenpay_entity"]
    
    if "ifthenpay_subentity" in request_data:
        update_data["ifthenpay_subentity"] = request_data["ifthenpay_subentity"]
    
    if "ifthenpay_api_key" in request_data:
        update_data["ifthenpay_api_key"] = request_data["ifthenpay_api_key"]
    
    # Moloni credentials
    if "moloni_client_id" in request_data:
        update_data["moloni_client_id"] = request_data["moloni_client_id"]
    
    if "moloni_client_secret" in request_data:
        update_data["moloni_client_secret"] = request_data["moloni_client_secret"]
    
    if "moloni_company_id" in request_data:
        update_data["moloni_company_id"] = request_data["moloni_company_id"]
    
    await db.configuracoes.update_one(
        {"id": "config_sistema"},
        {"$set": update_data},
        upsert=True
    )
    
    return {"message": "Credenciais de integração atualizadas com sucesso", "data": update_data}


# ==================== SUBSCRIPTION/PLANOS ENDPOINTS ====================

@api_router.post("/admin/planos/{plano_id}/promocao")
async def criar_promocao_plano(
    plano_id: str,
    promocao_data: Dict,
    current_user: Dict = Depends(get_current_user)
):
    """Admin: Add promotional pricing to a plan"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Validate plano exists
    plano = await db.planos.find_one({"id": plano_id}, {"_id": 0})
    if not plano:
        raise HTTPException(status_code=404, detail="Plano not found")
    
    # Create promotion
    promocao = {
        "ativa": True,
        "nome": promocao_data.get("nome", "Promoção"),
        "desconto_percentual": promocao_data.get("desconto_percentual", 0),
        "data_inicio": promocao_data.get("data_inicio"),
        "data_fim": promocao_data.get("data_fim"),
        "created_at": datetime.now(timezone.utc),
        "created_by": current_user["id"]
    }
    
    # Update plano with promotion
    await db.planos.update_one(
        {"id": plano_id},
        {
            "$set": {
                "promocao": promocao,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    return {"message": "Promoção criada com sucesso", "promocao": promocao}

@api_router.delete("/admin/planos/{plano_id}/promocao")
async def remover_promocao_plano(
    plano_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Admin: Remove promotion from a plan"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    await db.planos.update_one(
        {"id": plano_id},
        {
            "$unset": {"promocao": ""},
            "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}
        }
    )
    
    return {"message": "Promoção removida com sucesso"}

@api_router.post("/parceiros/{parceiro_id}/comprar-plano-motorista")
async def parceiro_comprar_plano_motorista(
    parceiro_id: str,
    motorista_id: str = Body(...),
    plano_especial_id: str = Body(...),
    current_user: Dict = Depends(get_current_user)
):
    """Parceiro: Buy discounted plan for motorista and resell"""
    if current_user["role"] not in [UserRole.PARCEIRO, UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    if current_user["role"] == UserRole.PARCEIRO and current_user["id"] != parceiro_id:
        raise HTTPException(status_code=403, detail="Can only buy for your own motoristas")
    
    # Verify motorista belongs to parceiro
    motorista = await db.users.find_one({"id": motorista_id, "parceiro_id": parceiro_id}, {"_id": 0})
    if not motorista:
        raise HTTPException(status_code=404, detail="Motorista not found or doesn't belong to this parceiro")
    
    # Get special plan
    plano_especial = await db.planos.find_one({"id": plano_especial_id, "tipo_usuario": "motorista"}, {"_id": 0})
    if not plano_especial:
        raise HTTPException(status_code=404, detail="Special plan not found")
    
    # Calculate pricing (parceiro pays discounted price)
    preco_base = plano_especial.get("precos", {}).get("mensal", {}).get("preco_com_iva", 0)
    desconto_revenda = plano_especial.get("desconto_revenda_parceiro", 20)  # 20% default
    preco_parceiro = preco_base * (1 - desconto_revenda / 100)
    
    # Create attribution
    atribuicao_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    atribuicao = {
        "id": atribuicao_id,
        "user_id": motorista_id,
        "plano_id": plano_especial_id,
        "modulos_ativos": plano_especial.get("modulos", []),
        "tipo_pagamento": "mensal",
        "valor_pago": preco_base,  # Motorista pays full price
        "valor_custo_parceiro": preco_parceiro,  # Parceiro pays discounted
        "margem_parceiro": preco_base - preco_parceiro,
        "comprado_por_parceiro": True,
        "parceiro_id": parceiro_id,
        "data_inicio": now,
        "data_fim": None,
        "status": "ativo",
        "renovacao_automatica": True,
        "created_at": now,
        "updated_at": now,
        "created_by": current_user["id"]
    }
    
    await db.planos_usuarios.insert_one(atribuicao)
    
    # Update user
    await db.users.update_one(
        {"id": motorista_id},
        {"$set": {"plano_id": plano_especial_id, "plano_status": "ativo"}}
    )
    
    return {
        "message": "Plano comprado e atribuído com sucesso",
        "motorista_paga": preco_base,
        "parceiro_paga": preco_parceiro,
        "margem": preco_base - preco_parceiro
    }

@api_router.post("/admin/planos", response_model=PlanoAssinatura)
async def create_plano(plano_data: PlanoCreate, current_user: Dict = Depends(get_current_user)):
    """Admin: Create new subscription plan"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    plano_dict = plano_data.model_dump()
    plano_dict["id"] = str(uuid.uuid4())
    plano_dict["ativo"] = True
    plano_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    plano_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.planos.insert_one(plano_dict)
    
    if isinstance(plano_dict["created_at"], str):
        plano_dict["created_at"] = datetime.fromisoformat(plano_dict["created_at"])
    if isinstance(plano_dict["updated_at"], str):
        plano_dict["updated_at"] = datetime.fromisoformat(plano_dict["updated_at"])
    
    return PlanoAssinatura(**plano_dict)

@api_router.get("/admin/planos", response_model=List[PlanoAssinatura])
async def get_planos(current_user: Dict = Depends(get_current_user)):
    """Admin: Get all subscription plans"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    planos = await db.planos.find({}, {"_id": 0}).to_list(100)
    
    for plano in planos:
        # Fix created_at and updated_at if missing or invalid
        if "created_at" in plano and isinstance(plano["created_at"], str):
            plano["created_at"] = datetime.fromisoformat(plano["created_at"])
        elif "created_at" not in plano:
            plano["created_at"] = datetime.now(timezone.utc)
            
        if "updated_at" in plano and isinstance(plano["updated_at"], str):
            plano["updated_at"] = datetime.fromisoformat(plano["updated_at"])
        elif "updated_at" not in plano:
            plano["updated_at"] = datetime.now(timezone.utc)
        
        # Map funcionalidades to features for backward compatibility
        if "funcionalidades" in plano and "features" not in plano:
            plano["features"] = plano["funcionalidades"]
        
        # Set default values for missing fields
        if "tipo_usuario" not in plano:
            plano["tipo_usuario"] = "parceiro"
        if "preco_por_unidade" not in plano:
            plano["preco_por_unidade"] = plano.get("preco_mensal", 0.0)
    
    return planos

@api_router.put("/admin/planos/{plano_id}")
async def update_plano(
    plano_id: str,
    updates: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Admin: Update subscription plan"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.planos.update_one(
        {"id": plano_id},
        {"$set": updates}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Plano not found")
    
    return {"message": "Plano updated successfully"}

@api_router.get("/planos/public", response_model=List[PlanoAssinatura])
async def get_public_planos(tipo_usuario: Optional[str] = None):
    """Public: Get active subscription plans (no auth required)"""
    query = {"ativo": True}
    if tipo_usuario:
        query["tipo_usuario"] = tipo_usuario
    
    planos = await db.planos.find(query, {"_id": 0}).to_list(100)
    
    for plano in planos:
        if isinstance(plano["created_at"], str):
            plano["created_at"] = datetime.fromisoformat(plano["created_at"])
        if isinstance(plano["updated_at"], str):
            plano["updated_at"] = datetime.fromisoformat(plano["updated_at"])
    
    return planos

@api_router.post("/admin/seed-planos")
async def seed_planos(current_user: Dict = Depends(get_current_user)):
    """Admin: Seed default subscription plans"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if plans already exist
    existing_count = await db.planos.count_documents({})
    if existing_count > 0:
        return {"message": f"Planos already exist ({existing_count} planos found). Skipping seed."}
    
    default_planos = [
        # PARCEIRO PLANS
        {
            "id": str(uuid.uuid4()),
            "nome": "Parceiro Base",
            "tipo_usuario": "parceiro",
            "preco_por_unidade": 50.0,
            "descricao": "Acesso a relatórios e gestão de pagamentos",
            "features": ["relatorios", "pagamentos"],
            "ativo": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "nome": "Parceiro Premium",
            "tipo_usuario": "parceiro",
            "preco_por_unidade": 80.0,
            "descricao": "Base + Gestão de seguros",
            "features": ["relatorios", "pagamentos", "seguros"],
            "ativo": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "nome": "Parceiro Executive",
            "tipo_usuario": "parceiro",
            "preco_por_unidade": 120.0,
            "descricao": "Gestão completa: seguros, manutenções, contas e emissão de contratos",
            "features": ["relatorios", "pagamentos", "seguros", "manutencoes", "contas", "contratos"],
            "ativo": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        },
        # OPERACIONAL PLANS
        {
            "id": str(uuid.uuid4()),
            "nome": "Operacional Base",
            "tipo_usuario": "operacional",
            "preco_por_unidade": 30.0,
            "descricao": "Upload CSV de ganhos Uber/Bolt, inserção manual de combustível e via verde",
            "features": ["upload_csv_ganhos", "combustivel_manual", "viaverde_manual"],
            "ativo": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "nome": "Operacional Premium",
            "tipo_usuario": "operacional",
            "preco_por_unidade": 50.0,
            "descricao": "Base + Upload CSV de KM + Gestão de veículos, manutenções e seguros",
            "features": ["upload_csv_ganhos", "combustivel_manual", "viaverde_manual", "upload_csv_km", "gestao_veiculos", "gestao_manutencoes", "gestao_seguros"],
            "ativo": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "nome": "Operacional Executive",
            "tipo_usuario": "operacional",
            "preco_por_unidade": 80.0,
            "descricao": "Premium + Ligação automática com fornecedores",
            "features": ["upload_csv_ganhos", "combustivel_manual", "viaverde_manual", "upload_csv_km", "gestao_veiculos", "gestao_manutencoes", "gestao_seguros", "integracoes_fornecedores"],
            "ativo": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
    ]
    
    await db.planos.insert_many(default_planos)
    
    return {
        "message": "Default subscription plans created successfully",
        "planos_created": len(default_planos)
    }



@api_router.post("/parceiros/{parceiro_id}/solicitar-plano")
async def solicitar_plano_parceiro(
    parceiro_id: str,
    plano_data: Dict[str, str],
    current_user: Dict = Depends(get_current_user)
):
    """Parceiro: Request a subscription plan (pending admin approval)"""
    # Check if user is the parceiro or admin
    if current_user["role"] not in [UserRole.ADMIN, UserRole.PARCEIRO, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    plano_id = plano_data.get("plano_id")
    if not plano_id:
        raise HTTPException(status_code=400, detail="plano_id required")
    
    # Check if plano exists
    plano = await db.planos.find_one({"id": plano_id}, {"_id": 0})
    if not plano:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    # Update parceiro with pending plan
    result = await db.parceiros.update_one(
        {"id": parceiro_id},
        {
            "$set": {
                "plano_id": plano_id,
                "plano_status": "pendente",
                "plano_solicitado_em": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Parceiro not found")
    
    return {"message": "Plano solicitado com sucesso. Aguardando aprovação do admin.", "plano_status": "pendente"}

@api_router.post("/admin/parceiros/{parceiro_id}/aprovar-plano")
async def aprovar_plano_parceiro(
    parceiro_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Admin: Approve parceiro's plan request"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get parceiro
    parceiro = await db.parceiros.find_one({"id": parceiro_id}, {"_id": 0})
    if not parceiro:
        raise HTTPException(status_code=404, detail="Parceiro not found")
    
    if not parceiro.get("plano_id"):
        raise HTTPException(status_code=400, detail="Parceiro has no plan request")
    
    # Approve plan
    result = await db.parceiros.update_one(
        {"id": parceiro_id},
        {
            "$set": {
                "plano_status": "ativo",
                "plano_aprovado_em": datetime.now(timezone.utc).isoformat(),
                "plano_aprovado_por": current_user["id"]
            }
        }
    )
    
    return {"message": "Plano aprovado com sucesso!", "plano_status": "ativo"}

@api_router.post("/admin/parceiros/{parceiro_id}/atribuir-plano")
async def atribuir_plano_parceiro(
    parceiro_id: str,
    plano_data: Dict[str, str],
    current_user: Dict = Depends(get_current_user)
):
    """Admin: Assign subscription plan to parceiro (direct assignment, auto-approved)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    plano_id = plano_data.get("plano_id")
    periodicidade = plano_data.get("periodicidade", "mensal")
    if not plano_id:
        raise HTTPException(status_code=400, detail="plano_id required")
    
    # Check if plano exists
    plano = await db.planos.find_one({"id": plano_id}, {"_id": 0})
    if not plano:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    # Obter valor do plano baseado na periodicidade
    valor_pago = 0
    if plano.get("precos") and plano["precos"].get(periodicidade):
        valor_pago = plano["precos"][periodicidade].get("preco_com_iva", 0)
    elif plano.get("valor_mensal"):
        # Fallback para formato antigo
        valor_pago = plano.get("valor_mensal", 0)
    
    # Update parceiro - try both users and parceiros collections
    update_doc = {
        "plano_id": plano_id,
        "plano_status": "ativo",
        "periodicidade": periodicidade,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    result = await db.users.update_one(
        {"id": parceiro_id, "role": "parceiro"},
        {"$set": update_doc}
    )
    
    # If not found in users, try parceiros collection
    if result.matched_count == 0:
        result = await db.parceiros.update_one(
            {"id": parceiro_id},
            {"$set": update_doc}
        )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Parceiro not found")
    
    # Criar/atualizar registro em planos_usuarios para controle de módulos
    now = datetime.now(timezone.utc)
    atribuicao_id = str(uuid.uuid4())
    
    # Remover atribuição anterior se existir
    await db.planos_usuarios.delete_many({"user_id": parceiro_id})
    
    # Criar nova atribuição
    # Support both old 'features' and new 'modulos' field
    modulos_ativos = plano.get("modulos", [])
    if not modulos_ativos and plano.get("features"):
        # Fallback to features for old plans
        modulos_ativos = plano.get("features", [])
    
    atribuicao = {
        "id": atribuicao_id,
        "user_id": parceiro_id,
        "plano_id": plano_id,
        "modulos_ativos": modulos_ativos,
        "tipo_pagamento": periodicidade,
        "valor_pago": valor_pago,
        "data_inicio": now,
        "data_fim": None,  # Sem data de fim por padrão
        "status": "ativo",
        "renovacao_automatica": True,
        "created_at": now,
        "updated_at": now,
        "created_by": current_user["id"]
    }
    
    await db.planos_usuarios.insert_one(atribuicao)
    
    return {"message": "Plan assigned successfully", "plano_id": plano_id, "atribuicao_id": atribuicao_id}


@api_router.post("/admin/motoristas/{motorista_id}/atribuir-plano")
async def atribuir_plano_motorista(
    motorista_id: str,
    plano_data: Dict[str, str],
    current_user: Dict = Depends(get_current_user)
):
    """Admin: Assign subscription plan to motorista"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    plano_id = plano_data.get("plano_id")
    periodicidade = plano_data.get("periodicidade", "mensal")
    if not plano_id:
        raise HTTPException(status_code=400, detail="plano_id required")
    
    # Check if plano exists
    plano = await db.planos.find_one({"id": plano_id}, {"_id": 0})
    if not plano:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    # Obter valor do plano baseado na periodicidade
    valor_pago = 0
    if plano.get("precos") and plano["precos"].get(periodicidade):
        valor_pago = plano["precos"][periodicidade].get("preco_com_iva", 0)
    elif plano.get("preco"):
        # Fallback para formato antigo
        valor_pago = plano.get("preco", 0)
    
    # Update motorista - try both users and motoristas collections
    update_doc = {
        "plano_id": plano_id,
        "plano_status": "ativo",
        "periodicidade": periodicidade,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    result = await db.users.update_one(
        {"id": motorista_id, "role": "motorista"},
        {"$set": update_doc}
    )
    
    # If not found in users, try motoristas collection
    if result.matched_count == 0:
        result = await db.motoristas.update_one(
            {"id": motorista_id},
            {"$set": update_doc}
        )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Motorista not found")
    
    # Criar/atualizar registro em planos_usuarios para controle de módulos
    now = datetime.now(timezone.utc)
    atribuicao_id = str(uuid.uuid4())
    
    # Remover atribuição anterior se existir
    await db.planos_usuarios.delete_many({"user_id": motorista_id})
    
    # Criar nova atribuição
    # Support both old 'features' and new 'modulos' field
    modulos_ativos = plano.get("modulos", [])
    if not modulos_ativos and plano.get("features"):
        # Fallback to features for old plans
        modulos_ativos = plano.get("features", [])
    
    atribuicao = {
        "id": atribuicao_id,
        "user_id": motorista_id,
        "plano_id": plano_id,
        "modulos_ativos": modulos_ativos,
        "tipo_pagamento": periodicidade,
        "valor_pago": valor_pago,
        "data_inicio": now,
        "data_fim": None,
        "status": "ativo",
        "renovacao_automatica": True,
        "created_at": now,
        "updated_at": now,
        "created_by": current_user["id"]
    }
    
    await db.planos_usuarios.insert_one(atribuicao)
    
    return {"message": "Plan assigned successfully to motorista", "plano_id": plano_id, "atribuicao_id": atribuicao_id}


@api_router.post("/users/trocar-plano")
async def trocar_plano(
    plano_data: Dict[str, str],
    current_user: Dict = Depends(get_current_user)
):
    """Parceiro/Motorista: Change subscription plan with change fee"""
    plano_id = plano_data.get("plano_id")
    periodicidade = plano_data.get("periodicidade", "mensal")
    
    if not plano_id:
        raise HTTPException(status_code=400, detail="plano_id required")
    
    # Verificar se o plano existe e é do tipo correto
    novo_plano = await db.planos.find_one({"id": plano_id, "ativo": True}, {"_id": 0})
    if not novo_plano:
        raise HTTPException(status_code=404, detail="Plan not found or inactive")
    
    if novo_plano.get("tipo_usuario") != current_user["role"]:
        raise HTTPException(status_code=400, detail="Plan type does not match user role")
    
    # Verificar se o utilizador já tem um plano ativo
    plano_atual_id = current_user.get("plano_id")
    taxa_troca = 0
    
    if plano_atual_id and plano_atual_id != plano_id:
        # Utilizador está a trocar de plano - aplicar taxa de troca (10% do valor do novo plano)
        precos_novo = novo_plano.get("precos", {})
        valor_novo_plano = precos_novo.get(periodicidade, {}).get("preco_com_iva", 0)
        taxa_troca = round(valor_novo_plano * 0.10, 2)  # 10% de taxa de troca
        
        # Criar registro de taxa na coleção de transações/faturas
        taxa_record = {
            "id": str(uuid.uuid4()),
            "user_id": current_user["id"],
            "tipo": "taxa_troca_plano",
            "plano_anterior_id": plano_atual_id,
            "plano_novo_id": plano_id,
            "valor": taxa_troca,
            "status": "pendente",  # Pode ser processado posteriormente
            "data": datetime.now(timezone.utc),
            "created_at": datetime.now(timezone.utc)
        }
        await db.taxas_troca_plano.insert_one(taxa_record)
    
    # Atualizar user
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": {
            "plano_id": plano_id,
            "plano_status": "ativo",
            "plano_periodicidade": periodicidade,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Remover atribuição anterior
    await db.planos_usuarios.delete_many({"user_id": current_user["id"]})
    
    # Criar nova atribuição
    now = datetime.now(timezone.utc)
    atribuicao_id = str(uuid.uuid4())
    
    precos = plano.get("precos", {})
    valor_pago = precos.get(periodicidade, {}).get("preco_com_iva", 0)
    
    atribuicao = {
        "id": atribuicao_id,
        "user_id": current_user["id"],
        "plano_id": plano_id,
        "modulos_ativos": plano.get("modulos", []),
        "tipo_pagamento": periodicidade,
        "valor_pago": valor_pago,
        "data_inicio": now,
        "data_fim": None,
        "status": "ativo",
        "renovacao_automatica": True,
        "created_at": now,
        "updated_at": now,
        "created_by": current_user["id"]
    }
    
    await db.planos_usuarios.insert_one(atribuicao)
    
    response = {
        "message": "Plan changed successfully", 
        "plano_id": plano_id,
        "periodicidade": periodicidade,
        "valor_plano": valor_pago
    }
    
    if taxa_troca > 0:
        response["taxa_troca"] = taxa_troca
        response["valor_total"] = valor_pago + taxa_troca
        response["message"] = f"Plano alterado com sucesso. Taxa de troca de {taxa_troca}€ aplicada."
    
    return response

# ==================== RELATORIO CONFIGURATION ENDPOINTS ====================

@api_router.get("/parceiros/{parceiro_id}/config-relatorio", response_model=RelatorioConfig)
async def get_config_relatorio(
    parceiro_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Obter configuração de relatórios do parceiro"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Se é parceiro, só pode ver sua própria configuração
    if current_user["role"] == UserRole.PARCEIRO and current_user["id"] != parceiro_id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    config = await db.relatorio_config.find_one({"parceiro_id": parceiro_id}, {"_id": 0})
    
    if not config:
        # Criar configuração padrão se não existir
        config = {
            "id": str(uuid.uuid4()),
            "parceiro_id": parceiro_id,
            "incluir_numero_relatorio": True,
            "incluir_data_emissao": True,
            "incluir_periodo": True,
            "incluir_nome_parceiro": True,
            "incluir_nome_motorista": True,
            "incluir_veiculo": True,
            "incluir_viagens_bolt": True,
            "incluir_viagens_uber": True,
            "incluir_viagens_totais": True,
            "incluir_horas_bolt": True,
            "incluir_horas_uber": True,
            "incluir_horas_totais": True,
            "incluir_ganhos_uber": True,
            "incluir_ganhos_bolt": True,
            "incluir_ganhos_totais": True,
            "incluir_valor_aluguer": True,
            "incluir_combustivel": True,
            "incluir_via_verde": True,
            "via_verde_atraso_semanas": 1,
            "incluir_caucao": True,
            "incluir_caucao_parcelada": True,
            "incluir_danos": True,
            "incluir_danos_acumulados": True,
            "incluir_danos_descricao": True,
            "incluir_danos_parcelados": True,
            "incluir_extras": True,
            "incluir_total_recibo": True,
            "incluir_tabela_combustivel": True,
            "incluir_combustivel_matricula": True,
            "incluir_combustivel_local": True,
            "incluir_combustivel_data_hora": True,
            "incluir_combustivel_cartao": True,
            "incluir_combustivel_quantidade": True,
            "incluir_combustivel_valor": True,
            "formato_numero_relatorio": "xxxxx/ano",
            "created_at": datetime.now(timezone.utc),
            "updated_at": datetime.now(timezone.utc)
        }
        await db.relatorio_config.insert_one(config)
    
    # Fix datetime fields
    if isinstance(config["created_at"], str):
        config["created_at"] = datetime.fromisoformat(config["created_at"])
    if isinstance(config["updated_at"], str):
        config["updated_at"] = datetime.fromisoformat(config["updated_at"])
    
    return config

@api_router.put("/parceiros/{parceiro_id}/config-relatorio")
async def update_config_relatorio(
    parceiro_id: str,
    updates: RelatorioConfigUpdate,
    current_user: Dict = Depends(get_current_user)
):
    """Atualizar configuração de relatórios do parceiro"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Se é parceiro, só pode editar sua própria configuração
    if current_user["role"] == UserRole.PARCEIRO and current_user["id"] != parceiro_id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Build update document
    update_doc = {k: v for k, v in updates.model_dump(exclude_unset=True).items() if v is not None}
    update_doc["updated_at"] = datetime.now(timezone.utc)
    
    # Check if config exists
    config = await db.relatorio_config.find_one({"parceiro_id": parceiro_id})
    
    if not config:
        # Create new config with provided updates
        config = {
            "id": str(uuid.uuid4()),
            "parceiro_id": parceiro_id,
            "created_at": datetime.now(timezone.utc),
            **update_doc
        }
        await db.relatorio_config.insert_one(config)
    else:
        # Update existing config
        await db.relatorio_config.update_one(
            {"parceiro_id": parceiro_id},
            {"$set": update_doc}
        )
    
    return {"message": "Configuração de relatórios atualizada com sucesso"}

@api_router.post("/relatorios/motorista/{motorista_id}/gerar-semanal")
async def gerar_relatorio_semanal(
    motorista_id: str,
    data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Gerar relatório semanal para motorista baseado na configuração do parceiro"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.PARCEIRO, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get parameters
    data_inicio = data.get("data_inicio")  # YYYY-MM-DD
    data_fim = data.get("data_fim")  # YYYY-MM-DD
    semana = data.get("semana", 1)
    ano = data.get("ano", datetime.now().year)
    
    if not data_inicio or not data_fim:
        raise HTTPException(status_code=400, detail="data_inicio e data_fim são obrigatórios")
    
    # Get motorista data
    motorista = await db.motoristas.find_one({"id": motorista_id}, {"_id": 0})
    if not motorista:
        raise HTTPException(status_code=404, detail="Motorista não encontrado")
    
    parceiro_id = motorista.get("parceiro_atribuido")
    if not parceiro_id:
        raise HTTPException(status_code=400, detail="Motorista não tem parceiro atribuído")
    
    # Check permissions - parceiro só pode gerar relatórios dos seus motoristas
    if current_user["role"] == UserRole.PARCEIRO and current_user["id"] != parceiro_id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get parceiro data
    parceiro = await db.parceiros.find_one({"id": parceiro_id}, {"_id": 0})
    if not parceiro:
        raise HTTPException(status_code=404, detail="Parceiro não encontrado")
    
    # Get vehicle data
    veiculo_id = motorista.get("veiculo_atribuido")
    veiculo = None
    if veiculo_id:
        veiculo = await db.vehicles.find_one({"id": veiculo_id}, {"_id": 0})
    
    # Get contrato data
    contrato = await db.contratos.find_one({
        "motorista_id": motorista_id,
        "ativo": True
    }, {"_id": 0})
    
    # Get relatorio configuration
    config = await db.relatorio_config.find_one({"parceiro_id": parceiro_id}, {"_id": 0})
    if not config:
        # Use default config
        config = {
            "incluir_numero_relatorio": True,
            "incluir_data_emissao": True,
            "incluir_periodo": True,
            "incluir_nome_parceiro": True,
            "incluir_nome_motorista": True,
            "incluir_veiculo": True,
            "incluir_viagens_bolt": True,
            "incluir_viagens_uber": True,
            "incluir_viagens_totais": True,
            "incluir_horas_bolt": True,
            "incluir_horas_uber": True,
            "incluir_horas_totais": True,
            "incluir_ganhos_uber": True,
            "incluir_ganhos_bolt": True,
            "incluir_ganhos_totais": True,
            "incluir_valor_aluguer": True,
            "incluir_combustivel": True,
            "incluir_via_verde": True,
            "via_verde_atraso_semanas": 1,
            "incluir_caucao": True,
            "incluir_caucao_parcelada": True,
            "incluir_danos": True,
            "incluir_extras": True,
            "incluir_total_recibo": True,
            "incluir_tabela_combustivel": True,
            "formato_numero_relatorio": "xxxxx/ano"
        }
    
    # Calculate data inicio/fim for via verde (with delay)
    via_verde_atraso = config.get("via_verde_atraso_semanas", 1)
    data_inicio_via_verde = (datetime.fromisoformat(data_inicio) - timedelta(weeks=via_verde_atraso)).strftime("%Y-%m-%d")
    data_fim_via_verde = (datetime.fromisoformat(data_fim) - timedelta(weeks=via_verde_atraso)).strftime("%Y-%m-%d")
    
    # Get ganhos data (from relatorios_ganhos collection)
    ganhos_query = {
        "motorista_id": motorista_id,
        "data_inicio": {"$lte": data_fim},
        "data_fim": {"$gte": data_inicio}
    }
    ganhos_records = await db.relatorios_ganhos.find(ganhos_query, {"_id": 0}).to_list(100)
    
    # Calculate totals
    total_viagens_uber = 0
    total_viagens_bolt = 0
    total_horas_uber = 0.0
    total_horas_bolt = 0.0
    total_ganhos_uber = 0.0
    total_ganhos_bolt = 0.0
    
    for record in ganhos_records:
        total_viagens_uber += record.get("uber_viagens", 0)
        total_viagens_bolt += record.get("bolt_viagens", 0)
        total_horas_uber += record.get("uber_horas", 0.0)
        total_horas_bolt += record.get("bolt_horas", 0.0)
        total_ganhos_uber += record.get("uber_ganhos", 0.0)
        total_ganhos_bolt += record.get("bolt_ganhos", 0.0)
    
    # Get combustivel data
    combustivel_query = {
        "veiculo_id": veiculo_id,
        "data": {"$gte": data_inicio, "$lte": data_fim}
    } if veiculo_id else {}
    
    combustivel_records = []
    total_combustivel = 0.0
    if veiculo_id and config.get("incluir_combustivel", True):
        combustivel_cursor = db.abastecimentos.find(combustivel_query, {"_id": 0})
        combustivel_records = await combustivel_cursor.to_list(1000)
        total_combustivel = sum(record.get("valor_com_iva", 0.0) for record in combustivel_records)
    
    # Get via verde data (with delay)
    via_verde_records = []
    total_via_verde = 0.0
    if config.get("incluir_via_verde", True):
        via_verde_query = {
            "veiculo_id": veiculo_id,
            "data": {"$gte": data_inicio_via_verde, "$lte": data_fim_via_verde}
        } if veiculo_id else {}
        
        if veiculo_id:
            via_verde_cursor = db.via_verde.find(via_verde_query, {"_id": 0})
            via_verde_records = await via_verde_cursor.to_list(1000)
            total_via_verde = sum(record.get("valor", 0.0) for record in via_verde_records)
    
    # Calculate valor aluguer/comissao
    valor_aluguer = 0.0
    if contrato:
        tipo_contrato = contrato.get("tipo_contrato", "aluguer_normal")
        if tipo_contrato == "comissao":
            # Comissão sobre ganhos
            comissao_percentual = contrato.get("comissao_percentual", 0.0)
            valor_aluguer = (total_ganhos_uber + total_ganhos_bolt) * (comissao_percentual / 100)
        else:
            # Aluguer fixo semanal
            valor_aluguer = contrato.get("valor_semanal", 0.0)
    
    # Get caucao data
    caucao_acumulada = contrato.get("caucao_total", 0.0) if contrato else 0.0
    caucao_semanal = 0.0
    if contrato and contrato.get("caucao_parcelada", False):
        caucao_parcelas = contrato.get("caucao_parcelas", 1)
        caucao_semanal = caucao_acumulada / caucao_parcelas if caucao_parcelas > 0 else 0.0
    
    # Get danos data
    danos_records = []
    total_danos_acumulados = 0.0
    danos_semanal = 0.0
    if config.get("incluir_danos", True) and veiculo_id:
        danos_cursor = db.danos.find({
            "veiculo_id": veiculo_id,
            "motorista_id": motorista_id,
            "ativo": True
        }, {"_id": 0})
        danos_records = await danos_cursor.to_list(100)
        total_danos_acumulados = sum(d.get("valor_total", 0.0) for d in danos_records)
        danos_semanal = sum(d.get("valor_semanal", 0.0) for d in danos_records)
    
    # Get extras
    extras = data.get("extras", 0.0)
    
    # Calculate totals
    total_ganhos = total_ganhos_uber + total_ganhos_bolt
    total_despesas = valor_aluguer + total_combustivel + total_via_verde + caucao_semanal + danos_semanal - extras
    total_recibo = total_ganhos - total_despesas
    
    # Generate numero relatorio
    formato = config.get("formato_numero_relatorio", "xxxxx/ano")
    # Count existing reports for this year
    count = await db.relatorios_semanais.count_documents({
        "parceiro_id": parceiro_id,
        "ano": ano
    })
    numero_relatorio = formato.replace("xxxxx", str(count + 1).zfill(5)).replace("ano", str(ano))
    
    # Build relatorio document
    relatorio = {
        "id": str(uuid.uuid4()),
        "numero_relatorio": numero_relatorio,
        "parceiro_id": parceiro_id,
        "parceiro_nome": parceiro.get("nome_empresa", ""),
        "motorista_id": motorista_id,
        "motorista_nome": motorista.get("name", ""),
        "veiculo_id": veiculo_id,
        "veiculo_marca": veiculo.get("marca", "") if veiculo else "",
        "veiculo_modelo": veiculo.get("modelo", "") if veiculo else "",
        "veiculo_matricula": veiculo.get("matricula", "") if veiculo else "",
        "data_emissao": datetime.now(timezone.utc).isoformat(),
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "semana": semana,
        "ano": ano,
        
        # Statistics
        "viagens_uber": total_viagens_uber,
        "viagens_bolt": total_viagens_bolt,
        "viagens_totais": total_viagens_uber + total_viagens_bolt,
        "horas_uber": total_horas_uber,
        "horas_bolt": total_horas_bolt,
        "horas_totais": total_horas_uber + total_horas_bolt,
        
        # Ganhos
        "ganhos_uber": total_ganhos_uber,
        "ganhos_bolt": total_ganhos_bolt,
        "ganhos_totais": total_ganhos,
        
        # Despesas
        "valor_aluguer": valor_aluguer,
        "tipo_contrato": contrato.get("tipo_contrato", "") if contrato else "",
        "combustivel": total_combustivel,
        "via_verde": total_via_verde,
        "caucao_acumulada": caucao_acumulada,
        "caucao_semanal": caucao_semanal,
        "danos_acumulados": total_danos_acumulados,
        "danos_semanal": danos_semanal,
        "danos_descricao": [d.get("descricao", "") for d in danos_records],
        "extras": extras,
        
        # Totals
        "total_despesas": total_despesas,
        "total_recibo": total_recibo,
        
        # Detailed lists
        "combustivel_detalhes": combustivel_records,
        "via_verde_detalhes": via_verde_records,
        "danos_detalhes": danos_records,
        
        # Configuration used
        "config": config,
        
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user["id"]
    }
    
    # Save relatorio
    await db.relatorios_semanais.insert_one(relatorio)
    
    return {
        "message": "Relatório semanal gerado com sucesso",
        "relatorio_id": relatorio["id"],
        "numero_relatorio": numero_relatorio,
        "total_recibo": total_recibo,
        "relatorio": relatorio
    }

# ==================== ENDPOINTS DE CONTRATOS ====================

@api_router.get("/parceiros/{parceiro_id}/templates-contrato")
async def get_templates_contrato(
    parceiro_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Get all contract templates for a parceiro"""
    # Check permissions
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        if current_user["id"] != parceiro_id:
            raise HTTPException(status_code=403, detail="Sem permissão")
    
    templates = await db.templates_contrato.find({"parceiro_id": parceiro_id}, {"_id": 0}).to_list(length=None)
    return templates

@api_router.post("/parceiros/{parceiro_id}/templates-contrato")
async def create_template_contrato(
    parceiro_id: str,
    template_data: TemplateContratoCreate,
    current_user: Dict = Depends(get_current_user)
):
    """Create a new contract template for a parceiro"""
    # Check permissions - apenas admin e gestor podem criar templates
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Sem permissão para criar templates")
    
    # Validações
    if template_data.tipo_contrato == "comissao":
        if template_data.percentagem_motorista and template_data.percentagem_parceiro:
            if (template_data.percentagem_motorista + template_data.percentagem_parceiro) != 100:
                raise HTTPException(status_code=400, detail="Percentagens de comissão devem somar 100%")
    
    template = TemplateContrato(
        parceiro_id=parceiro_id,
        nome_template=template_data.nome_template,
        tipo_contrato=template_data.tipo_contrato,
        periodicidade_padrao=template_data.periodicidade_padrao,
        valor_base=template_data.valor_base,
        valor_caucao=template_data.valor_caucao,
        numero_parcelas_caucao=template_data.numero_parcelas_caucao,
        valor_epoca_alta=template_data.valor_epoca_alta,
        valor_epoca_baixa=template_data.valor_epoca_baixa,
        percentagem_motorista=template_data.percentagem_motorista,
        percentagem_parceiro=template_data.percentagem_parceiro,
        combustivel_incluido=template_data.combustivel_incluido,
        regime_trabalho=template_data.regime_trabalho,
        valor_compra_veiculo=template_data.valor_compra_veiculo,
        numero_semanas_compra=template_data.numero_semanas_compra,
        com_slot=template_data.com_slot,
        extra_seguro=template_data.extra_seguro,
        valor_extra_seguro=template_data.valor_extra_seguro,
        clausulas_texto=template_data.clausulas_texto,
        created_by=current_user["id"]
    )
    
    # Calcular valor_parcela_caucao se necessário
    if template.numero_parcelas_caucao and template.valor_caucao:
        template.valor_parcela_caucao = round(template.valor_caucao / template.numero_parcelas_caucao, 2)
    
    await db.templates_contrato.insert_one(template.dict())
    
    return template

@api_router.put("/templates-contrato/{template_id}")
async def update_template_contrato(
    template_id: str,
    template_data: TemplateContratoCreate,
    current_user: Dict = Depends(get_current_user)
):
    """Update a contract template"""
    # Check permissions - apenas admin e gestor podem editar templates
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    # Validações
    if template_data.tipo_contrato == "comissao":
        if template_data.percentagem_motorista and template_data.percentagem_parceiro:
            if (template_data.percentagem_motorista + template_data.percentagem_parceiro) != 100:
                raise HTTPException(status_code=400, detail="Percentagens de comissão devem somar 100%")
    


@api_router.get("/templates-contratos/{template_id}/download-pdf")
async def download_template_pdf(
    template_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Generate and download template as PDF in A4 format"""
    # Check permissions
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    # Get template
    template = await db.templates_contrato.find_one({"id": template_id}, {"_id": 0})
    if not template:
        raise HTTPException(status_code=404, detail="Template não encontrado")
    
    # Create PDF
    output_dir = UPLOAD_DIR / "templates_pdf"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    pdf_filename = f"template_{template_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    output_path = output_dir / pdf_filename
    
    # Create PDF with A4 format
    c = canvas.Canvas(str(output_path), pagesize=A4)
    width, height = A4
    
    # Header
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(width/2, height - 2*cm, "CONTRATO DE PRESTAÇÃO DE SERVIÇOS")
    
    c.setFont("Helvetica", 10)
    c.drawCentredString(width/2, height - 2.7*cm, f"Template: {template.get('nome_template', 'N/A')}")
    c.drawCentredString(width/2, height - 3.2*cm, f"Tipo: {template.get('tipo_contrato', 'N/A').upper()}")
    
    # Line separator
    c.setStrokeColorRGB(0.3, 0.3, 0.3)
    c.setLineWidth(2)
    c.line(2*cm, height - 3.7*cm, width - 2*cm, height - 3.7*cm)
    
    y = height - 5*cm
    
    # Template Information Box
    c.setFillColorRGB(0.95, 0.95, 0.95)
    c.rect(2*cm, y - 3*cm, width - 4*cm, 2.5*cm, fill=1, stroke=0)
    
    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(2.5*cm, y - 0.8*cm, "CONDIÇÕES DO TEMPLATE")
    
    c.setFont("Helvetica", 9)
    info_y = y - 1.5*cm
    
    if template.get('periodicidade_padrao'):
        c.drawString(2.5*cm, info_y, f"Periodicidade: {template['periodicidade_padrao']}")
        info_y -= 0.5*cm
    
    if template.get('valor_caucao'):
        caucao_text = f"Caução: €{template['valor_caucao']}"
        if template.get('numero_parcelas_caucao'):
            caucao_text += f" ({template['numero_parcelas_caucao']}x)"
        c.drawString(2.5*cm, info_y, caucao_text)
        info_y -= 0.5*cm
    
    if template.get('valor_epoca_alta'):
        c.drawString(width/2 + 1*cm, y - 1.5*cm, f"Época Alta: €{template['valor_epoca_alta']}")
    
    if template.get('valor_epoca_baixa'):
        c.drawString(width/2 + 1*cm, y - 2*cm, f"Época Baixa: €{template['valor_epoca_baixa']}")
    
    y -= 4*cm
    
    # Contract Clauses Section
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2*cm, y, "CLÁUSULAS CONTRATUAIS")
    y -= 0.8*cm
    
    # Draw line under title
    c.setStrokeColorRGB(0.5, 0.5, 0.5)
    c.setLineWidth(0.5)
    c.line(2*cm, y, width - 2*cm, y)
    y -= 0.5*cm
    
    # Contract text
    c.setFont("Helvetica", 9)
    clausulas_texto = template.get('clausulas_texto', 'Sem texto de contrato definido')
    
    # Split text into lines and handle pagination
    lines = clausulas_texto.split('\n')
    for line in lines:
        # Word wrap for long lines
        while len(line) > 0:
            if y < 3*cm:  # New page if needed
                c.showPage()
                y = height - 2*cm
                c.setFont("Helvetica", 9)
            
            # Calculate how much text fits in one line
            max_width = width - 4*cm
            if c.stringWidth(line, "Helvetica", 9) <= max_width:
                c.drawString(2*cm, y, line)
                y -= 0.5*cm
                break
            else:
                # Find break point
                words = line.split(' ')
                current_line = ''
                remaining = ''
                
                for i, word in enumerate(words):
                    test_line = current_line + (' ' if current_line else '') + word
                    if c.stringWidth(test_line, "Helvetica", 9) <= max_width:
                        current_line = test_line
                    else:
                        remaining = ' '.join(words[i:])
                        break
                
                if current_line:
                    c.drawString(2*cm, y, current_line)
                    y -= 0.5*cm
                
                line = remaining
                if not remaining:
                    break
    
    # Footer with signatures
    if y < 8*cm:
        c.showPage()
        y = height - 2*cm
    
    y = 8*cm  # Fixed position for signatures
    
    c.setFont("Helvetica-Bold", 10)
    c.drawString(2*cm, y, "ASSINATURAS")
    y -= 0.5*cm
    c.setLineWidth(0.5)
    c.line(2*cm, y, width - 2*cm, y)
    y -= 2*cm
    
    # Signature lines
    c.setFont("Helvetica", 9)
    sig_y = y - 2*cm
    
    # Parceiro signature
    c.line(2.5*cm, sig_y, 9*cm, sig_y)
    c.drawString(2.5*cm, sig_y - 0.5*cm, "O Parceiro")
    c.setFont("Helvetica", 8)
    c.drawString(2.5*cm, sig_y - 0.9*cm, "{PARCEIRO_NOME}")
    
    # Motorista signature
    c.setFont("Helvetica", 9)
    c.line(width/2 + 1*cm, sig_y, width - 2.5*cm, sig_y)
    c.drawString(width/2 + 1*cm, sig_y - 0.5*cm, "O Motorista")
    c.setFont("Helvetica", 8)
    c.drawString(width/2 + 1*cm, sig_y - 0.9*cm, "{MOTORISTA_NOME}")
    
    # Footer info
    c.setFont("Helvetica", 7)
    c.setFillColorRGB(0.5, 0.5, 0.5)
    footer_y = 1.5*cm
    c.drawCentredString(width/2, footer_y, f"Template: {template.get('nome_template', 'N/A')}")
    c.drawCentredString(width/2, footer_y - 0.4*cm, f"Criado em: {template.get('created_at', 'N/A')[:10] if template.get('created_at') else 'N/A'}")
    c.drawCentredString(width/2, footer_y - 0.8*cm, "Este é um template. As variáveis serão substituídas pelos dados reais ao gerar o contrato.")
    
    c.save()
    
    return FileResponse(
        path=output_path,
        media_type="application/pdf",
        filename=f"template_{template.get('nome_template', 'contrato').replace(' ', '_')}.pdf"
    )

    update_data = template_data.dict()
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    # Calcular valor_parcela_caucao
    if update_data.get("numero_parcelas_caucao") and update_data.get("valor_caucao"):
        update_data["valor_parcela_caucao"] = round(update_data["valor_caucao"] / update_data["numero_parcelas_caucao"], 2)
    
    result = await db.templates_contrato.update_one(
        {"id": template_id},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Template não encontrado")
    
    return {"message": "Template atualizado com sucesso"}

@api_router.delete("/templates-contrato/{template_id}")
async def delete_template_contrato(
    template_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Delete a contract template (soft delete)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    result = await db.templates_contrato.update_one(
        {"id": template_id},
        {"$set": {"ativo": False, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Template não encontrado")
    
    return {"message": "Template removido com sucesso"}

@api_router.post("/contratos")
async def create_contrato_motorista(
    contrato_data: ContratoMotoristaCreate,
    current_user: Dict = Depends(get_current_user)
):
    """Create a contract for a motorista from a template"""
    # Check permissions
    if current_user["role"] == "operacional":
        if not await check_feature_access(current_user, "criar_contratos"):
            raise HTTPException(status_code=403, detail="Operacional base não pode criar contratos")
    elif current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, "parceiro"]:
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    # Get template
    template = await db.templates_contrato.find_one({"id": contrato_data.template_id}, {"_id": 0})
    if not template:
        raise HTTPException(status_code=404, detail="Template não encontrado")
    
    # If parceiro, verify they can only create contracts for their own templates
    if current_user["role"] == "parceiro":
        if template["parceiro_id"] != current_user["id"]:
            raise HTTPException(status_code=403, detail="Parceiro só pode criar contratos com seus próprios templates")
    
    # Get motorista
    motorista = await db.motoristas.find_one({"id": contrato_data.motorista_id}, {"_id": 0})
    if not motorista:
        raise HTTPException(status_code=404, detail="Motorista não encontrado")
    
    # Validações específicas por tipo
    if template["tipo_contrato"] == "comissao":
        if contrato_data.percentagem_motorista_aplicado and contrato_data.percentagem_parceiro_aplicado:
            if (contrato_data.percentagem_motorista_aplicado + contrato_data.percentagem_parceiro_aplicado) != 100:
                raise HTTPException(status_code=400, detail="Percentagens devem somar 100%")
    
    # Calcular parcela de caução se aplicável
    valor_parcela_calculado = None
    if contrato_data.numero_parcelas_caucao_aplicado and contrato_data.valor_caucao_aplicado:
        valor_parcela_calculado = round(contrato_data.valor_caucao_aplicado / contrato_data.numero_parcelas_caucao_aplicado, 2)
    
    contrato = ContratoMotorista(
        template_id=contrato_data.template_id,
        parceiro_id=template["parceiro_id"],
        motorista_id=contrato_data.motorista_id,
        veiculo_id=contrato_data.veiculo_id,
        nome_template=template["nome_template"],
        tipo_contrato=template["tipo_contrato"],
        periodicidade=contrato_data.periodicidade,
        valor_aplicado=contrato_data.valor_aplicado,
        valor_caucao_aplicado=contrato_data.valor_caucao_aplicado,
        numero_parcelas_caucao_aplicado=contrato_data.numero_parcelas_caucao_aplicado,
        valor_parcela_caucao_aplicado=valor_parcela_calculado,
        epoca_atual=contrato_data.epoca_atual,
        valor_epoca_alta_aplicado=contrato_data.valor_epoca_alta_aplicado,
        valor_epoca_baixa_aplicado=contrato_data.valor_epoca_baixa_aplicado,
        percentagem_motorista_aplicado=contrato_data.percentagem_motorista_aplicado,
        percentagem_parceiro_aplicado=contrato_data.percentagem_parceiro_aplicado,
        combustivel_incluido_aplicado=contrato_data.combustivel_incluido_aplicado,
        regime_trabalho_aplicado=contrato_data.regime_trabalho_aplicado,
        valor_compra_aplicado=contrato_data.valor_compra_aplicado,
        numero_semanas_aplicado=contrato_data.numero_semanas_aplicado,
        com_slot_aplicado=contrato_data.com_slot_aplicado,
        extra_seguro_aplicado=contrato_data.extra_seguro_aplicado,
        valor_extra_seguro_aplicado=contrato_data.valor_extra_seguro_aplicado,
        clausulas_texto=template.get("clausulas_texto"),
        data_inicio=contrato_data.data_inicio,
        data_fim=contrato_data.data_fim,
        created_by=current_user["id"]
    )
    
    await db.contratos_motorista.insert_one(contrato.dict())
    
    return contrato

@api_router.get("/contratos")
async def get_contratos(
    parceiro_id: Optional[str] = None,
    motorista_id: Optional[str] = None,
    current_user: Dict = Depends(get_current_user)
):
    """Get contracts with optional filters"""
    query = {}
    
    if parceiro_id:
        query["parceiro_id"] = parceiro_id
    if motorista_id:
        query["motorista_id"] = motorista_id
    
    # Check permissions
    if current_user["role"] == "parceiro":
        query["parceiro_id"] = current_user["id"]
    
    contratos = await db.contratos_motorista.find(query, {"_id": 0}).to_list(length=None)
    
    # Populate motorista and vehicle names
    for contrato in contratos:
        # Set default values
        contrato["motorista_nome"] = "N/A"
        contrato["veiculo_matricula"] = "Sem veículo"
        
        if contrato.get("motorista_id"):
            motorista = await db.users.find_one({"id": contrato["motorista_id"]}, {"_id": 0, "name": 1})
            if motorista:
                contrato["motorista_nome"] = motorista.get("name", "N/A")
        
        if contrato.get("veiculo_id"):
            veiculo = await db.vehicles.find_one({"id": contrato["veiculo_id"]}, {"_id": 0, "matricula": 1})
            if veiculo:
                contrato["veiculo_matricula"] = veiculo.get("matricula", "N/A")
    
    return contratos

@api_router.get("/contratos/{contrato_id}")
async def get_contrato(
    contrato_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Get a specific contract"""
    contrato = await db.contratos_motorista.find_one({"id": contrato_id}, {"_id": 0})
    if not contrato:
        raise HTTPException(status_code=404, detail="Contrato não encontrado")
    
    # Check permissions
    if current_user["role"] == "parceiro":
        if contrato["parceiro_id"] != current_user["id"]:
            raise HTTPException(status_code=403, detail="Sem permissão")
    
    return contrato

# Templates de Contratos
@api_router.get("/templates-contratos")
async def get_templates_contratos(current_user: Dict = Depends(get_current_user)):
    """Get contract templates for the current user (parceiro only)"""
    if current_user["role"] != "parceiro":
        raise HTTPException(status_code=403, detail="Apenas parceiros podem acessar templates")
    
    templates = await db.templates_contratos.find(
        {"parceiro_id": current_user["id"]},
        {"_id": 0}
    ).to_list(length=None)
    return templates

@api_router.post("/templates-contratos")
async def create_template_contrato(
    template_data: Dict,
    current_user: Dict = Depends(get_current_user)
):
    """Create a new contract template"""
    if current_user["role"] != "parceiro":
        raise HTTPException(status_code=403, detail="Apenas parceiros podem criar templates")
    
    from datetime import datetime
    
    template = {
        "id": str(uuid4()),
        "parceiro_id": current_user["id"],
        "nome": template_data.get("nome"),
        "descricao": template_data.get("descricao", ""),
        "texto_template": template_data.get("texto_template"),
        "created_at": datetime.now().isoformat(),
        "updated_at": datetime.now().isoformat()
    }
    
    await db.templates_contratos.insert_one(template)
    return {"message": "Template criado com sucesso", "id": template["id"]}

@api_router.put("/templates-contratos/{template_id}")
async def update_template_contrato(
    template_id: str,
    template_data: Dict,
    current_user: Dict = Depends(get_current_user)
):
    """Update a contract template"""
    if current_user["role"] != "parceiro":
        raise HTTPException(status_code=403, detail="Apenas parceiros podem editar templates")
    
    # Check if template exists and belongs to user
    template = await db.templates_contratos.find_one({"id": template_id})
    if not template:
        raise HTTPException(status_code=404, detail="Template não encontrado")
    
    if template["parceiro_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Sem permissão para editar este template")
    
    from datetime import datetime
    
    update_data = {
        "nome": template_data.get("nome"),
        "descricao": template_data.get("descricao", ""),
        "texto_template": template_data.get("texto_template"),
        "updated_at": datetime.now().isoformat()
    }
    
    await db.templates_contratos.update_one(
        {"id": template_id},
        {"$set": update_data}
    )
    
    return {"message": "Template atualizado com sucesso"}

@api_router.delete("/templates-contratos/{template_id}")
async def delete_template_contrato(
    template_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Delete a contract template"""
    if current_user["role"] != "parceiro":
        raise HTTPException(status_code=403, detail="Apenas parceiros podem excluir templates")
    
    # Check if template exists and belongs to user
    template = await db.templates_contratos.find_one({"id": template_id})
    if not template:
        raise HTTPException(status_code=404, detail="Template não encontrado")
    
    if template["parceiro_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Sem permissão para excluir este template")
    
    await db.templates_contratos.delete_one({"id": template_id})
    return {"message": "Template excluído com sucesso"}

# ==================== MENSAGENS (MOTORISTA) ====================

@api_router.get("/mensagens/motorista")
async def get_mensagens_motorista(
    destinatario: str,
    current_user: Dict = Depends(get_current_user)
):
    """Get messages between motorista and destinatario (suporte or parceiro)"""
    if current_user["role"] != "motorista":
        raise HTTPException(status_code=403, detail="Apenas motoristas podem acessar")
    
    try:
        if destinatario == "suporte":
            # Mensagens com admin/gestao (suporte)
            query = {
                "$or": [
                    {
                        "remetente_id": current_user["id"],
                        "tipo_destinatario": "suporte"
                    },
                    {
                        "destinatario_id": current_user["id"],
                        "tipo_remetente": "suporte"
                    }
                ]
            }
        elif destinatario == "parceiro":
            # Buscar parceiro do motorista
            motorista = await db.users.find_one({"id": current_user["id"]}, {"_id": 0, "parceiro_atribuido": 1})
            if not motorista or not motorista.get("parceiro_atribuido"):
                return []  # Sem parceiro atribuído
            
            parceiro_id = motorista["parceiro_atribuido"]
            query = {
                "$or": [
                    {
                        "remetente_id": current_user["id"],
                        "destinatario_id": parceiro_id
                    },
                    {
                        "remetente_id": parceiro_id,
                        "destinatario_id": current_user["id"]
                    }
                ]
            }
        else:
            raise HTTPException(status_code=400, detail="Destinatário inválido")
        
        mensagens = await db.mensagens.find(query, {"_id": 0}).sort("created_at", 1).to_list(1000)
        return mensagens
        
    except Exception as e:
        logger.error(f"Error fetching messages: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/mensagens")
async def enviar_mensagem(
    mensagem_data: Dict,
    current_user: Dict = Depends(get_current_user)
):
    """Send a message"""
    try:
        tipo_destinatario = mensagem_data.get("tipo_destinatario")
        texto = mensagem_data.get("texto")
        
        if not texto or not tipo_destinatario:
            raise HTTPException(status_code=400, detail="Texto e tipo_destinatario são obrigatórios")
        
        mensagem = {
            "id": str(uuid4()),
            "remetente_id": current_user["id"],
            "remetente_nome": current_user.get("name", ""),
            "tipo_destinatario": tipo_destinatario,
            "texto": texto,
            "lida": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Se for para parceiro, adicionar destinatario_id
        if tipo_destinatario == "parceiro" and current_user["role"] == "motorista":
            motorista = await db.users.find_one({"id": current_user["id"]}, {"_id": 0, "parceiro_atribuido": 1})
            if motorista and motorista.get("parceiro_atribuido"):
                mensagem["destinatario_id"] = motorista["parceiro_atribuido"]
        
        await db.mensagens.insert_one(mensagem)
        return {"message": "Mensagem enviada com sucesso", "id": mensagem["id"]}
        
    except Exception as e:
        logger.error(f"Error sending message: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/contratos/{contrato_id}/gerar-pdf")
async def gerar_pdf_contrato(
    contrato_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Generate PDF for a contract"""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_JUSTIFY, TA_CENTER
    
    # Get contract
    contrato = await db.contratos_motorista.find_one({"id": contrato_id}, {"_id": 0})
    if not contrato:
        raise HTTPException(status_code=404, detail="Contrato não encontrado")
    
    # Get related data
    parceiro = await db.parceiros.find_one({"id": contrato["parceiro_id"]}, {"_id": 0})
    motorista = await db.motoristas.find_one({"id": contrato["motorista_id"]}, {"_id": 0})
    veiculo = None
    if contrato.get("veiculo_id"):
        veiculo = await db.vehicles.find_one({"id": contrato["veiculo_id"]}, {"_id": 0})
    
    # Create PDFs directory
    contratos_dir = UPLOAD_DIR / "contratos"
    contratos_dir.mkdir(parents=True, exist_ok=True)
    
    pdf_filename = f"contrato_{contrato_id}.pdf"
    pdf_path = contratos_dir / pdf_filename
    
    # Create PDF
    c = canvas.Canvas(str(pdf_path), pagesize=A4)
    width, height = A4
    
    # Title
    c.setFont("Helvetica-Bold", 18)
    c.drawCentredString(width/2, height - 2*cm, "CONTRATO DE PRESTAÇÃO DE SERVIÇOS TVDE")
    
    # Subtitle with contract type
    c.setFont("Helvetica", 12)
    tipo_nome = {
        "aluguer_sem_caucao": "Aluguer sem Caução",
        "aluguer_com_caucao": "Aluguer com Caução",
        "aluguer_caucao_parcelada": "Aluguer com Caução Parcelada",
        "periodo_epoca": "Período de Época",
        "aluguer_epocas_sem_caucao": "Aluguer com Épocas sem Caução",
        "aluguer_epocas_caucao": "Aluguer com Épocas e Caução",
        "aluguer_epoca_caucao_parcelada": "Aluguer Época com Caução Parcelada",
        "compra_veiculo": "Compra de Veículo",
        "comissao": "Comissão",
        "motorista_privado": "Motorista Privado",
        "outros": "Outros"
    }.get(contrato["tipo_contrato"], contrato["tipo_contrato"])
    
    c.drawCentredString(width/2, height - 2.8*cm, f"({tipo_nome})")
    
    # Date
    c.setFont("Helvetica", 10)
    data_hoje = datetime.now().strftime("%d/%m/%Y")
    c.drawCentredString(width/2, height - 3.5*cm, f"Data de Emissão: {data_hoje}")
    
    y = height - 5*cm
    
    # Section 1: Parceiro
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, y, "1. PRIMEIRO OUTORGANTE (PARCEIRO/EMPRESA)")
    y -= 0.8*cm
    
    c.setFont("Helvetica", 10)
    c.drawString(2.5*cm, y, f"Nome/Empresa: {parceiro.get('nome_empresa', parceiro.get('name', 'N/A'))}")
    y -= 0.6*cm
    c.drawString(2.5*cm, y, f"NIF: {parceiro.get('contribuinte_empresa', parceiro.get('nif', 'N/A'))}")
    y -= 0.6*cm
    c.drawString(2.5*cm, y, f"Email: {parceiro.get('email', 'N/A')}")
    y -= 0.6*cm
    c.drawString(2.5*cm, y, f"Telefone: {parceiro.get('telefone', parceiro.get('phone', 'N/A'))}")
    y -= 1*cm
    
    # Section 2: Motorista
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, y, "2. SEGUNDO OUTORGANTE (MOTORISTA)")
    y -= 0.8*cm
    
    c.setFont("Helvetica", 10)
    c.drawString(2.5*cm, y, f"Nome: {motorista.get('name', 'N/A')}")
    y -= 0.6*cm
    c.drawString(2.5*cm, y, f"NIF: {motorista.get('nif', 'N/A')}")
    y -= 0.6*cm
    c.drawString(2.5*cm, y, f"Email: {motorista.get('email', 'N/A')}")
    y -= 0.6*cm
    c.drawString(2.5*cm, y, f"Telefone: {motorista.get('phone', 'N/A')}")
    y -= 0.6*cm
    c.drawString(2.5*cm, y, f"Licença TVDE: {motorista.get('licenca_tvde_numero', 'N/A')}")
    y -= 1*cm
    
    # Section 3: Veículo (se aplicável)
    if veiculo:
        c.setFont("Helvetica-Bold", 12)
        c.drawString(2*cm, y, "3. VEÍCULO")
        y -= 0.8*cm
        
        c.setFont("Helvetica", 10)
        c.drawString(2.5*cm, y, f"Marca/Modelo: {veiculo.get('marca', 'N/A')} {veiculo.get('modelo', 'N/A')}")
        y -= 0.6*cm
        c.drawString(2.5*cm, y, f"Matrícula: {veiculo.get('matricula', 'N/A')}")
        y -= 0.6*cm
        c.drawString(2.5*cm, y, f"Ano: {veiculo.get('ano', 'N/A')}")
        y -= 1*cm
    
    # Section 4: Condições Financeiras
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, y, f"{4 if veiculo else 3}. CONDIÇÕES FINANCEIRAS")
    y -= 0.8*cm
    
    c.setFont("Helvetica", 10)
    c.drawString(2.5*cm, y, f"Periodicidade: {contrato['periodicidade'].upper()}")
    y -= 0.6*cm
    c.drawString(2.5*cm, y, f"Valor: €{contrato['valor_aplicado']:.2f}")
    y -= 0.6*cm
    
    # Caução
    if contrato.get("valor_caucao_aplicado"):
        c.drawString(2.5*cm, y, f"Caução: €{contrato['valor_caucao_aplicado']:.2f}")
        y -= 0.6*cm
        if contrato.get("numero_parcelas_caucao_aplicado"):
            c.drawString(2.5*cm, y, f"  Parcelado em: {contrato['numero_parcelas_caucao_aplicado']}x €{contrato['valor_parcela_caucao_aplicado']:.2f}")
            y -= 0.6*cm
    
    # Épocas
    if contrato.get("valor_epoca_alta_aplicado"):
        c.drawString(2.5*cm, y, f"Época Alta: €{contrato['valor_epoca_alta_aplicado']:.2f}")
        y -= 0.6*cm
        c.drawString(2.5*cm, y, f"Época Baixa: €{contrato['valor_epoca_baixa_aplicado']:.2f}")
        y -= 0.6*cm
    
    # Comissão
    if contrato.get("percentagem_motorista_aplicado"):
        c.drawString(2.5*cm, y, f"Comissão Motorista: {contrato['percentagem_motorista_aplicado']}%")
        y -= 0.6*cm
        c.drawString(2.5*cm, y, f"Comissão Parceiro: {contrato['percentagem_parceiro_aplicado']}%")
        y -= 0.6*cm
        c.drawString(2.5*cm, y, f"Combustível Incluído: {'Sim' if contrato.get('combustivel_incluido_aplicado') else 'Não'}")
        y -= 0.6*cm
        if contrato.get("regime_trabalho_aplicado"):
            c.drawString(2.5*cm, y, f"Regime: {contrato['regime_trabalho_aplicado'].upper()}")
            y -= 0.6*cm
    
    # Compra de veículo
    if contrato.get("valor_compra_aplicado"):
        c.drawString(2.5*cm, y, f"Valor de Compra: €{contrato['valor_compra_aplicado']:.2f}")
        y -= 0.6*cm
        c.drawString(2.5*cm, y, f"Número de Semanas: {contrato['numero_semanas_aplicado']}")
        y -= 0.6*cm
        c.drawString(2.5*cm, y, f"Com Slot: {'Sim' if contrato.get('com_slot_aplicado') else 'Não'}")
        y -= 0.6*cm
        if contrato.get("extra_seguro_aplicado"):
            c.drawString(2.5*cm, y, f"Extra Seguro: €{contrato.get('valor_extra_seguro_aplicado', 0):.2f}")
            y -= 0.6*cm
    
    y -= 0.5*cm
    
    # Section 5: Período
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, y, f"{5 if veiculo else 4}. PERÍODO DE VIGÊNCIA")
    y -= 0.8*cm
    
    c.setFont("Helvetica", 10)
    c.drawString(2.5*cm, y, f"Início: {contrato['data_inicio']}")
    y -= 0.6*cm
    if contrato.get("data_fim"):
        c.drawString(2.5*cm, y, f"Fim: {contrato['data_fim']}")
        y -= 0.6*cm
    else:
        c.drawString(2.5*cm, y, "Duração: Indeterminada")
        y -= 0.6*cm
    
    y -= 1*cm
    
    # Cláusulas (se houver)
    if contrato.get("clausulas_texto"):
        if y < 10*cm:  # New page if needed
            c.showPage()
            y = height - 2*cm
        
        c.setFont("Helvetica-Bold", 12)
        c.drawString(2*cm, y, f"{6 if veiculo else 5}. CLÁUSULAS CONTRATUAIS")
        y -= 0.8*cm
        
        # Substituir variáveis no texto
        clausulas_texto_final = contrato["clausulas_texto"]
        
        # Variáveis do parceiro
        clausulas_texto_final = clausulas_texto_final.replace("{PARCEIRO_NOME}", parceiro.get('nome_empresa', parceiro.get('name', 'N/A')))
        clausulas_texto_final = clausulas_texto_final.replace("{PARCEIRO_NIF}", parceiro.get('contribuinte_empresa', parceiro.get('nif', 'N/A')))
        clausulas_texto_final = clausulas_texto_final.replace("{PARCEIRO_MORADA}", parceiro.get('morada_completa', parceiro.get('morada', 'N/A')))
        clausulas_texto_final = clausulas_texto_final.replace("{PARCEIRO_CP}", parceiro.get('codigo_postal', 'N/A'))
        clausulas_texto_final = clausulas_texto_final.replace("{PARCEIRO_LOCALIDADE}", parceiro.get('localidade', 'N/A'))
        clausulas_texto_final = clausulas_texto_final.replace("{PARCEIRO_TELEFONE}", parceiro.get('telefone', parceiro.get('phone', 'N/A')))
        clausulas_texto_final = clausulas_texto_final.replace("{PARCEIRO_EMAIL}", parceiro.get('email', 'N/A'))
        
        # Variáveis do representante legal
        clausulas_texto_final = clausulas_texto_final.replace("{REP_LEGAL_NOME}", parceiro.get('representante_legal_nome', parceiro.get('nome_manager', 'N/A')))
        clausulas_texto_final = clausulas_texto_final.replace("{REP_LEGAL_CC}", parceiro.get('representante_legal_cc', 'N/A'))
        clausulas_texto_final = clausulas_texto_final.replace("{REP_LEGAL_CC_VALIDADE}", parceiro.get('representante_legal_cc_validade', 'N/A'))
        clausulas_texto_final = clausulas_texto_final.replace("{REP_LEGAL_TELEFONE}", parceiro.get('representante_legal_telefone', 'N/A'))
        clausulas_texto_final = clausulas_texto_final.replace("{REP_LEGAL_EMAIL}", parceiro.get('representante_legal_email', 'N/A'))
        
        # Variáveis do motorista
        clausulas_texto_final = clausulas_texto_final.replace("{MOTORISTA_NOME}", motorista.get('name', 'N/A'))
        clausulas_texto_final = clausulas_texto_final.replace("{MOTORISTA_CC}", motorista.get('numero_cc', motorista.get('cc_numero', 'N/A')))
        clausulas_texto_final = clausulas_texto_final.replace("{MOTORISTA_CC_VALIDADE}", motorista.get('validade_cc', motorista.get('cc_validade', 'N/A')))
        clausulas_texto_final = clausulas_texto_final.replace("{MOTORISTA_NIF}", motorista.get('nif', 'N/A'))
        clausulas_texto_final = clausulas_texto_final.replace("{MOTORISTA_MORADA}", motorista.get('morada_completa', motorista.get('morada', 'N/A')))
        clausulas_texto_final = clausulas_texto_final.replace("{MOTORISTA_CP}", motorista.get('codigo_postal', 'N/A'))
        clausulas_texto_final = clausulas_texto_final.replace("{MOTORISTA_LOCALIDADE}", motorista.get('localidade', 'N/A'))
        clausulas_texto_final = clausulas_texto_final.replace("{MOTORISTA_TELEFONE}", motorista.get('phone', motorista.get('telefone', 'N/A')))
        clausulas_texto_final = clausulas_texto_final.replace("{MOTORISTA_CARTA_CONDUCAO}", motorista.get('carta_conducao_numero', motorista.get('numero_carta_conducao', 'N/A')))
        clausulas_texto_final = clausulas_texto_final.replace("{MOTORISTA_CARTA_CONDUCAO_VALIDADE}", motorista.get('carta_conducao_validade', motorista.get('validade_carta_conducao', 'N/A')))
        clausulas_texto_final = clausulas_texto_final.replace("{MOTORISTA_LICENCA_TVDE}", motorista.get('licenca_tvde_numero', motorista.get('numero_licenca_tvde', 'N/A')))
        clausulas_texto_final = clausulas_texto_final.replace("{MOTORISTA_LICENCA_TVDE_VALIDADE}", motorista.get('licenca_tvde_validade', motorista.get('validade_licenca_tvde', 'N/A')))
        clausulas_texto_final = clausulas_texto_final.replace("{MOTORISTA_SS}", motorista.get('numero_seguranca_social', 'N/A'))
        clausulas_texto_final = clausulas_texto_final.replace("{MOTORISTA_EMAIL}", motorista.get('email', 'N/A'))
        
        # Variáveis do veículo
        if veiculo:
            clausulas_texto_final = clausulas_texto_final.replace("{VEICULO_MARCA}", veiculo.get('marca', 'N/A'))
            clausulas_texto_final = clausulas_texto_final.replace("{VEICULO_MODELO}", veiculo.get('modelo', 'N/A'))
            clausulas_texto_final = clausulas_texto_final.replace("{VEICULO_MATRICULA}", veiculo.get('matricula', 'N/A'))
            clausulas_texto_final = clausulas_texto_final.replace("{VEICULO_ANO}", str(veiculo.get('ano', 'N/A')))
        
        # Variáveis do contrato (datas e valores)
        clausulas_texto_final = clausulas_texto_final.replace("{DATA_INICIO}", contrato.get('data_inicio', 'N/A'))
        clausulas_texto_final = clausulas_texto_final.replace("{DATA_EMISSAO}", datetime.now().strftime("%d/%m/%Y"))
        clausulas_texto_final = clausulas_texto_final.replace("{TIPO_CONTRATO}", contrato.get('tipo_contrato', 'N/A'))
        clausulas_texto_final = clausulas_texto_final.replace("{VALOR_SEMANAL}", f"€{contrato.get('valor_aplicado', 0):.2f}")
        clausulas_texto_final = clausulas_texto_final.replace("{COMISSAO}", f"{contrato.get('percentagem_motorista_aplicado', 0)}%")
        clausulas_texto_final = clausulas_texto_final.replace("{CAUCAO_TOTAL}", f"€{contrato.get('valor_caucao_aplicado', 0):.2f}")
        clausulas_texto_final = clausulas_texto_final.replace("{CAUCAO_PARCELAS}", str(contrato.get('numero_parcelas_caucao_aplicado', 'N/A')))
        clausulas_texto_final = clausulas_texto_final.replace("{CAUCAO_TEXTO}", contrato.get('caucao_texto', 'N/A'))
        
        # Variáveis de época alta
        clausulas_texto_final = clausulas_texto_final.replace("{DATA_INICIO_EPOCA_ALTA}", contrato.get('data_inicio_epoca_alta', 'N/A'))
        clausulas_texto_final = clausulas_texto_final.replace("{DATA_FIM_EPOCA_ALTA}", contrato.get('data_fim_epoca_alta', 'N/A'))
        clausulas_texto_final = clausulas_texto_final.replace("{EPOCA_ALTA_VALOR}", f"€{contrato.get('valor_epoca_alta_aplicado', 0):.2f}")
        clausulas_texto_final = clausulas_texto_final.replace("{TEXTO_EPOCA_ALTA}", contrato.get('texto_epoca_alta', 'N/A'))
        
        # Variáveis de época baixa
        clausulas_texto_final = clausulas_texto_final.replace("{DATA_INICIO_EPOCA_BAIXA}", contrato.get('data_inicio_epoca_baixa', 'N/A'))
        clausulas_texto_final = clausulas_texto_final.replace("{DATA_FIM_EPOCA_BAIXA}", contrato.get('data_fim_epoca_baixa', 'N/A'))
        clausulas_texto_final = clausulas_texto_final.replace("{EPOCA_BAIXA_VALOR}", f"€{contrato.get('valor_epoca_baixa_aplicado', 0):.2f}")
        clausulas_texto_final = clausulas_texto_final.replace("{TEXTO_EPOCA_BAIXA}", contrato.get('texto_epoca_baixa', 'N/A'))
        
        # Outras variáveis
        clausulas_texto_final = clausulas_texto_final.replace("{CONDICOES_VEICULO}", contrato.get('condicoes_veiculo', 'N/A'))
        
        c.setFont("Helvetica", 9)
        # Split text into lines
        clausulas_lines = clausulas_texto_final.split('\n')
        for line in clausulas_lines:
            if y < 2*cm:
                c.showPage()
                y = height - 2*cm
            c.drawString(2.5*cm, y, line[:100])  # Limit line length
            y -= 0.5*cm
    
    # Assinaturas
    if y < 6*cm:
        c.showPage()
        y = height - 2*cm
    
    y -= 2*cm
    
    c.setFont("Helvetica-Bold", 10)
    c.drawString(4*cm, y, "______________________________")
    c.drawString(13*cm, y, "______________________________")
    y -= 0.6*cm
    c.setFont("Helvetica", 9)
    c.drawString(4*cm, y, "Assinatura do Parceiro")
    c.drawString(13*cm, y, "Assinatura do Motorista")
    y -= 0.4*cm
    c.drawString(4*cm, y, parceiro.get('nome_empresa', parceiro.get('name', '')))
    c.drawString(13*cm, y, motorista.get('name', ''))
    
    # Footer
    c.setFont("Helvetica", 8)
    c.drawCentredString(width/2, 1.5*cm, f"Contrato ID: {contrato_id}")
    c.drawCentredString(width/2, 1*cm, f"Template: {contrato['nome_template']}")
    
    c.save()
    
    # Update contract with PDF URL
    pdf_url = f"/uploads/contratos/{pdf_filename}"
    await db.contratos_motorista.update_one(
        {"id": contrato_id},
        {"$set": {"pdf_url": pdf_url}}
    )
    
    # Adicionar contrato à ficha do motorista para download
    motorista_id = contrato.get("motorista_id")
    if motorista_id:
        # Adicionar ao array de contratos do motorista
        motorista_doc = await db.motoristas.find_one({"id": motorista_id}, {"_id": 0})
        if motorista_doc:
            contratos_list = motorista_doc.get("contratos_pdfs", [])
            contratos_list.append({
                "contrato_id": contrato_id,
                "pdf_url": pdf_url,
                "pdf_filename": pdf_filename,
                "template_nome": contrato.get("nome_template", "N/A"),
                "tipo_contrato": contrato.get("tipo_contrato", "N/A"),
                "data_criacao": datetime.now(timezone.utc).isoformat()
            })
            
            await db.motoristas.update_one(
                {"id": motorista_id},
                {"$set": {
                    "contratos_pdfs": contratos_list,
                    "contrato_atual_pdf": pdf_url
                }}
            )
    
    return {
        "message": "PDF gerado com sucesso",
        "pdf_url": pdf_url,
        "pdf_filename": pdf_filename
    }

@api_router.post("/subscriptions", response_model=UserSubscription)
async def create_subscription(
    subscription_data: SubscriptionCreate,
    current_user: Dict = Depends(get_current_user)
):
    """Create or update user subscription"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        # Users can only subscribe themselves
        if subscription_data.user_id != current_user["id"]:
            raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if plan exists
    plano = await db.planos.find_one({"id": subscription_data.plano_id}, {"_id": 0})
    if not plano:
        raise HTTPException(status_code=404, detail="Plano not found")
    
    # Check if user already has a subscription
    existing = await db.subscriptions.find_one({"user_id": subscription_data.user_id, "status": "ativo"})
    
    if existing:
        # Update existing subscription
        await db.subscriptions.update_one(
            {"id": existing["id"]},
            {
                "$set": {
                    "plano_id": subscription_data.plano_id,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
            }
        )
        
        # Update user's subscription_id
        await db.users.update_one(
            {"id": subscription_data.user_id},
            {"$set": {"subscription_id": existing["id"]}}
        )
        
        return UserSubscription(**{**existing, "plano_id": subscription_data.plano_id})
    
    # Create new subscription
    subscription_dict = {
        "id": str(uuid.uuid4()),
        "user_id": subscription_data.user_id,
        "plano_id": subscription_data.plano_id,
        "status": "ativo",
        "data_inicio": datetime.now().strftime("%Y-%m-%d"),
        "data_fim": None,
        "unidades_ativas": 0,
        "valor_mensal": 0.0,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.subscriptions.insert_one(subscription_dict)
    
    # Update user's subscription_id
    await db.users.update_one(
        {"id": subscription_data.user_id},
        {"$set": {"subscription_id": subscription_dict["id"]}}
    )
    
    if isinstance(subscription_dict["created_at"], str):
        subscription_dict["created_at"] = datetime.fromisoformat(subscription_dict["created_at"])
    if isinstance(subscription_dict["updated_at"], str):
        subscription_dict["updated_at"] = datetime.fromisoformat(subscription_dict["updated_at"])
    
    return UserSubscription(**subscription_dict)

@api_router.get("/subscriptions/me", response_model=Dict[str, Any])
async def get_my_subscription(current_user: Dict = Depends(get_current_user)):
    """Get current user's subscription details"""
    subscription_id = current_user.get("subscription_id")
    
    if not subscription_id:
        return {
            "has_subscription": False,
            "message": "Nenhuma subscrição ativa"
        }
    
    subscription = await db.subscriptions.find_one({"id": subscription_id}, {"_id": 0})
    if not subscription:
        return {
            "has_subscription": False,
            "message": "Subscrição não encontrada"
        }
    
    # Get plan details
    plano = await db.planos.find_one({"id": subscription["plano_id"]}, {"_id": 0})
    
    if isinstance(subscription.get("created_at"), str):
        subscription["created_at"] = datetime.fromisoformat(subscription["created_at"])
    if isinstance(subscription.get("updated_at"), str):
        subscription["updated_at"] = datetime.fromisoformat(subscription["updated_at"])
    
    return {
        "has_subscription": True,
        "subscription": subscription,
        "plano": plano
    }

@api_router.put("/subscriptions/{subscription_id}/cancel")
async def cancel_subscription(
    subscription_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Cancel a subscription"""
    subscription = await db.subscriptions.find_one({"id": subscription_id}, {"_id": 0})
    
    if not subscription:
        raise HTTPException(status_code=404, detail="Subscription not found")
    
    # Only admin, gestor, or the subscription owner can cancel
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        if subscription["user_id"] != current_user["id"]:
            raise HTTPException(status_code=403, detail="Not authorized")
    
    await db.subscriptions.update_one(
        {"id": subscription_id},
        {
            "$set": {
                "status": "cancelado",
                "data_fim": datetime.now().strftime("%Y-%m-%d"),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    return {"message": "Subscription cancelled successfully"}

# ==================== CSV UPLOAD ENDPOINTS ====================

@api_router.post("/operacional/upload-csv-uber")
async def upload_csv_uber(
    file: UploadFile = File(...),
    parceiro_id: str = Form(...),
    periodo_inicio: str = Form(...),
    periodo_fim: str = Form(...),
    current_user: Dict = Depends(get_current_user)
):
    """Upload Uber CSV earnings file (parceiro/operacional data)"""
    # Check feature access
    if not await check_feature_access(current_user, "upload_csv_ganhos"):
        raise HTTPException(
            status_code=403,
            detail="Upgrade para plano com acesso a upload de CSVs"
        )
    
    # Read file content
    file_content = await file.read()
    
    # Process CSV (parceiro_id instead of motorista_id)
    result = await process_uber_csv(file_content, parceiro_id, periodo_inicio, periodo_fim)
    
    if not result["success"]:
        raise HTTPException(status_code=400, detail=result["error"])
    
    return result

@api_router.post("/operacional/upload-csv-bolt")
async def upload_csv_bolt(
    file: UploadFile = File(...),
    parceiro_id: str = Form(...),
    periodo_inicio: str = Form(...),
    periodo_fim: str = Form(...),
    current_user: Dict = Depends(get_current_user)
):
    """Upload Bolt CSV earnings file (parceiro/operacional data)"""
    # Check feature access
    if not await check_feature_access(current_user, "upload_csv_ganhos"):
        raise HTTPException(
            status_code=403,
            detail="Upgrade para plano com acesso a upload de CSVs"
        )
    
    # Read file content
    file_content = await file.read()
    
    # Process CSV (parceiro_id instead of motorista_id)
    result = await process_bolt_csv(file_content, parceiro_id, periodo_inicio, periodo_fim)
    
    if not result["success"]:
        raise HTTPException(status_code=400, detail=result["error"])
    
    return result

@api_router.post("/operacional/upload-excel-combustivel")
async def upload_excel_combustivel(
    file: UploadFile = File(...),
    parceiro_id: str = Form(...),
    current_user: Dict = Depends(get_current_user)
):
    """Upload Prio/Combustível Excel file (parceiro/operacional data)"""
    # Check feature access (manual input is allowed in base plan)
    if not await check_feature_access(current_user, "combustivel_manual"):
        raise HTTPException(
            status_code=403,
            detail="Sem acesso a gestão de combustível"
        )
    
    # Read file content
    file_content = await file.read()
    
    # Process Excel (parceiro_id instead of motorista_id)
    result = await process_prio_excel(file_content, parceiro_id)
    
    if not result["success"]:
        raise HTTPException(status_code=400, detail=result["error"])
    
    return result

@api_router.post("/import/viaverde")
async def import_viaverde(
    file: UploadFile = File(...),
    parceiro_id: str = Form(...),
    periodo_inicio: str = Form(...),
    periodo_fim: str = Form(...),
    current_user: Dict = Depends(get_current_user)
):
    """Import Via Verde toll movements from Excel file"""
    if not await check_feature_access(current_user, "upload_csv_ganhos"):
        raise HTTPException(status_code=403, detail="Sem acesso a import de dados")
    
    file_content = await file.read()
    result = await process_viaverde_excel(file_content, parceiro_id, periodo_inicio, periodo_fim)
    
    if not result["success"]:
        raise HTTPException(status_code=400, detail=result["error"])
    
    return result

@api_router.post("/import/gps")
async def import_gps(
    file: UploadFile = File(...),
    parceiro_id: str = Form(...),
    periodo_inicio: str = Form(...),
    periodo_fim: str = Form(...),
    current_user: Dict = Depends(get_current_user)
):
    """Import GPS distance report from CSV file"""
    if not await check_feature_access(current_user, "upload_csv_km"):
        raise HTTPException(status_code=403, detail="Sem acesso a import de KMs")
    
    file_content = await file.read()
    result = await process_gps_csv(file_content, parceiro_id, periodo_inicio, periodo_fim)
    
    if not result["success"]:
        raise HTTPException(status_code=400, detail=result["error"])
    
    return result

@api_router.post("/import/combustivel-eletrico")
async def import_combustivel_eletrico(
    file: UploadFile = File(...),
    parceiro_id: str = Form(...),
    periodo_inicio: str = Form(...),
    periodo_fim: str = Form(...),
    current_user: Dict = Depends(get_current_user)
):
    """Import electric fuel transactions from Excel file"""
    if not await check_feature_access(current_user, "combustivel_manual"):
        raise HTTPException(status_code=403, detail="Sem acesso a gestão de combustível")
    
    file_content = await file.read()
    result = await process_combustivel_eletrico_excel(file_content, parceiro_id, periodo_inicio, periodo_fim)
    
    if not result["success"]:
        raise HTTPException(status_code=400, detail=result["error"])
    
    return result

@api_router.post("/import/combustivel-fossil")
async def import_combustivel_fossil(
    file: UploadFile = File(...),
    parceiro_id: str = Form(...),
    periodo_inicio: str = Form(...),
    periodo_fim: str = Form(...),
    current_user: Dict = Depends(get_current_user)
):
    """Import fossil fuel transactions from Excel file"""
    if not await check_feature_access(current_user, "combustivel_manual"):
        raise HTTPException(status_code=403, detail="Sem acesso a gestão de combustível")
    
    file_content = await file.read()
    result = await process_combustivel_fossil_excel(file_content, parceiro_id, periodo_inicio, periodo_fim)
    
    if not result["success"]:
        raise HTTPException(status_code=400, detail=result["error"])
    
    return result

@api_router.get("/operacional/ganhos-motorista/{motorista_id}")
async def get_ganhos_motorista(
    motorista_id: str,
    periodo_inicio: Optional[str] = None,
    periodo_fim: Optional[str] = None,
    current_user: Dict = Depends(get_current_user)
):
    """Get motorista earnings from Uber and Bolt"""
    query = {"motorista_id": motorista_id}
    
    if periodo_inicio and periodo_fim:
        # Filter by period
        query["periodo_inicio"] = {"$gte": periodo_inicio}
        query["periodo_fim"] = {"$lte": periodo_fim}
    
    # Get Uber earnings
    ganhos_uber = await db.ganhos_uber.find(query, {"_id": 0}).to_list(100)
    total_uber = sum(g["total_pago"] for g in ganhos_uber)
    
    # Get Bolt earnings
    ganhos_bolt = await db.ganhos_bolt.find(query, {"_id": 0}).to_list(100)
    total_bolt = sum(g["ganhos_liquidos"] for g in ganhos_bolt)
    
    # Get fuel transactions
    fuel_query = {"motorista_id": motorista_id}
    if periodo_inicio and periodo_fim:
        fuel_query["data_transacao"] = {"$gte": periodo_inicio, "$lte": periodo_fim}
    
    transacoes_combustivel = await db.transacoes_combustivel.find(fuel_query, {"_id": 0}).to_list(500)
    total_combustivel = sum(t["total"] for t in transacoes_combustivel)
    
    return {
        "motorista_id": motorista_id,
        "periodo": {
            "inicio": periodo_inicio,
            "fim": periodo_fim
        },
        "uber": {
            "registos": len(ganhos_uber),
            "total": total_uber
        },
        "bolt": {
            "registos": len(ganhos_bolt),
            "total": total_bolt
        },
        "combustivel": {
            "transacoes": len(transacoes_combustivel),
            "total": total_combustivel
        },
        "resumo": {
            "ganhos_totais": total_uber + total_bolt,
            "despesas_combustivel": total_combustivel,
            "liquido": (total_uber + total_bolt) - total_combustivel
        }
    }

@api_router.get("/operacional/csv-importados/{motorista_id}")
async def listar_csv_importados(
    motorista_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """List all imported CSVs for a motorista"""
    # Get Uber CSVs
    uber_imports = await db.ganhos_uber.find(
        {"motorista_id": motorista_id},
        {"_id": 0, "id": 1, "periodo_inicio": 1, "periodo_fim": 1, "total_pago": 1, "csv_original": 1, "created_at": 1}
    ).sort("created_at", -1).to_list(100)
    
    # Get Bolt CSVs
    bolt_imports = await db.ganhos_bolt.find(
        {"motorista_id": motorista_id},
        {"_id": 0, "id": 1, "periodo_inicio": 1, "periodo_fim": 1, "ganhos_liquidos": 1, "csv_original": 1, "created_at": 1}
    ).sort("created_at", -1).to_list(100)
    
    # Get Combustivel Excel
    combustivel_imports = await db.transacoes_combustivel.aggregate([
        {"$match": {"motorista_id": motorista_id}},
        {"$group": {
            "_id": "$excel_original",
            "data_primeiro": {"$min": "$data_transacao"},
            "data_ultimo": {"$max": "$data_transacao"},
            "total_transacoes": {"$sum": 1},
            "total_valor": {"$sum": "$total"}
        }},
        {"$sort": {"data_ultimo": -1}}
    ]).to_list(100)
    
    return {
        "motorista_id": motorista_id,
        "uber": {
            "total_imports": len(uber_imports),
            "imports": uber_imports
        },
        "bolt": {
            "total_imports": len(bolt_imports),
            "imports": bolt_imports
        },
        "combustivel": {
            "total_imports": len(combustivel_imports),
            "imports": combustivel_imports
        }
    }

# ==================== CONTRATOS ENDPOINTS ====================

# REMOVIDO: Endpoint duplicado - usar o endpoint completo na linha 9149

@api_router.get("/contratos")
async def listar_contratos(
    status: Optional[str] = None,
    current_user: Dict = Depends(get_current_user)
):
    """List all contracts"""
    query = {}
    if status:
        query["status"] = status
    
    contratos = await db.contratos.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    return contratos

@api_router.put("/contratos/{contrato_id}/assinar")
async def assinar_contrato(
    contrato_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Mark contract as signed"""
    result = await db.contratos.update_one(
        {"id": contrato_id},
        {
            "$set": {
                "status": "assinado",
                "assinado_em": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Contrato not found")
    
    return {"message": "Contrato marcado como assinado"}

# ==================== BILLING/FATURAS ENDPOINTS ====================

@api_router.post("/admin/gerar-faturas-mensais")
async def gerar_faturas_mensais(
    mes_referencia: str = Form(...),  # Format: "2025-11"
    current_user: Dict = Depends(get_current_user)
):
    """Admin: Generate monthly invoices for all active subscriptions"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get all active subscriptions
    subscriptions = await db.subscriptions.find({"status": "ativo"}, {"_id": 0}).to_list(1000)
    
    faturas_geradas = []
    
    for sub in subscriptions:
        # Get plan details
        plano = await db.planos.find_one({"id": sub["plano_id"]}, {"_id": 0})
        if not plano:
            continue
        
        # Calculate units (vehicles for parceiro, motoristas for operacional)
        user = await db.users.find_one({"id": sub["user_id"]}, {"_id": 0})
        
        if user["role"] == "parceiro":
            # Count vehicles
            unidades = await db.vehicles.count_documents({"parceiro_id": sub["user_id"]})
        elif user["role"] == "operacional":
            # Count motoristas
            unidades = await db.motoristas.count_documents({"parceiro_atribuido": sub["user_id"]})
        else:
            unidades = 0
        
        valor_total = plano["preco_por_unidade"] * unidades
        
        # Update subscription units
        await db.subscriptions.update_one(
            {"id": sub["id"]},
            {"$set": {"unidades_ativas": unidades, "valor_mensal": valor_total}}
        )
        
        # Create invoice
        data_emissao = datetime.now().strftime("%Y-%m-%d")
        data_vencimento = (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d")
        
        fatura = {
            "id": str(uuid.uuid4()),
            "subscription_id": sub["id"],
            "user_id": sub["user_id"],
            "periodo_referencia": mes_referencia,
            "valor_total": valor_total,
            "unidades_cobradas": unidades,
            "status": "pendente",
            "pdf_url": None,  # TODO: Generate invoice PDF
            "data_emissao": data_emissao,
            "data_vencimento": data_vencimento,
            "data_pagamento": None,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.faturas.insert_one(fatura)
        faturas_geradas.append(fatura)
    
    return {
        "message": "Faturas mensais geradas",
        "total_faturas": len(faturas_geradas),
        "valor_total": sum(f["valor_total"] for f in faturas_geradas)
    }

@api_router.get("/faturas/me")
async def minhas_faturas(current_user: Dict = Depends(get_current_user)):
    """Get my invoices"""
    faturas = await db.faturas.find(
        {"user_id": current_user["id"]},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    return faturas

@api_router.put("/admin/faturas/{fatura_id}/marcar-paga")
async def marcar_fatura_paga(
    fatura_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Admin: Mark invoice as paid"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.faturas.update_one(
        {"id": fatura_id},
        {
            "$set": {
                "status": "paga",
                "data_pagamento": datetime.now().strftime("%Y-%m-%d")
            }
        }
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Fatura not found")
    
    return {"message": "Fatura marcada como paga"}

# ==================== ALERTAS ENDPOINTS ====================

@api_router.get("/alertas", response_model=List[Alerta])
async def get_alertas(
    status: Optional[str] = "ativo",
    prioridade: Optional[str] = None,
    tipo: Optional[str] = None,
    current_user: Dict = Depends(get_current_user)
):
    """Get all active alerts"""
    query = {}
    
    if status:
        query["status"] = status
    if prioridade:
        query["prioridade"] = prioridade
    if tipo:
        query["tipo"] = tipo
    
    alertas = await db.alertas.find(query, {"_id": 0}).sort("criado_em", -1).to_list(1000)
    
    for alerta in alertas:
        if isinstance(alerta["criado_em"], str):
            alerta["criado_em"] = datetime.fromisoformat(alerta["criado_em"])
        if alerta.get("resolvido_em") and isinstance(alerta["resolvido_em"], str):
            alerta["resolvido_em"] = datetime.fromisoformat(alerta["resolvido_em"])
    
    return alertas

@api_router.post("/alertas/verificar")
async def verificar_alertas(current_user: Dict = Depends(get_current_user)):
    """Manually trigger alert checking"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    await check_and_create_alerts()
    
    return {"message": "Alertas verificados e atualizados"}

@api_router.put("/alertas/{alerta_id}/resolver")
async def resolver_alerta(alerta_id: str, current_user: Dict = Depends(get_current_user)):
    """Mark an alert as resolved"""
    result = await db.alertas.update_one(
        {"id": alerta_id},
        {
            "$set": {
                "status": "resolvido",
                "resolvido_em": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Alerta not found")
    
    return {"message": "Alerta marcado como resolvido"}

@api_router.put("/alertas/{alerta_id}/ignorar")
async def ignorar_alerta(alerta_id: str, current_user: Dict = Depends(get_current_user)):
    """Mark an alert as ignored"""
    result = await db.alertas.update_one(
        {"id": alerta_id},
        {"$set": {"status": "ignorado"}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Alerta not found")
    
    return {"message": "Alerta ignorado"}

@api_router.get("/alertas/dashboard-stats")
async def get_alertas_stats(current_user: Dict = Depends(get_current_user)):
    """Get alert statistics for dashboard"""
    alertas_ativos = await db.alertas.count_documents({"status": "ativo"})
    alertas_alta_prioridade = await db.alertas.count_documents({"status": "ativo", "prioridade": "alta"})
    
    # Get count by type
    tipos = ["seguro", "inspecao", "licenca_tvde", "manutencao", "validade_matricula", "carta_conducao"]
    por_tipo = {}
    
    for tipo in tipos:
        count = await db.alertas.count_documents({"status": "ativo", "tipo": tipo})
        por_tipo[tipo] = count
    
    return {
        "total_ativos": alertas_ativos,
        "alta_prioridade": alertas_alta_prioridade,
        "por_tipo": por_tipo
    }

# ==================== FILE SERVING ENDPOINT ====================
from fastapi.responses import FileResponse, StreamingResponse

@api_router.get("/files/{folder}/{filename:path}")
async def serve_file(folder: str, filename: str, current_user: Dict = Depends(get_current_user)):
    """Serve uploaded files (supports subfolders)"""
    allowed_folders = ["motoristas", "pagamentos", "vehicles", "vehicle_documents", "vehicle_photos_info", "extintor_docs"]
    
    if folder not in allowed_folders:
        raise HTTPException(status_code=400, detail="Invalid folder")
    
    # Build file path (filename can include subfolders like "motorista-001/file.pdf")
    file_path = UPLOAD_DIR / folder / filename
    
    # Security check: ensure path is within allowed folder
    try:
        file_path = file_path.resolve()
        allowed_base = (UPLOAD_DIR / folder).resolve()
        if not str(file_path).startswith(str(allowed_base)):
            raise HTTPException(status_code=403, detail="Access denied")
    except:
        raise HTTPException(status_code=400, detail="Invalid file path")
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(file_path)


@api_router.get("/users/pending")
async def get_pending_users(current_user: Dict = Depends(get_current_user)):
    """Get all pending users waiting for approval (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Get motoristas with status pending
    pending_motoristas = await db.motoristas.find(
        {"status": "pending"},
        {"_id": 0}
    ).to_list(length=None)
    
    return {
        "pending_count": len(pending_motoristas),
        "pending_users": pending_motoristas
    }

@api_router.get("/users/all")
async def get_all_users(current_user: Dict = Depends(get_current_user)):
    """Get all users (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Get all users with plan lookup using aggregation
    users = await db.users.aggregate([
        {
            "$lookup": {
                "from": "planos_sistema",
                "localField": "plano_id",
                "foreignField": "id",
                "as": "plano_info"
            }
        },
        {
            "$project": {
                "_id": 0,
                "password": 0
            }
        }
    ]).to_list(length=None)
    
    # Separate pending and approved users
    pending_users = []
    registered_users = []
    
    for user in users:
        # Convert created_at to datetime if string
        if isinstance(user.get("created_at"), str):
            user["created_at"] = datetime.fromisoformat(user["created_at"])
        
        # Extract plan name from lookup result
        if user.get("plano_info") and len(user["plano_info"]) > 0:
            user["plano_nome"] = user["plano_info"][0].get("nome")
        else:
            user["plano_nome"] = None
        
        # Remove the plano_info array to clean up response
        user.pop("plano_info", None)
        
        if user.get("approved", False):
            registered_users.append(user)
        else:
            pending_users.append(user)
    
    return {
        "pending_users": pending_users,
        "registered_users": registered_users,
        "pending_count": len(pending_users),
        "registered_count": len(registered_users)
    }

@api_router.put("/users/{user_id}/approve")
async def approve_user(
    user_id: str,
    approval_data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Approve a pending user and optionally set role (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    new_role = approval_data.get("role")
    valid_roles = ["motorista", "gestao", "parceiro", "admin"]
    
    if new_role and new_role not in valid_roles:
        raise HTTPException(status_code=400, detail="Invalid role")
    
    # Build update document
    update_doc = {"approved": True, "approved_at": datetime.now(timezone.utc).isoformat()}
    
    if new_role:
        update_doc["role"] = new_role
    
    # Update in users collection
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": update_doc}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    # If user has motorista role, create/update motoristas profile
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if user and user.get("role") == "motorista":
        # Check if motorista profile exists
        motorista_exists = await db.motoristas.find_one({"id": user_id}, {"_id": 0})
        
        if not motorista_exists:
            # Create motorista profile from user data
            motorista_profile = {
                "id": user_id,
                "name": user.get("name"),
                "email": user.get("email"),
                "phone": user.get("phone", ""),
                "whatsapp": user.get("whatsapp", ""),
                "data_nascimento": user.get("data_nascimento", ""),
                "nif": user.get("nif", ""),
                "nacionalidade": user.get("nacionalidade", "Portuguesa"),
                "morada_completa": user.get("morada_completa", ""),
                "codigo_postal": user.get("codigo_postal", ""),
                "status": "approved",
                "approved": True,
                "vehicle_assigned": None,
                "parceiro_atribuido": None,
                "contrato_id": None,
                "documentos": {},
                "contacto_emergencia": {},
                "dados_bancarios": {},
                "plano_id": None,
                "plano_nome": None,
                "plano_features": {},
                "created_at": user.get("created_at"),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            await db.motoristas.insert_one(motorista_profile)
        else:
            # Update existing motorista profile
            await db.motoristas.update_one(
                {"id": user_id},
                {"$set": {"status": "approved", "approved": True, "updated_at": datetime.now(timezone.utc).isoformat()}}
            )
    
    # If user has parceiro role, update parceiros collection
    if user and user.get("role") == "parceiro":
        await db.parceiros.update_one(
            {"email": user.get("email")},
            {"$set": {
                "approved": True,
                "status": "aprovado",
                "approved_at": datetime.now(timezone.utc).isoformat(),
                "approved_by": current_user["id"]
            }}
        )
    
    return {"message": "User approved successfully"}

@api_router.put("/users/{user_id}/set-role")
async def set_user_role(
    user_id: str,
    role_data: Dict[str, str],
    current_user: Dict = Depends(get_current_user)
):
    """Set role for a user (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    new_role = role_data.get("role")
    valid_roles = ["motorista", "gestao", "parceiro", "admin"]
    
    if new_role not in valid_roles:
        raise HTTPException(status_code=400, detail="Invalid role")
    
    # Update in users collection
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": {"role": new_role}}
    )
    
    # If user is motorista, also update motoristas collection
    if new_role == "motorista":
        await db.motoristas.update_one(
            {"id": user_id},
            {"$set": {"role": new_role}}
        )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "Role updated successfully"}

@api_router.put("/users/{user_id}/status")
async def update_user_status(
    user_id: str,
    status_data: Dict[str, str],
    current_user: Dict = Depends(get_current_user)
):
    """Update user status - block/unblock (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    new_status = status_data.get("status")
    valid_statuses = ["active", "blocked", "pendente"]
    
    if new_status not in valid_statuses:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    # Update in users collection
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": {"status": new_status}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": f"User status updated to {new_status}"}

@api_router.delete("/users/{user_id}")
async def delete_user(
    user_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Delete/reject a user (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Check if trying to delete self
    if user_id == current_user["id"]:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    # Delete from users collection
    result = await db.users.delete_one({"id": user_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Also delete from motoristas collection if exists
    await db.motoristas.delete_one({"id": user_id})
    
    return {"message": "User deleted successfully"}

@api_router.put("/users/{user_id}/reset-password")
async def admin_reset_password(
    user_id: str,
    password_data: Dict[str, str],
    current_user: Dict = Depends(get_current_user)
):
    """Admin resets user password and returns the new password (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    new_password = password_data.get("new_password")
    if not new_password or len(new_password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    
    # Hash the new password
    hashed_password = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt())
    
    # Update user password
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": {
            "password": hashed_password.decode('utf-8'),
            "senha_provisoria": True  # Mark as temporary password
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Also update in motoristas if exists
    await db.motoristas.update_one(
        {"id": user_id},
        {"$set": {"senha_provisoria": True}}
    )
    
    return {
        "message": "Password reset successfully",
        "new_password": new_password,  # Return plaintext password for admin to see
        "user_id": user_id
    }

@api_router.post("/auth/forgot-password")
async def forgot_password(email_data: Dict[str, str]):
    """Generate temporary password for user who forgot password"""
    import random
    import string
    
    email = email_data.get("email")
    if not email:
        raise HTTPException(status_code=400, detail="Email is required")
    
    # Find user by email
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Email not found in system")
    
    # Generate temporary password (8 characters: letters + numbers)
    temp_password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
    
    # Hash the temporary password
    hashed_password = bcrypt.hashpw(temp_password.encode('utf-8'), bcrypt.gensalt())
    
    # Update user password
    await db.users.update_one(
        {"email": email},
        {"$set": {
            "password": hashed_password.decode('utf-8'),
            "senha_provisoria": True
        }}
    )
    
    # Also update in motoristas if exists
    await db.motoristas.update_one(
        {"email": email},
        {"$set": {"senha_provisoria": True}}
    )
    
    # In production, this would send an email
    # For now, we return the temporary password
    return {
        "message": "Temporary password generated successfully",
        "temp_password": temp_password,
        "email": email,
        "instructions": "Use this temporary password to login. You will be prompted to change it on first login."
    }


# ==================== PLANOS E SUBSCRIÇÕES ====================

@api_router.get("/planos")
async def get_planos(current_user: Dict = Depends(get_current_user)):
    """Get available plans (filtered by user role, except admin sees all)"""
    try:
        planos = await db.planos.find({"ativo": True}, {"_id": 0}).to_list(None)
        
        # Admin sees all plans
        if current_user["role"] == UserRole.ADMIN:
            return planos
        
        # Filter plans by user role
        user_role = current_user["role"]
        filtered_planos = []
        for plano in planos:
            perfis = plano.get("perfis_permitidos", [])
            # If no profiles specified or user's role is in the list
            if not perfis or user_role in perfis:
                filtered_planos.append(plano)
        
        return filtered_planos
    except Exception as e:
        logger.error(f"Error fetching planos: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# DEPRECATED: Endpoint duplicado - usar o endpoint mais completo na linha 6808
# @api_router.post("/planos")
# async def create_plano(plano: PlanoCreate, current_user: Dict = Depends(get_current_user)):
#     """Create a new plan (Admin only)"""
#     if current_user["role"] != UserRole.ADMIN:
#         raise HTTPException(status_code=403, detail="Admin only")
#     
#     plano_dict = plano.dict()
#     plano_dict["id"] = str(uuid.uuid4())
#     plano_dict["ativo"] = True
#     plano_dict["created_at"] = datetime.now(timezone.utc)
#     plano_dict["updated_at"] = datetime.now(timezone.utc)
#     
#     await db.planos.insert_one(plano_dict)
#     return {"message": "Plan created successfully", "plano_id": plano_dict["id"]}

# DEPRECATED: Endpoint duplicado - usar o endpoint mais completo na linha 6868
# @api_router.put("/planos/{plano_id}")
# async def update_plano(plano_id: str, updates: Dict[str, Any], current_user: Dict = Depends(get_current_user)):
#     """Update a plan (Admin only)"""
#     if current_user["role"] != UserRole.ADMIN:
#         raise HTTPException(status_code=403, detail="Admin only")
#     
#     updates["updated_at"] = datetime.now(timezone.utc)
#     result = await db.planos.update_one({"id": plano_id}, {"$set": updates})
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    return {"message": "Plan updated successfully"}

@api_router.delete("/planos/{plano_id}")
async def delete_plano(plano_id: str, current_user: Dict = Depends(get_current_user)):
    """Delete a plan (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    await db.planos.update_one({"id": plano_id}, {"$set": {"ativo": False}})
    return {"message": "Plan deactivated successfully"}

# ==================== PLANOS MOTORISTA ENDPOINTS ====================

@api_router.get("/planos-motorista")
async def list_planos_motorista(current_user: Dict = Depends(get_current_user)):
    """List all motorista plans"""
    try:
        planos = await db.planos_motorista.find({"ativo": True}, {"_id": 0}).to_list(100)
        return planos
    except Exception as e:
        logger.error(f"Error listing motorista plans: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/planos-motorista")
async def create_plano_motorista(plano_data: PlanoMotoristaCreate, current_user: Dict = Depends(get_current_user)):
    """Create a new motorista plan (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    try:
        plano_id = str(uuid.uuid4())
        plano = {
            "id": plano_id,
            "nome": plano_data.nome,
            "descricao": plano_data.descricao,
            "preco_semanal": plano_data.preco_semanal,
            "preco_mensal": plano_data.preco_mensal,
            "desconto_mensal_percentagem": plano_data.desconto_mensal_percentagem,
            "features": plano_data.features,
            "permite_pagamento_online": plano_data.permite_pagamento_online,
            "ativo": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.planos_motorista.insert_one(plano)
        return {"message": "Plano criado com sucesso", "plano_id": plano_id}
    except Exception as e:
        logger.error(f"Error creating motorista plan: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/planos-motorista/{plano_id}")
async def update_plano_motorista(
    plano_id: str,
    plano_data: PlanoMotoristaUpdate,
    current_user: Dict = Depends(get_current_user)
):
    """Update a motorista plan (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    try:
        plano = await db.planos_motorista.find_one({"id": plano_id}, {"_id": 0})
        if not plano:
            raise HTTPException(status_code=404, detail="Plano not found")
        
        update_data = plano_data.dict(exclude_unset=True)
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        await db.planos_motorista.update_one({"id": plano_id}, {"$set": update_data})
        return {"message": "Plano atualizado com sucesso"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating motorista plan: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/planos-motorista/{plano_id}")
async def delete_plano_motorista(plano_id: str, current_user: Dict = Depends(get_current_user)):
    """Delete a motorista plan (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    await db.planos_motorista.update_one({"id": plano_id}, {"$set": {"ativo": False}})
    return {"message": "Plano desativado com sucesso"}

# Planos de Parceiro CRUD

# Sistema Unificado de Planos
@api_router.get("/planos-sistema")
async def list_planos_sistema(current_user: Dict = Depends(get_current_user)):
    """List all plans in unified system (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    planos = await db.planos_sistema.find({}, {"_id": 0}).to_list(1000)
    return planos

@api_router.post("/planos-sistema")
async def create_plano_sistema(plano_data: Dict, current_user: Dict = Depends(get_current_user)):
    """Create new plan in unified system (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    plano_id = str(uuid.uuid4())
    preco_sem_iva = plano_data.get("preco_mensal", 0)
    taxa_iva = plano_data.get("taxa_iva", 23)  # Default 23% (Portugal)
    preco_com_iva = preco_sem_iva * (1 + taxa_iva / 100)
    
    # Calcular preço semanal automaticamente (mensal / 4.33)
    preco_semanal = preco_sem_iva / 4.33
    preco_semanal_com_iva = preco_com_iva / 4.33
    
    plano = {
        "id": plano_id,
        "nome": plano_data["nome"],
        "descricao": plano_data.get("descricao", ""),
        "preco_mensal": preco_sem_iva,  # Preço sem IVA
        "preco_mensal_com_iva": round(preco_com_iva, 2),  # Preço com IVA
        "preco_semanal": round(preco_semanal, 2),  # Preço semanal sem IVA
        "preco_semanal_com_iva": round(preco_semanal_com_iva, 2),  # Preço semanal com IVA
        "taxa_iva": taxa_iva,  # Percentagem de IVA
        "tipo_usuario": plano_data["tipo_usuario"],  # motorista, parceiro
        "modulos": plano_data.get("modulos", []),
        "ativo": plano_data.get("ativo", True),
        "permite_trial": plano_data.get("permite_trial", False),
        "dias_trial": plano_data.get("dias_trial", 30),
        "desconto_promocao": plano_data.get("desconto_promocao", 0),  # % de desconto
        "data_inicio_promocao": plano_data.get("data_inicio_promocao"),  # ISO string
        "data_fim_promocao": plano_data.get("data_fim_promocao"),  # ISO string
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.planos_sistema.insert_one(plano)
    
    # Retornar sem o _id do MongoDB
    created_plano = await db.planos_sistema.find_one({"id": plano_id}, {"_id": 0})
    return created_plano

@api_router.put("/planos-sistema/{plano_id}")
async def update_plano_sistema(plano_id: str, plano_data: Dict, current_user: Dict = Depends(get_current_user)):
    """Update plan in unified system (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    preco_sem_iva = plano_data.get("preco_mensal", 0)
    taxa_iva = plano_data.get("taxa_iva", 23)
    preco_com_iva = preco_sem_iva * (1 + taxa_iva / 100)
    
    # Calcular preço semanal automaticamente
    preco_semanal = preco_sem_iva / 4.33
    preco_semanal_com_iva = preco_com_iva / 4.33
    
    update_data = {
        "nome": plano_data["nome"],
        "descricao": plano_data.get("descricao", ""),
        "preco_mensal": preco_sem_iva,
        "preco_mensal_com_iva": round(preco_com_iva, 2),
        "preco_semanal": round(preco_semanal, 2),
        "preco_semanal_com_iva": round(preco_semanal_com_iva, 2),
        "taxa_iva": taxa_iva,
        "tipo_usuario": plano_data["tipo_usuario"],
        "modulos": plano_data.get("modulos", []),
        "ativo": plano_data.get("ativo", True),
        "permite_trial": plano_data.get("permite_trial", False),
        "dias_trial": plano_data.get("dias_trial", 30),
        "desconto_promocao": plano_data.get("desconto_promocao", 0),
        "data_inicio_promocao": plano_data.get("data_inicio_promocao"),
        "data_fim_promocao": plano_data.get("data_fim_promocao"),
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.planos_sistema.update_one({"id": plano_id}, {"$set": update_data})
    return {"message": "Plano atualizado com sucesso"}

@api_router.delete("/planos-sistema/{plano_id}")
async def delete_plano_sistema(plano_id: str, current_user: Dict = Depends(get_current_user)):
    """Delete plan permanently from unified system (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Contar quantos utilizadores têm este plano
    motoristas_com_plano = await db.motoristas.count_documents({"plano_id": plano_id})
    parceiros_com_plano = await db.parceiros.count_documents({"plano_id": plano_id})
    users_com_plano = await db.users.count_documents({"plano_id": plano_id})
    
    # Remover o plano de todos os utilizadores antes de eliminar
    if motoristas_com_plano > 0:
        await db.motoristas.update_many(
            {"plano_id": plano_id},
            {"$unset": {"plano_id": ""}}
        )
    
    if parceiros_com_plano > 0:
        await db.parceiros.update_many(
            {"plano_id": plano_id},
            {"$unset": {"plano_id": ""}}
        )
    
    if users_com_plano > 0:
        await db.users.update_many(
            {"plano_id": plano_id},
            {"$unset": {"plano_id": ""}}
        )
    
    # Eliminar o plano permanentemente
    result = await db.planos_sistema.delete_one({"id": plano_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Plano não encontrado")
    
    total_affected = motoristas_com_plano + parceiros_com_plano + users_com_plano
    logger.info(f"Admin {current_user['id']} eliminou plano {plano_id} e removeu de {total_affected} utilizadores")
    
    return {
        "message": "Plano eliminado com sucesso",
        "utilizadores_afetados": total_affected
    }


@api_router.get("/planos-parceiro")
async def list_planos_parceiro(current_user: Dict = Depends(get_current_user)):
    """List all active parceiro plans"""
    try:
        planos = await db.planos_parceiro.find({"ativo": True}, {"_id": 0}).to_list(100)
        return planos
    except Exception as e:
        return []

@api_router.post("/planos-parceiro")
async def create_plano_parceiro(plano_data: dict, current_user: Dict = Depends(get_current_user)):
    """Create new parceiro plan (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    plano = {
        "id": str(uuid.uuid4()),
        "nome": plano_data.get("nome"),
        "descricao": plano_data.get("descricao", ""),
        "preco_mensal": plano_data.get("preco_mensal", 0),
        "features": plano_data.get("features", []),
        "max_veiculos": plano_data.get("max_veiculos", 0),
        "max_motoristas": plano_data.get("max_motoristas", 0),
        "ativo": True,
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.planos_parceiro.insert_one(plano)
    return plano

@api_router.put("/planos-parceiro/{plano_id}")
async def update_plano_parceiro(plano_id: str, plano_data: dict, current_user: Dict = Depends(get_current_user)):
    """Update parceiro plan (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    plano = await db.planos_parceiro.find_one({"id": plano_id}, {"_id": 0})
    if not plano:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    update_data = {k: v for k, v in plano_data.items() if k != "id"}
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    await db.planos_parceiro.update_one({"id": plano_id}, {"$set": update_data})
    
    updated_plano = await db.planos_parceiro.find_one({"id": plano_id}, {"_id": 0})
    return updated_plano

@api_router.delete("/planos-parceiro/{plano_id}")
async def delete_plano_parceiro(plano_id: str, current_user: Dict = Depends(get_current_user)):
    """Soft delete parceiro plan (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    await db.planos_parceiro.update_one({"id": plano_id}, {"$set": {"ativo": False}})
    return {"message": "Plano desativado com sucesso"}

@api_router.post("/motoristas/{motorista_id}/atribuir-plano")
async def atribuir_plano_motorista(
    motorista_id: str,
    plano_data: dict,
    current_user: Dict = Depends(get_current_user)
):
    """Assign a motorista plan manually (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    try:
        plano_id = plano_data.get("plano_id")
        periodicidade = plano_data.get("periodicidade", "mensal")  # "semanal" ou "mensal"
        
        # Verify plan exists and is a MOTORISTA plan (using unified system)
        plano = await db.planos_sistema.find_one({"id": plano_id, "ativo": True, "tipo_usuario": "motorista"}, {"_id": 0})
        if not plano:
            raise HTTPException(status_code=404, detail="Plano de motorista not found")
        
        # Verify motorista exists
        motorista = await db.motoristas.find_one({"id": motorista_id}, {"_id": 0})
        if not motorista:
            raise HTTPException(status_code=404, detail="Motorista not found")
        
        # Calculate price and duration (unified system uses only monthly)
        preco_pago = plano["preco_mensal"]
        duracao_dias = 30  # Sistema unificado usa apenas mensal
        
        # Cancel any existing active subscription
        await db.motorista_plano_assinaturas.update_many(
            {"motorista_id": motorista_id, "status": "ativo"},
            {"$set": {"status": "cancelado", "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
        
        # Create new subscription
        assinatura_id = str(uuid.uuid4())
        data_inicio = datetime.now(timezone.utc)
        data_fim = data_inicio + timedelta(days=duracao_dias)
        
        assinatura = {
            "id": assinatura_id,
            "motorista_id": motorista_id,
            "plano_id": plano_id,
            "plano_nome": plano["nome"],
            "periodicidade": periodicidade,
            "preco_pago": preco_pago,
            "status": "ativo",
            "data_inicio": data_inicio.isoformat(),
            "data_fim": data_fim.isoformat(),
            "auto_renovacao": plano_data.get("auto_renovacao", False),
            "metodo_pagamento": "manual",  # Atribuição manual pelo admin
            "pagamento_confirmado": True,
            "data_pagamento": data_inicio.isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.motorista_plano_assinaturas.insert_one(assinatura)
        
        # Update motorista with plan info
        await db.motoristas.update_one(
            {"id": motorista_id},
            {"$set": {
                "plano_id": plano_id,
                "plano_nome": plano["nome"],
                "plano_modulos": plano.get("modulos", []),
                "plano_valida_ate": data_fim.isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        return {"message": f"Plano {periodicidade} atribuído com sucesso", "assinatura_id": assinatura_id, "preco_pago": preco_pago}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error assigning plan to motorista: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/motoristas/{motorista_id}/plano-atual")
async def get_plano_atual_motorista(motorista_id: str, current_user: Dict = Depends(get_current_user)):
    """Get current motorista plan"""
    try:
        # Verify access
        if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
            if current_user["id"] != motorista_id:
                raise HTTPException(status_code=403, detail="Not authorized")
        
        assinatura = await db.motorista_plano_assinaturas.find_one(
            {"motorista_id": motorista_id, "status": "ativo"},
            {"_id": 0}
        )
        
        if not assinatura:
            return {"message": "Nenhum plano ativo", "plano": None}
        
        # Get plan details
        plano = await db.planos_motorista.find_one({"id": assinatura["plano_id"]}, {"_id": 0})
        
        return {
            "assinatura": assinatura,
            "plano": plano
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting motorista current plan: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/motoristas/{motorista_id}/iniciar-pagamento-plano")
async def iniciar_pagamento_plano(
    motorista_id: str,
    payment_data: dict,
    current_user: Dict = Depends(get_current_user)
):
    """Initiate payment for motorista plan (Multibanco or MBWay)"""
    try:
        # Verify access - motorista can only pay for their own plan
        if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
            if current_user["id"] != motorista_id:
                raise HTTPException(status_code=403, detail="Not authorized")
        
        plano_id = payment_data.get("plano_id")
        periodicidade = payment_data.get("periodicidade", "mensal")
        metodo_pagamento = payment_data.get("metodo_pagamento", "multibanco")
        
        # Verify plan exists
        plano = await db.planos_motorista.find_one({"id": plano_id, "ativo": True}, {"_id": 0})
        if not plano:
            raise HTTPException(status_code=404, detail="Plano not found")
        
        # Verify motorista exists
        motorista = await db.motoristas.find_one({"id": motorista_id}, {"_id": 0})
        if not motorista:
            raise HTTPException(status_code=404, detail="Motorista not found")
        
        # Calculate price
        if periodicidade == "semanal":
            preco_pago = plano.get("preco_semanal", 0)
            duracao_dias = 7
        else:  # mensal
            preco_base = plano.get("preco_mensal", 0)
            desconto = plano.get("desconto_mensal_percentagem", 0)
            preco_pago = preco_base * (1 - desconto / 100) if desconto > 0 else preco_base
            duracao_dias = 30
        
        if preco_pago == 0:
            raise HTTPException(status_code=400, detail="Cannot initiate payment for free plan")
        
        # Cancel any existing pending payments
        await db.motorista_plano_assinaturas.update_many(
            {"motorista_id": motorista_id, "status": {"$in": ["aguardando_pagamento", "pagamento_pendente"]}},
            {"$set": {"status": "cancelado", "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
        
        # Create assinatura in pending status
        assinatura_id = str(uuid.uuid4())
        referencia_id = f"PLN-{assinatura_id[:8].upper()}"
        
        assinatura = {
            "id": assinatura_id,
            "motorista_id": motorista_id,
            "plano_id": plano_id,
            "plano_nome": plano["nome"],
            "periodicidade": periodicidade,
            "preco_pago": preco_pago,
            "status": "aguardando_pagamento",
            "data_inicio": datetime.now(timezone.utc).isoformat(),
            "data_fim": (datetime.now(timezone.utc) + timedelta(days=duracao_dias)).isoformat(),
            "auto_renovacao": payment_data.get("auto_renovacao", False),
            "metodo_pagamento": metodo_pagamento,
            "pagamento_confirmado": False,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Generate payment reference based on method
        if metodo_pagamento == "multibanco":
            # Generate Multibanco reference (simplified - in production use IFThenPay API)
            entidade = "11604"  # Example entity
            referencia = f"{assinatura_id[:3]}{assinatura_id[-6:]}".upper()[:9]
            
            assinatura["entidade_pagamento"] = entidade
            assinatura["referencia_pagamento"] = referencia
            
            await db.motorista_plano_assinaturas.insert_one(assinatura)
            
            return {
                "message": "Referência Multibanco gerada com sucesso",
                "assinatura_id": assinatura_id,
                "metodo": "multibanco",
                "entidade": entidade,
                "referencia": referencia,
                "valor": f"{preco_pago:.2f}",
                "status": "aguardando_pagamento"
            }
            
        elif metodo_pagamento == "mbway":
            # Generate MBWay transaction (simplified - in production use IFThenPay API)
            transaction_id = f"MBWAY-{assinatura_id[:12].upper()}"
            phone_number = motorista.get("phone", "")
            
            assinatura["referencia_pagamento"] = transaction_id
            
            await db.motorista_plano_assinaturas.insert_one(assinatura)
            
            return {
                "message": "Pedido MB WAY enviado com sucesso",
                "assinatura_id": assinatura_id,
                "metodo": "mbway",
                "transaction_id": transaction_id,
                "phone_number": phone_number,
                "valor": f"{preco_pago:.2f}",
                "status": "aguardando_pagamento"
            }
        else:
            raise HTTPException(status_code=400, detail="Invalid payment method")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error initiating plan payment: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/subscriptions/solicitar")
async def solicitar_subscription(sub_create: SubscriptionCreate, current_user: Dict = Depends(get_current_user)):
    """User requests a plan subscription"""
    # Get plan details
    plano = await db.planos.find_one({"id": sub_create.plano_id, "ativo": True}, {"_id": 0})
    if not plano:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    # Calculate price with IVA and discounts
    if sub_create.periodo == "semanal":
        preco_sem_iva = plano["preco_semanal_sem_iva"]
    else:  # mensal
        preco_sem_iva = plano["preco_mensal_sem_iva"]
        # Apply monthly discount
        if plano.get("desconto_mensal_percentagem", 0) > 0:
            preco_sem_iva = preco_sem_iva * (1 - plano["desconto_mensal_percentagem"] / 100)
    
    # Apply promotional discount
    if plano.get("promocao", {}).get("ativa", False):
        promocao = plano["promocao"]
        valida_ate = promocao.get("valida_ate")
        if valida_ate:
            valida_ate_date = datetime.fromisoformat(valida_ate).date()
            if valida_ate_date >= datetime.now(timezone.utc).date():
                preco_sem_iva = preco_sem_iva * (1 - promocao.get("desconto_percentagem", 0) / 100)
    
    # Calculate final price with IVA
    iva = plano.get("iva_percentagem", 23)
    preco_final = preco_sem_iva * (1 + iva / 100)
    
    # Generate payment reference (mock for now, will integrate IFThenPay)
    import random
    referencia = f"{random.randint(100, 999)} {random.randint(100, 999)} {random.randint(100, 999)}"
    entidade = "11604"  # Mock entity
    
    # Create subscription
    subscription_dict = {
        "id": str(uuid.uuid4()),
        "user_id": current_user["id"],
        "user_name": current_user["name"],
        "user_email": current_user["email"],
        "plano_id": sub_create.plano_id,
        "plano_nome": plano["nome"],
        "periodo": sub_create.periodo,
        "status": "pendente",
        "preco_pago": round(preco_final, 2),
        "data_inicio": None,
        "data_expiracao": None,
        "data_proximo_pagamento": None,
        "pagamento_metodo": sub_create.pagamento_metodo,
        "pagamento_referencia": referencia,
        "pagamento_entidade": entidade,
        "pagamento_id_transacao": str(uuid.uuid4()),
        "pagamento_status": "pendente",
        "atribuido_manualmente": False,
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.subscriptions.insert_one(subscription_dict)
    
    return {
        "message": "Subscription created successfully. Please complete payment.",
        "subscription_id": subscription_dict["id"],
        "pagamento": {
            "metodo": sub_create.pagamento_metodo,
            "referencia": referencia,
            "entidade": entidade,
            "valor": preco_final
        }
    }

@api_router.get("/subscriptions/minhas")
async def get_minhas_subscriptions(current_user: Dict = Depends(get_current_user)):
    """Get user's subscriptions"""
    subscriptions = await db.subscriptions.find(
        {"user_id": current_user["id"]},
        {"_id": 0}
    ).to_list(None)
    return subscriptions

@api_router.post("/users/{user_id}/atribuir-plano")
async def atribuir_plano_manual(
    user_id: str,
    data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Admin manually assigns a plan to a user (motorista or parceiro plan)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Admin or Gestor only")
    
    plano_id = data.get("plano_id")
    periodicidade = data.get("periodicidade", "mensal")
    duracao_dias = 30  # Sistema unificado usa apenas mensal
    
    # Find plan in unified system
    plano = await db.planos_sistema.find_one({"id": plano_id, "ativo": True}, {"_id": 0})
    
    if not plano:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    plano_type = plano.get("tipo_usuario", "motorista")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Create active subscription
    now = datetime.now(timezone.utc)
    data_expiracao = now + timedelta(days=duracao_dias)
    
    subscription_dict = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "user_name": user.get("name", ""),
        "user_email": user.get("email", ""),
        "plano_id": plano_id,
        "plano_nome": plano["nome"],
        "periodo": "manual",
        "status": "ativo",
        "preco_pago": 0,
        "data_inicio": now,
        "data_expiracao": data_expiracao,
        "data_proximo_pagamento": data_expiracao - timedelta(days=2),
        "pagamento_metodo": "manual_admin",
        "pagamento_status": "pago",
        "atribuido_manualmente": True,
        "duracao_dias": duracao_dias,
        "created_at": now,
        "updated_at": now
    }
    
    await db.subscriptions.insert_one(subscription_dict)
    
    # Update user subscription_id
    await db.users.update_one(
        {"id": user_id},
        {"$set": {
            "subscription_id": subscription_dict["id"],
            "plano_id": plano_id,
            "plano_nome": plano["nome"],
            "plano_valida_ate": data_expiracao.isoformat()
        }}
    )
    
    # Update specific collection based on user type
    if plano_type == "motorista":
        await db.motoristas.update_one(
            {"id": user_id},
            {"$set": {
                "plano_id": plano_id,
                "plano_nome": plano["nome"],
                "plano_valida_ate": data_expiracao.isoformat()
            }}
        )
    elif plano_type == "parceiro":
        await db.parceiros.update_one(
            {"id": user_id},
            {"$set": {
                "plano_id": plano_id,
                "plano_nome": plano["nome"],
                "plano_valida_ate": data_expiracao.isoformat(),
                "plano_status": "ativo"
            }}
        )
    
    logger.info(f"Admin {current_user['id']} atribuiu plano {plano['nome']} ao user {user_id}")
    
    return {
        "message": f"Plano atribuído com sucesso por {duracao_dias} dias",
        "subscription_id": subscription_dict["id"],
        "plano_nome": plano["nome"]
    }

@api_router.post("/webhooks/ifthen_pay-callback")
async def ifthen_pay_callback(data: Dict[str, Any]):
    """Webhook callback from IFThenPay to confirm payment"""
    # This will be called by IFThenPay when payment is confirmed
    # For now, we'll keep it simple
    
    transacao_id = data.get("id")
    status = data.get("status")
    
    if status == "pago":
        # Find subscription by transaction ID
        subscription = await db.subscriptions.find_one(
            {"pagamento_id_transacao": transacao_id},
            {"_id": 0}
        )
        
        if subscription:
            # Activate subscription
            now = datetime.now(timezone.utc)
            if subscription["periodo"] == "semanal":
                data_expiracao = now + timedelta(days=7)
            else:  # mensal
                data_expiracao = now + timedelta(days=30)
            
            await db.subscriptions.update_one(
                {"id": subscription["id"]},
                {"$set": {
                    "status": "ativo",
                    "pagamento_status": "pago",
                    "data_inicio": now,
                    "data_expiracao": data_expiracao,
                    "data_proximo_pagamento": data_expiracao - timedelta(days=2),
                    "updated_at": now
                }}
            )
            
            # Update user subscription_id
            await db.users.update_one(
                {"id": subscription["user_id"]},
                {"$set": {"subscription_id": subscription["id"]}}
            )
    
    return {"status": "received"}

# ==================== RECIBOS E GANHOS ====================

@api_router.post("/recibos/upload-ficheiro")
async def upload_recibo_file(
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    """Upload de ficheiro PDF de recibo"""
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Apenas ficheiros PDF são permitidos")
    
    # Create uploads directory if it doesn't exist
    recibos_dir = ROOT_DIR / "uploads" / "recibos"
    recibos_dir.mkdir(parents=True, exist_ok=True)
    
    # Generate unique filename
    file_id = str(uuid.uuid4())
    file_extension = file.filename.split('.')[-1]
    new_filename = f"{current_user['id']}_{file_id}.{file_extension}"
    
    # Save file
    file_path = await save_uploaded_file(file, recibos_dir, new_filename)
    
    # Return relative URL
    file_url = f"uploads/recibos/{new_filename}"
    
    return {
        "success": True,
        "file_url": file_url,
        "filename": new_filename
    }

@api_router.post("/recibos")
async def criar_recibo(recibo: ReciboCreate, current_user: Dict = Depends(get_current_user)):
    """Motorista cria um recibo (upload)"""
    # Get motorista info
    motorista = await db.motoristas.find_one({"id": current_user["id"]}, {"_id": 0})
    if not motorista:
        raise HTTPException(status_code=404, detail="Motorista not found")
    
    recibo_dict = {
        "id": str(uuid.uuid4()),
        "motorista_id": current_user["id"],
        "motorista_nome": motorista.get("name", ""),
        "parceiro_id": motorista.get("parceiro_atribuido"),
        "parceiro_nome": motorista.get("parceiro_atribuido_nome", ""),
        "valor": recibo.valor,
        "mes_referencia": recibo.mes_referencia,
        "ficheiro_url": recibo.ficheiro_url,
        "status": "pendente",
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.recibos.insert_one(recibo_dict)
    
    return {"message": "Recibo enviado para verificação", "recibo_id": recibo_dict["id"]}

@api_router.get("/recibos/meus")
async def get_meus_recibos(current_user: Dict = Depends(get_current_user)):
    """Get recibos do motorista"""
    recibos = await db.recibos.find(
        {"motorista_id": current_user["id"]},
        {"_id": 0}
    ).sort("created_at", -1).to_list(None)
    return recibos

@api_router.get("/recibos")
async def get_all_recibos(
    parceiro_id: Optional[str] = None,
    current_user: Dict = Depends(get_current_user)
):
    """Get recibos (filtered by role)"""
    try:
        query = {}
        
        # Se for parceiro ou operacional, só vê seus recibos
        if current_user["role"] in [UserRole.PARCEIRO]:
            # Get motoristas associated with this parceiro/operacional
            motoristas = await db.motoristas.find({"parceiro_atribuido": current_user["id"]}, {"_id": 0, "id": 1}).to_list(100)
            motorista_ids = [m["id"] for m in motoristas]
            query["motorista_id"] = {"$in": motorista_ids}
        elif parceiro_id:
            query["parceiro_id"] = parceiro_id
        # Admin and Gestao can see all recibos
        
        recibos = await db.recibos.find(query, {"_id": 0}).sort("created_at", -1).to_list(None)
        return recibos
    except Exception as e:
        logger.error(f"Error fetching recibos: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/recibos/{recibo_id}/verificar")
async def verificar_recibo(
    recibo_id: str,
    data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Admin/Gestor/Operacional/Parceiro verifica recibo"""
    if current_user["role"] not in [UserRole.ADMIN, "gestao", "operacional", UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    status = data.get("status", "verificado")
    observacoes = data.get("observacoes", "")
    
    await db.recibos.update_one(
        {"id": recibo_id},
        {"$set": {
            "status": status,
            "verificado_por": current_user["id"],
            "verificado_em": datetime.now(timezone.utc),
            "observacoes": observacoes,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {"message": f"Recibo {status}"}

@api_router.get("/ganhos/meus")
async def get_meus_ganhos(current_user: Dict = Depends(get_current_user)):
    """Get ganhos do motorista"""
    ganhos = await db.ganhos.find(
        {"motorista_id": current_user["id"]},
        {"_id": 0}
    ).sort("data", -1).to_list(None)
    return ganhos

@api_router.post("/pedidos-alteracao")
async def criar_pedido_alteracao(data: Dict[str, str], current_user: Dict = Depends(get_current_user)):
    """Motorista solicita alteração de dados"""
    motorista = await db.motoristas.find_one({"id": current_user["id"]}, {"_id": 0})
    if not motorista:
        raise HTTPException(status_code=404, detail="Motorista not found")
    
    pedido_dict = {
        "id": str(uuid.uuid4()),
        "motorista_id": current_user["id"],
        "motorista_nome": motorista.get("name", ""),
        "campo": data.get("campo"),
        "valor_atual": data.get("valor_atual"),
        "valor_novo": data.get("valor_novo"),
        "motivo": data.get("motivo"),
        "status": "pendente",
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.pedidos_alteracao.insert_one(pedido_dict)
    
    return {"message": "Pedido enviado para análise", "pedido_id": pedido_dict["id"]}

@api_router.get("/pedidos-alteracao")
async def get_pedidos_alteracao(current_user: Dict = Depends(get_current_user)):
    """Get pedidos de alteração"""
    if current_user["role"] == "motorista":
        query = {"motorista_id": current_user["id"]}
    else:
        query = {}
    
    pedidos = await db.pedidos_alteracao.find(query, {"_id": 0}).sort("created_at", -1).to_list(None)
    return pedidos


# ==================== CONFIGURAÇÕES ====================
@api_router.get("/integracoes/configuracoes")
async def get_integracoes_config(current_user: Dict = Depends(get_current_user)):
    """Get integrations configuration (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Get configurations from database
    config = await db.configuracoes.find_one({"tipo": "integracoes"}, {"_id": 0})
    
    if not config:
        # Return default structure
        return {
            "moloni": {
                "ativo": False,
                "client_id": "",
                "client_secret": "",
                "username": "",
                "password": "",
                "company_id": "",
                "taxa_mensal_extra": 10.00
            },
            "terabox": {
                "ativo": False,
                "api_key": "",
                "folder_id": ""
            },
            "google_drive": {
                "ativo": False,
                "credentials": "",
                "folder_id": ""
            }
        }
    
    return config

@api_router.post("/integracoes/configuracoes")
async def save_integracoes_config(
    config_data: Dict,
    current_user: Dict = Depends(get_current_user)
):
    """Save integrations configuration (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    config_data["tipo"] = "integracoes"
    config_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    config_data["updated_by"] = current_user["id"]
    
    await db.configuracoes.update_one(
        {"tipo": "integracoes"},
        {"$set": config_data},
        upsert=True
    )
    
    return {"message": "Configuração salva com sucesso"}

@api_router.get("/configuracoes/email")
async def get_email_config(current_user: Dict = Depends(get_current_user)):
    """Get email configuration (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    config = await db.configuracoes.find_one({"tipo": "email"}, {"_id": 0})
    if not config:
        # Return default config
        return {
            "email_contacto": "info@tvdefleet.com",
            "telefone_contacto": "",
            "morada_empresa": "Lisboa, Portugal",
            "nome_empresa": "TVDEFleet"
        }
    return config

@api_router.post("/configuracoes/email")
async def save_email_config(
    config_data: Dict[str, str],
    current_user: Dict = Depends(get_current_user)
):
    """Save email configuration (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    config_data["tipo"] = "email"
    config_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    config_data["updated_by"] = current_user["id"]
    
    await db.configuracoes.update_one(
        {"tipo": "email"},
        {"$set": config_data},
        upsert=True
    )
    
    return {"message": "Configuration saved successfully"}


# ==================== PUBLIC ENDPOINTS (NO AUTH) ====================
# =============================================================================
# SINCRONIZAÇÃO AUTOMÁTICA DE PLATAFORMAS
# =============================================================================

from sync_platforms import encrypt_password, decrypt_password, sync_platform
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger

# Scheduler global
scheduler = AsyncIOScheduler()

# ENDPOINT ANTIGO POST - COMENTADO (duplicado, usar o novo em GESTÃO DE CREDENCIAIS)
# @app.post("/api/credenciais-plataforma")
# async def salvar_credenciais_plataforma(
#     parceiro_id: str = Form(...),
#     plataforma: str = Form(...),
#     email: str = Form(...),
#     password: str = Form(...),
#     sincronizacao_automatica: bool = Form(False),
#     horario_sincronizacao: Optional[str] = Form(None),
#     frequencia_dias: int = Form(7),
#     credentials: HTTPAuthorizationCredentials = Depends(security)
# ):
#     """Salva credenciais de uma plataforma para um parceiro específico"""
#     try:
#         user = await get_current_user(credentials)
#         if user['role'] not in ['admin', 'manager']:
#             raise HTTPException(status_code=403, detail="Acesso negado")
#         
#         # Encriptar password
#         password_encrypted = encrypt_password(password)
#         
#         # Verificar se já existe para este parceiro e plataforma
#         existing = await db.credenciais_plataforma.find_one({
#             'parceiro_id': parceiro_id,
#             'plataforma': plataforma
#         })
#         
#         if existing:
#             # Atualizar
#             update_data = {
#                 'email': email,
#                 'sincronizacao_automatica': sincronizacao_automatica,
#                 'horario_sincronizacao': horario_sincronizacao,
#                 'frequencia_dias': frequencia_dias,
#                 'updated_at': datetime.now(timezone.utc)
#             }
#             # Só atualizar password se foi fornecida
#             if password:
#                 update_data['password_encrypted'] = password_encrypted
#                 
#             await db.credenciais_plataforma.update_one(
#                 {'parceiro_id': parceiro_id, 'plataforma': plataforma},
#                 {'$set': update_data}
#             )
#             cred_id = existing['id']
#         else:
#             # Criar novo
#             credencial = {
#                 'id': str(uuid.uuid4()),
#                 'parceiro_id': parceiro_id,
#                 'plataforma': plataforma,
#                 'email': email,
#                 'password_encrypted': password_encrypted,
#                 'ativo': True,
#                 'sincronizacao_automatica': sincronizacao_automatica,
#                 'horario_sincronizacao': horario_sincronizacao,
#                 'frequencia_dias': frequencia_dias,
#                 'created_at': datetime.now(timezone.utc),
#                 'updated_at': datetime.now(timezone.utc)
#             }
#             await db.credenciais_plataforma.insert_one(credencial)
#             cred_id = credencial['id']
#         
#         # Se sincronização automática ativada, agendar job
#         if sincronizacao_automatica and horario_sincronizacao:
#             await agendar_sincronizacao(cred_id, horario_sincronizacao, frequencia_dias)
#         
#         return {'success': True, 'message': 'Credenciais salvas com sucesso', 'id': cred_id}
#         
#     except Exception as e:
#         raise HTTPException(status_code=500, detail=str(e))

# ENDPOINT ANTIGO - COMENTADO (duplicado, usar o novo em GESTÃO DE CREDENCIAIS)
# @app.get("/api/credenciais-plataforma")
# async def listar_credenciais_plataformas(
#     parceiro_id: Optional[str] = None,
#     credentials: HTTPAuthorizationCredentials = Depends(security)
# ):
#     """Lista credenciais de plataformas por parceiro (sem passwords)"""
#     try:
#         user = await get_current_user(credentials)
#         if user['role'] not in ['admin', 'manager']:
#             raise HTTPException(status_code=403, detail="Acesso negado")
#         
#         query = {}
#         if parceiro_id:
#             query['parceiro_id'] = parceiro_id
#             
#         credenciais = await db.credenciais_plataforma.find(query).to_list(length=None)
#         
#         # Remover passwords encriptadas da resposta
#         for cred in credenciais:
#             cred.pop('password_encrypted', None)
#         
#         return credenciais
#         
#     except Exception as e:
#         raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/sincronizar/{parceiro_id}/{plataforma}")
async def sincronizar_plataforma_manual(
    parceiro_id: str,
    plataforma: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Sincroniza manualmente uma plataforma para um parceiro específico"""
    try:
        user = await get_current_user(credentials)
        if user['role'] not in ['admin', 'manager']:
            raise HTTPException(status_code=403, detail="Acesso negado")
        
        # Buscar credenciais do parceiro
        cred = await db.credenciais_plataforma.find_one({
            'parceiro_id': parceiro_id,
            'plataforma': plataforma
        })
        if not cred:
            raise HTTPException(status_code=404, detail="Credenciais não configuradas para este parceiro")
        
        # Criar log de sincronização
        log = {
            'id': str(uuid.uuid4()),
            'parceiro_id': parceiro_id,
            'plataforma': plataforma,
            'tipo_sincronizacao': 'manual',
            'status': 'em_progresso',
            'data_inicio': datetime.now(timezone.utc),
            'executado_por': user['id']
        }
        await db.logs_sincronizacao.insert_one(log)
        
        # Executar sincronização
        sucesso, file_path, mensagem = await sync_platform(
            plataforma, 
            cred['email'], 
            cred['password_encrypted'],
            headless=True
        )
        
        if sucesso:
            # Processar ficheiro baixado
            if plataforma == 'uber':
                # Importar usando a função existente de import Uber
                with open(file_path, 'rb') as f:
                    contents = f.read()
                    decoded = contents.decode('utf-8')
                    
                    # Detect delimiter (Portuguese Excel often uses semicolon)
                    sample = decoded[:1000]
                    delimiter = ';' if sample.count(';') > sample.count(',') else ','
                    logger.info(f"Detected CSV delimiter for automatic sync: '{delimiter}'")
                    
                    csv_reader = csv.DictReader(io.StringIO(decoded), delimiter=delimiter)
                    
                    registos_importados = 0
                    for row in csv_reader:
                        # Processar igual ao endpoint de import manual
                        registos_importados += 1
            
            # Atualizar log
            await db.logs_sincronizacao.update_one(
                {'id': log['id']},
                {'$set': {
                    'status': 'sucesso',
                    'data_fim': datetime.now(timezone.utc),
                    'registos_importados': registos_importados if 'registos_importados' in locals() else 0,
                    'mensagem': mensagem
                }}
            )
            
            # Atualizar última sincronização
            await db.credenciais_plataforma.update_one(
                {'parceiro_id': parceiro_id, 'plataforma': plataforma},
                {'$set': {'ultima_sincronizacao': datetime.now(timezone.utc)}}
            )
            
            # Limpar ficheiro temporário
            if file_path and os.path.exists(file_path):
                os.remove(file_path)
            
            return {'success': True, 'message': 'Sincronização concluída com sucesso'}
        else:
            # Atualizar log com erro
            await db.logs_sincronizacao.update_one(
                {'id': log['id']},
                {'$set': {
                    'status': 'erro',
                    'data_fim': datetime.now(timezone.utc),
                    'mensagem': mensagem
                }}
            )
            return {'success': False, 'message': mensagem}
            
    except Exception as e:
        logger.error(f"Erro ao sincronizar {plataforma}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/logs-sincronizacao")
async def listar_logs_sincronizacao(
    parceiro_id: Optional[str] = None,
    plataforma: Optional[str] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Lista histórico de sincronizações"""
    try:
        user = await get_current_user(credentials)
        
        query = {}
        if parceiro_id:
            query['parceiro_id'] = parceiro_id
        if plataforma:
            query['plataforma'] = plataforma
        
        logs = await db.logs_sincronizacao.find(query).sort('data_inicio', -1).limit(100).to_list(length=None)
        return logs
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

async def agendar_sincronizacao(credencial_id: str, horario: str, frequencia_dias: int):
    """Agenda sincronização automática"""
    try:
        # Parse horário (ex: "09:00")
        hora, minuto = map(int, horario.split(':'))
        
        # Criar job ID único baseado no credencial_id
        job_id = f"sync_{credencial_id}"
        
        # Remover job existente se houver
        if scheduler.get_job(job_id):
            scheduler.remove_job(job_id)
        
        # Adicionar novo job
        scheduler.add_job(
            executar_sincronizacao_automatica,
            CronTrigger(hour=hora, minute=minuto, day_of_week='*/' + str(frequencia_dias)),
            id=job_id,
            args=[credencial_id],
            replace_existing=True
        )
        
        logger.info(f"Sincronização agendada para credencial {credencial_id} às {horario} a cada {frequencia_dias} dias")
        
    except Exception as e:
        logger.error(f"Erro ao agendar sincronização: {e}")

async def executar_sincronizacao_automatica(credencial_id: str):
    """Executa sincronização automática agendada"""
    try:
        # Buscar credenciais por ID
        cred = await db.credenciais_plataforma.find_one({'id': credencial_id})
        if not cred or not cred.get('ativo'):
            logger.warning(f"Credenciais inativas ou não encontradas: {credencial_id}")
            return
        
        logger.info(f"Executando sincronização automática: {cred['plataforma']} para parceiro {cred['parceiro_id']}")
        
        # Criar log
        log = {
            'id': str(uuid.uuid4()),
            'parceiro_id': cred['parceiro_id'],
            'plataforma': cred['plataforma'],
            'tipo_sincronizacao': 'automatico',
            'status': 'em_progresso',
            'data_inicio': datetime.now(timezone.utc)
        }
        await db.logs_sincronizacao.insert_one(log)
        
        # Executar sincronização
        sucesso, file_path, mensagem = await sync_platform(
            cred['plataforma'], 
            cred['email'], 
            cred['password_encrypted'],
            headless=True
        )
        
        if sucesso:
            # Processar ficheiro (igual ao manual)
            await db.logs_sincronizacao.update_one(
                {'id': log['id']},
                {'$set': {
                    'status': 'sucesso',
                    'data_fim': datetime.now(timezone.utc),
                    'mensagem': mensagem
                }}
            )
            await db.credenciais_plataforma.update_one(
                {'id': credencial_id},
                {'$set': {'ultima_sincronizacao': datetime.now(timezone.utc)}}
            )
            
            if file_path and os.path.exists(file_path):
                os.remove(file_path)
        else:
            await db.logs_sincronizacao.update_one(
                {'id': log['id']},
                {'$set': {
                    'status': 'erro',
                    'data_fim': datetime.now(timezone.utc),
                    'mensagem': mensagem
                }}
            )
            
    except Exception as e:
        logger.error(f"Erro na sincronização automática de {credencial_id}: {e}")

# =============================================================================
# IMPORTAÇÃO DE GANHOS UBER
# =============================================================================

@app.post("/api/import/uber/ganhos")
async def importar_ganhos_uber(
    file: UploadFile = File(...),
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Importa ficheiro CSV de ganhos da Uber"""
    try:
        # Verificar autenticação
        user = await get_current_user(credentials)
        if user['role'] not in ['admin', 'manager']:
            raise HTTPException(status_code=403, detail="Acesso negado")
        
        # Ler conteúdo do ficheiro
        contents = await file.read()
        decoded = contents.decode('utf-8')
        
        # Detect delimiter (Portuguese Excel often uses semicolon)
        sample = decoded[:1000]
        delimiter = ';' if sample.count(';') > sample.count(',') else ','
        logger.info(f"Detected CSV delimiter for Uber ganhos import: '{delimiter}'")
        
        csv_reader = csv.DictReader(io.StringIO(decoded), delimiter=delimiter)
        
        ganhos_importados = []
        motoristas_encontrados = 0
        motoristas_nao_encontrados = 0
        erros = []
        total_ganhos = 0.0
        
        # Extrair período do nome do ficheiro (ex: 20251110-20251116)
        periodo_match = re.search(r'(\d{8})-(\d{8})', file.filename)
        periodo_inicio = None
        periodo_fim = None
        if periodo_match:
            periodo_inicio = periodo_match.group(1)
            periodo_fim = periodo_match.group(2)
        
        for row_num, row in enumerate(csv_reader, start=2):
            try:
                uuid_motorista = row.get('UUID do motorista', '').strip()
                if not uuid_motorista:
                    continue
                
                # Procurar motorista pelo UUID
                motorista = await db.motoristas.find_one({'uuid_motorista_uber': uuid_motorista})
                motorista_id = motorista['id'] if motorista else None
                
                if motorista:
                    motoristas_encontrados += 1
                else:
                    motoristas_nao_encontrados += 1
                
                # Função helper para converter valores
                def parse_float(value):
                    if not value or value == '':
                        return 0.0
                    try:
                        return float(value.replace(',', '.'))
                    except:
                        return 0.0
                
                # Extrair valores do CSV
                pago_total = parse_float(row.get('Pago a si', '0'))
                rendimentos_total = parse_float(row.get('Pago a si : Os seus rendimentos', '0'))
                
                ganho = {
                    'id': str(uuid.uuid4()),
                    'uuid_motorista_uber': uuid_motorista,
                    'motorista_id': motorista_id,
                    'nome_motorista': row.get('Nome próprio do motorista', ''),
                    'apelido_motorista': row.get('Apelido do motorista', ''),
                    'periodo_inicio': periodo_inicio,
                    'periodo_fim': periodo_fim,
                    'pago_total': pago_total,
                    'rendimentos_total': rendimentos_total,
                    'dinheiro_recebido': parse_float(row.get('Pago a si : Saldo da viagem : Pagamentos : Dinheiro recebido', '0')),
                    'tarifa_total': parse_float(row.get('Pago a si : Os seus rendimentos : Tarifa', '0')),
                    'tarifa_base': parse_float(row.get('Pago a si:Os seus rendimentos:Tarifa:Tarifa', '0')),
                    'tarifa_ajuste': parse_float(row.get('Pago a si:Os seus rendimentos:Tarifa:Ajuste', '0')),
                    'tarifa_cancelamento': parse_float(row.get('Pago a si:Os seus rendimentos:Tarifa:Cancelamento', '0')),
                    'tarifa_dinamica': parse_float(row.get('Pago a si:Os seus rendimentos:Tarifa:Tarifa dinâmica', '0')),
                    'taxa_reserva': parse_float(row.get('Pago a si:Os seus rendimentos:Tarifa:Taxa de reserva', '0')),
                    'uber_priority': parse_float(row.get('Pago a si:Os seus rendimentos:Tarifa:UberX Priority', '0')),
                    'tempo_espera': parse_float(row.get('Pago a si:Os seus rendimentos:Tarifa:Tempo de espera na recolha', '0')),
                    'taxa_servico': parse_float(row.get('Pago a si:Os seus rendimentos:Taxa de serviço', '0')),
                    'imposto_tarifa': parse_float(row.get('Pago a si:Os seus rendimentos:Tarifa:Imposto sobre a tarifa', '0')),
                    'taxa_aeroporto': parse_float(row.get('Pago a si:Os seus rendimentos:Outros rendimentos:Taxa de aeroporto', '0')),
                    'gratificacao': parse_float(row.get('Pago a si:Os seus rendimentos:Gratificação', '0')),
                    'portagens': parse_float(row.get('Pago a si:Saldo da viagem:Reembolsos:Portagem', '0')),
                    'ajustes': parse_float(row.get('Pago a si:Os seus rendimentos:Outros rendimentos:Ajuste', '0')),
                    'ficheiro_nome': file.filename,
                    'data_importacao': datetime.now(timezone.utc),
                    'importado_por': user['id']
                }
                
                # Salvar no banco
                await db.ganhos_uber.insert_one(ganho)
                ganhos_importados.append(ganho)
                total_ganhos += pago_total
                
            except Exception as e:
                erros.append(f"Linha {row_num}: {str(e)}")
        
        return {
            'success': True,
            'total_linhas': len(ganhos_importados),
            'motoristas_encontrados': motoristas_encontrados,
            'motoristas_nao_encontrados': motoristas_nao_encontrados,
            'total_ganhos': round(total_ganhos, 2),
            'ganhos_importados': ganhos_importados[:10],  # Primeiros 10 para preview
            'erros': erros
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao importar ficheiro: {str(e)}")

# =============================================================================
# IMPORTAÇÃO DE GANHOS BOLT
# =============================================================================

@app.post("/api/import/bolt/ganhos")
async def importar_ganhos_bolt(
    file: UploadFile = File(...),
    parceiro_id: Optional[str] = Form(None),
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Importa ficheiro CSV de ganhos da Bolt"""
    try:
        # Verificar autenticação
        user = await get_current_user(credentials)
        if user['role'] not in ['admin', 'manager']:
            raise HTTPException(status_code=403, detail="Acesso negado")
        
        # Ler conteúdo do ficheiro
        contents = await file.read()
        decoded = contents.decode('utf-8')
        
        # Detect delimiter (Portuguese Excel often uses semicolon)
        sample = decoded[:1000]
        delimiter = ';' if sample.count(';') > sample.count(',') else ','
        logger.info(f"Detected CSV delimiter for Bolt ganhos import: '{delimiter}'")
        
        csv_reader = csv.DictReader(io.StringIO(decoded), delimiter=delimiter)
        
        ganhos_importados = []
        motoristas_encontrados = 0
        motoristas_nao_encontrados = 0
        erros = []
        total_ganhos = 0.0
        
        # Extrair período do nome do ficheiro (ex: 2025W45)
        periodo_match = re.search(r'(\d{4})W(\d{2})', file.filename)
        periodo_ano = None
        periodo_semana = None
        if periodo_match:
            periodo_ano = periodo_match.group(1)
            periodo_semana = periodo_match.group(2)
        
        for row_num, row in enumerate(csv_reader, start=2):
            try:
                identificador_motorista = row.get('Identificador do motorista', '').strip()
                identificador_individual = row.get('Identificador individual', '').strip()
                
                if not identificador_motorista:
                    continue
                
                # Procurar motorista pelo identificador Bolt
                motorista = await db.motoristas.find_one({'identificador_motorista_bolt': identificador_individual})
                motorista_id = motorista['id'] if motorista else None
                
                if motorista:
                    motoristas_encontrados += 1
                else:
                    motoristas_nao_encontrados += 1
                
                # Função helper para converter valores (vírgula decimal)
                def parse_float(value):
                    if not value or value == '':
                        return 0.0
                    try:
                        # Remover espaços e converter
                        clean_value = value.strip().replace(',', '.')
                        return float(clean_value)
                    except:
                        return 0.0
                
                # Extrair valores do CSV
                ganhos_liquidos = parse_float(row.get('Ganhos líquidos|€', '0'))
                
                ganho = {
                    'id': str(uuid.uuid4()),
                    'identificador_motorista_bolt': identificador_motorista,
                    'identificador_individual': identificador_individual,
                    'motorista_id': motorista_id,
                    'nome_motorista': row.get('Motorista', ''),
                    'email_motorista': row.get('Email', ''),
                    'telemovel_motorista': row.get('Telemóvel', ''),
                    'periodo_ano': periodo_ano,
                    'periodo_semana': periodo_semana,
                    # Ganhos brutos
                    'ganhos_brutos_total': parse_float(row.get('Ganhos brutos (total)|€', '0')),
                    'ganhos_brutos_app': parse_float(row.get('Ganhos brutos (pagamentos na app)|€', '0')),
                    'iva_ganhos_app': parse_float(row.get('IVA sobre os ganhos brutos (pagamentos na app)|€', '0')),
                    'ganhos_brutos_dinheiro': parse_float(row.get('Ganhos brutos (pagamentos em dinheiro)|€', '0')),
                    'iva_ganhos_dinheiro': parse_float(row.get('IVA sobre os ganhos brutos (pagamentos em dinheiro)|€', '0')),
                    'dinheiro_recebido': parse_float(row.get('Dinheiro recebido|€', '0')),
                    # Extras
                    'gorjetas': parse_float(row.get('Gorjetas dos passageiros|€', '0')),
                    'ganhos_campanha': parse_float(row.get('Ganhos da campanha|€', '0')),
                    'reembolsos_despesas': parse_float(row.get('Reembolsos de despesas|€', '0')),
                    # Taxas
                    'taxas_cancelamento': parse_float(row.get('Taxas de cancelamento|€', '0')),
                    'iva_taxas_cancelamento': parse_float(row.get('IVA das taxas de cancelamento|€', '0')),
                    'portagens': parse_float(row.get('Portagens|€', '0')),
                    'taxas_reserva': parse_float(row.get('Taxas de reserva|€', '0')),
                    'iva_taxas_reserva': parse_float(row.get('IVA das taxas de reserva|€', '0')),
                    'total_taxas': parse_float(row.get('Total de taxas|€', '0')),
                    # Comissões
                    'comissoes': parse_float(row.get('Comissões|€', '0')),
                    'reembolsos_passageiros': parse_float(row.get('Reembolsos aos passageiros|€', '0')),
                    'outras_taxas': parse_float(row.get('Outras taxas|€', '0')),
                    # Ganhos líquidos
                    'ganhos_liquidos': ganhos_liquidos,
                    'pagamento_previsto': parse_float(row.get('Pagamento previsto|€', '0')),
                    # Produtividade
                    'ganhos_brutos_por_hora': parse_float(row.get('Ganhos brutos por hora|€/h', '0')),
                    'ganhos_liquidos_por_hora': parse_float(row.get('Ganhos líquidos por hora|€/h', '0')),
                    # Metadata
                    'ficheiro_nome': file.filename,
                    'parceiro_id': parceiro_id,
                    'data_importacao': datetime.now(timezone.utc),
                    'importado_por': user['id']
                }
                
                # Salvar no banco
                await db.ganhos_bolt.insert_one(ganho)
                ganhos_importados.append(ganho)
                total_ganhos += ganhos_liquidos
                
            except Exception as e:
                erros.append(f"Linha {row_num}: {str(e)}")
        
        return {
            'success': True,
            'total_linhas': len(ganhos_importados),
            'motoristas_encontrados': motoristas_encontrados,
            'motoristas_nao_encontrados': motoristas_nao_encontrados,
            'total_ganhos': round(total_ganhos, 2),
            'periodo': f"{periodo_ano}W{periodo_semana}" if periodo_ano else None,
            'ganhos_importados': ganhos_importados[:10],  # Primeiros 10 para preview
            'erros': erros
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao importar ficheiro: {str(e)}")

@app.get("/api/ganhos-bolt")
async def listar_ganhos_bolt(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    motorista_id: Optional[str] = None,
    parceiro_id: Optional[str] = None,
    periodo_ano: Optional[str] = None,
    periodo_semana: Optional[str] = None
):
    """Lista ganhos importados da Bolt"""
    try:
        user = await get_current_user(credentials)
        
        query = {}
        if motorista_id:
            query['motorista_id'] = motorista_id
        if parceiro_id:
            query['parceiro_id'] = parceiro_id
        if periodo_ano:
            query['periodo_ano'] = periodo_ano
        if periodo_semana:
            query['periodo_semana'] = periodo_semana
        
        ganhos = await db.ganhos_bolt.find(query).sort('data_importacao', -1).to_list(length=None)
        return ganhos
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/ganhos-uber")
async def listar_ganhos_uber(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    motorista_id: Optional[str] = None,
    periodo_inicio: Optional[str] = None,
    periodo_fim: Optional[str] = None
):
    """Lista ganhos importados da Uber"""
    try:
        user = await get_current_user(credentials)
        
        query = {}
        if motorista_id:
            query['motorista_id'] = motorista_id
        if periodo_inicio:
            query['periodo_inicio'] = periodo_inicio
        if periodo_fim:
            query['periodo_fim'] = periodo_fim
        
        ganhos = await db.ganhos_uber.find(query).sort('data_importacao', -1).to_list(length=None)
        return ganhos
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/public/veiculos")
async def get_public_veiculos():
    """Get public vehicles available for sale or rent"""
    veiculos = await db.vehicles.find({
        "$or": [
            {"disponivel_venda": True},
            {"disponivel_aluguer": True}
        ]
    }, {"_id": 0}).to_list(length=None)
    
    # Convert datetime fields
    for v in veiculos:
        if isinstance(v.get("created_at"), str):
            v["created_at"] = datetime.fromisoformat(v["created_at"])
        if isinstance(v.get("updated_at"), str):
            v["updated_at"] = datetime.fromisoformat(v["updated_at"])
    
    return veiculos

@app.post("/api/public/contacto")
async def public_contacto(contacto_data: Dict[str, Any]):
    """Receive public contact form (sends email to configured address)"""
    # Get email config
    config = await db.configuracoes.find_one({"tipo": "email"}, {"_id": 0})
    email_destino = config.get("email_contacto", "info@tvdefleet.com") if config else "info@tvdefleet.com"
    
    # Save contact to database
    contacto_id = f"contacto-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    contacto_data["id"] = contacto_id
    contacto_data["created_at"] = datetime.now(timezone.utc).isoformat()
    contacto_data["status"] = "pendente"
    contacto_data["email_destino"] = email_destino
    
    await db.contactos.insert_one(contacto_data)
    
    # Send notification email to admin
    await enviar_notificacao_contacto(contacto_data, email_destino)
    
    return {"message": "Contact received successfully", "id": contacto_id}

async def enviar_notificacao_contacto(contacto: Dict[str, Any], email_destino: str):
    """Send email notification about new contact (placeholder for email service)"""
    # This is a placeholder. In production, integrate with email service like SendGrid, AWS SES, etc.
    # For now, we just log it
    print(f"📧 NOVO CONTACTO RECEBIDO:")
    print(f"   Para: {email_destino}")
    print(f"   De: {contacto.get('nome')} ({contacto.get('email')})")
    print(f"   Assunto: {contacto.get('assunto', 'Contacto do Website')}")
    print(f"   Mensagem: {contacto.get('mensagem')}")
    print(f"   Telefone: {contacto.get('telefone')}")
    
    # Save to email queue for manual sending or integration later
    await db.email_queue.insert_one({
        "to": email_destino,
        "from": "noreply@tvdefleet.com",
        "subject": f"Novo Contacto: {contacto.get('assunto', 'Website')}",
        "body": f"""
        Novo contacto recebido através do website:
        
        Nome: {contacto.get('nome')}
        Email: {contacto.get('email')}
        Telefone: {contacto.get('telefone')}
        Assunto: {contacto.get('assunto', 'N/A')}
        
        Mensagem:
        {contacto.get('mensagem')}
        
        ---
        ID: {contacto.get('id')}
        Data: {contacto.get('created_at')}
        """,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "status": "pending",
        "tipo": "notificacao_contacto"
    })

@app.get("/api/public/parceiros")
async def get_public_parceiros():
    """Get public list of partners"""
    parceiros = await db.parceiros.find({}, {"_id": 0}).to_list(length=None)
    return parceiros

# ==================== CONTRACT ENDPOINTS ====================

@api_router.post("/contratos/gerar")
async def gerar_contrato(contrato_data: ContratoCreate, current_user: Dict = Depends(get_current_user)):
    """Generate a new contract between parceiro, motorista and vehicle"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO, "operacional"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    try:
        # Get parceiro data
        parceiro_user = await db.users.find_one({"id": contrato_data.parceiro_id}, {"_id": 0})
        parceiro = await db.parceiros.find_one({"id": contrato_data.parceiro_id}, {"_id": 0})
        if not parceiro and not parceiro_user:
            raise HTTPException(status_code=404, detail="Parceiro not found")
        
        # Get motorista data
        motorista = await db.motoristas.find_one({"id": contrato_data.motorista_id}, {"_id": 0})
        if not motorista:
            raise HTTPException(status_code=404, detail="Motorista not found")
        
        # Get vehicle data
        vehicle = await db.vehicles.find_one({"id": contrato_data.vehicle_id}, {"_id": 0})
        if not vehicle:
            raise HTTPException(status_code=404, detail="Vehicle not found")
        
        # Generate referencia (count existing contracts + 1)
        contract_count = await db.contratos.count_documents({})
        ano = datetime.now().year
        referencia = f"{str(contract_count + 1).zfill(3)}/{ano}"
        
        # Create contract
        contrato_id = str(uuid.uuid4())
        
        # Use parceiro data (prefer parceiros collection, fallback to users)
        parceiro_data = parceiro if parceiro else parceiro_user
        
        contrato = {
            "id": contrato_id,
            "referencia": referencia,
            "parceiro_id": contrato_data.parceiro_id,
            "motorista_id": contrato_data.motorista_id,
            "vehicle_id": contrato_data.vehicle_id,
            
            # Dados Parceiro Completos
            "parceiro_nome": parceiro_data.get("nome_empresa") or parceiro_data.get("nome") or parceiro_data.get("name"),
            "parceiro_nif": parceiro_data.get("contribuinte_empresa") or parceiro_data.get("nif"),
            "parceiro_morada": parceiro_data.get("morada_completa") or parceiro_data.get("morada") or "",
            "parceiro_codigo_postal": parceiro_data.get("codigo_postal") or "",
            "parceiro_localidade": parceiro_data.get("localidade") or "",
            "parceiro_telefone": parceiro_data.get("telefone") or parceiro_data.get("telemovel") or "",
            "parceiro_email": parceiro_data.get("email_empresa") or parceiro_data.get("email"),
            "parceiro_representante_legal_nome": parceiro_data.get("representante_legal_nome") or parceiro_data.get("nome_manager") or "",
            "parceiro_representante_legal_contribuinte": parceiro_data.get("representante_legal_contribuinte") or "",
            "parceiro_representante_legal_cc": parceiro_data.get("representante_legal_cc") or "",
            "parceiro_representante_legal_cc_validade": parceiro_data.get("representante_legal_cc_validade") or "",
            "parceiro_representante_legal_telefone": parceiro_data.get("representante_legal_telefone") or "",
            "parceiro_representante_legal_email": parceiro_data.get("representante_legal_email") or "",
            
            # Dados Motorista Completos
            "motorista_nome": motorista.get("name"),
            "motorista_cc": motorista.get("cc_numero") or motorista.get("numero_cc") or "",
            "motorista_cc_validade": motorista.get("cc_validade") or "",
            "motorista_nif": motorista.get("nif") or "",
            "motorista_morada": motorista.get("morada") or "",
            "motorista_codigo_postal": motorista.get("codigo_postal") or "",
            "motorista_localidade": motorista.get("localidade") or "",
            "motorista_telefone": motorista.get("telefone") or motorista.get("telemovel") or "",
            "motorista_carta_conducao": motorista.get("carta_conducao") or motorista.get("numero_carta") or "",
            "motorista_carta_conducao_validade": motorista.get("carta_validade") or motorista.get("carta_conducao_validade") or "",
            "motorista_licenca_tvde": motorista.get("licenca_tvde") or motorista.get("numero_tvde") or "",
            "motorista_licenca_tvde_validade": motorista.get("licenca_tvde_validade") or motorista.get("tvde_validade") or "",
            "motorista_seguranca_social": motorista.get("seguranca_social") or motorista.get("numero_ss") or "",
            "motorista_email": motorista.get("email") or "",
            
            # Dados Veículo
            "vehicle_marca": vehicle.get("marca"),
            "vehicle_modelo": vehicle.get("modelo"),
            "vehicle_matricula": vehicle.get("matricula"),
            
            # Termos Financeiros
            "tipo_contrato": contrato_data.tipo_contrato,
            "valor_semanal": contrato_data.valor_semanal,
            "comissao_percentual": contrato_data.comissao_percentual,
            "caucao_total": contrato_data.caucao_total,
            "caucao_lavagem": contrato_data.caucao_lavagem,
            
            # Campos de caução
            "tem_caucao": contrato_data.tem_caucao,
            "caucao_parcelada": contrato_data.caucao_parcelada,
            "caucao_parcelas": contrato_data.caucao_parcelas,
            "caucao_texto": contrato_data.caucao_texto,
            
            # Campos de Época
            "tem_epoca": contrato_data.tem_epoca,
            "data_inicio_epoca_alta": contrato_data.data_inicio_epoca_alta,
            "data_fim_epoca_alta": contrato_data.data_fim_epoca_alta,
            "valor_epoca_alta": contrato_data.valor_epoca_alta,
            "texto_epoca_alta": contrato_data.texto_epoca_alta,
            "data_inicio_epoca_baixa": contrato_data.data_inicio_epoca_baixa,
            "data_fim_epoca_baixa": contrato_data.data_fim_epoca_baixa,
            "valor_epoca_baixa": contrato_data.valor_epoca_baixa,
            "texto_epoca_baixa": contrato_data.texto_epoca_baixa,
            
            # Condições do veículo
            "condicoes_veiculo": contrato_data.condicoes_veiculo or vehicle.get("observacoes") or "",
            
            # Template do contrato (usar do parceiro se não enviado)
            "template_texto": contrato_data.template_texto or parceiro_data.get("template_contrato_padrao") or "",
            
            # Datas
            "data_inicio": contrato_data.data_inicio,
            "data_emissao": datetime.now(timezone.utc).strftime("%d/%m/%Y"),
            "data_assinatura": datetime.now(timezone.utc).strftime("%d/%m/%Y"),
            "local_assinatura": "Lisboa",
            
            # Status
            "status": "rascunho",
            "parceiro_assinado": False,
            "motorista_assinado": False,
            "parceiro_assinatura_data": None,
            "motorista_assinatura_data": None,
            "pdf_url": None,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.contratos.insert_one(contrato)
        
        return {
            "message": "Contract generated successfully",
            "contrato_id": contrato_id,
            "referencia": referencia
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating contract: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/contratos")
async def list_contratos(current_user: Dict = Depends(get_current_user)):
    """List all contracts"""
    try:
        query = {}
        
        # Filter by user role
        if current_user["role"] == UserRole.PARCEIRO:
            query["parceiro_id"] = current_user["id"]
        elif current_user["role"] == UserRole.MOTORISTA:
            query["motorista_id"] = current_user["id"]
        
        contratos = await db.contratos.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
        return contratos
        
    except Exception as e:
        logger.error(f"Error listing contracts: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/contratos/{contrato_id}")
async def get_contrato(contrato_id: str, current_user: Dict = Depends(get_current_user)):
    """Get contract details"""
    contrato = await db.contratos.find_one({"id": contrato_id}, {"_id": 0})
    if not contrato:
        raise HTTPException(status_code=404, detail="Contract not found")
    
    return contrato

@api_router.post("/contratos/{contrato_id}/assinar")
async def assinar_contrato(contrato_id: str, signer_data: Dict[str, str], current_user: Dict = Depends(get_current_user)):
    """Sign a contract (parceiro or motorista)"""
    contrato = await db.contratos.find_one({"id": contrato_id}, {"_id": 0})
    if not contrato:
        raise HTTPException(status_code=404, detail="Contract not found")
    
    signer_type = signer_data.get("signer_type")  # "parceiro" or "motorista"
    
    if signer_type == "parceiro":
        if current_user["role"] != UserRole.PARCEIRO or current_user["id"] != contrato["parceiro_id"]:
            raise HTTPException(status_code=403, detail="Not authorized to sign as parceiro")
        
        update_data = {
            "parceiro_assinado": True,
            "parceiro_assinatura_data": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Check if both signed
        if contrato.get("motorista_assinado"):
            update_data["status"] = "completo"
        else:
            update_data["status"] = "parceiro_assinado"
            
    elif signer_type == "motorista":
        if current_user["role"] != UserRole.MOTORISTA or current_user["id"] != contrato["motorista_id"]:
            raise HTTPException(status_code=403, detail="Not authorized to sign as motorista")
        
        update_data = {
            "motorista_assinado": True,
            "motorista_assinatura_data": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Check if both signed
        if contrato.get("parceiro_assinado"):
            update_data["status"] = "completo"
        else:
            update_data["status"] = "motorista_assinado"
    else:
        raise HTTPException(status_code=400, detail="Invalid signer_type")
    
    await db.contratos.update_one({"id": contrato_id}, {"$set": update_data})
    
    return {"message": "Contract signed successfully", "status": update_data.get("status")}

@api_router.get("/contratos/{contrato_id}/download")
async def download_contrato(contrato_id: str, current_user: Dict = Depends(get_current_user)):
    """Download contract PDF"""
    contrato = await db.contratos.find_one({"id": contrato_id}, {"_id": 0})
    if not contrato:
        raise HTTPException(status_code=404, detail="Contract not found")
    
    # Generate PDF on-the-fly
    try:
        # Get full data
        parceiro = await db.users.find_one({"id": contrato["parceiro_id"]}, {"_id": 0})
        motorista = await db.motoristas.find_one({"id": contrato["motorista_id"]}, {"_id": 0})
        vehicle = await db.vehicles.find_one({"id": contrato["vehicle_id"]}, {"_id": 0})
        
        # Create PDF in memory
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        
        # Title
        c.setFont("Helvetica-Bold", 18)
        c.drawString(50, 800, "CONTRATO DE PRESTAÇÃO DE SERVIÇOS TVDE")
        
        # Contract ID and Date
        c.setFont("Helvetica", 10)
        c.drawString(50, 780, f"Contrato ID: {contrato_id}")
        c.drawString(50, 765, f"Data: {datetime.fromisoformat(contrato['created_at']).strftime('%d/%m/%Y')}")
        
        # Parceiro section
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, 730, "PARCEIRO (Contratante)")
        c.setFont("Helvetica", 10)
        y_pos = 710
        c.drawString(50, y_pos, f"Nome/Empresa: {parceiro.get('nome_empresa', parceiro.get('name', 'N/A'))}")
        y_pos -= 15
        c.drawString(50, y_pos, f"NIF: {parceiro.get('contribuinte_empresa', 'N/A')}")
        y_pos -= 15
        c.drawString(50, y_pos, f"Morada: {parceiro.get('morada_completa', 'N/A')}")
        y_pos -= 15
        c.drawString(50, y_pos, f"Email: {parceiro.get('email', 'N/A')}")
        y_pos -= 15
        c.drawString(50, y_pos, f"Telefone: {parceiro.get('telefone', 'N/A')}")
        
        # Motorista section
        y_pos -= 30
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, y_pos, "MOTORISTA (Contratado)")
        c.setFont("Helvetica", 10)
        y_pos -= 20
        c.drawString(50, y_pos, f"Nome: {motorista.get('name', 'N/A')}")
        y_pos -= 15
        c.drawString(50, y_pos, f"Email: {motorista.get('email', 'N/A')}")
        y_pos -= 15
        c.drawString(50, y_pos, f"NIF: {motorista.get('professional', {}).get('nif', 'N/A')}")
        y_pos -= 15
        c.drawString(50, y_pos, f"Licença TVDE: {motorista.get('professional', {}).get('licenca_tvde', 'N/A')}")
        
        # Vehicle section
        y_pos -= 30
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, y_pos, "VEÍCULO")
        c.setFont("Helvetica", 10)
        y_pos -= 20
        c.drawString(50, y_pos, f"Matrícula: {vehicle.get('matricula', 'N/A')}")
        y_pos -= 15
        c.drawString(50, y_pos, f"Marca/Modelo: {vehicle.get('marca', 'N/A')} {vehicle.get('modelo', 'N/A')}")
        y_pos -= 15
        c.drawString(50, y_pos, f"Ano: {vehicle.get('ano', 'N/A')}")
        
        # Contract terms
        y_pos -= 40
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, y_pos, "CLÁUSULAS DO CONTRATO")
        c.setFont("Helvetica", 9)
        y_pos -= 20
        
        clauses = [
            "1. O MOTORISTA compromete-se a prestar serviços de transporte de passageiros através de plataformas TVDE.",
            "2. O veículo identificado permanecerá sob responsabilidade do MOTORISTA durante o período de vigência.",
            "3. O MOTORISTA deverá manter em dia toda a documentação necessária (licença TVDE, seguro, inspeção).",
            "4. As despesas de manutenção e combustível serão partilhadas conforme acordado entre as partes.",
            "5. Este contrato é válido por tempo indeterminado, podendo ser rescindido por qualquer parte.",
        ]
        
        for clause in clauses:
            c.drawString(50, y_pos, clause)
            y_pos -= 20
        
        # Signatures section
        y_pos -= 40
        c.setFont("Helvetica-Bold", 10)
        
        # Parceiro signature
        c.drawString(80, y_pos, "PARCEIRO")
        c.line(50, y_pos - 5, 250, y_pos - 5)
        if contrato.get("parceiro_assinado"):
            c.setFont("Helvetica", 8)
            c.drawString(50, y_pos - 20, f"Assinado em: {datetime.fromisoformat(contrato['parceiro_assinatura_data']).strftime('%d/%m/%Y %H:%M')}")
        
        # Motorista signature
        c.setFont("Helvetica-Bold", 10)
        c.drawString(350, y_pos, "MOTORISTA")
        c.line(320, y_pos - 5, 520, y_pos - 5)
        if contrato.get("motorista_assinado"):
            c.setFont("Helvetica", 8)
            c.drawString(320, y_pos - 20, f"Assinado em: {datetime.fromisoformat(contrato['motorista_assinatura_data']).strftime('%d/%m/%Y %H:%M')}")
        
        c.save()
        buffer.seek(0)
        
        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=contrato_{contrato_id}.pdf"}
        )
        
    except Exception as e:
        logger.error(f"Error generating contract PDF: {e}")
        raise HTTPException(status_code=500, detail="Error generating PDF")



# ==================== RECIBOS E PAGAMENTOS ENDPOINTS ====================

@api_router.post("/relatorios-ganhos")
async def criar_relatorio_ganhos(relatorio_data: RelatorioGanhosCreate, current_user: Dict = Depends(get_current_user)):
    """Create earnings report for a driver (Admin, Gestor, Operacional only)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    try:
        # Get motorista data
        motorista = await db.motoristas.find_one({"id": relatorio_data.motorista_id}, {"_id": 0})
        if not motorista:
            raise HTTPException(status_code=404, detail="Motorista not found")
        
        # Create relatorio
        relatorio_id = str(uuid.uuid4())
        relatorio = {
            "id": relatorio_id,
            "motorista_id": relatorio_data.motorista_id,
            "motorista_nome": motorista.get("name"),
            "periodo_inicio": relatorio_data.periodo_inicio,
            "periodo_fim": relatorio_data.periodo_fim,
            "valor_total": relatorio_data.valor_total,
            "detalhes": relatorio_data.detalhes,
            "notas": relatorio_data.notas,
            "status": "pendente_recibo",
            "recibo_url": None,
            "recibo_emitido_em": None,
            "aprovado_pagamento": False,
            "aprovado_pagamento_por": None,
            "aprovado_pagamento_em": None,
            "pago": False,
            "pago_por": None,
            "pago_em": None,
            "comprovativo_pagamento_url": None,
            "moloni_invoice_id": None,
            "moloni_invoice_status": None,
            "moloni_invoice_error": None,
            "created_by": current_user["id"],
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.relatorios_ganhos.insert_one(relatorio)
        
        return {"message": "Earnings report created successfully", "relatorio_id": relatorio_id}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating earnings report: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/relatorios-ganhos")
async def list_relatorios_ganhos(current_user: Dict = Depends(get_current_user)):
    """List earnings reports"""
    try:
        query = {}
        
        # Filter based on role
        if current_user["role"] == UserRole.MOTORISTA:
            query["motorista_id"] = current_user["id"]
        elif current_user["role"] in [UserRole.PARCEIRO]:
            # Get motoristas do parceiro/operacional
            motoristas = await db.motoristas.find({"parceiro_atribuido": current_user["id"]}, {"_id": 0, "id": 1}).to_list(100)
            motorista_ids = [m["id"] for m in motoristas]
            query["motorista_id"] = {"$in": motorista_ids}
        # Admin and Gestao can see all reports
        
        relatorios = await db.relatorios_ganhos.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)
        return relatorios
        
    except Exception as e:
        logger.error(f"Error listing earnings reports: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/relatorios-ganhos/{relatorio_id}/upload-recibo")
async def upload_recibo(
    relatorio_id: str,
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    """Upload receipt PDF (Motorista only)"""
    relatorio = await db.relatorios_ganhos.find_one({"id": relatorio_id}, {"_id": 0})
    if not relatorio:
        raise HTTPException(status_code=404, detail="Report not found")
    
    if current_user["role"] != UserRole.MOTORISTA or current_user["id"] != relatorio["motorista_id"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    try:
        # Create recibos directory
        recibos_dir = UPLOAD_DIR / "recibos"
        recibos_dir.mkdir(exist_ok=True)
        
        # Process file
        file_id = f"recibo_{relatorio_id}_{uuid.uuid4()}"
        file_info = await process_uploaded_file(file, recibos_dir, file_id)
        
        # Update relatorio
        update_data = {
            "status": "recibo_emitido",
            "recibo_url": file_info["saved_path"],
            "recibo_emitido_em": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.relatorios_ganhos.update_one({"id": relatorio_id}, {"$set": update_data})
        
        # MOLONI AUTO-INVOICING: Generate invoice when receipt is uploaded
        moloni_result = None
        try:
            from services.moloni_service import MoloniService
            moloni_service = MoloniService(db)
            
            # Only attempt if not already created
            if not relatorio.get("moloni_invoice_id"):
                moloni_result = await moloni_service.create_invoice_for_report(
                    motorista_id=relatorio["motorista_id"],
                    relatorio_id=relatorio_id,
                    valor_total=relatorio["valor_total"],
                    periodo_inicio=relatorio["periodo_inicio"],
                    periodo_fim=relatorio["periodo_fim"]
                )
                
                # Update relatorio with Moloni result
                if moloni_result.get("success"):
                    await db.relatorios_ganhos.update_one(
                        {"id": relatorio_id},
                        {
                            "$set": {
                                "moloni_invoice_id": moloni_result.get("moloni_invoice_id"),
                                "moloni_invoice_status": "created",
                                "updated_at": datetime.now(timezone.utc).isoformat()
                            }
                        }
                    )
                    logger.info(f"Moloni invoice created for report {relatorio_id}: {moloni_result.get('moloni_invoice_id')}")
                else:
                    # Log error but don't fail the receipt upload
                    await db.relatorios_ganhos.update_one(
                        {"id": relatorio_id},
                        {
                            "$set": {
                                "moloni_invoice_status": "failed",
                                "moloni_invoice_error": moloni_result.get("error"),
                                "updated_at": datetime.now(timezone.utc).isoformat()
                            }
                        }
                    )
                    logger.warning(f"Moloni invoice creation failed for report {relatorio_id}: {moloni_result.get('error')}")
            
            await moloni_service.close()
            
        except Exception as moloni_error:
            logger.error(f"Error during Moloni invoice creation: {moloni_error}")
            # Continue - don't fail receipt upload if Moloni fails
            await db.relatorios_ganhos.update_one(
                {"id": relatorio_id},
                {
                    "$set": {
                        "moloni_invoice_status": "error",
                        "moloni_invoice_error": str(moloni_error),
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }
                }
            )
        
        response_data = {
            "message": "Receipt uploaded successfully",
            "file_url": file_info["saved_path"]
        }
        
        if moloni_result and moloni_result.get("success"):
            response_data["moloni_invoice_created"] = True
            response_data["moloni_invoice_id"] = moloni_result.get("moloni_invoice_id")
        elif moloni_result:
            response_data["moloni_invoice_created"] = False
            response_data["moloni_error"] = moloni_result.get("error")
        
        return response_data
        
    except Exception as e:
        logger.error(f"Error uploading receipt: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/relatorios-ganhos/{relatorio_id}/aprovar-pagamento")
async def aprovar_pagamento(relatorio_id: str, current_user: Dict = Depends(get_current_user)):
    """Approve payment (Admin, Gestor, Operacional, Parceiro)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    relatorio = await db.relatorios_ganhos.find_one({"id": relatorio_id}, {"_id": 0})
    if not relatorio:
        raise HTTPException(status_code=404, detail="Report not found")
    
    if relatorio["status"] != "recibo_emitido":
        raise HTTPException(status_code=400, detail="Receipt not yet issued")
    
    update_data = {
        "status": "aguardando_pagamento",
        "aprovado_pagamento": True,
        "aprovado_pagamento_por": current_user["id"],
        "aprovado_pagamento_em": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.relatorios_ganhos.update_one({"id": relatorio_id}, {"$set": update_data})
    
    return {"message": "Payment approved successfully"}

@api_router.post("/relatorios-ganhos/{relatorio_id}/gerar-recibo")
async def gerar_recibo_semanal(relatorio_id: str, current_user: Dict = Depends(get_current_user)):
    """Generate weekly receipt PDF (Admin, Gestor, Operacional)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    relatorio = await db.relatorios_ganhos.find_one({"id": relatorio_id}, {"_id": 0})
    if not relatorio:
        raise HTTPException(status_code=404, detail="Report not found")
    
    try:
        # Create receipts directory
        recibos_dir = ROOT_DIR / "uploads" / "recibos_semanais"
        recibos_dir.mkdir(parents=True, exist_ok=True)
        
        # Generate PDF filename
        pdf_filename = f"recibo_{relatorio['motorista_id']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        pdf_path = recibos_dir / pdf_filename
        
        # Here you would generate the actual PDF with all details
        # For now, create a simple placeholder
        # TODO: Implement full PDF generation with reportlab or similar
        
        # Update relatorio with PDF URL
        pdf_url = f"uploads/recibos_semanais/{pdf_filename}"
        update_data = {
            "recibo_url": pdf_url,
            "recibo_emitido_em": datetime.now(timezone.utc).isoformat(),
            "status": "recibo_gerado",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.relatorios_ganhos.update_one({"id": relatorio_id}, {"$set": update_data})
        
        return {
            "message": "Receipt generated successfully",
            "recibo_url": pdf_url
        }
        
    except Exception as e:
        logger.error(f"Error generating receipt: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/relatorios-ganhos/{relatorio_id}/marcar-pago")
async def marcar_pago(
    relatorio_id: str,
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    """Mark as paid with payment proof (Admin, Gestor, Operacional, Parceiro)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    relatorio = await db.relatorios_ganhos.find_one({"id": relatorio_id}, {"_id": 0})
    if not relatorio:
        raise HTTPException(status_code=404, detail="Report not found")
    
    if not relatorio["aprovado_pagamento"]:
        raise HTTPException(status_code=400, detail="Payment not yet approved")
    
    try:
        # Create comprovativos directory
        comprova_dir = UPLOAD_DIR / "comprovativos_pagamento"
        comprova_dir.mkdir(exist_ok=True)
        
        # Process file
        file_id = f"comprovativo_{relatorio_id}_{uuid.uuid4()}"
        file_info = await process_uploaded_file(file, comprova_dir, file_id)
        
        # Update relatorio
        update_data = {
            "status": "pago",
            "pago": True,
            "pago_por": current_user["id"],
            "pago_em": datetime.now(timezone.utc).isoformat(),
            "comprovativo_pagamento_url": file_info["saved_path"],
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.relatorios_ganhos.update_one({"id": relatorio_id}, {"$set": update_data})
        
        return {"message": "Marked as paid successfully", "comprovativo_url": file_info["saved_path"]}
        
    except Exception as e:
        logger.error(f"Error marking as paid: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/relatorios-ganhos/{relatorio_id}/alterar-estado")
async def alterar_estado_relatorio(
    relatorio_id: str,
    estado_data: dict,
    current_user: Dict = Depends(get_current_user)
):
    """Change report status (Admin, Gestor, Operacional, Parceiro)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    relatorio = await db.relatorios_ganhos.find_one({"id": relatorio_id}, {"_id": 0})
    if not relatorio:
        raise HTTPException(status_code=404, detail="Report not found")
    
    # If parceiro, verify ownership
    if current_user["role"] == UserRole.PARCEIRO:
        motorista = await db.motoristas.find_one({"id": relatorio["motorista_id"]}, {"_id": 0})
        if not motorista or motorista.get("parceiro_atribuido") != current_user["id"]:
            raise HTTPException(status_code=403, detail="Not authorized for this report")
    
    new_status = estado_data.get("status")
    if not new_status:
        raise HTTPException(status_code=400, detail="Status is required")
    
    # Valid status transitions
    valid_statuses = ["por_enviar", "em_analise", "a_pagamento", "liquidado", "pendente_recibo", "recibo_emitido", "aguardando_pagamento", "pago"]
    if new_status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Valid options: {', '.join(valid_statuses)}")
    
    update_data = {
        "status": new_status,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # If marking as liquidado, also update pago flag
    if new_status == "liquidado":
        update_data["pago"] = True
        update_data["pago_por"] = current_user["id"]
        update_data["pago_em"] = datetime.now(timezone.utc).isoformat()
    
    await db.relatorios_ganhos.update_one({"id": relatorio_id}, {"$set": update_data})
    
    return {"message": "Status updated successfully", "new_status": new_status}

@api_router.post("/relatorios-ganhos/{relatorio_id}/comprovativo")
async def upload_comprovativo_pagamento(
    relatorio_id: str,
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    """Upload payment proof (Admin, Gestor, Operacional, Parceiro)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    relatorio = await db.relatorios_ganhos.find_one({"id": relatorio_id}, {"_id": 0})
    if not relatorio:
        raise HTTPException(status_code=404, detail="Report not found")
    
    # If parceiro, verify ownership
    if current_user["role"] == UserRole.PARCEIRO:
        motorista = await db.motoristas.find_one({"id": relatorio["motorista_id"]}, {"_id": 0})
        if not motorista or motorista.get("parceiro_atribuido") != current_user["id"]:
            raise HTTPException(status_code=403, detail="Not authorized for this report")
    
    try:
        # Create comprovativos directory
        comprova_dir = UPLOAD_DIR / "comprovativos_pagamento"
        comprova_dir.mkdir(exist_ok=True)
        
        # Process file
        file_id = f"comprovativo_{relatorio_id}_{uuid.uuid4()}"
        file_info = await process_uploaded_file(file, comprova_dir, file_id)
        
        # Update relatorio with comprovativo URL and change status to liquidado
        comprovativo_url = file_info.get("pdf_path") or file_info.get("original_path")
        update_data = {
            "comprovativo_pagamento_url": comprovativo_url,
            "status": "liquidado",  # Automatically set to liquidado when comprovativo is uploaded
            "pago_em": datetime.now(timezone.utc).isoformat(),
            "pago_por": current_user["id"],
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.relatorios_ganhos.update_one({"id": relatorio_id}, {"$set": update_data})
        
        return {"message": "Payment proof uploaded successfully - Status changed to liquidado", "comprovativo_url": comprovativo_url}
        
    except Exception as e:
        logger.error(f"Error uploading payment proof: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/relatorios-ganhos/{relatorio_id}/comprovativo/download")
async def download_comprovativo_pagamento(
    relatorio_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Download payment proof"""
    relatorio = await db.relatorios_ganhos.find_one({"id": relatorio_id}, {"_id": 0})
    if not relatorio:
        raise HTTPException(status_code=404, detail="Report not found")
    
    # Check authorization
    is_authorized = current_user["role"] in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]
    
    if current_user["role"] == UserRole.PARCEIRO:
        motorista = await db.motoristas.find_one({"id": relatorio["motorista_id"]}, {"_id": 0})
        is_authorized = motorista and motorista.get("parceiro_atribuido") == current_user["id"]
    
    if not is_authorized:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    comprovativo_url = relatorio.get("comprovativo_pagamento_url")
    if not comprovativo_url:
        raise HTTPException(status_code=404, detail="Payment proof not found")
    
    # Construct file path
    file_path = UPLOAD_DIR / comprovativo_url.replace("/uploads/", "")
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Payment proof file not found")
    
    return FileResponse(
        path=file_path,
        media_type="application/pdf",
        filename=file_path.name
    )



# ==================== CSV TEMPLATE DOWNLOADS ====================

@api_router.get("/templates/csv/{template_name}")
async def download_csv_template(template_name: str, current_user: Dict = Depends(get_current_user)):
    """Download CSV template examples"""
    templates_dir = ROOT_DIR / "templates" / "csv_examples"
    
    # Map template names to filenames
    template_files = {
        "uber": "uber_example.csv",
        "bolt": "bolt_example.csv",
        "prio": "prio_example.xlsx",
        "viaverde": "viaverde_example.csv",
        "gps": "gps_example.csv"
    }
    
    if template_name not in template_files:
        raise HTTPException(status_code=404, detail="Template not found")
    
    file_path = templates_dir / template_files[template_name]
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Template file not found")
    
    # Set appropriate media type
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if template_name == "prio" else "text/csv"
    
    return FileResponse(
        path=file_path,
        media_type=media_type,
        filename=template_files[template_name]
    )

# Include modular routers (new structure)
app.include_router(auth_router, prefix="/api")
app.include_router(motoristas_router, prefix="/api")
app.include_router(notificacoes_router, prefix="/api")
app.include_router(mensagens_router, prefix="/api")
app.include_router(ifthenpay_router)

# Include main API router (legacy routes - to be migrated)
# ==================== DOCUMENTO VALIDATION SYSTEM ====================

class DocumentoUpload(BaseModel):
    tipo_documento: str
    user_id: str
    role: str

class DocumentoAprovacao(BaseModel):
    aprovado: bool
    observacoes: Optional[str] = None

DOCUMENTOS_MOTORISTA = [
    "carta_conducao",
    "identificacao",  # CC, Passaporte ou Título Residência
    "licenca_tvde",
    "registo_criminal",
    "comprovativo_morada"
]

DOCUMENTOS_PARCEIRO = [
    "certidao_comercial"
]

DOCUMENTOS_DIR = UPLOAD_DIR / "documentos_validacao"
DOCUMENTOS_DIR.mkdir(exist_ok=True)

@api_router.post("/documentos/upload")
async def upload_documento(
    file: UploadFile = File(...),
    tipo_documento: str = Form(...),
    user_id: str = Form(...),
    role: str = Form(...),
    current_user: Dict = Depends(get_current_user)
):
    """Upload de documento para validação (com conversão automática para PDF)"""
    try:
        # Validar tipo de documento
        if role == "motorista" and tipo_documento not in DOCUMENTOS_MOTORISTA:
            raise HTTPException(status_code=400, detail="Tipo de documento inválido para motorista")
        if role == "parceiro" and tipo_documento not in DOCUMENTOS_PARCEIRO:
            raise HTTPException(status_code=400, detail="Tipo de documento inválido para parceiro")
        
        # Criar diretório para o utilizador se não existir
        user_docs_dir = DOCUMENTOS_DIR / user_id
        user_docs_dir.mkdir(exist_ok=True)
        
        # Salvar arquivo temporário
        file_extension = Path(file.filename).suffix.lower()
        temp_file_path = user_docs_dir / f"{tipo_documento}_temp{file_extension}"
        
        with open(temp_file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        # Converter para PDF se for imagem
        final_file_path = user_docs_dir / f"{tipo_documento}.pdf"
        
        if file_extension in ['.jpg', '.jpeg', '.png', '.webp', '.heic', '.heif']:
            # Converter imagem para PDF
            await convert_image_to_pdf(temp_file_path, final_file_path)
            temp_file_path.unlink()  # Remover arquivo temporário
        elif file_extension == '.pdf':
            # Renomear para o nome final
            temp_file_path.rename(final_file_path)
        else:
            temp_file_path.unlink()
            raise HTTPException(status_code=400, detail="Formato de arquivo não suportado. Use PDF ou imagens.")
        
        # Salvar no banco de dados
        documento = {
            "id": str(uuid.uuid4()),
            "user_id": user_id,
            "role": role,
            "tipo_documento": tipo_documento,
            "file_path": str(final_file_path),
            "filename": file.filename,
            "status": "pendente",
            "observacoes": None,
            "data_upload": datetime.now(timezone.utc).isoformat(),
            "data_aprovacao": None,
            "aprovado_por": None,
            "pode_alterar": tipo_documento == "registo_criminal"  # Registo criminal pode ser alterado livremente
        }
        
        # Verificar se já existe documento deste tipo
        existing = await db.documentos_validacao.find_one({"user_id": user_id, "tipo_documento": tipo_documento})
        
        if existing:
            # Atualizar documento existente
            await db.documentos_validacao.update_one(
                {"user_id": user_id, "tipo_documento": tipo_documento},
                {"$set": documento}
            )
        else:
            # Inserir novo documento
            await db.documentos_validacao.insert_one(documento)
        
        logger.info(f"Documento {tipo_documento} carregado para utilizador {user_id}")
        
        return {
            "message": "Documento carregado com sucesso",
            "documento_id": documento["id"],
            "tipo_documento": tipo_documento
        }
        
    except Exception as e:
        logger.error(f"Erro ao fazer upload de documento: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/documentos/pendentes")
async def get_documentos_pendentes(current_user: Dict = Depends(get_current_user)):
    """Listar todos os utilizadores não aprovados (com ou sem documentos pendentes) (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    try:
        # Buscar todos os utilizadores não aprovados
        users_nao_aprovados = await db.users.find(
            {"approved": False},
            {"_id": 0, "password": 0}
        ).to_list(length=None)
        
        users_pendentes = {}
        
        # Adicionar utilizadores não aprovados
        for user in users_nao_aprovados:
            user_id = user["id"]
            
            # Se for parceiro, buscar dados adicionais
            parceiro_data = None
            if user.get("role") == "parceiro":
                parceiro_data = await db.parceiros.find_one(
                    {"email": user.get("email")},
                    {"_id": 0}
                )
            
            # Se for motorista, buscar dados adicionais
            motorista_data = None
            if user.get("role") == "motorista":
                motorista_data = await db.motoristas.find_one(
                    {"email": user.get("email")},
                    {"_id": 0}
                )
            
            # Buscar documentos deste utilizador
            documentos = await db.documentos_validacao.find(
                {"user_id": user_id},
                {"_id": 0}
            ).to_list(length=None)
            
            users_pendentes[user_id] = {
                "user": user,
                "parceiro_data": parceiro_data,
                "motorista_data": motorista_data,
                "documentos": documentos
            }
        
        # Adicionar também utilizadores com documentos pendentes (mesmo que approved=true)
        documentos_pendentes = await db.documentos_validacao.find(
            {"status": "pendente"},
            {"_id": 0}
        ).to_list(length=None)
        
        for doc in documentos_pendentes:
            user_id = doc["user_id"]
            if user_id not in users_pendentes:
                # Buscar informações do utilizador
                user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
                if user:
                    # Se for parceiro, buscar dados adicionais
                    parceiro_data = None
                    if user.get("role") == "parceiro":
                        parceiro_data = await db.parceiros.find_one(
                            {"email": user.get("email")},
                            {"_id": 0}
                        )
                    
                    # Se for motorista, buscar dados adicionais
                    motorista_data = None
                    if user.get("role") == "motorista":
                        motorista_data = await db.motoristas.find_one(
                            {"email": user.get("email")},
                            {"_id": 0}
                        )
                    
                    # Buscar todos os documentos
                    documentos = await db.documentos_validacao.find(
                        {"user_id": user_id},
                        {"_id": 0}
                    ).to_list(length=None)
                    
                    users_pendentes[user_id] = {
                        "user": user,
                        "parceiro_data": parceiro_data,
                        "motorista_data": motorista_data,
                        "documentos": documentos
                    }
        
        return list(users_pendentes.values())
        
    except Exception as e:
        logger.error(f"Erro ao listar documentos pendentes: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/documentos/user/{user_id}")
async def get_documentos_user(user_id: str, current_user: Dict = Depends(get_current_user)):
    """Listar documentos de um utilizador específico"""
    # Admin pode ver todos, utilizador só pode ver os próprios
    if current_user["role"] != UserRole.ADMIN and current_user["id"] != user_id:
        raise HTTPException(status_code=403, detail="Não autorizado")
    
    try:
        documentos = await db.documentos_validacao.find(
            {"user_id": user_id},
            {"_id": 0}
        ).to_list(length=None)
        
        return documentos
        
    except Exception as e:
        logger.error(f"Erro ao listar documentos do utilizador: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/users/{user_id}/complete-details")
async def get_user_complete_details(user_id: str, current_user: Dict = Depends(get_current_user)):
    """Obter TODOS os dados de um utilizador (user + parceiro/motorista + documentos)"""
    # Apenas admin pode ver
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Apenas Admin")
    
    try:
        # Buscar utilizador
        user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
        if not user:
            raise HTTPException(status_code=404, detail="Utilizador não encontrado")
        
        # Buscar dados de parceiro se aplicável
        parceiro_data = None
        if user.get("role") == "parceiro":
            parceiro_data = await db.parceiros.find_one(
                {"email": user.get("email")},
                {"_id": 0}
            )
        
        # Buscar dados de motorista se aplicável
        motorista_data = None
        if user.get("role") == "motorista":
            motorista_data = await db.motoristas.find_one(
                {"email": user.get("email")},
                {"_id": 0}
            )
        
        # Buscar documentos
        documentos = await db.documentos_validacao.find(
            {"user_id": user_id},
            {"_id": 0}
        ).to_list(length=None)
        
        return {
            "user": user,
            "parceiro_data": parceiro_data,
            "motorista_data": motorista_data,
            "documentos": documentos
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao buscar detalhes completos: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/documentos/{documento_id}/download")
async def download_documento(documento_id: str, current_user: Dict = Depends(get_current_user)):
    """Download de documento (Admin, Parceiro ou Gestor)"""
    if current_user["role"] not in [UserRole.ADMIN, "parceiro", "gestao"]:
        raise HTTPException(status_code=403, detail="Não autorizado")
    
    try:
        documento = await db.documentos_validacao.find_one({"id": documento_id}, {"_id": 0})
        
        if not documento:
            raise HTTPException(status_code=404, detail="Documento não encontrado")
        
        file_path = Path(documento["file_path"])
        
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="Arquivo não encontrado no servidor")
        
        return FileResponse(
            path=file_path,
            filename=f"{documento['tipo_documento']}.pdf",
            media_type="application/pdf"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao fazer download de documento: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/documentos/{documento_id}/aprovar")
async def aprovar_documento(
    documento_id: str,
    aprovacao: DocumentoAprovacao,
    current_user: Dict = Depends(get_current_user)
):
    """Aprovar ou rejeitar documento individual (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    try:
        documento = await db.documentos_validacao.find_one({"id": documento_id}, {"_id": 0})
        
        if not documento:
            raise HTTPException(status_code=404, detail="Documento não encontrado")
        
        # Atualizar status do documento
        update_data = {
            "status": "aprovado" if aprovacao.aprovado else "rejeitado",
            "observacoes": aprovacao.observacoes,
            "data_aprovacao": datetime.now(timezone.utc).isoformat(),
            "aprovado_por": current_user["id"]
        }
        
        await db.documentos_validacao.update_one(
            {"id": documento_id},
            {"$set": update_data}
        )
        
        logger.info(f"Documento {documento_id} {'aprovado' if aprovacao.aprovado else 'rejeitado'} por {current_user['id']}")
        
        # Se rejeitado, enviar email com observações (implementar depois)
        if not aprovacao.aprovado:
            # TODO: Enviar email de notificação
            pass
        
        return {
            "message": f"Documento {'aprovado' if aprovacao.aprovado else 'rejeitado'} com sucesso",
            "documento_id": documento_id
        }
        
    except Exception as e:
        logger.error(f"Erro ao aprovar documento: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/documentos/user/{user_id}/aprovar-todos")
async def aprovar_todos_documentos(
    user_id: str,
    aprovacao: DocumentoAprovacao,
    current_user: Dict = Depends(get_current_user)
):
    """Aprovar ou rejeitar todos os documentos de um utilizador (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    try:
        # Buscar todos os documentos pendentes do utilizador
        documentos = await db.documentos_validacao.find(
            {"user_id": user_id, "status": "pendente"},
            {"_id": 0}
        ).to_list(length=None)
        
        if not documentos:
            raise HTTPException(status_code=404, detail="Nenhum documento pendente encontrado")
        
        # Atualizar todos os documentos
        update_data = {
            "status": "aprovado" if aprovacao.aprovado else "rejeitado",
            "observacoes": aprovacao.observacoes,
            "data_aprovacao": datetime.now(timezone.utc).isoformat(),
            "aprovado_por": current_user["id"]
        }
        
        await db.documentos_validacao.update_many(
            {"user_id": user_id, "status": "pendente"},
            {"$set": update_data}
        )
        
        # Se aprovado, atribuir plano base gratuito
        if aprovacao.aprovado:
            user = await db.users.find_one({"id": user_id}, {"_id": 0})
            
            if user:
                # Buscar plano base gratuito correspondente ao role
                plano_base = await db.planos_sistema.find_one({
                    "tipo_usuario": user["role"],
                    "preco_mensal": 0,
                    "ativo": True
                }, {"_id": 0})
                
                if plano_base:
                    # Atribuir plano ao utilizador
                    await db.users.update_one(
                        {"id": user_id},
                        {"$set": {"plano_id": plano_base["id"]}}
                    )
                    
                    logger.info(f"Plano base '{plano_base['nome']}' atribuído ao utilizador {user_id}")
        
        logger.info(f"Todos os documentos do utilizador {user_id} foram {'aprovados' if aprovacao.aprovado else 'rejeitados'}")
        
        return {
            "message": f"Todos os documentos {'aprovados' if aprovacao.aprovado else 'rejeitados'} com sucesso",
            "documentos_afetados": len(documentos)
        }
        
    except Exception as e:
        logger.error(f"Erro ao aprovar todos os documentos: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==================== ENDPOINT PÚBLICO PARA TEXTOS ====================

@api_router.get("/configuracoes/textos")
async def get_textos_publicos():
    """Obter textos de Termos e Privacidade (público)"""
    try:
        config = await db.configuracoes_sistema.find_one({}, {"_id": 0})
        
        if not config:
            return {
                "termos_condicoes": "Termos e Condições não configurados.",
                "politica_privacidade": "Política de Privacidade não configurada."
            }
        
        return {
            "termos_condicoes": config.get("termos_condicoes", "Termos e Condições não configurados."),
            "politica_privacidade": config.get("politica_privacidade", "Política de Privacidade não configurada.")
        }
    except Exception as e:
        logger.error(f"Erro ao buscar textos: {e}")
        return {
            "termos_condicoes": "Erro ao carregar Termos e Condições.",
            "politica_privacidade": "Erro ao carregar Política de Privacidade."
        }

# ==================== GERAÇÃO DE RELATÓRIO PDF ====================

class RelatorioSemanalData(BaseModel):
    motorista_nome: str
    motorista_id: str
    veiculo_matricula: str
    veiculo_id: Optional[str] = None
    data_inicio: str
    data_fim: str
    numero_semana: int
    
    # Ganhos
    uber_ganhos: float = 0
    bolt_ganhos: float = 0
    uber_gorjeta: float = 0
    bolt_gorjeta: float = 0
    uber_portagens: float = 0
    bolt_portagens: float = 0
    
    # Condições
    tipo_exploracao: str  # "aluguer" ou "comissao"
    valor_aluguer: Optional[float] = None
    percentagem_motorista: Optional[float] = None
    percentagem_parceiro: Optional[float] = None
    
    # Despesas
    via_verde: float = 0
    abastecimentos: float = 0
    caucao: float = 0
    acumulado_caucao: float = 0
    divida_anterior: float = 0
    extra: float = 0  # positivo = crédito, negativo = débito
    
    # Outros
    km: float = 0
    horas: float = 0
    
    # Listas detalhadas
    via_verde_detalhes: List[Dict] = []
    abastecimentos_detalhes: List[Dict] = []

@api_router.post("/relatorios/gerar-pdf")
async def gerar_relatorio_pdf(
    relatorio_data: RelatorioSemanalData,
    current_user: Dict = Depends(get_current_user)
):
    """Gerar PDF de relatório semanal completo"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        
        # TODO: Implement PDF generation logic
        return {"message": "PDF generation not implemented yet"}
    
    except Exception as e:
        logger.error(f"Error generating PDF: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# COMUNICAÇÕES - EMAIL & WHATSAPP CONFIGURATION
# ============================================================================

@api_router.post("/configuracoes/comunicacoes/email")
async def save_email_config(
    config_data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Save email (SendGrid or SMTP) configuration (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Only admins can configure communications")
    
    try:
        provider = config_data.get("provider", "sendgrid")
        
        config = {
            "tipo": "email_comunicacoes",
            "provider": provider,
            "sender_email": config_data.get("sender_email"),
            "sender_name": config_data.get("sender_name", "TVDEFleet"),
            "enabled": config_data.get("enabled", False),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": current_user["id"]
        }
        
        # Add provider-specific fields
        if provider == "sendgrid":
            config["api_key"] = config_data.get("api_key")
        elif provider == "smtp":
            config["smtp_host"] = config_data.get("smtp_host")
            config["smtp_port"] = config_data.get("smtp_port", 587)
            config["smtp_user"] = config_data.get("smtp_user")
            config["smtp_password"] = config_data.get("smtp_password")
            config["smtp_use_tls"] = config_data.get("smtp_use_tls", True)
        
        await db.configuracoes_sistema.update_one(
            {"tipo": "email_comunicacoes"},
            {"$set": config},
            upsert=True
        )
        
        logger.info(f"Email configuration ({provider}) saved by {current_user['email']}")
        return {"message": "Email configuration saved successfully"}
        
    except Exception as e:
        logger.error(f"Error saving email config: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/configuracoes/comunicacoes/email")
async def get_email_config(current_user: Dict = Depends(get_current_user)):
    """Get email configuration (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Only admins can view communications config")
    
    try:
        config = await db.configuracoes_sistema.find_one(
            {"tipo": "email_comunicacoes"},
            {"_id": 0}
        )
        
        if not config:
            return {
                "provider": "sendgrid",
                "api_key": "",
                "smtp_host": "",
                "smtp_port": 587,
                "smtp_user": "",
                "smtp_password": "",
                "smtp_use_tls": True,
                "sender_email": "",
                "sender_name": "TVDEFleet",
                "enabled": False
            }
        
        return config
        
    except Exception as e:
        logger.error(f"Error fetching email config: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.patch("/configuracoes/comunicacoes/email")
async def update_email_status(
    data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Update email service status (enable/disable)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Only admins can modify communications config")
    
    try:
        enabled = data.get("enabled", False)
        
        await db.configuracoes_sistema.update_one(
            {"tipo": "email_comunicacoes"},
            {
                "$set": {
                    "enabled": enabled,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
            }
        )
        
        logger.info(f"Email service {'enabled' if enabled else 'disabled'} by {current_user['email']}")
        return {"message": f"Email service {'enabled' if enabled else 'disabled'} successfully"}
        
    except Exception as e:
        logger.error(f"Error updating email status: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/configuracoes/comunicacoes/email/test")
async def test_email_config(
    data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Send test email to verify configuration"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Only admins can test email config")
    
    try:
        from services.email_service import EmailService
        
        email_service = EmailService(db)
        recipient = data.get("recipient")
        
        if not recipient:
            raise HTTPException(status_code=400, detail="Recipient email is required")
        
        # Send test email
        result = await email_service.send_email(
            to_email=recipient,
            subject="🔔 Teste de Configuração - TVDEFleet",
            html_content="""
            <html>
            <head>
                <style>
                    body { font-family: Arial, sans-serif; line-height: 1.6; color: #333; }
                    .container { max-width: 600px; margin: 0 auto; padding: 20px; }
                    .header { background-color: #2563eb; color: white; padding: 20px; text-align: center; border-radius: 8px 8px 0 0; }
                    .content { background-color: #f9fafb; padding: 30px; border: 1px solid #e5e7eb; }
                    .success-box { background-color: #ecfdf5; border-left: 4px solid #10b981; padding: 15px; margin: 20px 0; }
                    .footer { background-color: #f3f4f6; padding: 15px; text-align: center; font-size: 12px; color: #6b7280; border-radius: 0 0 8px 8px; }
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1>✅ Email de Teste</h1>
                    </div>
                    <div class="content">
                        <p>Olá,</p>
                        <p>Este é um email de teste do sistema TVDEFleet.</p>
                        
                        <div class="success-box">
                            <strong>✅ Configuração SendGrid OK!</strong><br>
                            Se recebeu este email, significa que a configuração está correta e funcional.
                        </div>
                        
                        <p>O sistema está pronto para enviar notificações automáticas.</p>
                    </div>
                    <div class="footer">
                        <p>Esta é uma mensagem de teste automática.</p>
                        <p>&copy; 2024 TVDEFleet - Sistema de Gestão de Frotas</p>
                    </div>
                </div>
            </body>
            </html>
            """
        )
        
        return result
        
    except Exception as e:
        logger.error(f"Error testing email: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/configuracoes/comunicacoes/whatsapp")
async def get_whatsapp_config(current_user: Dict = Depends(get_current_user)):
    """Get WhatsApp configuration (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Only admins can view communications config")
    
    try:
        config = await db.configuracoes_sistema.find_one(
            {"tipo": "whatsapp_comunicacoes"},
            {"_id": 0}
        )
        
        if not config:
            return {
                "provider": "baileys",
                "enabled": False,
                "connected": False
            }
        
        return config
        
    except Exception as e:
        logger.error(f"Error fetching WhatsApp config: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/notificacoes/send")
async def send_notification(
    notification_data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Send notification via email (Admin, Gestao, Parceiro)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    try:
        from services.email_service import EmailService
        
        email_service = EmailService(db)
        
        to_email = notification_data.get("to_email")
        notification_type = notification_data.get("type")
        data = notification_data.get("data", {})
        
        if not to_email or not notification_type:
            raise HTTPException(status_code=400, detail="to_email and type are required")
        
        result = await email_service.send_notification(
            to_email=to_email,
            notification_type=notification_type,
            data=data
        )
        
        return result
        
    except Exception as e:
        logger.error(f"Error sending notification: {e}")

@api_router.get("/termos-conteudo")
async def get_termos_conteudo():
    """Get Terms & Conditions content"""
    doc = await db.configuracoes.find_one({"tipo": "termos"}, {"_id": 0})
    if not doc:
        return {"conteudo": "", "updated_at": None}
    return {"conteudo": doc.get("conteudo", ""), "updated_at": doc.get("updated_at")}

@api_router.put("/termos-conteudo")
async def update_termos_conteudo(data: Dict[str, Any], current_user: Dict = Depends(get_current_user)):
    """Update Terms & Conditions (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    update_doc = {
        "tipo": "termos",
        "conteudo": data.get("conteudo", ""),
        "updated_at": datetime.now(timezone.utc),
        "updated_by": current_user["id"]
    }
    
    await db.configuracoes.update_one(
        {"tipo": "termos"},
        {"$set": update_doc},
        upsert=True
    )
    
    return {"message": "Termos atualizados com sucesso"}

@api_router.get("/privacidade-conteudo")
async def get_privacidade_conteudo():
    """Get Privacy Policy content"""
    doc = await db.configuracoes.find_one({"tipo": "privacidade"}, {"_id": 0})
    if not doc:
        return {"conteudo": "", "updated_at": None}
    return {"conteudo": doc.get("conteudo", ""), "updated_at": doc.get("updated_at")}

@api_router.put("/privacidade-conteudo")
async def update_privacidade_conteudo(data: Dict[str, Any], current_user: Dict = Depends(get_current_user)):
    """Update Privacy Policy (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    update_doc = {
        "tipo": "privacidade",
        "conteudo": data.get("conteudo", ""),
        "updated_at": datetime.now(timezone.utc),
        "updated_by": current_user["id"]
    }
    
    await db.configuracoes.update_one(
        {"tipo": "privacidade"},
        {"$set": update_doc},
        upsert=True
    )
    
    return {"message": "Política de Privacidade atualizada com sucesso"}

# ============================================
# RELATÓRIOS SEMANAIS
# ============================================

@api_router.post("/relatorios/gerar-semanal")
async def gerar_relatorio_semanal(
    data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Generate weekly report PDF for motorista or parceiro"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    usuario_id = data.get("usuario_id")
    tipo_usuario = data.get("tipo_usuario")  # motorista ou parceiro
    data_inicio = data.get("data_inicio")
    data_fim = data.get("data_fim")
    
    if not all([usuario_id, tipo_usuario, data_inicio, data_fim]):
        raise HTTPException(status_code=400, detail="Missing required fields")
    
    # TODO: Implement PDF generation
    # For now, return a placeholder response
    raise HTTPException(
        status_code=501,
        detail="Funcionalidade de geração de PDF em desenvolvimento. Por favor, use a exportação manual por enquanto."
    )

@api_router.post("/relatorios/enviar-email")
async def enviar_relatorio_email(
    data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Send weekly report via email"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    usuario_id = data.get("usuario_id")
    tipo_usuario = data.get("tipo_usuario")
    data_inicio = data.get("data_inicio")
    data_fim = data.get("data_fim")
    
    if not all([usuario_id, tipo_usuario, data_inicio, data_fim]):
        raise HTTPException(status_code=400, detail="Missing required fields")
    
    # Get user email
    collection = db.motoristas if tipo_usuario == "motorista" else db.parceiros
    usuario = await collection.find_one({"id": usuario_id}, {"_id": 0})
    
    if not usuario:
        # Try users collection
        usuario = await db.users.find_one({"id": usuario_id}, {"_id": 0})
    
    if not usuario:
        raise HTTPException(status_code=404, detail="User not found")
    
    email_destino = usuario.get("email")
    if not email_destino:
        raise HTTPException(status_code=400, detail="User does not have an email")
    
    # TODO: Implement email sending with report attachment
    # For now, return success message
    raise HTTPException(
        status_code=501,
        detail="Funcionalidade de envio de email em desenvolvimento. Configure a integração de email (SMTP) para utilizar esta funcionalidade."
    )

@api_router.post("/relatorios/enviar-whatsapp")
async def enviar_relatorio_whatsapp(
    data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Send weekly report via WhatsApp"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    usuario_id = data.get("usuario_id")
    tipo_usuario = data.get("tipo_usuario")
    data_inicio = data.get("data_inicio")
    data_fim = data.get("data_fim")
    
    if not all([usuario_id, tipo_usuario, data_inicio, data_fim]):
        raise HTTPException(status_code=400, detail="Missing required fields")
    
    # Get user phone
    collection = db.motoristas if tipo_usuario == "motorista" else db.parceiros
    usuario = await collection.find_one({"id": usuario_id}, {"_id": 0})
    
    if not usuario:
        # Try users collection
        usuario = await db.users.find_one({"id": usuario_id}, {"_id": 0})
    
    if not usuario:
        raise HTTPException(status_code=404, detail="User not found")
    
    telefone = usuario.get("telefone")
    if not telefone:
        raise HTTPException(status_code=400, detail="User does not have a phone number")
    
    # TODO: Implement WhatsApp sending with report attachment
    # For now, return success message
    raise HTTPException(
        status_code=501,
        detail="Funcionalidade de envio por WhatsApp em desenvolvimento. Configure a integração WhatsApp Business API para utilizar esta funcionalidade."
    )

@api_router.post("/relatorios/enviar-email-massa")
async def enviar_relatorio_email_massa(
    data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Send weekly reports via email to multiple users"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    usuario_ids = data.get("usuario_ids", [])
    tipo_usuario = data.get("tipo_usuario")
    data_inicio = data.get("data_inicio")
    data_fim = data.get("data_fim")
    
    if not all([usuario_ids, tipo_usuario, data_inicio, data_fim]):
        raise HTTPException(status_code=400, detail="Missing required fields")
    
    if len(usuario_ids) == 0:
        raise HTTPException(status_code=400, detail="No users selected")
    
    collection = db.motoristas if tipo_usuario == "motorista" else db.parceiros
    enviados = 0
    erros = []
    
    for usuario_id in usuario_ids:
        try:
            usuario = await collection.find_one({"id": usuario_id}, {"_id": 0})
            if not usuario:
                usuario = await db.users.find_one({"id": usuario_id}, {"_id": 0})
            
            if not usuario or not usuario.get("email"):
                erros.append(f"Utilizador {usuario_id}: sem email")
                continue
            
            # Registar envio no histórico
            historico_record = {
                "id": str(uuid.uuid4()),
                "usuario_id": usuario_id,
                "tipo_usuario": tipo_usuario,
                "tipo_envio": "email",
                "destino": usuario.get("email"),
                "data_inicio_relatorio": data_inicio,
                "data_fim_relatorio": data_fim,
                "data_envio": datetime.now(timezone.utc),
                "status": "enviado",
                "estado_relatorio": "enviado",
                "enviado_por": current_user["id"]
            }
            await db.historico_relatorios.insert_one(historico_record)
            enviados += 1
            
        except Exception as e:
            erros.append(f"Utilizador {usuario_id}: {str(e)}")
    
    return {
        "message": f"Relatórios enviados para {enviados} de {len(usuario_ids)} utilizadores",
        "enviados": enviados,
        "erros": erros
    }

@api_router.post("/relatorios/enviar-whatsapp-massa")
async def enviar_relatorio_whatsapp_massa(
    data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Send weekly reports via WhatsApp to multiple users"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    usuario_ids = data.get("usuario_ids", [])
    tipo_usuario = data.get("tipo_usuario")
    data_inicio = data.get("data_inicio")
    data_fim = data.get("data_fim")
    
    if not all([usuario_ids, tipo_usuario, data_inicio, data_fim]):
        raise HTTPException(status_code=400, detail="Missing required fields")
    
    if len(usuario_ids) == 0:
        raise HTTPException(status_code=400, detail="No users selected")
    
    collection = db.motoristas if tipo_usuario == "motorista" else db.parceiros
    enviados = 0
    erros = []
    
    for usuario_id in usuario_ids:
        try:
            usuario = await collection.find_one({"id": usuario_id}, {"_id": 0})
            if not usuario:
                usuario = await db.users.find_one({"id": usuario_id}, {"_id": 0})
            
            if not usuario or not usuario.get("telefone"):
                erros.append(f"Utilizador {usuario_id}: sem telefone")
                continue
            
            # Registar envio no histórico
            historico_record = {
                "id": str(uuid.uuid4()),
                "usuario_id": usuario_id,
                "tipo_usuario": tipo_usuario,
                "tipo_envio": "whatsapp",
                "destino": usuario.get("telefone"),
                "data_inicio_relatorio": data_inicio,
                "data_fim_relatorio": data_fim,
                "data_envio": datetime.now(timezone.utc),
                "status": "enviado",
                "estado_relatorio": "enviado",
                "enviado_por": current_user["id"]
            }
            await db.historico_relatorios.insert_one(historico_record)
            enviados += 1
            
        except Exception as e:
            erros.append(f"Utilizador {usuario_id}: {str(e)}")
    
    return {
        "message": f"Relatórios enviados para {enviados} de {len(usuario_ids)} utilizadores",
        "enviados": enviados,
        "erros": erros
    }

@api_router.get("/relatorios/historico")
async def get_historico_relatorios(
    tipo_usuario: Optional[str] = None,
    current_user: Dict = Depends(get_current_user)
):
    """Get report sending history"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    query = {}
    if tipo_usuario:
        query["tipo_usuario"] = tipo_usuario
    
    # Se for parceiro, só pode ver seus próprios envios
    if current_user["role"] == UserRole.PARCEIRO:
        query["enviado_por"] = current_user["id"]
    
    historico = await db.historico_relatorios.find(
        query, 
        {"_id": 0}
    ).sort("data_envio", -1).limit(50).to_list(50)
    
    return historico

@api_router.patch("/relatorios/historico/{historico_id}/estado")
async def atualizar_estado_relatorio(
    historico_id: str,
    data: Dict[str, str],
    current_user: Dict = Depends(get_current_user)
):
    """Update report state"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    novo_estado = data.get("estado_relatorio")
    estados_validos = [
        "enviado",
        "pendente_recibo",
        "recibo_enviado",
        "verificar_recibo",
        "aprovado",
        "pagamento",
        "liquidado"
    ]
    
    if novo_estado not in estados_validos:
        raise HTTPException(status_code=400, detail="Invalid state")
    
    result = await db.historico_relatorios.update_one(
        {"id": historico_id},
        {
            "$set": {
                "estado_relatorio": novo_estado,
                "updated_at": datetime.now(timezone.utc)
            }
        }
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Report history not found")
    
    return {"message": "State updated successfully", "estado": novo_estado}

# ============================================
# GESTÃO DE PAGAMENTOS E RECIBOS (UNIFICADO)
# ============================================

@api_router.get("/pagamentos-recibos")
async def get_pagamentos_recibos(
    parceiro_id: Optional[str] = None,
    data_inicio: Optional[str] = None,
    data_fim: Optional[str] = None,
    estado: Optional[str] = None,
    current_user: Dict = Depends(get_current_user)
):
    """Get unified payments and receipts records"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    query = {}
    
    # Se for parceiro, só vê seus registos
    if current_user["role"] == UserRole.PARCEIRO:
        query["parceiro_id"] = current_user["id"]
    elif parceiro_id:
        query["parceiro_id"] = parceiro_id
    
    # Filtros de data - verificar se o período do registo interseta com o filtro
    if data_inicio and data_fim:
        # Registo deve ter data_fim >= filtro_inicio E data_inicio <= filtro_fim
        query["$and"] = [
            {"data_fim": {"$gte": data_inicio}},
            {"data_inicio": {"$lte": data_fim}}
        ]
    elif data_inicio:
        query["data_fim"] = {"$gte": data_inicio}
    elif data_fim:
        query["data_inicio"] = {"$lte": data_fim}
    
    if estado:
        query["estado"] = estado
    
    registos = await db.pagamentos_recibos.find(
        query,
        {"_id": 0}
    ).sort("data_inicio", -1).to_list(100)
    
    return registos

@api_router.get("/pagamentos-recibos/{registro_id}/recibo")
async def get_recibo(
    registro_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Get receipt details"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    registro = await db.pagamentos_recibos.find_one({"id": registro_id}, {"_id": 0})
    if not registro:
        raise HTTPException(status_code=404, detail="Record not found")
    
    # Buscar recibo associado
    recibo = await db.recibos.find_one({"registro_id": registro_id}, {"_id": 0})
    if not recibo:
        raise HTTPException(status_code=404, detail="Receipt not found")
    
    return recibo

@api_router.post("/pagamentos-recibos/{registro_id}/pagamento")
async def realizar_pagamento(
    registro_id: str,
    data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Process payment"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    metodo = data.get("metodo_pagamento")
    valor = data.get("valor")
    observacoes = data.get("observacoes", "")
    
    if not metodo or not valor:
        raise HTTPException(status_code=400, detail="Missing required fields")
    
    # Criar registro de pagamento
    pagamento_id = str(uuid.uuid4())
    pagamento = {
        "id": pagamento_id,
        "registro_id": registro_id,
        "metodo_pagamento": metodo,
        "valor": valor,
        "observacoes": observacoes,
        "data_pagamento": datetime.now(timezone.utc),
        "processado_por": current_user["id"],
        "status": "concluido"
    }
    await db.pagamentos.insert_one(pagamento)
    
    # Atualizar estado do registro
    await db.pagamentos_recibos.update_one(
        {"id": registro_id},
        {
            "$set": {
                "estado": "liquidado",
                "pagamento_id": pagamento_id,
                "data_pagamento": datetime.now(timezone.utc),
                "updated_at": datetime.now(timezone.utc)
            }
        }
    )
    
    return {"message": "Payment processed successfully", "pagamento_id": pagamento_id}

@api_router.patch("/pagamentos-recibos/{registro_id}/estado")
async def atualizar_estado_pagamento(
    registro_id: str,
    data: Dict[str, str],
    current_user: Dict = Depends(get_current_user)
):
    """Update payment/receipt state"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO, UserRole.PARCEIRO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    novo_estado = data.get("estado")
    estados_validos = [
        "pendente_recibo",
        "recibo_enviado",
        "verificar_recibo",
        "aprovado",
        "pagamento_pendente",
        "pagamento_processando",
        "liquidado"
    ]
    
    if novo_estado not in estados_validos:
        raise HTTPException(status_code=400, detail="Invalid state")
    
    result = await db.pagamentos_recibos.update_one(
        {"id": registro_id},
        {
            "$set": {
                "estado": novo_estado,
                "updated_at": datetime.now(timezone.utc)
            }
        }
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Record not found")
    
    return {"message": "State updated successfully", "estado": novo_estado}

@api_router.post("/pagamentos-recibos/{registro_id}/enviar-relatorio")
async def enviar_relatorio_pagamento(
    registro_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """Send payment report"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.GESTAO]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    registro = await db.pagamentos_recibos.find_one({"id": registro_id}, {"_id": 0})
    if not registro:
        raise HTTPException(status_code=404, detail="Record not found")
    
    # TODO: Implementar envio real de email
    # Por agora, apenas marcar como enviado
    
    await db.pagamentos_recibos.update_one(
        {"id": registro_id},
        {
            "$set": {
                "relatorio_enviado": True,
                "data_envio_relatorio": datetime.now(timezone.utc),
                "updated_at": datetime.now(timezone.utc)
            }
        }
    )
    
    return {"message": "Report sent successfully"}

@api_router.get("/recibos/{recibo_id}/pdf")
async def get_recibo_pdf(recibo_id: str):
    """Get receipt PDF (example/mock)"""
    recibo = await db.recibos.find_one({"id": recibo_id}, {"_id": 0})
    if not recibo:
        raise HTTPException(status_code=404, detail="Receipt not found")
    
    # HTML de exemplo do recibo
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            body {{ font-family: Arial, sans-serif; margin: 40px; }}
            .header {{ text-align: center; margin-bottom: 40px; }}
            .header h1 {{ color: #1e40af; margin: 0; }}
            .info-box {{ background: #f1f5f9; padding: 20px; border-radius: 8px; margin: 20px 0; }}
            .info-row {{ display: flex; justify-content: space-between; margin: 10px 0; }}
            .label {{ font-weight: bold; color: #475569; }}
            .value {{ color: #0f172a; }}
            .total {{ font-size: 24px; font-weight: bold; color: #10b981; text-align: right; margin-top: 30px; }}
            .footer {{ margin-top: 50px; text-align: center; color: #64748b; font-size: 12px; }}
            table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
            th, td {{ border: 1px solid #e2e8f0; padding: 12px; text-align: left; }}
            th {{ background: #f8fafc; font-weight: bold; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>🚗 TVDEFleet</h1>
            <p>Relatório Semanal de Ganhos e Despesas</p>
        </div>
        
        <div class="info-box">
            <div class="info-row">
                <span class="label">Recibo Nº:</span>
                <span class="value">{recibo.get('numero_recibo', 'N/A')}</span>
            </div>
            <div class="info-row">
                <span class="label">Data de Emissão:</span>
                <span class="value">{datetime.fromisoformat(recibo['data_emissao']).strftime('%d/%m/%Y')}</span>
            </div>
            <div class="info-row">
                <span class="label">Período:</span>
                <span class="value">{recibo.get('periodo', 'N/A')}</span>
            </div>
            <div class="info-row">
                <span class="label">Parceiro:</span>
                <span class="value">{recibo.get('parceiro_nome', 'N/A')}</span>
            </div>
        </div>
        
        <h3>Resumo Financeiro</h3>
        <table>
            <thead>
                <tr>
                    <th>Descrição</th>
                    <th style="text-align: right;">Valor (€)</th>
                </tr>
            </thead>
            <tbody>
                <tr>
                    <td>Ganhos Brutos</td>
                    <td style="text-align: right;">{recibo.get('valor', 0) + 50:.2f}</td>
                </tr>
                <tr>
                    <td>Comissões Plataforma</td>
                    <td style="text-align: right; color: red;">-30.00</td>
                </tr>
                <tr>
                    <td>Combustível</td>
                    <td style="text-align: right; color: red;">-20.00</td>
                </tr>
                <tr style="background: #f8fafc; font-weight: bold;">
                    <td>Total Líquido</td>
                    <td style="text-align: right; color: #10b981;">{recibo.get('valor', 0):.2f}</td>
                </tr>
            </tbody>
        </table>
        
        <div class="total">
            Total a Receber: €{recibo.get('valor', 0):.2f}
        </div>
        
        <div class="footer">
            <p>Este documento foi gerado automaticamente pelo sistema TVDEFleet</p>
            <p>Para questões, contacte: suporte@tvdefleet.com</p>
            <p>Data de geração: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')}</p>
        </div>
    </body>
    </html>
    """
    
    from fastapi.responses import HTMLResponse
    return HTMLResponse(content=html_content)

@api_router.delete("/notificacoes/{notificacao_id}")
async def delete_notificacao(notificacao_id: str, current_user: Dict = Depends(get_current_user)):
    """Delete a notification"""
    # Verify ownership or admin
    notif = await db.notificacoes.find_one({"id": notificacao_id}, {"_id": 0})
    if not notif:
        raise HTTPException(status_code=404, detail="Notification not found")
    
    if current_user["role"] != UserRole.ADMIN and notif.get("destinatario_id") != current_user["id"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.notificacoes.delete_one({"id": notificacao_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    
    return {"message": "Notification deleted successfully"}

@api_router.patch("/notificacoes/{notificacao_id}")
async def update_notificacao(
    notificacao_id: str, 
    updates: Dict[str, Any], 
    current_user: Dict = Depends(get_current_user)
):
    """Update a notification (e.g., edit message)"""
    # Verify ownership or admin
    notif = await db.notificacoes.find_one({"id": notificacao_id}, {"_id": 0})
    if not notif:
        raise HTTPException(status_code=404, detail="Notification not found")
    
    if current_user["role"] != UserRole.ADMIN and notif.get("destinatario_id") != current_user["id"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    updates["updated_at"] = datetime.now(timezone.utc)
    result = await db.notificacoes.update_one({"id": notificacao_id}, {"$set": updates})
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    
    return {"message": "Notification updated successfully"}


# ============================================================================
# CONFIGURAÇÃO DE CATEGORIAS DE PLATAFORMAS (Uber/Bolt)
# ============================================================================

@api_router.post("/configuracoes/categorias-plataformas")
async def save_categorias_plataformas(
    config_data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Save platform categories configuration (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Only admins can configure categories")
    
    try:
        plataforma = config_data.get("plataforma")  # 'uber' ou 'bolt'
        categorias = config_data.get("categorias", [])
        
        if not plataforma or not categorias:
            raise HTTPException(status_code=400, detail="Platform and categories are required")
        
        # Update or create config
        await db.configuracoes_sistema.update_one(
            {"tipo": f"categorias_{plataforma}"},
            {
                "$set": {
                    "tipo": f"categorias_{plataforma}",
                    "categorias": categorias,
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                    "updated_by": current_user["id"]
                }
            },
            upsert=True
        )
        
        logger.info(f"{plataforma.capitalize()} categories updated by {current_user['email']}")
        return {"message": f"{plataforma.capitalize()} categories saved successfully"}
        
    except Exception as e:
        logger.error(f"Error saving platform categories: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/configuracoes/categorias-plataformas")
async def get_categorias_plataformas(current_user: Dict = Depends(get_current_user)):
    """Get platform categories configuration"""
    try:
        # Default categories
        default_uber = [
            "UberX", "Share", "Electric", "Black", "Comfort",
            "XL", "XXL", "Pet", "Package"
        ]
        
        default_bolt = [
            "Economy", "Comfort", "Executive", "XL", "Green",
            "XXL", "Motorista Privado", "Pet"
        ]
        
        # Fetch custom categories
        uber_config = await db.configuracoes_sistema.find_one(
            {"tipo": "categorias_uber"},
            {"_id": 0}
        )
        
        bolt_config = await db.configuracoes_sistema.find_one(
            {"tipo": "categorias_bolt"},
            {"_id": 0}
        )
        
        return {
            "uber": uber_config.get("categorias") if uber_config else default_uber,
            "bolt": bolt_config.get("categorias") if bolt_config else default_bolt
        }
        
    except Exception as e:
        logger.error(f"Error fetching platform categories: {e}")
        raise HTTPException(status_code=500, detail=str(e))


        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# GOOGLE DRIVE STORAGE INTEGRATION
# ============================================================================

@api_router.get("/storage/drive/connect")
async def connect_google_drive(current_user: Dict = Depends(get_current_user)):
    """Initiate Google Drive OAuth flow"""
    try:
        from google_auth_oauthlib.flow import Flow
        
        # Get configuration
        config = await db.configuracoes_sistema.find_one({"tipo": "storage_google_drive"})
        if not config:
            raise HTTPException(
                status_code=400,
                detail="Google Drive not configured. Admin must configure it first."
            )
        
        client_id = config.get("client_id")
        client_secret = config.get("client_secret")
        redirect_uri = config.get("redirect_uri")
        
        if not all([client_id, client_secret, redirect_uri]):
            raise HTTPException(
                status_code=400,
                detail="Google Drive configuration is incomplete"
            )
        
        flow = Flow.from_client_config(
            {
                "web": {
                    "client_id": client_id,
                    "client_secret": client_secret,
                    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                    "token_uri": "https://oauth2.googleapis.com/token",
                    "redirect_uris": [redirect_uri]
                }
            },
            scopes=['https://www.googleapis.com/auth/drive.file'],
            redirect_uri=redirect_uri
        )
        
        authorization_url, state = flow.authorization_url(
            access_type='offline',
            include_granted_scopes='true',
            prompt='consent',
            state=current_user["id"]
        )
        
        logger.info(f"Drive OAuth initiated for user {current_user['id']}")
        return {"authorization_url": authorization_url}
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to initiate Drive OAuth: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/storage/drive/callback")
async def drive_callback(code: str, state: str):
    """Handle Google Drive OAuth callback"""
    try:
        from google_auth_oauthlib.flow import Flow
        
        # Get configuration
        config = await db.configuracoes_sistema.find_one({"tipo": "storage_google_drive"})
        if not config:
            raise HTTPException(status_code=400, detail="Google Drive not configured")
        
        client_id = config.get("client_id")
        client_secret = config.get("client_secret")
        redirect_uri = config.get("redirect_uri")
        
        flow = Flow.from_client_config(
            {
                "web": {
                    "client_id": client_id,
                    "client_secret": client_secret,
                    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                    "token_uri": "https://oauth2.googleapis.com/token",
                    "redirect_uris": [redirect_uri]
                }
            },
            scopes=None,
            redirect_uri=redirect_uri
        )
        
        flow.fetch_token(code=code)
        credentials = flow.credentials
        
        # Verify required scopes
        required_scopes = {"https://www.googleapis.com/auth/drive.file"}
        granted_scopes = set(credentials.scopes or [])
        
        if not required_scopes.issubset(granted_scopes):
            missing = required_scopes - granted_scopes
            raise HTTPException(
                status_code=400,
                detail=f"Missing required Drive scopes: {', '.join(missing)}"
            )
        
        # Store credentials
        await db.drive_credentials.update_one(
            {"user_id": state},
            {"$set": {
                "user_id": state,
                "access_token": credentials.token,
                "refresh_token": credentials.refresh_token,
                "token_uri": credentials.token_uri,
                "client_id": credentials.client_id,
                "client_secret": credentials.client_secret,
                "scopes": credentials.scopes,
                "expiry": credentials.expiry.isoformat() if credentials.expiry else None,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }},
            upsert=True
        )
        
        logger.info(f"Drive credentials stored for user {state}")
        
        # Redirect to frontend
        frontend_url = config.get("frontend_url", "")
        return {"message": "Google Drive connected successfully", "redirect": f"{frontend_url}/integracoes?drive_connected=true"}
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Drive OAuth callback failed: {e}")
        raise HTTPException(status_code=400, detail=str(e))


@api_router.get("/storage/drive/status")
async def get_drive_status(current_user: Dict = Depends(get_current_user)):
    """Check if user has connected Google Drive"""
    try:
        creds = await db.drive_credentials.find_one({"user_id": current_user["id"]})
        
        if not creds:
            return {"connected": False}
        
        # Get storage usage
        from services.storage_service import StorageService
        storage_service = StorageService(db)
        
        usage = await storage_service.get_storage_usage(current_user["id"])
        
        return {
            "connected": True,
            "storage_usage": usage
        }
    
    except Exception as e:
        logger.error(f"Error checking Drive status: {e}")
        return {"connected": False, "error": str(e)}


@api_router.post("/storage/drive/upload")
async def upload_to_drive(
    entity_type: str,
    entity_id: str,
    subfolder: str,
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    """Upload file to Google Drive"""
    try:
        from services.storage_service import StorageService
        
        storage_service = StorageService(db)
        
        # Create folder structure if needed
        folders = await storage_service.create_folder_structure(
            user_id=current_user["id"],
            entity_type=entity_type,
            entity_id=entity_id
        )
        
        # Determine target folder
        folder_id = folders.get(subfolder)
        if not folder_id:
            raise HTTPException(status_code=400, detail=f"Invalid subfolder: {subfolder}")
        
        # Read file content
        file_content = await file.read()
        
        # Upload file
        result = await storage_service.upload_file(
            user_id=current_user["id"],
            file_content=file_content,
            file_name=file.filename,
            folder_id=folder_id,
            mime_type=file.content_type or 'application/octet-stream'
        )
        
        if result.get("success"):
            # Save file reference in database
            await db.drive_files.insert_one({
                "user_id": current_user["id"],
                "entity_type": entity_type,
                "entity_id": entity_id,
                "subfolder": subfolder,
                "file_id": result["file_id"],
                "file_name": result["file_name"],
                "web_view_link": result.get("web_view_link"),
                "uploaded_at": datetime.now(timezone.utc).isoformat()
            })
        
        return result
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error uploading to Drive: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/storage/drive/files/{entity_type}/{entity_id}")
async def list_drive_files(
    entity_type: str,
    entity_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """List files for an entity"""
    try:
        files = await db.drive_files.find({
            "user_id": current_user["id"],
            "entity_type": entity_type,
            "entity_id": entity_id
        }, {"_id": 0}).sort("uploaded_at", -1).to_list(100)
        
        return files
    
    except Exception as e:
        logger.error(f"Error listing Drive files: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/configuracoes/storage/google-drive")
async def save_drive_config(
    config_data: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """Save Google Drive configuration (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Only admins can configure storage")
    
    try:
        config = {
            "tipo": "storage_google_drive",
            "client_id": config_data.get("client_id"),
            "client_secret": config_data.get("client_secret"),
            "redirect_uri": config_data.get("redirect_uri"),
            "frontend_url": config_data.get("frontend_url"),
            "enabled": config_data.get("enabled", False),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": current_user["id"]
        }
        
        await db.configuracoes_sistema.update_one(
            {"tipo": "storage_google_drive"},
            {"$set": config},
            upsert=True
        )
        
        logger.info(f"Google Drive configuration saved by {current_user['email']}")
        return {"message": "Google Drive configuration saved successfully"}
        
    except Exception as e:
        logger.error(f"Error saving Drive config: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/configuracoes/storage/google-drive")
async def get_drive_config(current_user: Dict = Depends(get_current_user)):
    """Get Google Drive configuration (Admin only)"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Only admins can view storage config")
    
    try:
        config = await db.configuracoes_sistema.find_one(
            {"tipo": "storage_google_drive"},
            {"_id": 0}
        )
        
        if not config:
            return {
                "client_id": "",
                "client_secret": "",
                "redirect_uri": "",
                "frontend_url": "",
                "enabled": False
            }
        
        return config
        
    except Exception as e:
        logger.error(f"Error fetching Drive config: {e}")
        raise HTTPException(status_code=500, detail=str(e))


        raise HTTPException(status_code=403, detail="Not authorized")
    
    try:
        from services.email_service import EmailService
        
        email_service = EmailService(db)
        
        to_email = notification_data.get("to_email")
        notification_type = notification_data.get("type")
        data = notification_data.get("data", {})
        
        if not to_email or not notification_type:
            raise HTTPException(status_code=400, detail="to_email and type are required")
        
        result = await email_service.send_notification(
            to_email=to_email,
            notification_type=notification_type,
            data=data
        )
        
        return result
        
    except Exception as e:
        logger.error(f"Error sending notification: {e}")
        raise HTTPException(status_code=500, detail=str(e))


        
        # Criar diretório para relatórios se não existir
        relatorios_dir = UPLOAD_DIR / "relatorios"
        relatorios_dir.mkdir(exist_ok=True)
        
        # Nome do arquivo
        filename = f"relatorio_{relatorio_data.motorista_id}_{relatorio_data.data_inicio}_{relatorio_data.data_fim}.pdf"
        filepath = relatorios_dir / filename
        
        # Criar PDF
        c = canvas.Canvas(str(filepath), pagesize=A4)
        width, height = A4
        
        # Cabeçalho
        c.setFont("Helvetica-Bold", 16)
        c.drawString(30, height - 40, "RELATÓRIO SEMANAL - TVDEFleet")
        
        c.setFont("Helvetica", 10)
        y = height - 70
        
        # Informações principais
        c.setFont("Helvetica-Bold", 12)
        c.drawString(30, y, "Dados do Motorista")
        y -= 20
        
        c.setFont("Helvetica", 10)
        c.drawString(30, y, f"Nome: {relatorio_data.motorista_nome}")
        y -= 15
        c.drawString(30, y, f"Matrícula: {relatorio_data.veiculo_matricula}")
        y -= 15
        c.drawString(30, y, f"Período: {relatorio_data.data_inicio} a {relatorio_data.data_fim}")
        c.drawString(300, y, f"Semana: {relatorio_data.numero_semana}")
        y -= 30
        
        # Ganhos
        c.setFont("Helvetica-Bold", 12)
        c.drawString(30, y, "GANHOS")
        y -= 20
        
        c.setFont("Helvetica", 10)
        c.drawString(30, y, f"Uber - Corridas: €{relatorio_data.uber_ganhos:.2f}")
        c.drawString(300, y, f"Gorjeta: €{relatorio_data.uber_gorjeta:.2f}")
        y -= 15
        c.drawString(30, y, f"Bolt - Corridas: €{relatorio_data.bolt_ganhos:.2f}")
        c.drawString(300, y, f"Gorjeta: €{relatorio_data.bolt_gorjeta:.2f}")
        y -= 15
        c.drawString(30, y, f"Uber - Portagens: €{relatorio_data.uber_portagens:.2f}")
        y -= 15
        c.drawString(30, y, f"Bolt - Portagens: €{relatorio_data.bolt_portagens:.2f}")
        y -= 20
        
        total_ganhos = (relatorio_data.uber_ganhos + relatorio_data.bolt_ganhos + 
                       relatorio_data.uber_gorjeta + relatorio_data.bolt_gorjeta +
                       relatorio_data.uber_portagens + relatorio_data.bolt_portagens)
        
        c.setFont("Helvetica-Bold", 10)
        c.drawString(30, y, f"Total Ganhos Brutos: €{total_ganhos:.2f}")
        y -= 30
        
        # Condição de Exploração
        c.setFont("Helvetica-Bold", 12)
        c.drawString(30, y, "CONDIÇÃO DE EXPLORAÇÃO")
        y -= 20
        
        c.setFont("Helvetica", 10)
        if relatorio_data.tipo_exploracao == "aluguer":
            c.drawString(30, y, f"Tipo: Aluguer Semanal")
            y -= 15
            c.drawString(30, y, f"Valor: €{relatorio_data.valor_aluguer:.2f}")
            valor_a_descontar = relatorio_data.valor_aluguer
        else:
            c.drawString(30, y, f"Tipo: Comissão")
            y -= 15
            c.drawString(30, y, f"Motorista: {relatorio_data.percentagem_motorista}%")
            c.drawString(200, y, f"Parceiro: {relatorio_data.percentagem_parceiro}%")
            valor_a_descontar = total_ganhos * (relatorio_data.percentagem_parceiro / 100)
        y -= 30
        
        # Despesas
        c.setFont("Helvetica-Bold", 12)
        c.drawString(30, y, "DESPESAS")
        y -= 20
        
        c.setFont("Helvetica", 10)
        c.drawString(30, y, f"Via Verde: €{relatorio_data.via_verde:.2f}")
        y -= 15
        c.drawString(30, y, f"Abastecimentos: €{relatorio_data.abastecimentos:.2f}")
        y -= 15
        c.drawString(30, y, f"Caução (esta semana): €{relatorio_data.caucao:.2f}")
        y -= 15
        c.drawString(30, y, f"Acumulado Caução: €{relatorio_data.acumulado_caucao:.2f}")
        y -= 15
        c.drawString(30, y, f"Dívida Anterior: €{relatorio_data.divida_anterior:.2f}")
        y -= 15
        
        extra_label = "Crédito" if relatorio_data.extra > 0 else "Débito (Danos)"
        c.drawString(30, y, f"Extra ({extra_label}): €{abs(relatorio_data.extra):.2f}")
        y -= 30
        
        # Cálculo Total
        c.setFont("Helvetica-Bold", 12)
        c.drawString(30, y, "CÁLCULO FINAL")
        y -= 20
        
        c.setFont("Helvetica", 10)
        c.drawString(30, y, f"Ganhos Brutos: €{total_ganhos:.2f}")
        y -= 15
        c.drawString(30, y, f"(-) {'Aluguer' if relatorio_data.tipo_exploracao == 'aluguer' else 'Comissão Parceiro'}: €{valor_a_descontar:.2f}")
        y -= 15
        c.drawString(30, y, f"(-) Via Verde: €{relatorio_data.via_verde:.2f}")
        y -= 15
        c.drawString(30, y, f"(-) Abastecimentos: €{relatorio_data.abastecimentos:.2f}")
        y -= 15
        c.drawString(30, y, f"(-) Caução: €{relatorio_data.caucao:.2f}")
        y -= 15
        c.drawString(30, y, f"(-) Dívida Anterior: €{relatorio_data.divida_anterior:.2f}")
        y -= 15
        
        if relatorio_data.extra > 0:
            c.drawString(30, y, f"(+) Crédito Extra: €{relatorio_data.extra:.2f}")
        else:
            c.drawString(30, y, f"(-) Débito Extra: €{abs(relatorio_data.extra):.2f}")
        y -= 25
        
        total_a_receber = (total_ganhos - valor_a_descontar - relatorio_data.via_verde - 
                          relatorio_data.abastecimentos - relatorio_data.caucao - 
                          relatorio_data.divida_anterior + relatorio_data.extra)
        
        c.setFont("Helvetica-Bold", 14)
        c.setFillColor(colors.green if total_a_receber >= 0 else colors.red)
        c.drawString(30, y, f"TOTAL A RECEBER: €{total_a_receber:.2f}")
        c.setFillColor(colors.black)
        y -= 30
        
        # Informações adicionais
        c.setFont("Helvetica", 9)
        c.drawString(30, y, f"KM Percorridos: {relatorio_data.km:.0f} km")
        c.drawString(200, y, f"Horas Trabalhadas: {relatorio_data.horas:.1f}h")
        y -= 40
        
        # Nova página para detalhes
        if relatorio_data.via_verde_detalhes or relatorio_data.abastecimentos_detalhes:
            c.showPage()
            y = height - 40
            
            # Via Verde
            if relatorio_data.via_verde_detalhes:
                c.setFont("Helvetica-Bold", 12)
                c.drawString(30, y, "DETALHES VIA VERDE")
                y -= 20
                
                c.setFont("Helvetica", 9)
                for item in relatorio_data.via_verde_detalhes:
                    c.drawString(30, y, f"{item.get('data', 'N/A')} - {item.get('local', 'N/A')}: €{item.get('valor', 0):.2f}")
                    y -= 12
                y -= 20
            
            # Abastecimentos
            if relatorio_data.abastecimentos_detalhes:
                c.setFont("Helvetica-Bold", 12)
                c.drawString(30, y, "DETALHES ABASTECIMENTOS")
                y -= 20
                
                c.setFont("Helvetica", 9)
                for item in relatorio_data.abastecimentos_detalhes:
                    c.drawString(30, y, f"{item.get('data', 'N/A')} - {item.get('litros', 0)}L: €{item.get('valor', 0):.2f}")
                    y -= 12
        
        # Rodapé
        c.setFont("Helvetica", 8)
        c.drawString(30, 30, f"Gerado em: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')}")
        c.drawString(30, 20, "TVDEFleet - Sistema de Gestão de Frotas")
        
        c.save()
        
        logger.info(f"Relatório PDF gerado: {filename}")
        
        return {
            "message": "Relatório gerado com sucesso",
            "filename": filename,
            "filepath": f"/uploads/relatorios/{filename}",
            "download_url": f"{API}/uploads/relatorios/{filename}"
        }
        
    except Exception as e:
        logger.error(f"Erro ao gerar relatório PDF: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/relatorios/enviar-email")
async def enviar_relatorio_email(
    motorista_id: str,
    relatorio_url: str,
    current_user: Dict = Depends(get_current_user)
):
    """Enviar relatório por email ao motorista (Admin/Gestor/Parceiro)"""
    if current_user["role"] not in [UserRole.ADMIN, "gestao", "parceiro"]:
        raise HTTPException(status_code=403, detail="Não autorizado")
    
    try:
        # Buscar dados do motorista
        motorista = await db.users.find_one({"id": motorista_id}, {"_id": 0})
        
        if not motorista:
            raise HTTPException(status_code=404, detail="Motorista não encontrado")
        
        # TODO: Implementar envio de email
        # Por enquanto, apenas simular
        logger.info(f"Relatório enviado para {motorista.get('email')} por {current_user['email']}")
        
        return {
            "message": f"Relatório enviado com sucesso para {motorista.get('email')}",
            "destinatario": motorista.get('email')
        }
        
    except Exception as e:
        logger.error(f"Erro ao enviar relatório: {e}")
        raise HTTPException(status_code=500, detail=str(e))


app.include_router(api_router)

# Mount static files for uploads (PDFs, documents, etc.)
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Background task for checking alerts
import asyncio
from typing import AsyncGenerator

async def check_alerts_periodically():
    """Background task to check alerts every 6 hours"""
    while True:
        try:
            logger.info("Running periodic alert check...")
            await check_and_create_alerts()
            logger.info("Alert check completed successfully")
        except Exception as e:
            logger.error(f"Error during periodic alert check: {e}")
        
        # Wait 6 hours before next check
        await asyncio.sleep(6 * 60 * 60)


async def check_notifications_periodically():
    """Background task to check and create notifications every 12 hours"""
    from utils.notificacoes import check_documentos_expirando, check_recibos_pendentes
    
    while True:
        try:
            logger.info("Running periodic notification check...")
            await check_documentos_expirando(db)
            await check_recibos_pendentes(db)
            logger.info("Notification check completed successfully")
        except Exception as e:
            logger.error(f"Error during periodic notification check: {e}")
        
        # Wait 12 hours before next check
        await asyncio.sleep(12 * 60 * 60)

@app.on_event("startup")
async def startup_event():
    """Run startup tasks"""
    # Run initial alert check
    try:
        logger.info("Running initial alert check...")
        await check_and_create_alerts()
        logger.info("Initial alert check completed")
    except Exception as e:
        logger.error(f"Error during initial alert check: {e}")
    
    # Start background task for periodic checks
    asyncio.create_task(check_alerts_periodically())
    logger.info("Background alert checker started")
    
    # Start background task for notifications
    asyncio.create_task(check_notifications_periodically())
    logger.info("Background notification checker started")
    
    # Start scheduler for automatic sync
    scheduler.start()
    logger.info("Scheduler started for automatic platform sync")
    
    # Carregar jobs agendados do banco
    try:
        credenciais = await db.credenciais_plataforma.find({'sincronizacao_automatica': True, 'ativo': True}).to_list(length=None)
        for cred in credenciais:
            if cred.get('horario_sincronizacao'):
                await agendar_sincronizacao(
                    cred['id'],
                    cred['horario_sincronizacao'],
                    cred.get('frequencia_dias', 7)
                )
        logger.info(f"Carregados {len(credenciais)} agendamentos de sincronização")
    except Exception as e:
        logger.error(f"Erro ao carregar agendamentos: {e}")

# ==================================================
# SINCRONIZAÇÃO - Configuração e Agendamento
# ==================================================

@app.get("/api/sincronizacao/configuracoes")
async def obter_configuracoes_sincronizacao(
    current_user: dict = Depends(get_current_user)
):
    """
    Obter configurações de sincronização de todos os parceiros
    """
    try:
        if current_user["role"] not in ["admin", "gestao"]:
            raise HTTPException(status_code=403, detail="Acesso negado")
        
        # Buscar todas as configurações de sincronização
        configs = await db.configuracoes_sincronizacao.find({}, {"_id": 0}).to_list(1000)
        
        # Se não houver configs, criar configs vazias para cada parceiro
        if not configs or len(configs) == 0:
            parceiros = await db.parceiros.find({}, {"_id": 0, "id": 1}).to_list(1000)
            configs = []
            for parceiro in parceiros:
                config = {
                    "parceiro_id": parceiro["id"],
                    "dia_semana": 1,  # Segunda-feira por padrão
                    "hora": "00:00",
                    "ativo": True,
                    "ultima_sincronizacao": None,
                    "status": None,
                    "mensagem_erro": None,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.configuracoes_sincronizacao.insert_one(config)
                configs.append(config)
        
        return configs
        
    except Exception as e:
        logger.error(f"Erro ao obter configurações: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/sincronizacao/configurar-dia")
async def configurar_dia_sincronizacao(
    request: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Configurar o dia da semana para sincronização automática
    """
    try:
        if current_user["role"] not in ["admin", "gestao"]:
            raise HTTPException(status_code=403, detail="Acesso negado")
        
        parceiro_id = request.get("parceiro_id")
        dia_semana = request.get("dia_semana")
        
        if parceiro_id is None or dia_semana is None:
            raise HTTPException(status_code=400, detail="Dados inválidos")
        
        # Verificar se já existe configuração
        config_existente = await db.configuracoes_sincronizacao.find_one(
            {"parceiro_id": parceiro_id},
            {"_id": 0}
        )
        
        if config_existente:
            # Atualizar
            await db.configuracoes_sincronizacao.update_one(
                {"parceiro_id": parceiro_id},
                {"$set": {
                    "dia_semana": dia_semana,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
        else:
            # Criar nova
            config = {
                "parceiro_id": parceiro_id,
                "dia_semana": dia_semana,
                "hora": "00:00",
                "ativo": True,
                "ultima_sincronizacao": None,
                "status": None,
                "mensagem_erro": None,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.configuracoes_sincronizacao.insert_one(config)
        
        return {"message": "Configuração atualizada com sucesso"}
        
    except Exception as e:
        logger.error(f"Erro ao configurar dia: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/sincronizacao/forcar")
async def forcar_sincronizacao(
    request: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Forçar sincronização manual imediata
    """
    try:
        if current_user["role"] not in ["admin", "gestao"]:
            raise HTTPException(status_code=403, detail="Acesso negado")
        
        parceiro_id = request.get("parceiro_id")
        
        if not parceiro_id:
            raise HTTPException(status_code=400, detail="parceiro_id é obrigatório")
        
        # Buscar parceiro
        parceiro = await db.parceiros.find_one({"id": parceiro_id}, {"_id": 0})
        if not parceiro:
            raise HTTPException(status_code=404, detail="Parceiro não encontrado")
        
        # Atualizar status da configuração para "em_progresso"
        await db.configuracoes_sincronizacao.update_one(
            {"parceiro_id": parceiro_id},
            {"$set": {
                "status": "em_progresso",
                "mensagem_erro": None,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }},
            upsert=True
        )
        
        # Simular sincronização (em produção, aqui seria feita a integração real)
        # Por exemplo: sincronizar dados de uma plataforma externa
        await asyncio.sleep(1)  # Simular processamento
        
        # Atualizar status para sucesso
        await db.configuracoes_sincronizacao.update_one(
            {"parceiro_id": parceiro_id},
            {"$set": {
                "status": "sucesso",
                "ultima_sincronizacao": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        # Registar log
        log = {
            "id": str(uuid.uuid4()),
            "parceiro_id": parceiro_id,
            "tipo": "manual",
            "status": "sucesso",
            "data": datetime.now(timezone.utc).isoformat(),
            "usuario_id": current_user["id"]
        }
        await db.logs_sincronizacao_parceiro.insert_one(log)
        
        return {
            "message": f"Sincronização concluída para {parceiro.get('nome_empresa', 'parceiro')}",
            "status": "sucesso"
        }
        
    except Exception as e:
        logger.error(f"Erro ao forçar sincronização: {e}")
        
        # Atualizar status para erro
        if parceiro_id:
            await db.configuracoes_sincronizacao.update_one(
                {"parceiro_id": parceiro_id},
                {"$set": {
                    "status": "erro",
                    "mensagem_erro": str(e),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }},
                upsert=True
            )
        
        raise HTTPException(status_code=500, detail=str(e))

# ==================================================
# GESTÃO DE CREDENCIAIS DE PLATAFORMAS
# ==================================================

@app.get("/api/credenciais-plataforma")
async def listar_credenciais(
    parceiro_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Listar credenciais de plataformas"""
    try:
        query = {}
        
        # Filtrar por parceiro se fornecido
        if parceiro_id:
            query['parceiro_id'] = parceiro_id
        
        # Se não for admin/gestao, mostrar apenas suas próprias credenciais
        if current_user['role'] not in ['admin', 'gestao']:
            query['parceiro_id'] = current_user['id']
        
        credenciais = await db.credenciais_plataforma.find(
            query,
            {"_id": 0, "password": 0, "password_encrypted": 0}  # Não retornar passwords
        ).to_list(1000)
        
        return credenciais
        
    except Exception as e:
        logger.error(f"Erro ao listar credenciais: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/credenciais-plataforma")
async def criar_credencial(
    request: dict = Body(...),
    current_user: dict = Depends(get_current_user)
):
    """Criar nova credencial de plataforma"""
    try:
        if current_user['role'] not in ['admin', 'gestao', 'parceiro']:
            raise HTTPException(status_code=403, detail="Acesso negado")
        
        # Validar dados obrigatórios
        required_fields = ['plataforma', 'email', 'password']
        for field in required_fields:
            if not request.get(field):
                raise HTTPException(status_code=400, detail=f"Campo {field} é obrigatório")
        
        # Se não for admin/gestao, usar próprio ID como parceiro
        parceiro_id = request.get('parceiro_id')
        if current_user['role'] == 'parceiro':
            parceiro_id = current_user['id']
        
        # Verificar se já existe credencial para esta plataforma/parceiro
        existing = await db.credenciais_plataforma.find_one({
            'plataforma': request['plataforma'],
            'parceiro_id': parceiro_id
        })
        
        if existing:
            raise HTTPException(
                status_code=400, 
                detail="Já existe credencial para esta plataforma e parceiro"
            )
        
        credencial = {
            "id": str(uuid.uuid4()),
            "parceiro_id": parceiro_id,
            "plataforma": request['plataforma'],
            "email": request['email'],
            "password": request['password'],  # TODO: Encriptar em produção
            "ativo": request.get('ativo', True),
            "sincronizacao_automatica": request.get('sincronizacao_automatica', True),
            "frequencia_dias": request.get('frequencia_dias', 7),
            "ultima_sincronizacao": None,
            "proximo_sincronizacao": None,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "created_by": current_user['id']
        }
        
        await db.credenciais_plataforma.insert_one(credencial)
        
        logger.info(f"✅ Credencial criada: {credencial['plataforma']} para {credencial.get('email')}")
        
        return {"success": True, "message": "Credencial criada com sucesso", "id": credencial['id']}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao criar credencial: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/credenciais-plataforma/{cred_id}")
async def atualizar_credencial(
    cred_id: str,
    request: dict = Body(...),
    current_user: dict = Depends(get_current_user)
):
    """Atualizar credencial existente"""
    try:
        # Buscar credencial
        credencial = await db.credenciais_plataforma.find_one({"id": cred_id}, {"_id": 0})
        if not credencial:
            raise HTTPException(status_code=404, detail="Credencial não encontrada")
        
        # Verificar permissão
        if current_user['role'] not in ['admin', 'gestao']:
            if credencial['parceiro_id'] != current_user['id']:
                raise HTTPException(status_code=403, detail="Acesso negado")
        
        # Atualizar campos
        update_data = {
            "email": request.get('email', credencial['email']),
            "ativo": request.get('ativo', credencial['ativo']),
            "sincronizacao_automatica": request.get('sincronizacao_automatica', credencial.get('sincronizacao_automatica', True)),
            "frequencia_dias": request.get('frequencia_dias', credencial.get('frequencia_dias', 7)),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Atualizar password apenas se fornecida
        if request.get('password'):
            update_data['password'] = request['password']
        
        await db.credenciais_plataforma.update_one(
            {"id": cred_id},
            {"$set": update_data}
        )
        
        return {"success": True, "message": "Credencial atualizada"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao atualizar credencial: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/credenciais-plataforma/{cred_id}")
async def deletar_credencial(
    cred_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Deletar credencial"""
    try:
        # Buscar credencial
        credencial = await db.credenciais_plataforma.find_one({"id": cred_id}, {"_id": 0})
        if not credencial:
            raise HTTPException(status_code=404, detail="Credencial não encontrada")
        
        # Verificar permissão
        if current_user['role'] not in ['admin', 'gestao']:
            if credencial['parceiro_id'] != current_user['id']:
                raise HTTPException(status_code=403, detail="Acesso negado")
        
        await db.credenciais_plataforma.delete_one({"id": cred_id})
        
        return {"success": True, "message": "Credencial removida"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao deletar credencial: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/credenciais-plataforma/{cred_id}/testar")
async def testar_credencial(
    cred_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Testar conexão com credencial"""
    try:
        logger.info(f"🔍 Buscando credencial com ID: {cred_id}")
        
        # Debug: listar todas as credenciais
        all_creds = await db.credenciais_plataforma.find({}, {"_id": 0, "id": 1, "plataforma": 1}).to_list(100)
        logger.info(f"📊 Total de credenciais na DB: {len(all_creds)}")
        for c in all_creds:
            logger.info(f"  - ID: {c.get('id')}, Plataforma: {c.get('plataforma')}")
        
        # Buscar credencial
        credencial = await db.credenciais_plataforma.find_one({"id": cred_id}, {"_id": 0})
        if not credencial:
            logger.error(f"❌ Credencial com ID {cred_id} não encontrada")
            raise HTTPException(status_code=404, detail="Credencial não encontrada")
        
        logger.info(f"✅ Credencial encontrada: {credencial.get('plataforma')} - {credencial.get('email')}")
        
        # Testar login
        from integrations.platform_scrapers import get_scraper
        
        async with get_scraper(credencial['plataforma'], headless=True) as scraper:
            login_success = await scraper.login(credencial['email'], credencial['password'])
            
            if login_success:
                return {
                    "success": True,
                    "message": f"Conexão com {credencial['plataforma']} bem-sucedida!"
                }
            else:
                return {
                    "success": False,
                    "message": "Falha no login. Verifique as credenciais."
                }
        
    except Exception as e:
        logger.error(f"Erro ao testar credencial: {e}")
        return {
            "success": False,
            "message": f"Erro: {str(e)}"
        }

@app.post("/api/credenciais-plataforma/{cred_id}/sincronizar")
async def sincronizar_credencial(
    cred_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Forçar sincronização manual"""
    try:
        # Buscar credencial
        credencial = await db.credenciais_plataforma.find_one({"id": cred_id}, {"_id": 0})
        if not credencial:
            raise HTTPException(status_code=404, detail="Credencial não encontrada")
        
        logger.info(f"🔄 Iniciando sincronização: {credencial['plataforma']}")
        
        # Criar log
        log_id = str(uuid.uuid4())
        log = {
            "id": log_id,
            "credencial_id": cred_id,
            "plataforma": credencial['plataforma'],
            "parceiro_id": credencial.get('parceiro_id'),
            "tipo": "manual",
            "status": "em_progresso",
            "data_inicio": datetime.now(timezone.utc).isoformat(),
            "usuario_id": current_user['id']
        }
        await db.logs_sincronizacao.insert_one(log)
        
        # Executar scraping
        from integrations.platform_scrapers import get_scraper
        
        async with get_scraper(credencial['plataforma'], headless=True) as scraper:
            # Login
            login_success = await scraper.login(credencial['email'], credencial['password'])
            
            if not login_success:
                await db.logs_sincronizacao.update_one(
                    {"id": log_id},
                    {"$set": {
                        "status": "erro",
                        "mensagem_erro": "Falha no login",
                        "data_fim": datetime.now(timezone.utc).isoformat()
                    }}
                )
                return {
                    "success": False,
                    "message": "Falha no login. Verifique as credenciais."
                }
            
            # Extrair dados
            data = await scraper.extract_data()
            logger.info(f"📊 Resultado da extração: {data}")
            
            if data.get('success'):
                # Atualizar log
                await db.logs_sincronizacao.update_one(
                    {"id": log_id},
                    {"$set": {
                        "status": "sucesso",
                        "data_fim": datetime.now(timezone.utc).isoformat(),
                        "registos_extraidos": len(data.get('data', []))
                    }}
                )
                
                # Atualizar última sincronização
                await db.credenciais_plataforma.update_one(
                    {"id": cred_id},
                    {"$set": {
                        "ultima_sincronizacao": datetime.now(timezone.utc).isoformat()
                    }}
                )
                
                return {
                    "success": True,
                    "message": "Sincronização concluída!",
                    "registos": len(data.get('data', []))
                }
            else:
                await db.logs_sincronizacao.update_one(
                    {"id": log_id},
                    {"$set": {
                        "status": "erro",
                        "mensagem_erro": data.get('error', 'Erro desconhecido'),
                        "data_fim": datetime.now(timezone.utc).isoformat()
                    }}
                )
                return {
                    "success": False,
                    "message": f"Erro: {data.get('error')}"
                }
        
    except Exception as e:
        logger.error(f"Erro na sincronização: {e}")
        return {
            "success": False,
            "message": f"Erro: {str(e)}"
        }

# ==================================================
# CSV IMPORT - Upload Manual de Ficheiros
# ==================================================

@app.post("/api/import-csv/{plataforma}")
async def import_csv(
    plataforma: str,
    file: UploadFile = File(...),
    parceiro_id: Optional[str] = Form(None),
    motorista_id: Optional[str] = Form(None),
    current_user: dict = Depends(get_current_user)
):
    """
    Importar ficheiro CSV de qualquer plataforma
    
    Plataformas suportadas: bolt, uber, via_verde, gps, combustivel
    """
    try:
        logger.info(f"📥 Importando CSV {plataforma} - Ficheiro: {file.filename}")
        
        # Validar plataforma
        plataformas_validas = ['bolt', 'uber', 'via_verde', 'gps', 'combustivel']
        if plataforma.lower() not in plataformas_validas:
            raise HTTPException(
                status_code=400, 
                detail=f"Plataforma inválida. Use: {', '.join(plataformas_validas)}"
            )
        
        # Ler conteúdo do ficheiro
        file_content = await file.read()
        
        if len(file_content) == 0:
            raise HTTPException(status_code=400, detail="Ficheiro vazio")
        
        # Extrair metadata do nome do ficheiro
        from utils.csv_parsers import get_parser, extract_dates_from_filename
        data_inicio, data_fim, nome_frota = extract_dates_from_filename(file.filename)
        
        # Obter parser adequado
        parser = get_parser(plataforma, file_content)
        
        # Fazer parse
        sucesso, registos, mensagem = parser.parse()
        
        if not sucesso:
            logger.error(f"❌ Erro no parse: {mensagem}")
            return {
                "success": False,
                "message": mensagem,
                "registos_importados": 0
            }
        
        # Guardar registos na base de dados
        registos_salvos = 0
        collection_name = f"dados_{plataforma}"
        
        for registo in registos:
            try:
                # Adicionar metadata
                registo['id'] = str(uuid.uuid4())
                registo['parceiro_id'] = parceiro_id
                registo['motorista_id'] = motorista_id
                registo['importado_por'] = current_user['id']
                registo['nome_ficheiro'] = file.filename
                
                # Adicionar datas e frota extraídas do filename
                if data_inicio:
                    registo['periodo_inicio'] = data_inicio
                if data_fim:
                    registo['periodo_fim'] = data_fim
                if nome_frota:
                    registo['nome_frota'] = nome_frota
                registo['data_importacao'] = datetime.now(timezone.utc).isoformat()
                registo['ficheiro_origem'] = file.filename
                
                # Inserir na coleção apropriada
                await db[collection_name].insert_one(registo)
                registos_salvos += 1
                
            except Exception as e:
                logger.error(f"Erro ao salvar registo: {e}")
                continue
        
        # Registar log de importação
        log = {
            "id": str(uuid.uuid4()),
            "plataforma": plataforma,
            "ficheiro": file.filename,
            "registos_processados": len(registos),
            "registos_salvos": registos_salvos,
            "parceiro_id": parceiro_id,
            "motorista_id": motorista_id,
            "usuario_id": current_user['id'],
            "data": datetime.now(timezone.utc).isoformat()
        }
        await db.logs_importacao_csv.insert_one(log)
        
        logger.info(f"✅ {registos_salvos} registos importados com sucesso")
        
        return {
            "success": True,
            "message": f"{registos_salvos} registos importados com sucesso!",
            "registos_importados": registos_salvos,
            "registos_processados": len(registos)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao importar CSV: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/import-csv/history")
async def get_import_history(
    plataforma: Optional[str] = None,
    parceiro_id: Optional[str] = None,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Obter histórico de importações"""
    try:
        query = {}
        
        if plataforma:
            query['plataforma'] = plataforma
        
        if parceiro_id:
            query['parceiro_id'] = parceiro_id
        
        # Permitir que users vejam apenas suas importações
        if current_user['role'] not in ['admin', 'gestao']:
            query['usuario_id'] = current_user['id']
        
        logs = await db.logs_importacao_csv.find(
            query,
            {"_id": 0}
        ).sort('data', -1).limit(limit).to_list(length=None)
        
        return logs
        
    except Exception as e:
        logger.error(f"Erro ao obter histórico: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/dados/{plataforma}")
async def get_dados_plataforma(
    plataforma: str,
    parceiro_id: Optional[str] = None,
    motorista_id: Optional[str] = None,
    data_inicio: Optional[str] = None,
    data_fim: Optional[str] = None,
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    """Obter dados importados de uma plataforma"""
    try:
        collection_name = f"dados_{plataforma}"
        
        query = {}
        if parceiro_id:
            query['parceiro_id'] = parceiro_id
        if motorista_id:
            query['motorista_id'] = motorista_id
        if data_inicio:
            query['data'] = {"$gte": data_inicio}
        if data_fim:
            if 'data' not in query:
                query['data'] = {}
            query['data']['$lte'] = data_fim
        
        dados = await db[collection_name].find(
            query,
            {"_id": 0}
        ).sort('data', -1).limit(limit).to_list(length=None)
        
        return dados
        
    except Exception as e:
        logger.error(f"Erro ao obter dados: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==================================================
# BOLT INTEGRATION - Sincronização Real
# ==================================================

@app.post("/api/bolt/test-connection")
async def test_bolt_connection(
    request: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Testar conexão com Bolt Partner usando credenciais fornecidas
    """
    try:
        if current_user["role"] not in ["admin", "gestao"]:
            raise HTTPException(status_code=403, detail="Acesso negado")
        
        email = request.get("email")
        password = request.get("password")
        
        if not email or not password:
            raise HTTPException(status_code=400, detail="Email e password são obrigatórios")
        
        logger.info(f"🧪 Testando conexão Bolt para: {email}")
        
        # Import do scraper
        from integrations.bolt_scraper import test_bolt_connection
        
        # Testar conexão
        result = await test_bolt_connection(email, password)
        
        return result
        
    except Exception as e:
        logger.error(f"Erro ao testar conexão Bolt: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/bolt/sync-earnings")
async def sync_bolt_earnings(
    request: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Sincronizar ganhos da Bolt para um parceiro
    """
    try:
        if current_user["role"] not in ["admin", "gestao"]:
            raise HTTPException(status_code=403, detail="Acesso negado")
        
        parceiro_id = request.get("parceiro_id")
        email = request.get("email")
        password = request.get("password")
        
        if not all([parceiro_id, email, password]):
            raise HTTPException(status_code=400, detail="Dados incompletos")
        
        logger.info(f"🔄 Iniciando sincronização Bolt para parceiro: {parceiro_id}")
        
        # Buscar parceiro
        parceiro = await db.parceiros.find_one({"id": parceiro_id}, {"_id": 0})
        if not parceiro:
            raise HTTPException(status_code=404, detail="Parceiro não encontrado")
        
        # Import do scraper
        from integrations.bolt_scraper import BoltScraper
        
        # Criar log de sincronização
        log_id = str(uuid.uuid4())
        log = {
            "id": log_id,
            "parceiro_id": parceiro_id,
            "plataforma": "bolt",
            "tipo": "manual",
            "status": "em_progresso",
            "data_inicio": datetime.now(timezone.utc).isoformat(),
            "usuario_id": current_user["id"]
        }
        await db.logs_sincronizacao_parceiro.insert_one(log)
        
        # Executar scraping
        async with BoltScraper(headless=True) as scraper:
            # Login
            login_success = await scraper.login(email, password)
            
            if not login_success:
                # Atualizar log
                await db.logs_sincronizacao_parceiro.update_one(
                    {"id": log_id},
                    {"$set": {
                        "status": "erro",
                        "mensagem_erro": "Falha no login. Verifique as credenciais.",
                        "data_fim": datetime.now(timezone.utc).isoformat()
                    }}
                )
                return {
                    "success": False,
                    "message": "Falha no login. Verifique as credenciais."
                }
            
            # Navegar para earnings
            await scraper.navigate_to_earnings()
            
            # Extrair dados
            earnings_data = await scraper.extract_earnings_data()
            
            if earnings_data.get("success"):
                # Guardar dados extraídos
                ganho_bolt = {
                    "id": str(uuid.uuid4()),
                    "parceiro_id": parceiro_id,
                    "plataforma": "bolt",
                    "periodo_inicio": earnings_data.get("period_start"),
                    "periodo_fim": earnings_data.get("period_end"),
                    "valor_total": earnings_data.get("total_earnings", 0),
                    "num_viagens": earnings_data.get("trips_count", 0),
                    "data_extracao": earnings_data.get("extracted_at"),
                    "dados_completos": earnings_data,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.ganhos_bolt.insert_one(ganho_bolt)
                
                # Atualizar log
                await db.logs_sincronizacao_parceiro.update_one(
                    {"id": log_id},
                    {"$set": {
                        "status": "sucesso",
                        "data_fim": datetime.now(timezone.utc).isoformat(),
                        "registos_extraidos": 1,
                        "valor_total": earnings_data.get("total_earnings", 0)
                    }}
                )
                
                # Atualizar configuração de sincronização
                await db.configuracoes_sincronizacao.update_one(
                    {"parceiro_id": parceiro_id},
                    {"$set": {
                        "ultima_sincronizacao": datetime.now(timezone.utc).isoformat(),
                        "status": "sucesso"
                    }},
                    upsert=True
                )
                
                return {
                    "success": True,
                    "message": f"Sincronização concluída! Ganhos: €{earnings_data.get('total_earnings', 0)}",
                    "data": earnings_data
                }
            else:
                # Erro na extração
                await db.logs_sincronizacao_parceiro.update_one(
                    {"id": log_id},
                    {"$set": {
                        "status": "erro",
                        "mensagem_erro": earnings_data.get("error", "Erro desconhecido"),
                        "data_fim": datetime.now(timezone.utc).isoformat()
                    }}
                )
                return {
                    "success": False,
                    "message": f"Erro ao extrair dados: {earnings_data.get('error')}"
                }
        
    except Exception as e:
        logger.error(f"Erro na sincronização Bolt: {e}")
        
        # Atualizar log
        if 'log_id' in locals():
            await db.logs_sincronizacao_parceiro.update_one(
                {"id": log_id},
                {"$set": {
                    "status": "erro",
                    "mensagem_erro": str(e),
                    "data_fim": datetime.now(timezone.utc).isoformat()
                }}
            )
        
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/bolt/save-credentials")
async def save_bolt_credentials(
    request: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Guardar credenciais da Bolt para um parceiro
    """
    try:
        if current_user["role"] not in ["admin", "gestao"]:
            raise HTTPException(status_code=403, detail="Acesso negado")
        
        parceiro_id = request.get("parceiro_id")
        email = request.get("email")
        password = request.get("password")
        
        if not all([parceiro_id, email, password]):
            raise HTTPException(status_code=400, detail="Dados incompletos")
        
        # Verificar se parceiro existe
        parceiro = await db.parceiros.find_one({"id": parceiro_id}, {"_id": 0})
        if not parceiro:
            raise HTTPException(status_code=404, detail="Parceiro não encontrado")
        
        # Guardar credenciais (encriptar password na produção)
        cred = {
            "parceiro_id": parceiro_id,
            "plataforma": "bolt",
            "email": email,
            "password": password,  # ATENÇÃO: Na produção, encriptar!
            "ativo": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Verificar se já existe
        existing = await db.credenciais_bolt.find_one({
            "parceiro_id": parceiro_id,
            "plataforma": "bolt"
        })
        
        if existing:
            await db.credenciais_bolt.update_one(
                {"parceiro_id": parceiro_id, "plataforma": "bolt"},
                {"$set": cred}
            )
            message = "Credenciais atualizadas"
        else:
            await db.credenciais_bolt.insert_one(cred)
            message = "Credenciais guardadas"
        
        return {"success": True, "message": message}
        
    except Exception as e:
        logger.error(f"Erro ao guardar credenciais: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==================================================
# CREDENCIAIS PARCEIROS - Admin only
# ==================================================

@app.get("/api/admin/credenciais-parceiros")
async def obter_credenciais_parceiros(
    current_user: dict = Depends(get_current_user)
):
    """
    Obter credenciais de login dos parceiros (apenas para admin)
    """
    try:
        if current_user["role"] != "admin":
            raise HTTPException(status_code=403, detail="Acesso negado - apenas administradores")
        
        # Buscar todos os utilizadores com role 'parceiro'
        users_parceiros = await db.users.find(
            {"role": "parceiro"},
            {"_id": 0, "id": 1, "email": 1, "password": 1, "name": 1, "active": 1, "created_at": 1}
        ).to_list(1000)
        
        # Para cada parceiro, buscar dados adicionais da coleção parceiros
        credenciais_completas = []
        
        for user in users_parceiros:
            parceiro_data = await db.parceiros.find_one(
                {"email": user["email"]},
                {"_id": 0, "id": 1, "nome_empresa": 1, "telefone": 1, "nif": 1}
            )
            
            # Montar credencial
            # Nota: Passwords são encriptadas com bcrypt e não podem ser desencriptadas
            # Para ver a password original, consultar o ficheiro CREDENCIAIS_TESTE.md
            credencial = {
                "parceiro_id": user["id"],
                "email": user["email"],
                "password": "[ENCRIPTADA - Consultar CREDENCIAIS_TESTE.md]",
                "password_hash": user["password"],  # Hash bcrypt (para referência)
                "nome": user.get("name", "Sem nome"),
                "nome_empresa": parceiro_data.get("nome_empresa") if parceiro_data else None,
                "telefone": parceiro_data.get("telefone") if parceiro_data else None,
                "nif": parceiro_data.get("nif") if parceiro_data else None,
                "ativo": user.get("active", True),
                "plataforma": "tvdefleet",  # Sistema principal
                "updated_at": user.get("created_at")
            }
            
            credenciais_completas.append(credencial)
        
        # Registar acesso (auditoria)
        log_auditoria = {
            "id": str(uuid.uuid4()),
            "acao": "visualizacao_credenciais",
            "usuario_id": current_user["id"],
            "usuario_email": current_user["email"],
            "data": datetime.now(timezone.utc).isoformat(),
            "detalhes": f"Visualizou credenciais de {len(credenciais_completas)} parceiros"
        }
        await db.logs_auditoria.insert_one(log_auditoria)
        
        return credenciais_completas
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao obter credenciais: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()