async def importar_carregamentos_excel(
    file_content: bytes,
    current_user: dict,
    periodo_inicio: str,
    periodo_fim: str
):
    """
    Importar carregamentos elétricos de Excel (.xlsx)
    - Formato: Excel (.xlsx) com colunas portuguesas
    - Identificador: Nº. CARTÃO (CardCode) → cartao_frota_eletric_id
    - Associa motorista via veículo (motorista_atribuido)
    - NÃO usa email do motorista
    """
    try:
        import openpyxl
        from io import BytesIO
        from datetime import datetime, timezone
        
        # Carregar workbook
        wb = openpyxl.load_workbook(BytesIO(file_content))
        sheet = wb.active
        
        sucesso = 0
        erros = 0
        erros_detalhes = []
        
        # Ler linha 1 como cabeçalho
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        header = [str(cell).strip() if cell else '' for cell in header_row]
        
        logger.info(f"📄 Cabeçalho Excel Carregamentos: {header}")
        
        # Mapear nomes de colunas (podem variar)
        col_map = {}
        for i, col in enumerate(header):
            col_lower = col.lower()
            if 'cartão' in col_lower or 'cartao' in col_lower:
                col_map['card_code'] = col
            elif 'matrícula' in col_lower or 'matricula' in col_lower:
                col_map['matricula'] = col
            elif 'data' in col_lower and 'descrição' not in col_lower:
                col_map['data'] = col
            elif 'duração' in col_lower or 'duracao' in col_lower:
                col_map['duracao'] = col
            elif 'posto' in col_lower and 'energia' in col_lower:
                col_map['posto'] = col
            elif 'total' in col_lower and 'iva' in col_lower:
                col_map['valor_total'] = col
            elif 'custo' in col_lower:
                col_map['custo'] = col
            elif 'energia' in col_lower and 'posto' not in col_lower:
                col_map['energia'] = col
        
        logger.info(f"🗺️ Mapeamento de colunas: {col_map}")
        
        # Processar linhas a partir da linha 2
        for row_num, row_values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Criar dicionário da linha
                row = dict(zip(header, row_values))
                
                # Extrair CardCode (Nº. CARTÃO)
                card_code = str(row.get(col_map.get('card_code', 'Nº. CARTÃO'), '')).strip() if row.get(col_map.get('card_code', 'Nº. CARTÃO')) else None
                
                if not card_code or card_code == 'None':
                    erros += 1
                    erros_detalhes.append(f"Linha {row_num}: Nº. CARTÃO não encontrado")
                    continue
                
                # BUSCAR VEÍCULO POR CardCode → cartao_frota_eletric_id
                vehicle = await db.vehicles.find_one(
                    {"cartao_frota_eletric_id": card_code},
                    {"_id": 0}
                )
                
                if not vehicle:
                    erros += 1
                    erros_detalhes.append(
                        f"Linha {row_num}: Veículo não encontrado com CardCode '{card_code}'. "
                        f"Preencher 'Cartão Frota Elétrico ID (Carregamentos)' no veículo."
                    )
                    continue
                
                logger.info(f"✅ Excel Carregamento - Veículo encontrado: {vehicle.get('matricula')} (CardCode: {card_code})")
                
                # Buscar motorista atribuído ao veículo
                motorista = None
                motorista_email = ""
                if vehicle.get('motorista_atribuido'):
                    motorista = await db.motoristas.find_one(
                        {"id": vehicle['motorista_atribuido']},
                        {"_id": 0}
                    )
                    if motorista:
                        motorista_email = motorista.get("email", "")
                        logger.info(f"✅ Excel Carregamento - Motorista: {motorista.get('name')}")
                
                # Processar data (pode ser serial do Excel)
                data_valor = row.get(col_map.get('data', 'DATA'))
                if isinstance(data_valor, (int, float)):
                    # Converter serial date do Excel para datetime
                    from datetime import datetime, timedelta
                    excel_epoch = datetime(1899, 12, 30)
                    data_dt = excel_epoch + timedelta(days=data_valor)
                    data = data_dt.strftime('%Y-%m-%d')
                    hora = data_dt.strftime('%H:%M:%S')
                elif isinstance(data_valor, datetime):
                    data = data_valor.strftime('%Y-%m-%d')
                    hora = data_valor.strftime('%H:%M:%S')
                else:
                    data = datetime.now(timezone.utc).strftime('%Y-%m-%d')
                    hora = '00:00:00'
                
                # Extrair outros campos
                duracao = float(row.get(col_map.get('duracao', 'DURAÇÃO'), 0) or 0)
                posto = str(row.get(col_map.get('posto', 'POSTO ENERGIA'), '') or '').strip()
                valor_total = float(row.get(col_map.get('valor_total', 'TOTAL c/ IVA'), 0) or 0)
                energia = float(row.get(col_map.get('energia', 'ENERGIA'), 0) or 0)
                
                # Se não tem energia mas tem custo, usar custo
                if energia == 0 and col_map.get('custo'):
                    energia = float(row.get(col_map.get('custo'), 0) or 0)
                
                # Criar documento
                documento = {
                    "id": str(uuid.uuid4()),
                    "vehicle_id": vehicle["id"],
                    "motorista_id": motorista.get("id") if motorista else None,
                    "motorista_email": motorista_email,
                    "matricula": vehicle.get("matricula"),
                    "card_code": card_code,
                    "estacao_id": posto,
                    "duracao_minutos": duracao,
                    "energia_kwh": energia,
                    "valor_total_com_taxas": valor_total,
                    "data": data,
                    "hora": hora,
                    "tipo_transacao": "carregamento_eletrico",
                    "plataforma": "viaverde",
                    "periodo_inicio": periodo_inicio,
                    "periodo_fim": periodo_fim,
                    "ano": int(periodo_inicio.split('-')[0]) if periodo_inicio else datetime.now(timezone.utc).year,
                    "semana": datetime.strptime(periodo_inicio, '%Y-%m-%d').isocalendar()[1] if periodo_inicio else datetime.now(timezone.utc).isocalendar()[1],
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "created_by": current_user["id"]
                }
                
                # Inserir no MongoDB
                await db.portagens_viaverde.insert_one(documento)
                sucesso += 1
                
            except Exception as e:
                erros += 1
                erros_detalhes.append(f"Linha {row_num}: {str(e)}")
                logger.error(f"❌ Erro linha {row_num}: {str(e)}")
                continue
        
        # Criar relatórios de rascunho
        info_rascunhos = None
        if sucesso > 0:
            try:
                info_rascunhos = await criar_relatorios_rascunho_apos_importacao(
                    'viaverde',
                    current_user["id"],
                    periodo_inicio,
                    periodo_fim,
                    db=db
                )
            except Exception as e:
                logger.error(f"Erro ao criar rascunhos: {str(e)}")
        
        return {
            "message": "Importação de carregamentos Excel concluída",
            "sucesso": sucesso,
            "erros": erros,
            "erros_detalhes": erros_detalhes[:20],
            "relatorios_criados": info_rascunhos
        }
        
    except Exception as e:
        logger.error(f"❌ Erro ao processar Excel de carregamentos: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Erro ao processar Excel: {str(e)}")